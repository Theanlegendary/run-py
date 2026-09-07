# -*- coding: utf-8 -*-
"""
generate_dual_august_tracking_reports.py
Generates 2 focused Tracking Reports with Full Standard Template (110–600) & VAS:
  1. Delivered Bills Report (410 - Giao thành công)
  2. Returned Bills Report (520 - Đã trả hàng & Return stages)
"""

import os
import re
import json
import requests
import pandas as pd
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
CONFIG_PATH = os.path.join(HERE, "config.json")

# Complete Full Standard Template of all TMS status codes 110-600 in operational order
FULL_STATUS_CODES_TEMPLATE = [
    # 1xx: Booking & Pickup assignment
    "110", "120", "130", "140", "150",
    # 2xx: Picked up & Received at Store/Agent
    "200", "201", "202", "205", "210", "220", "230", "240", "250",
    # 3xx: Sacks, Packing, Transit & Hub Handover
    "300", "302", "304", "306", "308", "309", "310", "311", "312", "315", "320", "330",
    # 4xx: Out for Delivery, Appointment, Redelivery, Forwarding
    "400", "401", "402", "403", "404", "405", "406", "408", "415", "417", "420", "421", "422", "425", "428",
    "430", "431", "432", "440", "450", "460", "470", "471", "472", "475", "480", "485", "490", "495",
    # 5xx: Exceptions, Problem parcel, Returns & Investigations
    "500", "501", "502", "505", "510", "512", "515", "520", "525", "530", "540", "550", "560", "570", "580", "590",
    # 6xx: Final Audit & Archival
    "600",
    # 410: Done Delivered
    "410",
    # 99: System Cancellation
    "99"
]

def fetch_and_build_report(df_subset, report_title, sheet_name, out_paths, max_workers=75):
    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        cfg = json.load(f)

    headers_api = {
        "Authorization": "Bearer " + cfg["api"]["bearer_token"],
        "Accept": "application/json, text/plain, */*",
        "User-Agent": "Mozilla/5.0"
    }

    service_map = {}
    svc_col = next((c for c in df_subset.columns if c in ["SERVICE", "SERVICE TYPE", "SERVICE_TYPE", "SERVICE NAME", "SERVICETYPE"]), None)
    if svc_col:
        for _, r in df_subset.iterrows():
            oid = str(r.get("ORDER ID", "")).strip()
            val = str(r.get(svc_col, "") or "").strip()
            if oid and val and val.lower() != "nan":
                service_map[oid] = val

    action_user_map = {}
    au_col = next((c for c in df_subset.columns if c in ["ACTION USER", "ACTION_USER", "LAST ACTION USER", "LAST USER", "USER"]), None)
    if au_col:
        for _, r in df_subset.iterrows():
            oid = str(r.get("ORDER ID", "")).strip()
            val = str(r.get(au_col, "") or "").strip()
            if oid and val and val.lower() != "nan":
                action_user_map[oid] = val

    vas_fee_map = {}
    vf_col = next((c for c in df_subset.columns if "VAS FEE" in c or "VAS" in c), None)
    if vf_col:
        for _, r in df_subset.iterrows():
            oid = str(r.get("ORDER ID", "")).strip()
            val = r.get(vf_col, 0)
            try:
                if float(val) > 0:
                    vas_fee_map[oid] = float(val)
            except Exception:
                pass

    def fetch_bill(oid):
        oid_str = str(oid).strip()
        if not oid_str or oid_str.lower() == "nan":
            return None

        try:
            r_tr = requests.get(
                "https://gw-express.metfone.com.kh/tms-tracking/api/v1/order-tracking",
                params={"order_id": oid_str},
                headers=headers_api,
                timeout=12
            )
            if r_tr.status_code != 200:
                return None
            data_tr = r_tr.json()
            trips = data_tr.get("trackingTrips", [])
            if not trips:
                return None

            vas_val = ""
            try:
                r_sc = requests.get(
                    "https://gw-express.metfone.com.kh/tms-receiving/api/v1/orders/search",
                    params={"order_code": oid_str},
                    headers=headers_api,
                    timeout=8
                )
                if r_sc.status_code == 200:
                    data_sc = r_sc.json()
                    vas_val = str(data_sc.get("added_service_code") or "").strip()
            except Exception:
                pass

            if not vas_val and oid_str in vas_fee_map:
                vas_val = "VTT"

            trips_sorted = list(reversed(trips))
            svc_type = service_map.get(oid_str, "")
            if not svc_type and isinstance(data_tr, dict):
                svc_type = str(data_tr.get("serviceType") or data_tr.get("serviceName") or data_tr.get("service") or "").strip()

            row_dict = {
                "BILL ID": oid_str,
                "SERVICE TYPE": svc_type,
                "VAS": vas_val,
                "_trips_map": {}
            }

            latest_st = ""
            latest_unit = ""
            latest_time = ""
            latest_user = ""

            for t in trips_sorted:
                st = str(t.get("status", "") or "").lstrip("S").strip()
                po = t.get("postOffice") or {}
                unit = t.get("postcode") or (po.get("code") if isinstance(po, dict) else "") or ""
                if not unit and "handoverInfo" in t:
                    unit = t.get("handoverInfo", {}).get("departmentCode", "")

                dt_raw = t.get("updatedAt", "")
                dt_str = ""
                if dt_raw:
                    try:
                        dt_obj = pd.to_datetime(dt_raw)
                        dt_str = dt_obj.strftime("%d/%m/%Y %H:%M:%S")
                    except Exception:
                        dt_str = str(dt_raw)

                upd_obj = t.get("updatedBy")
                upd_user = ""
                if isinstance(upd_obj, dict):
                    upd_user = str(upd_obj.get("name") or "").strip()
                elif isinstance(upd_obj, str):
                    upd_user = upd_obj.strip()

                if st:
                    row_dict["_trips_map"][st] = (unit, dt_str)

                latest_st = st
                latest_unit = unit
                latest_time = dt_str
                if upd_user:
                    latest_user = upd_user

            if not latest_user:
                latest_user = action_user_map.get(oid_str, "")

            row_dict["LATEST_STATUS"] = latest_st
            row_dict["LATEST_UNIT"] = latest_unit
            row_dict["LATEST_TIME"] = latest_time
            row_dict["LATEST_USER"] = latest_user
            return row_dict
        except Exception:
            return None

    unique_oids = df_subset["ORDER ID"].dropna().unique()
    print(f"\n[INFO] {report_title}: Processing {len(unique_oids)} bills with {max_workers} threads...")
    results = []

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {executor.submit(fetch_bill, oid): oid for oid in unique_oids}
        count = 0
        total = len(unique_oids)
        for future in as_completed(futures):
            res = future.result()
            count += 1
            if count % 2000 == 0 or count == total:
                print(f"[PROGRESS] {count}/{total} bills processed ({len(results)} valid)...")
            if res:
                results.append(res)

    if not results:
        print(f"[WARN] No tracking results for {report_title}")
        return

    print(f"[INFO] Formatting Excel ({len(FULL_STATUS_CODES_TEMPLATE)} status codes)...")
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    font_family = "Calibri"
    f_title = Font(name=font_family, size=14, bold=True, color="FFFFFF")
    f_header = Font(name=font_family, size=10, bold=True, color="FFFFFF")
    f_data = Font(name=font_family, size=9)
    f_log = Font(name=font_family, size=9, bold=True, color="1E3A8A")
    f_vas = Font(name=font_family, size=9, bold=True, color="047857")

    fill_title = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    fill_hdr1  = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    fill_hdr2  = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
    fill_alt   = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

    thin = Side(border_style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["BILL ID", "SERVICE TYPE", "VAS"]
    for sc in FULL_STATUS_CODES_TEMPLATE:
        headers.extend([f"LOG {sc}", f"UNIT {sc}", f"TIME {sc}"])
    headers.extend(["LATEST STATUS", "LATEST UNIT", "LATEST TIME", "LAST USER"])

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    t_cell = ws.cell(1, 1, f"{report_title} ({len(results)} Bills)")
    t_cell.font = f_title
    t_cell.fill = fill_title
    t_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.row_dimensions[3].height = 26
    for c_idx, h_text in enumerate(headers, 1):
        cell = ws.cell(3, c_idx, h_text)
        cell.font = f_header
        cell.fill = fill_hdr1 if c_idx in (1, 2, 3) or ((c_idx - 4) // 3) % 2 == 0 else fill_hdr2
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    results.sort(key=lambda x: str(x.get("BILL ID", "")))

    r_idx = 4
    for r_data in results:
        row_fill = fill_alt if r_idx % 2 == 0 else None
        ws.row_dimensions[r_idx].height = 20

        c_bill = ws.cell(r_idx, 1, str(r_data.get("BILL ID", "")))
        c_bill.font = f_data
        c_bill.border = border
        c_bill.alignment = Alignment(horizontal="left", vertical="center")

        c_svc = ws.cell(r_idx, 2, str(r_data.get("SERVICE TYPE", "")))
        c_svc.font = f_data
        c_svc.border = border
        if row_fill: c_svc.fill = row_fill
        c_svc.alignment = Alignment(horizontal="center", vertical="center")

        c_vas = ws.cell(r_idx, 3, str(r_data.get("VAS", "")))
        c_vas.font = f_vas if r_data.get("VAS") else f_data
        c_vas.border = border
        if row_fill: c_vas.fill = row_fill
        c_vas.alignment = Alignment(horizontal="center", vertical="center")

        trips_map = r_data.get("_trips_map", {})
        col_pos = 4
        for sc in FULL_STATUS_CODES_TEMPLATE:
            unit_val, time_val = trips_map.get(sc, ("", ""))
            log_val = sc if unit_val or time_val else ""

            c_log = ws.cell(r_idx, col_pos, log_val)
            c_log.font = f_log
            c_log.border = border
            if row_fill: c_log.fill = row_fill
            c_log.alignment = Alignment(horizontal="center", vertical="center")
            col_pos += 1

            c_unit = ws.cell(r_idx, col_pos, unit_val)
            c_unit.font = f_data
            c_unit.border = border
            if row_fill: c_unit.fill = row_fill
            c_unit.alignment = Alignment(horizontal="center", vertical="center")
            col_pos += 1

            c_time = ws.cell(r_idx, col_pos, time_val)
            c_time.font = f_data
            c_time.border = border
            if row_fill: c_time.fill = row_fill
            c_time.alignment = Alignment(horizontal="center", vertical="center")
            col_pos += 1

        for val in [r_data.get("LATEST_STATUS", ""), r_data.get("LATEST_UNIT", ""), r_data.get("LATEST_TIME", ""), r_data.get("LATEST_USER", "")]:
            c_last = ws.cell(r_idx, col_pos, val)
            c_last.font = f_data
            c_last.border = border
            if row_fill: c_last.fill = row_fill
            c_last.alignment = Alignment(horizontal="center", vertical="center")
            col_pos += 1

        r_idx += 1

    ws.freeze_panes = "D4"
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col[:80])
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 11)

    for p in out_paths:
        wb.save(p)
        print(f"[SUCCESS] Saved {report_title} to: {p}")

def main():
    detail_path = os.path.join(HERE, "latest_august_detail.xlsx")
    print(f"[INFO] Reading {detail_path}...")
    df = pd.read_excel(detail_path)
    df.columns = [str(c).strip().upper() for c in df.columns]

    st_col = next((c for c in df.columns if "CURRENT STATUS" in c or "STATUS" in c), None)

    # 1. Delivered Bills (410)
    df_del = df[df[st_col].astype(str).str.contains("410", na=False)].copy() if st_col else df
    deliv_paths = [
        r"C:\Users\DELL\Desktop\Bill_Tracking_Status_Logs_Done_Delivered_01Aug_31Aug_20260904.xlsx",
        os.path.join(HERE, "Bill_Tracking_Status_Logs_01Aug_31Aug_20260904_1050.xlsx")
    ]
    fetch_and_build_report(df_del, "DELIVERED BILL TRACKING STATUS LOGS REPORT — AUGUST 2026", "Delivered Tracking Logs", deliv_paths, max_workers=75)

    # 2. Returned Bills (520 / 5xx)
    df_ret = df[df[st_col].astype(str).str.contains("520|500|501|502|510|512|515|530", na=False)].copy() if st_col else pd.DataFrame()
    ret_paths = [
        r"C:\Users\DELL\Desktop\Bill_Tracking_Status_Logs_Returned_520_01Aug_31Aug_20260904.xlsx",
        os.path.join(HERE, "Bill_Tracking_Status_Logs_Returned_520_August.xlsx")
    ]
    fetch_and_build_report(df_ret, "RETURNED (520) BILL TRACKING STATUS LOGS REPORT — AUGUST 2026", "Returned 520 Tracking Logs", ret_paths, max_workers=75)

    print("\n[ALL DONE] Both Delivered and Returned 520 Tracking Reports generated successfully!")

if __name__ == "__main__":
    main()
