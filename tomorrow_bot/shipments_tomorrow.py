import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime

def build_shipments_tomorrow_report(src_xlsx, out_xlsx, target_label="Zone 1", start_time="00:00", end_time="06:00"):
    """
    Builds CEO-Level Executive SHIPMENTS TOMORROW REPORT Excel file:
      - Sheet 1: SHIPMENTS TOMORROW REPORT (Left main table + Right Executive Summary table)
      - Sheet 2: base (Raw order dataset)
    Filters orders created on current date between start_time (00:00) and end_time (06:00).
    """
    import pandas as pd
    wb_src = openpyxl.load_workbook(src_xlsx, data_only=True)
    ws_src = wb_src.active

    headers = [str(ws_src.cell(1, c).value or "").strip() for c in range(1, ws_src.max_column + 1)]
    col_map = {h.upper(): idx for idx, h in enumerate(headers, 1)}

    ci_order   = col_map.get("ORDER ID", col_map.get("ORDER_NUMBER", 3))
    ci_origin_br = col_map.get("ACTION POST OFFICE", col_map.get("ORIGIN_BRANCH", 36))
    ci_origin_po = col_map.get("CURRENT POST OFFICE", col_map.get("ORIGIN_POST", 16))
    ci_dest_prov = col_map.get("DELIVERY PROVINCE", col_map.get("DESTINATION_BRANCH", 13))
    ci_dest_po   = col_map.get("DELIVERY POST OFFICE", col_map.get("DESTINATION_POST", 14))
    ci_created   = col_map.get("CREATED DATE", col_map.get("CREATED_AT", 2))
    ci_fee       = col_map.get("TOTAL FEE (USD)", col_map.get("TOTAL_AMOUNT (USD)", 20))
    ci_vas_fee   = col_map.get("VAS FEE (USD) (2)", 18)
    ci_cod       = col_map.get("COD (USD)", 21)
    ci_weight    = col_map.get("WEIGHT (G)", col_map.get("ACTUAL_WEIGHT (G)", 8))
    ci_status    = col_map.get("CURRENT STATUS", col_map.get("STATUES", 24))
    ci_receiver  = col_map.get("RECEIVER", 5)
    ci_service   = col_map.get("SERVICE", 6)
    ci_pay_meth  = col_map.get("PAYMENT METHOD", 22)

    base_rows = []
    r_idx = 1
    for r in range(2, ws_src.max_row + 1):
        order_id = str(ws_src.cell(r, ci_order).value or "").strip()
        if not order_id or order_id == "None":
            continue

        dest_prov = str(ws_src.cell(r, ci_dest_prov).value or "").strip().upper()
        dest_po   = str(ws_src.cell(r, ci_dest_po).value or "").strip().upper()
        orig_br   = str(ws_src.cell(r, ci_origin_br).value or "").strip().upper()
        orig_po   = str(ws_src.cell(r, ci_origin_po).value or "").strip().upper()
        created   = str(ws_src.cell(r, ci_created).value or "").strip()
        status    = str(ws_src.cell(r, ci_status).value or "").strip()
        receiver  = str(ws_src.cell(r, ci_receiver).value or "").strip()
        service   = str(ws_src.cell(r, ci_service).value or "").strip().upper()
        pay_meth  = str(ws_src.cell(r, ci_pay_meth).value or "").strip()

        try:
            weight = float(ws_src.cell(r, ci_weight).value or 0)
        except (ValueError, TypeError):
            weight = 0.0

        try:
            fee = float(ws_src.cell(r, ci_fee).value or 0)
        except (ValueError, TypeError):
            fee = 0.0

        try:
            vas_fee = float(ws_src.cell(r, ci_vas_fee).value or 0)
        except (ValueError, TypeError):
            vas_fee = 0.0

        try:
            cod = float(ws_src.cell(r, ci_cod).value or 0)
        except (ValueError, TypeError):
            cod = 0.0

        # Date & Time Filter: Today between start_time (00:00) and end_time (06:00)
        today_str = datetime.now().strftime("%Y%m%d")
        val_created = ws_src.cell(r, ci_created).value
        row_date_str = ""
        row_dt = None

        if isinstance(val_created, datetime):
            row_date_str = val_created.strftime("%Y%m%d")
            row_dt = val_created
        elif val_created:
            s = str(val_created).strip()
            try:
                parsed_dt = pd.to_datetime(s, dayfirst=True, format='mixed', errors='coerce')
                if pd.notna(parsed_dt):
                    row_date_str = parsed_dt.strftime("%Y%m%d")
                    row_dt = parsed_dt.to_pydatetime()
            except Exception:
                pass

        if row_date_str and row_date_str != today_str:
            continue

        # Time interval check (00:00 to 06:00)
        if row_dt is not None and start_time and end_time:
            try:
                st_h, st_m = [int(x) for x in start_time.split(":")]
                et_h, et_m = [int(x) for x in end_time.split(":")]
                row_minutes = row_dt.hour * 60 + row_dt.minute
                start_minutes = st_h * 60 + st_m
                end_minutes = et_h * 60 + et_m
                if not (start_minutes <= row_minutes <= end_minutes):
                    continue
            except Exception:
                pass

        # Filter active transit orders (Status 306, 309, 302, 310, 311)
        sc = status.split(" - ")[0].split()[0] if status else ""
        if sc not in ("306", "309", "302", "310", "311"):
            continue

        # Zone, Branch, or Specific Post Office Target Filtering
        tgt = target_label.upper().replace(" ", "")
        zone_by_prefix = {
            "KAN": "ZONE1", "PNP": "ZONE1", "PRE": "ZONE1", "SVA": "ZONE1",
            "KAM": "ZONE2", "KOH": "ZONE2", "SIH": "ZONE2", "SPE": "ZONE2", "TAK": "ZONE2",
            "BAN": "ZONE3", "BAT": "ZONE3", "CHH": "ZONE3", "PUR": "ZONE3",
            "ODD": "ZONE4", "PRH": "ZONE4", "SIE": "ZONE4", "THO": "ZONE4",
            "CHA": "ZONE5", "KRA": "ZONE5", "TBK": "ZONE5", "ROT": "ZONE5", "MON": "ZONE5", "STU": "ZONE5"
        }

        if tgt.startswith("ZONE"):
            target_zone_name = tgt if len(tgt) > 4 else "ZONE1" # e.g. ZONE1
            item_zone = zone_by_prefix.get(dest_prov, "ZONE1")
            if item_zone != target_zone_name:
                continue

        elif len(tgt) >= 7 or (len(tgt) > 3 and tgt[3:4] in ("P", "A", "S")):
            # Specific Post Office handle (e.g. PNPP014, PNPP010, SVAP001) -> MUST match dest_po exactly!
            if dest_po != tgt:
                continue

        elif tgt not in ("ALL", "TOTAL", "MEGA"):
            # 3-Letter Branch code (e.g. PNP, PRE, SVA, BAT)
            branch_prefix = tgt[:3]
            if dest_prov != branch_prefix and not dest_po.startswith(branch_prefix):
                continue


        # Determine VAS Code & VAS Khmer Description
        vas_codes = []
        vas_khmer_list = []

        if "NGƯỜI GỬI" in pay_meth.upper() or "SENDER" in pay_meth.upper():
            vas_codes.append("NTN")
            vas_khmer_list.append("អ្នកផ្ញើបង់")
        if vas_fee > 0:
            vas_codes.append("VBH")
            vas_khmer_list.append("ធានារ៉ាប់រង")
        if cod > 0:
            vas_codes.append("VBP")
            vas_khmer_list.append("ទារប្រាក់")
        if not vas_codes or service in ("CCN", "CLT"):
            vas_codes.append("VTT")
            vas_khmer_list.append("ដឹកដល់ផ្ទះ")

        vas_code_str = ", ".join(vas_codes)
        vas_khmer_str = ", ".join(vas_khmer_list)

        district_map = {
            # Phnom Penh (PNP)
            'PNPA002': 'Boeng Keng Kang', 'PNPP001': 'Boeng Keng Kang', 'PNPP007': 'Boeng Keng Kang', 'PNPS007': 'Boeng Keng Kang',
            'PNPP005': 'Chbar Ampov', 'PNPP010': 'Chraoy Chongvar', 'PNPP011': 'Dangkao', 'PNPP014': 'Doun Penh',
            'PNPA016': 'Kamboul', 'PNPP012': 'Kamboul', 'PNPA029': 'Mean Chey', 'PNPA055': 'Mean Chey', 'PNPP002': 'Mean Chey', 'PNPP003': 'Mean Chey',
            'PNPA028': 'Pou Saen Chey', 'PNPP008': 'Pou Saen Chey', 'PNPP009': 'Pou Saen Chey', 'PNPP004': 'Preaek Pnov',
            'PNPA040': 'Saen Sokh', 'PNPP013': 'Saen Sokh', 'PNPA036': 'Tuol Kouk', 'PNPP006': 'Tuol Kouk',

            # Kandal (KAN)
            'KANA024': 'Kandal Stueng', 'KANA028': 'Kandal Stueng', 'KANA049': 'Kandal Stueng',
            'KANS003': 'Kaoh Thum', 'KANA031': 'Khsach Kandal', 'KANA012': 'Kien Svay', 'KANA013': 'Kien Svay',
            'KANA008': 'Leuk Daek', 'KANS004': 'Mukh Kampul', 'KANA019': 'Ponhea Lueu',
            'KANA007': "S'ang", 'KANA020': "S'ang", 'KANA026': "S'ang", 'KANA040': 'Sampov Pun',
            'KANA023': 'Ta Khmau', 'KANP001': 'Ta Khmau',

            # Prey Veng (PRE)
            'PREA024': 'Ba Phnum', 'PREA023': 'Peam Ro', 'PREA020': 'Preah Sdach', 'PREA029': 'Preah Sdach', 'PREA035': 'Preah Sdach',
            'PREP001': 'Prey Veng', 'PRES001': 'Prey Veng', 'PREA002': 'Pur Rieng', 'PREA028': 'Sithor Kandal',

            # Svay Rieng (SVA)
            'SVAP001': 'Svay Rieng', 'SVAS002': 'Bavet', 'SVAA001': 'Bavet', 'SVAA002': 'Bavet', 'SVAA003': 'Romeas Haek', 'SVAA004': 'Rumduol', 'SVAA005': 'Svay Chrum',

            # Battambang (BAT)
            'BATA003': 'Banan', 'BATA016': 'Banan',
            'BATA001': 'Battambang', 'BATA009': 'Battambang', 'BATA040': 'Battambang', 'BATA042': 'Battambang', 'BATP001': 'Battambang',
            'BATA011': 'Kamrieng', 'BATS007': 'Moung Ruessei', 'BATA017': 'Samlout',
            'BATA010': 'Sampov Lun', 'BATA028': 'Sampov Lun', 'BATA004': 'Sangkae', 'BATA008': 'Sangkae',
            'BATA023': 'Thma Koul', 'BATA025': 'Thma Koul',

            # Siem Reap (SIE)
            'SIEP001': 'Siem Reap', 'SIEA001': 'Angkor Chum', 'SIEA002': 'Angkor Thon', 'SIEA003': 'Banteay Srei', 'SIEA004': 'Chi Kraeng',

            # Sihanoukville (SIH)
            'SIHP001': 'Sihanoukville', 'SIHA001': 'Preah Sihanouk', 'SIHA002': 'Stung Hav', 'SIHA003': 'Kampong Seila'
        }
        if dest_po in district_map:
            dist_name = district_map[dest_po]
        else:
            prov_map = {
                'PNP': 'Phnom Penh',
                'KAN': 'Kandal',
                'PRE': 'Prey Veng',
                'SVA': 'Svay Rieng',
                'BAT': 'Battambang',
                'SIE': 'Siem Reap',
                'SIH': 'Sihanoukville',
                'KOH': 'Koh Kong',
                'KAM': 'Kampot',
                'TAK': 'Takeo',
                'PUR': 'Pursat',
                'PRH': 'Preah Vihear',
                'TBK': 'Tbong Khmum',
                'THO': 'Kampong Thom',
                'CHA': 'Kampong Cham',
                'KRA': 'Kratie',
                'BAN': 'Banteay Meanchey',
                'CHH': 'Kampong Chhnang',
                'MON': 'Mondulkiri',
                'ROT': 'Ratanakiri',
                'STU': 'Stung Treng',
                'ODD': 'Oddar Meanchey',
                'KEP': 'Kep',
                'PAI': 'Pailin',
            }
            br_prefix = dest_prov.upper()[:3] if dest_prov else (dest_po[:3] if dest_po else '')
            dist_name = prov_map.get(br_prefix, dest_prov if dest_prov and dest_prov.upper() != 'NONE' else 'General District')




        base_rows.append({
            "no": r_idx,
            "order_number": order_id,
            "customer": receiver[:30],
            "origin_branch": orig_br,
            "origin_post": orig_po,
            "destination_branch": dest_prov,
            "destination_post": dest_po,
            "created_at": created,
            "fee": fee,
            "cod": cod,
            "weight_g": weight,
            "status": status,
            "receiver": receiver,
            "vas_code": vas_code_str,
            "vas_khmer": vas_khmer_str,
            "district": dist_name,
            "zone": "Zone 3" if dest_prov in ("BAT", "SIE", "PUR") else ("Zone 5" if dest_prov in ("SIH", "KOH", "TAK") else "Zone 1")
        })
        r_idx += 1

    wb = openpyxl.Workbook()

    # Sheet 1: SHIPMENTS TOMORROW REPORT
    ws1 = wb.active
    ws1.title = "SHIPMENTS TOMORROW REPORT"
    ws1.views.sheetView[0].showGridLines = True

    # Executive CEO Color Palette (Subtle, High-Contrast, Professional)
    fill_title_left  = PatternFill("solid", fgColor="0F172A") # Deep Slate Navy
    fill_title_right = PatternFill("solid", fgColor="0F766E") # Deep Teal Slate
    fill_hdr_left    = PatternFill("solid", fgColor="1E293B") # Executive Navy Slate
    fill_hdr_right   = PatternFill("solid", fgColor="0F766E") # Deep Teal Slate
    fill_row_alt     = PatternFill("solid", fgColor="F8FAFC") # Subtle Zebra Tint
    fill_left_tot    = PatternFill("solid", fgColor="CBD5E1") # Refined Slate Grey Total
    fill_sum_tot     = PatternFill("solid", fgColor="CCFBF1") # Refined Soft Teal Total

    border_clean = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0")
    )

    # Double-line Accounting Bottom Border for CEO Grand Total
    tot_border_accounting = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="64748B"),
        bottom=Side(style="double", color="0F172A") # Executive Double Line
    )

    font_banner = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_hdr    = Font(name="Segoe UI", size=9,  bold=True, color="FFFFFF")
    font_data   = Font(name="Segoe UI", size=9,  color="0F172A")
    font_data_b = Font(name="Segoe UI", size=9,  bold=True, color="0F172A")
    font_tot    = Font(name="Segoe UI", size=10, bold=True, color="0F172A")
    font_tot_red= Font(name="Segoe UI", size=10, bold=True, color="991B1B")

    # Row 1: Title Banners (Height 36)
    stamp_date = datetime.now().strftime("%d.%m")
    target_clean = target_label.upper()
    title_left_txt = f"SHIPMENTS TOMORROW REPORT {stamp_date} (Báo cáo hàng đến {target_clean})"
    title_right_txt= f"EXECUTIVE SUMMARY ({target_clean} / {target_clean[:3]})"

    ws1.merge_cells("A1:H1")
    ws1.cell(1, 1, title_left_txt).font = font_banner
    ws1.cell(1, 1).alignment = Alignment(horizontal="left", vertical="center")
    for c in range(1, 9):
        ws1.cell(1, c).fill = fill_title_left

    ws1.merge_cells("J1:N1")
    ws1.cell(1, 10, title_right_txt).font = font_banner
    ws1.cell(1, 10).alignment = Alignment(horizontal="center", vertical="center")
    for c in range(10, 15):
        ws1.cell(1, c).fill = fill_title_right

    ws1.row_dimensions[1].height = 36.0

    # Row 2: Header Rows (Height 32)
    headers_left = [
        "DESTINATION\n(សាខា)",
        "District\n(ស្រុក/ខណ្ឌ)",
        "DESTINATION_POS\n(បូស្តិ៍គោលដៅ)",
        "ORDER_NUMBER\n(លេខវិក្កយបត្រ)",
        "Receiver\n(អ្នកទទួល)",
        "SUM ACTUAL_WEIGHT (G)\n(ទម្ងន់សរុប g)",
        "VAS\n(សេវា)",
        "VAS Description\n(ឈ្មោះសេវាបន្ថែម)"
    ]
    headers_right = [
        "ZONE\n(តំបន់)",
        "DESTINATION_BRANCH\n(សាខា)",
        "District\n(ស្រុក/ខណ្ឌ)",
        "Bill\n(ចំនួនប័ណ្ណ)",
        "SUM ACTUAL_WEIGHT (G)\n(ទម្ងន់សរុប g)"
    ]

    ws1.row_dimensions[2].height = 32.0
    for ci, h in enumerate(headers_left, 1):
        cell = ws1.cell(2, ci, h)
        cell.font = font_hdr
        cell.fill = fill_hdr_left
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_clean

    for ci, h in enumerate(headers_right, 10):
        cell = ws1.cell(2, ci, h)
        cell.font = font_hdr
        cell.fill = fill_hdr_right
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_clean

    # Populate Left Data Rows (Single pure white background for all data rows)
    summary_data = {}
    total_bills = 0
    total_weight = 0.0

    r_curr = 3
    for idx_row, item in enumerate(base_rows):
        ws1.row_dimensions[r_curr].height = 20.0

        vals = [
            item["destination_branch"],
            item["district"],
            item["destination_post"],
            item["order_number"],
            item["receiver"],
            item["weight_g"],
            item["vas_code"],
            item["vas_khmer"]
        ]
        for ci, val in enumerate(vals, 1):
            cell = ws1.cell(r_curr, ci, val)
            cell.font = font_data_b if ci == 1 else font_data
            cell.border = border_clean


            if ci in (1, 2, 3, 4):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif ci in (5, 7, 8):
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif ci == 6:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = "#,##0"

        key = (item["zone"], item["destination_branch"], item["district"])
        summary_data.setdefault(key, {"bills": 0, "weight": 0.0})
        summary_data[key]["bills"] += 1
        summary_data[key]["weight"] += item["weight_g"]

        total_bills += 1
        total_weight += item["weight_g"]
        r_curr += 1

    # Left Grand Total Row (Refined CEO Double-Line Accounting Finish)
    ws1.row_dimensions[r_curr].height = 25.0
    ws1.merge_cells(start_row=r_curr, start_column=1, end_row=r_curr, end_column=5)
    gt_left = ws1.cell(r_curr, 1, "Grand Total / សរុប")
    gt_left.font = font_tot
    gt_left.alignment = Alignment(horizontal="left", vertical="center")
    for c in range(1, 6):
        cell = ws1.cell(r_curr, c)
        cell.fill = fill_left_tot
        cell.border = tot_border_accounting

    gt_w_cell = ws1.cell(r_curr, 6, total_weight)
    gt_w_cell.font = font_tot_red
    gt_w_cell.fill = fill_left_tot
    gt_w_cell.border = tot_border_accounting
    gt_w_cell.alignment = Alignment(horizontal="right", vertical="center")
    gt_w_cell.number_format = "#,##0"

    for c in (7, 8):
        cell = ws1.cell(r_curr, c)
        cell.fill = fill_left_tot
        cell.border = tot_border_accounting

    # Populate Executive Summary Table on Right with Province/Branch Subtotals
    r_sum = 3
    branch_groups = {}
    for (zone_str, br, dist), stats in sorted(summary_data.items()):
        if br not in branch_groups:
            branch_groups[br] = []
        branch_groups[br].append((zone_str, br, dist, stats))

    sub_fill = PatternFill("solid", fgColor="E0F2FE")
    sub_border = Border(
        top=Side(style="thin", color="94A3B8"),
        bottom=Side(style="thin", color="94A3B8"),
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0")
    )

    for br in sorted(branch_groups.keys()):
        br_items = branch_groups[br]
        br_bills = 0
        br_weight = 0

        for zone_str, b_code, dist, stats in br_items:
            ws1.row_dimensions[r_sum].height = 20.0
            s_vals = [zone_str, b_code, dist, stats["bills"], stats["weight"]]
            for ci, val in enumerate(s_vals, 10):
                cell = ws1.cell(r_sum, ci, val)
                cell.font = font_data
                cell.border = border_clean
                if ci in (10, 11, 12):
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif ci in (13, 14):
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    if ci == 14:
                        cell.number_format = "#,##0"
            br_bills += stats["bills"]
            br_weight += stats["weight"]
            r_sum += 1

        # Branch Subtotal Row (e.g. KAN Total, PNP Total, PRE Total, SVA Total)
        ws1.row_dimensions[r_sum].height = 22.0
        ws1.merge_cells(start_row=r_sum, start_column=10, end_row=r_sum, end_column=12)
        sub_lbl = ws1.cell(r_sum, 10, f"{br} Total")
        sub_lbl.font = Font(name="Segoe UI", size=10, bold=True, color="0F172A")
        sub_lbl.alignment = Alignment(horizontal="left", vertical="center")

        for c in range(10, 13):
            cell = ws1.cell(r_sum, c)
            cell.fill = sub_fill
            cell.border = sub_border

        sub_b_cell = ws1.cell(r_sum, 13, br_bills)
        sub_b_cell.font = Font(name="Segoe UI", size=10, bold=True, color="0F172A")
        sub_b_cell.fill = sub_fill
        sub_b_cell.border = sub_border
        sub_b_cell.alignment = Alignment(horizontal="right", vertical="center")
        sub_b_cell.number_format = "#,##0"

        sub_w_cell = ws1.cell(r_sum, 14, br_weight)
        sub_w_cell.font = Font(name="Segoe UI", size=10, bold=True, color="991B1B")
        sub_w_cell.fill = sub_fill
        sub_w_cell.border = sub_border
        sub_w_cell.alignment = Alignment(horizontal="right", vertical="center")
        sub_w_cell.number_format = "#,##0"
        r_sum += 1

    # Right Summary Total Row (CEO Double-Line Accounting Finish)
    ws1.row_dimensions[r_sum].height = 25.0
    ws1.merge_cells(start_row=r_sum, start_column=10, end_row=r_sum, end_column=12)
    tot_label_cell = ws1.cell(r_sum, 10, f"{target_clean[:3]} Total")
    tot_label_cell.font = font_tot
    tot_label_cell.alignment = Alignment(horizontal="left", vertical="center")
    for c in range(10, 13):
        cell = ws1.cell(r_sum, c)
        cell.fill = fill_sum_tot
        cell.border = tot_border_accounting

    tot_b_cell = ws1.cell(r_sum, 13, total_bills)
    tot_b_cell.font = font_tot
    tot_b_cell.fill = fill_sum_tot
    tot_b_cell.border = tot_border_accounting
    tot_b_cell.alignment = Alignment(horizontal="right", vertical="center")

    tot_w_cell = ws1.cell(r_sum, 14, total_weight)
    tot_w_cell.font = font_tot_red
    tot_w_cell.fill = fill_sum_tot
    tot_w_cell.border = tot_border_accounting
    tot_w_cell.alignment = Alignment(horizontal="right", vertical="center")
    tot_w_cell.number_format = "#,##0"

    # Exact Column Widths matching example file
    exact_widths = {
        'A': 16.0, 'B': 18.0, 'C': 20.0, 'D': 18.0, 'E': 35.0, 'F': 24.0,
        'G': 14.0, 'H': 26.0, 'I': 4.0,  'J': 12.0, 'K': 22.0, 'L': 18.0,
        'M': 12.0, 'N': 24.0
    }
    for col_let, w in exact_widths.items():
        ws1.column_dimensions[col_let].width = w

    # Sheet 2: base (Raw order dataset)
    ws2 = wb.create_sheet(title="base")
    ws2.views.sheetView[0].showGridLines = True

    base_headers = [
        "No", "ORDER_NUMBER", "CUSTOMER", "ORIGIN_BRANCH", "ORIGIN_POST",
        "DESTINATION_BRANCH", "DESTINATION_POST", "CREATED_BY", "CREATED_AT",
        "PAYMENT_METHOD", "SHIPPING_FEE", "DISCOUNT", "SERVICE_FEE",
        "TOTAL_FEE", "COD", "WEIGHT_G", "LENGTH", "WIDTH",
        "VAS_CODE", "VAS_DESCRIPTION", "DESTINATION_PROVINCE", "DESTINATION_DISTRICT",
        "STATUS", "RECEIVER_NAME"
    ]
    ws2.append(base_headers)
    ws2.row_dimensions[1].height = 24.0
    for c in range(1, len(base_headers) + 1):
        cell = ws2.cell(1, c)
        cell.font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0F172A")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for item in base_rows:
        row_data = [
            item["no"],
            item["order_number"],
            item["customer"],
            item["origin_branch"],
            item["origin_post"],
            item["destination_branch"],
            item["destination_post"],
            item["customer"],
            item["created_at"],
            "Sender",
            item["fee"],
            0.0,
            0.0,
            item["fee"],
            item["cod"],
            item["weight_g"],
            "",
            "",
            item["vas_code"],
            item["vas_khmer"],
            item.get("province", item.get("destination_branch", "")),
            item["district"],
            item["status"],
            item["receiver"]
        ]
        ws2.append(row_data)


    wb.save(out_xlsx)
    return total_bills, total_weight


def render_executive_summary_image(out_xlsx):
    """Renders ONLY the small right Executive Summary table to a pixel-perfect PNG image."""
    import tempfile, copy, openpyxl, excel_to_image
    from openpyxl.utils import get_column_letter

    wb = openpyxl.load_workbook(out_xlsx)
    ws = wb['SHIPMENTS TOMORROW REPORT']

    # Create 1-table Executive Summary workbook
    wb_sum = openpyxl.Workbook()
    ws_sum = wb_sum.active
    ws_sum.title = 'Executive Summary'
    ws_sum.views.sheetView[0].showGridLines = True

    # Find max row in summary table (Col J=10, Col M=13)
    max_r = 1
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, 10).value is not None or ws.cell(r, 13).value is not None:
            max_r = r

    # Copy Cols J to N (10 to 14) into Cols A to E (1 to 5)
    for r in range(1, max_r + 1):
        if ws.row_dimensions[r].height:
            ws_sum.row_dimensions[r].height = ws.row_dimensions[r].height
        for c_idx in range(5):
            orig_c = 10 + c_idx
            target_c = 1 + c_idx
            cell_orig = ws.cell(r, orig_c)
            cell_tgt  = ws_sum.cell(r, target_c, cell_orig.value)
            
            if cell_orig.has_style:
                cell_tgt.font = copy.copy(cell_orig.font)
                cell_tgt.fill = copy.copy(cell_orig.fill)
                cell_tgt.border = copy.copy(cell_orig.border)
                cell_tgt.alignment = copy.copy(cell_orig.alignment)
                cell_tgt.number_format = cell_orig.number_format

    # Copy all merged ranges for Cols J..N (10..14) -> A..E (1..5)
    for m_range in ws.merged_cells.ranges:
        if m_range.min_col >= 10 and m_range.max_col <= 14:
            new_min_c = m_range.min_col - 9
            new_max_c = m_range.max_col - 9
            ws_sum.merge_cells(
                start_row=m_range.min_row,
                end_row=m_range.max_row,
                start_column=new_min_c,
                end_column=new_max_c
            )

    # Column Widths
    col_widths = [14, 22, 18, 12, 24]
    for ci, w in enumerate(col_widths, 1):
        ws_sum.column_dimensions[get_column_letter(ci)].width = w

    tmp_sum_dir = tempfile.mkdtemp()
    tmp_sum_xlsx = os.path.join(tmp_sum_dir, 'exec_summary_only.xlsx')
    wb_sum.save(tmp_sum_xlsx)

    # Render to pixel-perfect cropped image using excel_to_image
    return excel_to_image.excel_to_image(tmp_sum_xlsx)
