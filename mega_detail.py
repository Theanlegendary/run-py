"""mega_detail.py - Clean & Standard Detail Excel for MEGA / DVC Hubs.

Layout:
  Sheet 1 — ⚠ Urgent (Hold time >= 1 day at MEGA Hub)
  Sheet 2 — ✓ Normal (Hold time < 1 day at MEGA Hub)

Clean Design:
  - Dark Blue Table Header (#1F4E78)
  - Pure White Data Rows (#FFFFFF) with subtle gridlines
  - Formatted Money ($#,##0.00) in green text (#196B24)
  - Bold Order IDs (#1F3864)
  - Clean Khmer column headers & Status Codes
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

STATUS_KM = {
    "110": "មិនទាន់បញ្ជូន",
    "120": "កំពុងប្រមូល",
    "200": "បានទទួល",
    "210": "កំពុងដឹក",
    "300": "កំពុងដឹក",
    "302": "បែងចែក",
    "306": "ដល់មណ្ឌល",
    "309": "ស្កែន",
    "310": "បានទទួល",
    "311": "បន្តដឹក",
    "400": "រង់ចាំ",
    "401": "ដឹកជូន",
    "402": "ដឹកជូន",
    "420": "រង់នៅហាង",
    "430": "ដឹកឡើងវិញ",
    "460": "ដឹកឡើងវិញ",
    "472": "រក្សាទុក",
    "480": "កែអាសយដ្ឋាន",
    "500": "ត្រឡប់",
    "520": "បានត្រឡប់",
    "540": "ត្រឡប់បរាជ័យ",
}

# Column indices in source Excel (0-based)
CI_CREATED     = 1   # CREATED DATE
CI_ORDER       = 2   # ORDER ID
CI_SENDER      = 3   # SENDER
CI_RECEIVER    = 4   # RECEIVER
CI_RECV_PO     = 11  # RECEIVE POST OFFICE  (origin)
CI_DELIV_PO    = 13  # DELIVERY POST OFFICE (destination)
CI_CURRENT_PO  = 15  # CURRENT POST OFFICE
CI_TOTAL_FEE   = 19  # TOTAL FEE (USD)
CI_COD         = 20  # COD (USD)
CI_STATUS      = 23  # CURRENT STATUS
CI_ACTION_USER = 25  # ACTION USER (Staff ID & Name)
CI_HUB_CODE    = 35  # ACTION POST OFFICE at ORIGIN HUB (MEGA1 / DVCMEGA1)


def _sc(val):
    if val is None:
        return ""
    s = str(val).strip()
    if " - " in s:
        s = s.split(" - ", 1)[0]
    return s.split()[0].strip() if s else ""


def _parse_created(val, dt_cls):
    if val is None:
        return None
    if isinstance(val, dt_cls):
        return val.date()
    s = str(val).strip().split(" ")[0]
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y", "%d-%m-%Y"):
        try:
            return dt_cls.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _fmt_money(val):
    """Return float or '' for display."""
    try:
        f = float(val)
        return round(f, 2) if f > 0 else ""
    except Exception:
        return ""


def build_mega_detail(source_path, out_path, cfg):
    """Build Clean Detail Excel with 2 separate sheets.
    Returns (total_orders, urgent_count).
    """
    from datetime import datetime as _dt

    excl_test = cfg.get("pivot", {}).get("exclude_test", False)
    test_kw   = cfg.get("pivot", {}).get("test_keywords", ["test"])

    if isinstance(source_path, list):
        data_rows = source_path
    else:
        wb_src = openpyxl.load_workbook(source_path, data_only=True)
        ws_src = wb_src.active
        all_rows = list(ws_src.iter_rows(values_only=True))
        if not all_rows:
            return 0, 0
        data_rows = all_rows[1:]

    today = _dt.now().date()

    urgent_rows = []
    normal_rows = []

    hub_statuses = {"306", "309", "302", "311", "310"}

    # ── DEDUPLICATION: keep only the LATEST scan row per ORDER ID ─────────────
    # TMS export has one row per scan event. Order that moved DVCMEGA1→MEGA1
    # appears in both with status 306. Keep only the most recent row.
    from datetime import timedelta

    def _parse_ts_detail(row):
        for col in (24, CI_CREATED):
            val = row[col] if len(row) > col else None
            if val is None:
                continue
            if isinstance(val, _dt):
                return val
            s = str(val).strip()
            for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y",
                        "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
                try:
                    return _dt.strptime(s, fmt)
                except ValueError:
                    continue
        return _dt.min

    latest_by_order = {}
    for row in data_rows:
        if not row or len(row) <= CI_STATUS:
            continue
        if row[CI_ORDER] is None or str(row[CI_ORDER]).strip() == "":
            continue
        sc = _sc(row[CI_STATUS])
        po = str(row[CI_CURRENT_PO] or "").strip().upper()
        if po == "MEGA1":
            if sc not in ("306", "309"):
                continue
        elif po in ("DVCMEGA1", "DVCMEGA", "DVMEGA") or "DVCMEGA" in po or "DVMEGA" in po:
            if sc != "306":
                continue
        else:
            continue

        if excl_test:
            blob = " ".join(str(row[c] or "") for c in (CI_SENDER, CI_RECEIVER)).lower()
            if any(k.lower() in blob for k in test_kw):
                continue
        oid = str(row[CI_ORDER]).strip()
        ts = _parse_ts_detail(row)
        existing = latest_by_order.get(oid)
        if existing is None or ts > _parse_ts_detail(existing):
            latest_by_order[oid] = row

    for row in latest_by_order.values():
        sc = _sc(row[CI_STATUS])
        po = str(row[CI_CURRENT_PO] or "").strip().upper()

        if po == "MEGA1":
            hub_label = "MEGA1"
        else:
            hub_label = "DVMEGA"


        # Date parsing: Use Action Date (Col 24) or Created Date
        action_val = row[24] if len(row) > 24 and row[24] else row[CI_CREATED]
        action_date = _parse_created(action_val, _dt)
        if not action_date:
            action_date = _parse_created(row[CI_CREATED], _dt)

        if not action_date:
            continue

        # 14-day cutoff (starts from 20/08 when today is 03/09)
        cutoff_date = today - timedelta(days=14)
        if action_date < cutoff_date:
            continue

        age = (today - action_date).days
        # MEGA SLA Rule: age 0 days is normal, 1 or more is urgent
        is_urgent = age >= 1

        action_user = str(row[CI_ACTION_USER] or "").strip() if len(row) > CI_ACTION_USER else ""

        record = {
            "hub":         hub_label,
            "status_km":   f"{sc} - {STATUS_KM.get(sc, sc)}" if sc in STATUS_KM else str(row[CI_STATUS] or sc),
            "action_user": action_user,
            "origin":      str(row[CI_RECV_PO]  or "").strip(),
            "dest":        str(row[CI_DELIV_PO] or "").strip(),
            "order_id":    str(row[CI_ORDER]    or "").strip(),
            "fee":         _fmt_money(row[CI_TOTAL_FEE] if len(row) > CI_TOTAL_FEE else None),
            "cod":         _fmt_money(row[CI_COD]       if len(row) > CI_COD       else None),
            "created":     str(row[CI_CREATED]  or "").strip(),
            "age":         age,
        }

        if is_urgent:
            urgent_rows.append(record)
        else:
            normal_rows.append(record)

    # Sort: Hub (MEGA1 first, DVMEGA second), then age desc, then order_id
    _hub_rank = {"MEGA1": 0, "DVMEGA": 1}
    urgent_rows.sort(key=lambda r: (_hub_rank.get(r["hub"], 9), -r["age"], r["order_id"]))
    normal_rows.sort(key=lambda r: (_hub_rank.get(r["hub"], 9), -r["age"], r["order_id"]))

    # ── Build Excel with 2 Sheets ─────────────────────────────────────────────
    wb = openpyxl.Workbook()

    # Sheet 1: Urgent
    ws_urg = wb.active
    ws_urg.title = "⚠ Urgent"

    # Sheet 2: Normal
    ws_norm = wb.create_sheet(title="✓ Normal")

    # Clean Styles
    thin        = Side(style="thin", color="BFBFBF")
    bdr         = Border(left=thin, right=thin, top=thin, bottom=thin)
    ctr         = Alignment(horizontal="center", vertical="center", wrap_text=False)
    left_align  = Alignment(horizontal="left",   vertical="center", wrap_text=False)
    right_align = Alignment(horizontal="right",  vertical="center", wrap_text=False)

    hdr_fill   = PatternFill("solid", fgColor="1F4E78") # Clean Dark Blue Header
    urgent_sec = PatternFill("solid", fgColor="C00000") # Clean Red Title Bar
    normal_sec = PatternFill("solid", fgColor="375623") # Clean Green Title Bar
    white_bg   = PatternFill("solid", fgColor="FFFFFF") # Clean White Rows

    hdr_font  = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    sec_font  = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    data_font = Font(name="Segoe UI", size=10, color="1F1F1F")
    id_font   = Font(name="Segoe UI", size=10, color="1F3864", bold=True)
    red_font  = Font(name="Segoe UI", size=10, color="C00000", bold=True)
    gray_font = Font(name="Segoe UI", size=10, color="595959")
    fee_font  = Font(name="Segoe UI", size=10, color="196B24")

    HDR = [
        "មណ្ឌល",         # 1 — Hub (MEGA1 / DVMEGA)
        "ស្ថានភាព",      # 2 — Status KM
        "បុគ្គលិក",       # 3 — Action User
        "ប.ស.ចេញ",       # 4 — Origin PO
        "ប.ស.ទៅ",        # 5 — Dest PO
        "លេខបញ្ជា",      # 6 — Order ID
        "ថ្លៃ (USD)",    # 7 — Fee
        "COD (USD)",     # 8 — COD
        "កាលបរិច្ឆេទ",  # 9 — Created Date
        "ថ្ងៃ",          # 10 — Age
    ]
    COL_WIDTHS = [12, 26, 32, 12, 12, 16, 12, 12, 18, 6]
    NCOLS = len(HDR)

    def _populate_clean_sheet(ws, rows_data, sec_label, sec_fill, is_urg):
        ws.views.sheetView[0].showGridLines = True
        
        # Row 1: Clean Section Header Banner
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NCOLS)
        c = ws.cell(1, 1, sec_label)
        c.fill = sec_fill
        c.font = sec_font
        c.alignment = left_align
        c.border = bdr
        ws.row_dimensions[1].height = 20

        # Row 2: Table Column Headers
        ws.row_dimensions[2].height = 22
        for ci, h in enumerate(HDR, 1):
            cell = ws.cell(2, ci, h)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = ctr
            cell.border = bdr

        # Data Rows
        curr = 3
        row_fill = PatternFill("solid", fgColor="FFEBEB") if is_urg else PatternFill("solid", fgColor="FFFFFF")
        for rec in rows_data:
            ws.row_dimensions[curr].height = 18
            vals = [
                rec["hub"],
                rec["status_km"],
                rec["action_user"],
                rec["origin"],
                rec["dest"],
                rec["order_id"],
                rec["fee"],
                rec["cod"],
                rec["created"],
                rec["age"],
            ]
            for ci, val in enumerate(vals, 1):
                cell = ws.cell(curr, ci, val)
                cell.fill = row_fill
                cell.border = bdr


                if ci == 1:         # Hub
                    cell.font = data_font; cell.alignment = ctr
                elif ci == 2:       # Status
                    cell.font = data_font; cell.alignment = left_align
                elif ci == 3:       # Action User
                    cell.font = gray_font; cell.alignment = left_align
                elif ci in (4, 5):  # Origin / Dest
                    cell.font = data_font; cell.alignment = ctr
                elif ci == 6:       # Order ID
                    cell.font = id_font; cell.alignment = ctr
                elif ci in (7, 8):  # Fee / COD
                    cell.font = fee_font; cell.alignment = right_align
                    if val != "":
                        cell.number_format = "$#,##0.00"
                elif ci == 9:       # Created Date
                    cell.font = gray_font; cell.alignment = ctr
                elif ci == 10:      # Age
                    cell.font = red_font if is_urg else data_font
                    cell.alignment = ctr

            curr += 1

        for ci, w in enumerate(COL_WIDTHS, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

    _populate_clean_sheet(ws_urg, urgent_rows, f"⚠ ប្រញាប់ (Urgent) — {len(urgent_rows)} records", urgent_sec, True)
    _populate_clean_sheet(ws_norm, normal_rows, f"✓ ធម្មតា (Normal) — {len(normal_rows)} records", normal_sec, False)

    wb.save(out_path)
    total_orders = len(urgent_rows) + len(normal_rows)
    return total_orders, len(urgent_rows)
