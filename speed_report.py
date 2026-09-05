"""
speed_report.py — Fast Delivery Speed & Courier Commission Report
===================================================================
- Filters delivered orders (Status 410).
- Identifies true VTT (Door Delivery) orders using:
    1) VAS_SERVICE from latest_revenue.xlsx (status 110/120/210 extra services)
    2) extra service / note columns containing 'VTT'
    3) SERVICE column containing 'VTT'
- Left Table (Cols A-I):
    No | Order Number | Customer | VAS Service | Post Office | Duration | Hours | Speed Tier | Commission ($)
- Right Table (Cols K-U):
    No | Post Office | Need Deliver | Delivered | < 2 Hrs | 2–4 Hrs | 4–8 Hrs | > 8 Hrs | % Within 8h | % Over 8h | Commission ($)
- Sheet 2 (base): Complete audit dataset with timestamps and courier details.
- render_speed_summary_image: Executive summary dashboard image.
"""

import os
import copy
import tempfile
from datetime import datetime, timedelta
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import excel_to_image

MAIN_36_BRANCHES = [
    "BANP001", "BATP001", "CHAP001", "CHHP001", "KAMP001", "KANP001", "KOHP001", "KRAP001",
    "MONP001", "ODDP001", "PNPP001", "PNPP002", "PNPP003", "PNPP004", "PNPP005", "PNPP006",
    "PNPP007", "PNPP008", "PNPP009", "PNPP010", "PNPP011", "PNPP012", "PNPP013", "PNPP014",
    "PREP001", "PRHP001", "PURP001", "ROTP001", "SIEP001", "SIHP001", "SPEP001", "STUP001",
    "SVAP001", "TAKP001", "TBKP001", "THOP001"
]

ZONE_BRANCHES = {
    "ZONE1": [
        "PNPP001", "PNPP002", "PNPP003", "PNPP004", "PNPP005", "PNPP006", "PNPP007",
        "PNPP008", "PNPP009", "PNPP010", "PNPP011", "PNPP012", "PNPP013", "PNPP014",
        "KANP001", "PREP001", "SVAP001"
    ],
    "ZONE2": ["KAMP001", "KOHP001", "SIHP001", "SPEP001", "TAKP001"],
    "ZONE3": ["BANP001", "BATP001", "CHHP001", "PURP001"],
    "ZONE4": ["ODDP001", "PRHP001", "SIEP001", "THOP001"],
    "ZONE5": ["CHAP001", "KRAP001", "TBKP001", "ROTP001", "MONP001", "STUP001"]
}


def parse_time(val):
    if val is None or pd.isna(val):
        return None
    s = str(val).strip()
    if not s or s.lower() == 'nan':
        return None
    for fmt in ('%d/%m/%Y %H:%M:%S', '%d/%m/%Y %H:%M', '%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%d-%m-%Y %H:%M:%S'):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    try:
        dt = pd.to_datetime(val, dayfirst=True, format='mixed', errors='coerce')
        if pd.notna(dt):
            return dt.to_pydatetime()
    except Exception:
        pass
    return None


def parse_date(val):
    t = parse_time(val)
    return t.date() if t else None


_tracking_trips_cache = {}

def get_tracking_trips(order_id, cfg=None):
    if not order_id:
        return []
    order_id_str = str(order_id).strip()
    if order_id_str in _tracking_trips_cache:
        return _tracking_trips_cache[order_id_str]

    token = None
    if cfg and isinstance(cfg, dict):
        token = cfg.get("api", {}).get("bearer_token")
    if not token:
        base_dirs = [
            os.path.dirname(os.path.abspath(__file__)),
            os.getcwd(),
            os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        ]
        for d in base_dirs:
            cfg_p = os.path.join(d, "config.json")
            if os.path.exists(cfg_p):
                try:
                    import json
                    with open(cfg_p, "r", encoding="utf-8") as f:
                        c = json.load(f)
                    token = c.get("api", {}).get("bearer_token")
                    if token:
                        break
                except Exception:
                    pass

    if not token:
        _tracking_trips_cache[order_id_str] = []
        return []

    try:
        import requests
        url = "https://gw-express.metfone.com.kh/tms-tracking/api/v1/order-tracking"
        headers = {"Authorization": f"Bearer {token}", "Accept": "application/json"}
        r = requests.get(url, params={"order_id": order_id_str}, headers=headers, timeout=6)
        if r.status_code == 200:
            trips = r.json().get("trackingTrips", [])
            _tracking_trips_cache[order_id_str] = trips
            return trips
    except Exception:
        pass

    _tracking_trips_cache[order_id_str] = []
    return []


def find_po_arrival_time_from_trips(trips, po, action_user, deliv_time):
    """
    Find earliest arrival/processing scan (306, 309, 311, 400, 401, 402) at target branch 'po'
    or by 'action_user' prior to delivery time on the delivery date.
    """
    if not trips or not deliv_time:
        return None
    deliv_date = deliv_time.date()
    cand_times = []
    user_clean = str(action_user or '').split('(')[0].split('-')[-1].strip().lower()

    for t in trips:
        st = str(t.get('status', ''))
        if st in ('S306', 'S309', 'S311', 'S400', 'S401', 'S402', '306', '309', '311', '400', '401', '402'):
            ts_str = t.get('updatedAt')
            if not ts_str:
                continue
            try:
                ts_clean = ts_str.replace('Z', '+00:00')
                dt = datetime.fromisoformat(ts_clean).replace(tzinfo=None)
            except Exception:
                dt = parse_time(ts_str)
            if not dt or dt > deliv_time:
                continue

            t_po = (t.get('postOffice') or {}).get('postOfficeCode') or t.get('postOfficeCode') or (t.get('currentPostOffice') or {}).get('postOfficeCode') or ''
            t_user = (t.get('updatedBy') or {}).get('name') or t.get('shipperName') or ''
            t_user_clean = str(t_user).split('(')[0].split('-')[-1].strip().lower()

            matches_po = bool(po and str(t_po).strip().upper() == po.upper())
            matches_user = bool(user_clean and (user_clean in t_user_clean or t_user_clean in user_clean))

            if dt.date() == deliv_date and (matches_po or matches_user):
                cand_times.append(dt)
            elif matches_po:
                cand_times.append(dt)

    if cand_times:
        return min(cand_times)
    return None


def load_speed_context(src_xlsx, revenue_path=None):
    """
    Reads src_xlsx and revenue_path once into memory.
    Builds vas_mapping from revenue export and note/service columns.
    """
    df = pd.read_excel(src_xlsx)
    df.columns = [str(c).strip() for c in df.columns]

    vas_mapping = {}
    if not revenue_path:
        cache_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "cache")
        possible_rev = os.path.join(cache_dir, "latest_revenue.xlsx")
        if os.path.exists(possible_rev):
            revenue_path = possible_rev

    if revenue_path and os.path.exists(revenue_path):
        try:
            for skiprows in range(5):
                rev_df = pd.read_excel(revenue_path, skiprows=skiprows)
                rev_order_col = next((c for c in rev_df.columns if str(c).strip().upper() in ['ORDER ID', 'ORDER_NUMBER', 'MÃ ĐƠN HÀNG', 'BILL', 'MÃ ĐƠN']), None)
                if rev_order_col and 'VAS_SERVICE' in rev_df.columns:
                    for _, row in rev_df.dropna(subset=[rev_order_col]).iterrows():
                        oid = str(row[rev_order_col]).strip()
                        vas = str(row['VAS_SERVICE']).strip()
                        if vas.lower() not in ['nan', 'none', '']:
                            vas_mapping[oid] = vas
                    break
        except Exception as e:
            print(f"Failed to load revenue data in speed report: {e}")

    return df, vas_mapping


import re

def load_test_bills(cfg=None):
    """Load ignored/test bill IDs from test_bills.txt, delayed_bills.json, and test_receipts Excel file."""
    test_ids = set()
    base_dirs = [
        os.getcwd(),
        os.path.dirname(os.path.abspath(__file__)),
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    ]
    for d in base_dirs:
        txt_path = os.path.join(d, "test_bills.txt")
        if os.path.exists(txt_path):
            try:
                with open(txt_path, "r", encoding="utf-8") as f:
                    for line in f:
                        val = line.strip()
                        if val and not val.startswith("#"):
                            clean_val = str(val).strip().upper()
                            clean_val = re.sub(r'\.0$', '', clean_val)
                            if clean_val:
                                test_ids.add(clean_val)
            except Exception:
                pass

    for d in base_dirs:
        delay_path = os.path.join(d, "delayed_bills.json")
        if os.path.exists(delay_path):
            try:
                import json
                with open(delay_path, "r", encoding="utf-8") as f:
                    delayed = json.load(f)
                today_d = datetime.now().date()
                for bill_id, exp_date_str in delayed.items():
                    try:
                        exp_date = datetime.strptime(exp_date_str, "%Y-%m-%d").date()
                        if today_d < exp_date:
                            clean_val = str(bill_id).strip().upper()
                            clean_val = re.sub(r'\.0$', '', clean_val)
                            if clean_val:
                                test_ids.add(clean_val)
                    except Exception:
                        pass
            except Exception:
                pass

    excel_paths = []
    if cfg and isinstance(cfg, dict):
        tr_cfg = cfg.get("test_receipts", {})
        if tr_cfg.get("enabled") and tr_cfg.get("path"):
            excel_paths.append(tr_cfg.get("path"))
    else:
        for d in base_dirs:
            cfg_path = os.path.join(d, "config.json")
            if os.path.exists(cfg_path):
                try:
                    import json
                    with open(cfg_path, "r", encoding="utf-8") as f:
                        c = json.load(f)
                    tr_cfg = c.get("test_receipts", {})
                    if tr_cfg.get("enabled") and tr_cfg.get("path"):
                        excel_paths.append(tr_cfg.get("path"))
                        break
                except Exception:
                    pass

    for d in base_dirs:
        tx_path = os.path.join(d, "test.xlsx")
        if os.path.exists(tx_path) and tx_path not in excel_paths:
            excel_paths.append(tx_path)

    for ep in excel_paths:
        if os.path.exists(ep):
            try:
                import pandas as pd
                if ep.lower().endswith((".xlsx", ".xls")):
                    df_test = pd.read_excel(ep, dtype=str)
                else:
                    df_test = pd.read_csv(ep, dtype=str, keep_default_na=False)
                df_test = df_test.fillna("")
                if not df_test.empty:
                    order_col = next(
                        (
                            c for c in df_test.columns
                            if any(k in str(c).lower() for k in ("order", "code", "bill", "phi", "shipment"))
                        ),
                        df_test.columns[0]
                    )
                    for val in df_test[order_col].tolist():
                        clean_val = str(val).strip().upper()
                        clean_val = re.sub(r'\.0$', '', clean_val)
                        if clean_val and clean_val != "NAN":
                            test_ids.add(clean_val)
            except Exception:
                pass

    return test_ids


def build_speed_report(src_xlsx, out_xlsx, target_label="ALL", report_date=None, revenue_path=None, preloaded_df=None, preloaded_vas_map=None):
    os.makedirs(os.path.dirname(os.path.abspath(out_xlsx)), exist_ok=True)

    if preloaded_df is not None:
        df = preloaded_df.copy()
    else:
        df = pd.read_excel(src_xlsx)
        df.columns = [str(c).strip() for c in df.columns]

    col_order = next((c for c in df.columns if 'ORDER ID' in c or 'ORDER' in c), 'ORDER ID')

    # Exclude all testing bills from test_bills.txt and delayed_bills.json
    test_bills = load_test_bills()
    if test_bills and col_order in df.columns:
        order_series = df[col_order].astype(str).str.strip().str.upper().str.replace(r'\.0$', '', regex=True)
        df = df[~order_series.isin(test_bills)].copy()

    # Exclude orders with test keywords in Order ID, Sender, Receiver, Remark, Note, or Description
    test_keywords = ['test', 'kiểm thử', 'kiem thu', 'demo', 'trial', 'sample', 'dummy', 'thử nghiệm', 'thu nghiem']
    test_mask = pd.Series(False, index=df.index)

    if col_order in df.columns:
        ord_lower = df[col_order].astype(str).str.lower()
        for kw in test_keywords:
            test_mask |= ord_lower.str.contains(kw, na=False)

    check_cols = [c for c in df.columns if any(k in str(c).upper() for k in ('SENDER', 'RECEIVER', 'NOTE', 'REMARK', 'GOODS', 'DESC', 'COMMODITY', 'ITEM', 'CUSTOMER'))]
    for c in check_cols:
        if c in df.columns:
            s_lower = df[c].astype(str).str.lower()
            for kw in test_keywords:
                test_mask |= s_lower.str.contains(kw, na=False)

    if test_mask.any():
        df = df[~test_mask].copy()

    if preloaded_vas_map is not None:
        vas_mapping = preloaded_vas_map
    else:
        _, vas_mapping = load_speed_context(src_xlsx, revenue_path=revenue_path)

    col_order = next((c for c in df.columns if 'ORDER ID' in c or 'ORDER' in c), 'ORDER ID')
    col_dest_prov = next((c for c in df.columns if 'DELIVERY PROVINCE' in c or 'DESTINATION_BRANCH' in c), 'DELIVERY PROVINCE')
    col_dest_po = next((c for c in df.columns if 'DELIVERY POST' in c or 'DESTINATION_POST' in c), 'DELIVERY POST OFFICE')
    col_orig_br = next((c for c in df.columns if 'ACTION POST OFFICE' in c or 'ORIGIN_BRANCH' in c), 'ACTION POST OFFICE')
    col_orig_po = next((c for c in df.columns if 'CURRENT POST OFFICE' in c or 'ORIGIN_POST' in c), 'CURRENT POST OFFICE')
    col_status = next((c for c in df.columns if 'CURRENT STATUS' in c or 'STATUS' in c), 'CURRENT STATUS')
    col_created = next((c for c in df.columns if 'CREATED DATE' in c), 'CREATED DATE')
    col_receiver = next((c for c in df.columns if 'RECEIVER' in c), 'RECEIVER')
    col_action_time = next((c for c in df.columns if 'ACTION TIME' in c or 'CURRENT TIME' in c), 'CURRENT TIME')
    
    col_306_store_last = next((c for c in df.columns if '306' in c and ('STORE' in c or 'AGENT' in c) and 'LAST' in c), None)
    col_306_store_hub = next((c for c in df.columns if '306' in c and ('STORE' in c or 'AGENT' in c) and 'FROM HUB' in c), None)
    col_306_store_any = next((c for c in df.columns if '306' in c and ('STORE' in c or 'AGENT' in c) and 'ORIGIN' not in c), None)

    cols_list = list(df.columns)
    col_306_po_last = None
    if col_306_store_last:
        idx = cols_list.index(col_306_store_last)
        for j in range(idx + 1, min(idx + 4, len(cols_list))):
            if 'ACTION POST OFFICE' in cols_list[j]:
                col_306_po_last = cols_list[j]
                break

    col_306_po_first = None
    if col_306_store_hub:
        idx = cols_list.index(col_306_store_hub)
        for j in range(idx + 1, min(idx + 4, len(cols_list))):
            if 'ACTION POST OFFICE' in cols_list[j]:
                col_306_po_first = cols_list[j]
                break

    col_400_time = next((c for c in df.columns if '400' in c or 'OUT' in c or 'ASSIGN' in c), None)
    col_210_time = next((c for c in df.columns if '210' in c), None)
    col_service = next((c for c in df.columns if 'SERVICE' in c), 'SERVICE')
    col_note = next((c for c in df.columns if 'NOTE' in c), 'NOTE')
    col_user = next((c for c in df.columns if 'ACTION USER' in c or 'USER' in c), 'ACTION USER')
    note_cols = [c for c in df.columns if any(k in str(c).upper() for k in ('NOTE', 'VAS', 'EXTRA', 'SERVICE', 'DESCRIPTION'))]

    today = report_date or datetime.now().date()
    tgt = "".join(c for c in str(target_label).upper() if c.isalnum() or c in ("-", "_")).strip()
    if not tgt:
        tgt = "ALL"

    if tgt.startswith("ZONE"):
        target_branches = ZONE_BRANCHES.get(tgt, [])
    elif tgt in ("ALL", "TOTAL"):
        target_branches = MAIN_36_BRANCHES
    elif tgt in MAIN_36_BRANCHES:
        target_branches = [tgt]
    else:
        target_branches = [tgt]

    df['status_code_clean'] = df[col_status].astype(str).str.extract(r'^(\d{3})')[0]
    df['curr_po_clean'] = df[col_orig_po].astype(str).str.strip().str.upper()
    df['deliv_po_clean'] = df[col_dest_po].astype(str).str.strip().str.upper()
    df['deliv_date'] = df[col_action_time].apply(parse_date)

    summary_data = {}
    for b in target_branches:
        summary_data[b] = {
            "po": b,
            "need_deliver": 0,
            "total_delivered": 0,
            "under_2h": 0,
            "between_2_4h": 0,
            "between_4_8h": 0,
            "over_8h": 0,
            "total_commission": 0.0
        }

    # 1. Need Deliver Count (Pending VTT delivery bills)
    pending_statuses = ('306', '309', '311', '400', '401', '402', '420', '430', '471', '472', '480')
    need_deliv_df = df[df['status_code_clean'].isin(pending_statuses)]
    for row in need_deliv_df.to_dict('records'):
        order_id = str(row.get(col_order, '')).strip()
        svc_val = str(row.get(col_service, '') or '').strip().upper()
        combined_notes = ' '.join(str(row.get(col, '') or '') for col in note_cols).upper()
        vas_mapped = str(vas_mapping.get(order_id, '')).upper()

        is_vtt = ('VTT' in vas_mapped) or ('VTT' in combined_notes) or ('VTT' in svc_val)
        if not is_vtt:
            continue

        deliv_po = str(row.get('deliv_po_clean', '')).strip()
        curr_po = str(row.get('curr_po_clean', '')).strip()
        po = deliv_po if deliv_po and deliv_po != 'NAN' else curr_po
        if po in summary_data:
            summary_data[po]["need_deliver"] += 1

    # 2. Delivered Today (Status 410)
    delivered_df = df[(df['status_code_clean'] == '410') & (df['deliv_date'] == today)]

    base_rows = []
    r_idx = 1

    for row in delivered_df.to_dict('records'):
        deliv_po = str(row.get('deliv_po_clean', '')).strip()
        curr_po = str(row.get('curr_po_clean', '')).strip()
        raw_po = deliv_po if deliv_po and deliv_po != 'NAN' else curr_po

        if not raw_po or raw_po == 'NAN':
            continue

        if raw_po not in summary_data:
            continue
        po = raw_po

        order_id = str(row.get(col_order, '')).strip()
        svc_val = str(row.get(col_service, '') or '').strip().upper()
        if svc_val == 'NAN':
            svc_val = ''
        action_user_val = str(row.get(col_user, '') or '').strip()

        # Strict VTT Check: Verify order has VTT at Status 110/120/210 or in Revenue VAS export
        combined_notes = ' '.join(str(row.get(col, '') or '') for col in note_cols).upper()
        vas_mapped = str(vas_mapping.get(order_id, '')).upper()

        is_vtt = ('VTT' in vas_mapped) or ('VTT' in combined_notes) or ('VTT' in svc_val)
        if not is_vtt:
            continue  # Strictly VTT bills only

        t410 = parse_time(row.get(col_action_time)) or parse_time(row.get(col_created))
        
        t_start = None
        po_306_first = str(row.get(col_306_po_first, '') or '').strip().upper() if col_306_po_first else ""
        po_306_last = str(row.get(col_306_po_last, '') or '').strip().upper() if col_306_po_last else ""

        # 1. If FIRST TIME arrival from hub was at THIS delivering branch (po), use it (anti-cheat against rapid re-scans)
        if col_306_store_hub and pd.notna(row.get(col_306_store_hub)) and (not po_306_first or po_306_first == po):
            t_start = parse_time(row.get(col_306_store_hub))
        # 2. If parcel was redirected/forwarded from another branch (change address), use arrival at THIS branch (po)
        elif col_306_store_last and pd.notna(row.get(col_306_store_last)) and (not po_306_last or po_306_last == po):
            t_start = parse_time(row.get(col_306_store_last))

        # 3. If arrival at THIS branch is not in the export columns (e.g. inter-branch transfer or po_306_first is from origin branch):
        # Query live tracking trips API to find the physical arrival/processing scan (306/309/311/402) at 'po' by the courier:
        if not t_start or (po_306_first and po_306_first != po):
            trips = get_tracking_trips(order_id)
            if trips:
                t_trip = find_po_arrival_time_from_trips(trips, po, action_user_val, t410)
                if t_trip:
                    t_start = t_trip

        # 4. Fallback checks: Only use other 306 columns if they actually belong to THIS delivering branch (po)
        if not t_start:
            if col_306_store_last and pd.notna(row.get(col_306_store_last)) and (not po_306_last or po_306_last == po):
                t_start = parse_time(row.get(col_306_store_last))
            elif col_306_store_hub and pd.notna(row.get(col_306_store_hub)) and (not po_306_first or po_306_first == po):
                t_start = parse_time(row.get(col_306_store_hub))
            elif col_306_store_any and pd.notna(row.get(col_306_store_any)) and (not po_306_first or po_306_first == po):
                t_start = parse_time(row.get(col_306_store_any))
            elif col_400_time and pd.notna(row.get(col_400_time)):
                t_start = parse_time(row.get(col_400_time))
            elif col_210_time and pd.notna(row.get(col_210_time)):
                t_start = parse_time(row.get(col_210_time))

        if not t_start:
            # If the parcel came from another branch and no local arrival scan exists,
            # use delivery time minus 1.5h standard rather than penalizing for inter-branch transit:
            if po_306_first and po_306_first != po and t410:
                t_start = max(t410 - timedelta(hours=1.5), datetime(t410.year, t410.month, t410.day, 8, 0, 0))
            else:
                t_start = parse_time(row.get(col_created))

        if t410 and t_start and t410 >= t_start:
            duration_hours = (t410 - t_start).total_seconds() / 3600.0
        elif t410 and pd.notna(row.get(col_created)):
            t_created = parse_time(row.get(col_created))
            duration_hours = (t410 - t_created).total_seconds() / 3600.0 if (t_created and t410 >= t_created) else 4.0
        else:
            duration_hours = 4.0

        summary_data[po]["total_delivered"] += 1

        if duration_hours < 2.0:
            tier = "< 2 Hours (+50%)"
            rate_usd = 0.30
            summary_data[po]["under_2h"] += 1
            tag_color = "GREEN"
        elif duration_hours <= 4.0:
            tier = "2 - 4 Hours (+25%)"
            rate_usd = 0.25
            summary_data[po]["between_2_4h"] += 1
            tag_color = "BLUE"
        elif duration_hours <= 8.0:
            tier = "4 - 8 Hours (Normal)"
            rate_usd = 0.20
            summary_data[po]["between_4_8h"] += 1
            tag_color = "NORMAL"
        else:
            tier = "> 8 Hours (-25% Fine)"
            rate_usd = 0.15
            summary_data[po]["over_8h"] += 1
            tag_color = "RED"

        summary_data[po]["total_commission"] += rate_usd
        dur_str = f"{int(duration_hours)}h {int((duration_hours%1)*60):02d}m" if duration_hours else "N/A"

        base_rows.append({
            "no": r_idx,
            "order_number": order_id,
            "customer": str(row.get(col_receiver, ''))[:28],
            "vas_service": "VTT",
            "service": svc_val,
            "origin_branch": str(row.get(col_orig_br, '')),
            "origin_post": curr_po,
            "destination_branch": str(row.get(col_dest_prov, '')),
            "destination_post": deliv_po,
            "assigned_branch": po,
            "shipper": action_user_val,
            "created_at": str(row.get(col_created, '')),
            "t_start": str(t_start or ""),
            "t410": str(t410 or ""),
            "duration": dur_str,
            "duration_hours": round(duration_hours, 2),
            "tier": tier,
            "rate_usd": rate_usd,
            "tag_color": tag_color
        })
        r_idx += 1

    # Totals across all target branches
    tot_del = sum(s["total_delivered"] for s in summary_data.values())
    tot_u2 = sum(s["under_2h"] for s in summary_data.values())
    tot_24 = sum(s["between_2_4h"] for s in summary_data.values())
    tot_48 = sum(s["between_4_8h"] for s in summary_data.values())
    tot_o8 = sum(s["over_8h"] for s in summary_data.values())
    tot_need = sum(s["need_deliver"] for s in summary_data.values())
    tot_pay = sum(s["total_commission"] for s in summary_data.values())

    # Build Excel Workbook
    wb = openpyxl.Workbook()

    # Sheet 1: DELIVERY SPEED REPORT
    ws1 = wb.active
    ws1.title = "DELIVERY SPEED REPORT"
    ws1.views.sheetView[0].showGridLines = True

    # Styling Palette
    _TITLE_LEFT = PatternFill("solid", fgColor="0F172A") # Midnight Navy
    _TITLE_FILL = PatternFill("solid", fgColor="0B132B") # Deep Midnight
    _HDR_LEFT   = PatternFill("solid", fgColor="1E293B") # Slate Header
    _HDR_FILL   = PatternFill("solid", fgColor="1C3D82") # Royal Sapphire Blue
    _TOT_FILL   = PatternFill("solid", fgColor="E2E8F0") # Slate Total
    _BORDER     = Border(
        left=Side(style="thin", color="CBD5E1"), right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"), bottom=Side(style="thin", color="CBD5E1")
    )
    _DOUBLE_BOT = Border(
        left=Side(style="thin", color="CBD5E1"), right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="94A3B8"), bottom=Side(style="double", color="0B132B")
    )

    font_title = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_hdr   = Font(name="Segoe UI", size=8.5, bold=True, color="FFFFFF")
    font_data  = Font(name="Segoe UI", size=8.5, color="0F172A")
    font_bold  = Font(name="Segoe UI", size=8.5, bold=True, color="0F172A")
    font_tot   = Font(name="Segoe UI", size=9.5, bold=True, color="0F172A")
    font_vtt   = Font(name="Segoe UI", size=8.5, bold=True, color="059669") # Emerald Green for VTT
    font_green = Font(name="Segoe UI", size=8.5, bold=True, color="16A34A")
    font_red   = Font(name="Segoe UI", size=8.5, bold=True, color="DC2626")
    font_tot_comm = Font(name="Segoe UI", size=9.5, bold=True, color="047857")

    date_str = today.strftime('%d/%m/%Y')
    label_upper = target_label.upper().strip()

    # 1. Executive Summary Title Banner (Cols A-K: 1-11)
    ws1.merge_cells("A1:K1")
    ws1.cell(1, 1, f"SPEED & COMMISSION — {label_upper} • {date_str}").font = font_title
    ws1.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, 12):
        ws1.cell(1, c).fill = _TITLE_FILL
    ws1.row_dimensions[1].height = 36.0

    # 2. Detail Title Banner (Cols M-U: 13-21)
    ws1.merge_cells("M1:U1")
    ws1.cell(1, 13, f"METFONE EXPRESS — DAILY DELIVERY SPEED DETAIL — {date_str} (ព័ត៌មានលម្អិតល្បឿនដឹក)").font = font_title
    ws1.cell(1, 13).alignment = Alignment(horizontal="left", vertical="center")
    for c in range(13, 22):
        ws1.cell(1, c).fill = _TITLE_LEFT

    # Row 2: Headers (Option C: Concise, clean, executive headers)
    headers_summary = [
        "NO", "BRANCH", "PENDING", "DELIVERED", "< 2h (+50%)",
        "2-4h (+25%)", "4-8h (Normal)", "> 8h (-25%)",
        "% < 8h", "% > 8h", "COMMISSION"
    ]
    headers_detail = [
        "No\n(ល.រ)", "Order Number\n(លេខប័ណ្ណ)", "Customer\n(អតិថិជន)", "VAS Service\n(ប្រភេទសេវា)", "Post Office\n(សាខា)",
        "Duration\n(រយៈពេល)", "Hours\n(ម៉ោង)", "Speed Tier\n(កម្រិតល្បឿន)", "Commission\n(កម្រៃ $)"
    ]

    ws1.row_dimensions[2].height = 34.0
    _HDR_UNIFIED = PatternFill("solid", fgColor="0F172A") # Single Executive Dark Navy Header

    # Executive Summary Headers (Cols A to K: 1 to 11) - One single unified professional header
    for ci, h in enumerate(headers_summary, 1):
        cell = ws1.cell(2, ci, h)
        cell.font = font_hdr
        cell.fill = _HDR_UNIFIED
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER

    # Detail Headers (Cols M to U: 13 to 21)
    for ci, h in enumerate(headers_detail, 13):
        cell = ws1.cell(2, ci, h)
        cell.font = font_hdr
        cell.fill = _HDR_UNIFIED
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER

    # Populate Executive Summary Rows (Cols A to K: 1 to 11)
    r_sum = 3
    n_idx = 1

    def calc_speed_sort_key(stats):
        tot_b = stats["total_delivered"]
        has_delivery = 1 if tot_b > 0 else 0
        need_cnt = stats.get("need_deliver", 0)
        total_orders = need_cnt + tot_b
        
        # Fixed percentage calculation: include pending orders in denominator
        on_time_delivered = stats["under_2h"] + stats["between_2_4h"] + stats["between_4_8h"]
        pct_within_8h = (on_time_delivered / total_orders * 100.0) if total_orders > 0 else 0.0
        return (-has_delivery, -pct_within_8h, -tot_b, stats["po"])

    sorted_branches = sorted(summary_data.values(), key=calc_speed_sort_key)

    # Vivid text fonts & highlights matching reference image
    font_po        = Font(name="Segoe UI", size=8.5, bold=True, color="0F172A")    # Post office (dark blue bold)
    font_delivered = Font(name="Segoe UI", size=8.5, bold=True, color="2563EB")    # Delivered (410) (vivid blue)
    font_u2h       = Font(name="Segoe UI", size=8.5, bold=True, color="16A34A")    # < 2 Hours (+50%) (emerald green)
    font_2_4h      = Font(name="Segoe UI", size=8.5, bold=True, color="0284C7")    # 2-4 Hours (+25%) (sky blue / cyan)
    font_4_8h      = Font(name="Segoe UI", size=8.5, bold=False, color="64748B")   # 4-8 Hours (Normal) (slate gray)
    font_o8h       = Font(name="Segoe UI", size=8.5, bold=True, color="DC2626")    # > 8 Hours (-25%) (vivid red)
    font_pct_good  = Font(name="Segoe UI", size=8.5, bold=True, color="16A34A")    # % < 8h (emerald green)
    font_pct_bad   = Font(name="Segoe UI", size=8.5, bold=True, color="DC2626")    # % > 8h (vivid red)
    font_comm      = Font(name="Segoe UI", size=8.5, bold=True, color="16A34A")    # Commission ($) (emerald green)

    for stats in sorted_branches:
        ws1.row_dimensions[r_sum].height = 19.0
        need_cnt = stats.get("need_deliver", 0)
        tot_b = stats["total_delivered"]
        total_orders = need_cnt + tot_b

        # Fixed percentage calculation: include pending orders in denominator
        on_time_delivered = stats["under_2h"] + stats["between_2_4h"] + stats["between_4_8h"]
        slow_delivered = stats["over_8h"]
        
        # For correct KPI: % <8h = delivered_fast / (pending + delivered)
        # % >8h = (delivered_slow + pending_orders) / (pending + delivered)  
        # Note: All pending orders are considered "slow" since they haven't been completed yet
        pct_under_8h = (on_time_delivered / total_orders * 100.0) if total_orders > 0 else 0.0
        pct_over_8h = ((slow_delivered + need_cnt) / total_orders * 100.0) if total_orders > 0 else 0.0

        str_pct_u8 = f"{pct_under_8h:.1f}%"
        str_pct_o8 = f"{pct_over_8h:.1f}%"

        s_vals = [
            n_idx,
            stats["po"],
            need_cnt,
            stats["total_delivered"],
            stats["under_2h"],
            stats["between_2_4h"],
            stats["between_4_8h"],
            stats["over_8h"],
            str_pct_u8,
            str_pct_o8,
            f"${stats['total_commission']:.2f}"
        ]

        row_bg = PatternFill("solid", fgColor="F8FAFC" if (n_idx % 2 == 0) else "FFFFFF")

        for ci, val in enumerate(s_vals, 1):
            cell = ws1.cell(r_sum, ci, val)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _BORDER
            cell.fill = row_bg

            if ci == 1:          # NO
                cell.font = font_data
            elif ci == 2:        # POST OFFICE
                cell.font = font_po
            elif ci == 3:        # NEED DELIVER
                cell.font = font_data
            elif ci == 4:        # DELIVERED (410)
                cell.font = font_delivered
            elif ci == 5:        # < 2 HOURS (+50%)
                cell.font = font_u2h if isinstance(val, (int, float)) and val > 0 else font_data
            elif ci == 6:        # 2-4 HOURS (+25%)
                cell.font = font_2_4h if isinstance(val, (int, float)) and val > 0 else font_data
            elif ci == 7:        # 4-8 HOURS (NORMAL)
                cell.font = font_4_8h
            elif ci == 8:        # > 8 HOURS (-25%)
                cell.font = font_o8h
            elif ci == 9:        # % < 8 HOUR
                if tot_b > 0 and pct_under_8h >= 85.0:
                    cell.font = font_pct_good
                elif tot_b > 0 and pct_under_8h < 85.0:
                    cell.font = font_pct_bad
                else:
                    cell.font = font_data
            elif ci == 10:       # % > 8 HOUR
                if tot_b > 0 and stats["over_8h"] > 0:
                    cell.font = font_pct_bad
                else:
                    cell.font = font_data
            elif ci == 11:       # COMMISSION ($)
                if stats["total_commission"] > 0:
                    cell.font = font_comm
                else:
                    cell.font = font_data

        r_sum += 1
        n_idx += 1

    # Executive Summary Grand Total (Cols A to K) - Clean Unified Accounting Finish
    _TOT_SUM_FILL = PatternFill("solid", fgColor="E2E8F0") # Soft slate accounting row
    ws1.row_dimensions[r_sum].height = 24.0
    ws1.merge_cells(start_row=r_sum, start_column=1, end_row=r_sum, end_column=2)
    tot_lbl_r = ws1.cell(r_sum, 1, "Grand Total")
    tot_lbl_r.font = Font(name="Segoe UI", size=9.5, bold=True, color="0F172A")
    tot_lbl_r.alignment = Alignment(horizontal="center", vertical="center")
    ws1.cell(r_sum, 1).fill = _TOT_SUM_FILL
    ws1.cell(r_sum, 1).border = _DOUBLE_BOT
    ws1.cell(r_sum, 2).fill = _TOT_SUM_FILL
    ws1.cell(r_sum, 2).border = _DOUBLE_BOT

    tot_total_orders = tot_need + tot_del
    on_time_tot = tot_u2 + tot_24 + tot_48
    
    # Fixed Grand Total percentage calculation
    tot_pct_u8_num = (on_time_tot / tot_total_orders * 100.0) if tot_total_orders > 0 else 0.0
    tot_pct_u8 = f"{tot_pct_u8_num:.1f}%"
    # Include pending orders in >8h percentage (all pending are considered "slow")
    tot_pct_o8_num = ((tot_o8 + tot_need) / tot_total_orders * 100.0) if tot_total_orders > 0 else 0.0
    tot_pct_o8 = f"{tot_pct_o8_num:.1f}%"

    t_vals = [
        tot_need, tot_del, tot_u2, tot_24, tot_48, tot_o8,
        tot_pct_u8, tot_pct_o8, f"${tot_pay:.2f}"
    ]
    for ci, val in enumerate(t_vals, 3):
        cell = ws1.cell(r_sum, ci, val)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = _TOT_SUM_FILL
        cell.border = _DOUBLE_BOT

        if ci == 4:   # DELIVERED
            cell.font = Font(name="Segoe UI", size=9.5, bold=True, color="2563EB")
        elif ci == 8: # > 8 HOURS
            cell.font = Font(name="Segoe UI", size=9.5, bold=True, color="DC2626" if tot_o8 > 0 else "0F172A")
        elif ci == 11:# COMMISSION
            cell.font = Font(name="Segoe UI", size=9.5, bold=True, color="166534")
        else:
            cell.font = Font(name="Segoe UI", size=9.5, bold=True, color="0F172A")

    # Populate Detail Rows (Cols M to U: 13 to 21)
    r_curr = 3
    tot_pay_left = 0.0

    for idx, item in enumerate(base_rows, 1):
        ws1.row_dimensions[r_curr].height = 19.0
        tot_pay_left += item["rate_usd"]
        row_vals = [
            idx,
            item["order_number"],
            item["customer"],
            item["vas_service"],
            item["assigned_branch"],
            item["duration"],
            item["duration_hours"],
            item["tier"],
            f"${item['rate_usd']:.2f}"
        ]
        for ci, val in enumerate(row_vals, 13):
            cell = ws1.cell(r_curr, ci, val)
            cell.border = _BORDER
            if ci in (13, 14, 16, 17, 18, 19):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif ci == 15:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")

            if ci == 14:
                cell.font = font_bold
            elif ci == 16:
                cell.font = font_vtt
            elif ci in (20, 21):
                cell.font = font_green if item["tag_color"] == "GREEN" else (font_red if item["tag_color"] == "RED" else font_data)
            else:
                cell.font = font_data
        r_curr += 1

    # Detail Grand Total
    if base_rows:
        ws1.row_dimensions[r_curr].height = 22.0
        ws1.merge_cells(start_row=r_curr, start_column=13, end_row=r_curr, end_column=20)
        tot_lbl = ws1.cell(r_curr, 13, "Grand Total")
        tot_lbl.font = font_tot
        tot_lbl.alignment = Alignment(horizontal="left", vertical="center")
        for ci in range(13, 21):
            ws1.cell(r_curr, ci).fill = _TOT_FILL
            ws1.cell(r_curr, ci).border = _DOUBLE_BOT

        tot_val = ws1.cell(r_curr, 21, f"${tot_pay_left:.2f}")
        tot_val.font = font_tot_comm
        tot_val.fill = _TOT_FILL
        tot_val.border = _DOUBLE_BOT
        tot_val.alignment = Alignment(horizontal="center", vertical="center")
        r_curr += 1

    # Auto-Fit Column Widths for Sheet 1
    col_widths_s1 = {
        1: 5,   # Summary: NO
        2: 13,  # Summary: BRANCH
        3: 12,  # Summary: PENDING
        4: 13,  # Summary: DELIVERED
        5: 14,  # Summary: < 2h (+50%)
        6: 14,  # Summary: 2-4h (+25%)
        7: 14,  # Summary: 4-8h (Normal)
        8: 13,  # Summary: > 8h (-25%)
        9: 11,  # Summary: % < 8h
        10: 11, # Summary: % > 8h
        11: 15, # Summary: COMMISSION
        12: 4,  # Spacer Gap
        13: 6,  # Detail: No
        14: 16, # Detail: Order Number
        15: 25, # Detail: Customer
        16: 14, # Detail: VAS Service
        17: 14, # Detail: Post Office
        18: 12, # Detail: Duration
        19: 10, # Detail: Hours
        20: 24, # Detail: Speed Tier
        21: 15  # Detail: Commission ($)
    }
    for c, w in col_widths_s1.items():
        ws1.column_dimensions[get_column_letter(c)].width = w

    # Sheet 2: base
    ws2 = wb.create_sheet(title="base")
    ws2.views.sheetView[0].showGridLines = True

    base_headers = [
        "No", "Order Number", "Customer", "VAS Service", "Origin Branch", "Origin Post",
        "Destination Branch", "Destination Post", "Assigned Branch", "Raw Service",
        "Shipper (Action User)", "Created At",
        "Branch Arrival (306/400)", "Delivered (410)", "Duration", "Hours (Dec)",
        "Speed Category", "Rate ($/Bill)"
    ]

    ws2.row_dimensions[1].height = 22
    for col_idx, h in enumerate(base_headers, 1):
        c = ws2.cell(row=1, column=col_idx, value=h)
        c.font = font_hdr
        c.fill = _HDR_LEFT
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _BORDER

    for idx, item in enumerate(base_rows, 1):
        r_num = idx + 1
        ws2.row_dimensions[r_num].height = 18
        row_data = [
            idx,
            item["order_number"],
            item["customer"],
            item["vas_service"],
            item["origin_branch"],
            item["origin_post"],
            item["destination_branch"],
            item["destination_post"],
            item["assigned_branch"],
            item["service"],
            item["shipper"],
            item["created_at"],
            item["t_start"],
            item["t410"],
            item["duration"],
            item["duration_hours"],
            item["tier"],
            item["rate_usd"]
        ]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws2.cell(row=r_num, column=col_idx, value=val)
            cell.font = font_data
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _BORDER

    for col in ws2.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws2.column_dimensions[col_letter].width = max(max_len + 3, 11)

    wb.save(out_xlsx)
    return tot_del, tot_u2, tot_24, tot_o8, tot_pay


def render_speed_summary_image(out_xlsx):
    """
    Renders the Right Table summary into a clean executive summary image.
    """
    wb = openpyxl.load_workbook(out_xlsx)
    ws = wb['DELIVERY SPEED REPORT']

    wb_sum = openpyxl.Workbook()
    ws_sum = wb_sum.active
    ws_sum.title = 'Executive Summary'
    ws_sum.views.sheetView[0].showGridLines = True

    max_r = 1
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, 1).value is not None or ws.cell(r, 11).value is not None:
            max_r = r

    for r in range(1, max_r + 1):
        if ws.row_dimensions[r].height:
            ws_sum.row_dimensions[r].height = ws.row_dimensions[r].height
        for c_idx in range(11):
            orig_c = 1 + c_idx
            tgt_c = 1 + c_idx
            cell_orig = ws.cell(r, orig_c)
            cell_tgt = ws_sum.cell(r, tgt_c, cell_orig.value)
            if cell_orig.has_style:
                cell_tgt.font = copy.copy(cell_orig.font)
                cell_tgt.fill = copy.copy(cell_orig.fill)
                cell_tgt.border = copy.copy(cell_orig.border)
                cell_tgt.alignment = copy.copy(cell_orig.alignment)

    col_widths = {1: 5, 2: 13, 3: 12, 4: 13, 5: 14, 6: 14, 7: 14, 8: 13, 9: 11, 10: 11, 11: 15}
    for c, w in col_widths.items():
        ws_sum.column_dimensions[get_column_letter(c)].width = w

    ws_sum.merge_cells("A1:K1")
    ws_sum.merge_cells(start_row=max_r, start_column=1, end_row=max_r, end_column=2)

    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_f:
        tmp_path = tmp_f.name
    wb_sum.save(tmp_path)

    try:
        buf = excel_to_image.excel_to_image(tmp_path)
        return buf
    finally:
        try:
            os.remove(tmp_path)
        except Exception:
            pass
