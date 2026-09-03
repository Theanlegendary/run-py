"""
pivot.py
Doc file Excel chi tiet don (export-detail) va dung bang pivot
"PENDING BILL CHECK" giong mau:
  - Loc: CURRENT STATUS thuoc danh sach pending_status_codes
         (+ loai test neu exclude_test)
  - Hang (rows): ZONE > CURRENT POST OFFICE > ORDER ID
  - Cot (columns): MONTH / DAY theo CREATED DATE
  - Gia tri: Count of ORDER ID
  - Co dong/cot Grand Total
"""

from datetime import datetime
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ---- vi tri cot trong file nguon (0-based) ----
COL_CREATED_DATE    = 1   # CREATED DATE  (dd/mm/yyyy HH:MM:SS)
COL_ORDER_ID        = 2   # ORDER ID
COL_DELIVERY_PROV   = 12  # DELIVERY PROVINCE (e.g. MON, BAT, PNP, SIE)
COL_CURRENT_PO      = 15  # CURRENT POST OFFICE
COL_CURRENT_STATUS  = 23  # CURRENT STATUS  ("110 - Chua tiep nhan")
COL_SENDER          = 3
COL_RECEIVER        = 4
COL_ACTION_PO_HUB   = 35  # ACTION POST OFFICE at ORIGIN HUB (col 35)
COL_TOTAL_FEE       = 19  # TOTAL FEE (USD)
COL_COD             = 20  # COD (USD)


def _status_code(value):
    """Lay ma so dau chuoi trang thai: '110 - Chua tiep nhan' -> '110'."""
    if value is None:
        return ""
    s = str(value).strip()
    if " - " in s:
        s = s.split(" - ", 1)[0]
    return s.split()[0].strip() if s else ""


def _parse_day(value):
    """Tra ve (month, day) tu CREATED DATE. Ho tro dd/mm/yyyy ..."""
    if value is None:
        return None, None
    if isinstance(value, datetime):
        return value.month, value.day
    s = str(value).strip()
    date_part = s.split(" ")[0]
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y", "%d-%m-%Y"):
        try:
            dt = datetime.strptime(date_part, fmt)
            return dt.month, dt.day
        except ValueError:
            continue
    return None, None


def _zone_for(po_code, zone_cfg):
    if not po_code:
        return zone_cfg.get("default_zone", "Khac")
    po = str(po_code).strip()
    by_po = zone_cfg.get("by_post_office", {})
    if po in by_po:
        return by_po[po]
    by_prefix = zone_cfg.get("by_prefix", {})
    for prefix, zone in by_prefix.items():
        if po.upper().startswith(prefix.upper()):
            return zone
    return zone_cfg.get("default_zone", "Khac")


def _is_test_row(row, test_keywords):
    blob = " ".join(str(row[c] or "") for c in (COL_SENDER, COL_RECEIVER)).lower()
    return any(k.lower() in blob for k in test_keywords)


def read_source(path):
    """Doc file Excel nguon -> list rows (tuple), bo qua header."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = []
    for r in ws.iter_rows(values_only=True):
        rows.append(r)
    return rows[1:] if rows else []


def build_pivot(rows, pivot_cfg, zone_cfg):
    """
    Builds a standard pivot tree:
      zone -> po -> {order_id: (month, day)}
    """
    pending = set(str(c).strip() for c in pivot_cfg.get("pending_status_codes", []))
    exclude_test = pivot_cfg.get("exclude_test", False)
    test_keywords = pivot_cfg.get("test_keywords", ["test"])

    tree = defaultdict(lambda: defaultdict(dict))
    today = datetime.now().date()
    day_keys_seen = {(today.month, today.day)}

    for row in rows:
        if not row or row[COL_ORDER_ID] in (None, ""):
            continue
        if pending and _status_code(row[COL_CURRENT_STATUS]) not in pending:
            continue
        if exclude_test and _is_test_row(row, test_keywords):
            continue

        month, day = _parse_day(row[COL_CREATED_DATE])
        if day is None:
            continue
        po = str(row[COL_CURRENT_PO] or "").strip() or "(trong)"
        zone = _zone_for(po, zone_cfg)
        order_id = str(row[COL_ORDER_ID]).strip()

        key = (month, day)
        tree[zone][po][order_id] = key
        day_keys_seen.add(key)

    day_keys = sorted(day_keys_seen)
    month_val = day_keys[0][0] if day_keys else None
    return tree, day_keys, month_val


# ---- styling matching /push zone ----
_HDR_FILL  = PatternFill("solid", fgColor="1E293B")   # Dark Slate Header
_ZONE_FILL = PatternFill("solid", fgColor="EAEAEA")   # Grey Zone/Hub Header
_TOT_FILL  = PatternFill("solid", fgColor="F1F5F9")   # Soft Grey Total
_THIN   = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT   = Alignment(horizontal="left", vertical="center")
_RIGHT  = Alignment(horizontal="right", vertical="center")


def export_pivot(tree, day_keys, month_val, pivot_cfg, out_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = pivot_cfg.get("is_label", "Pivot Report")

    fn = "Segoe UI"
    hdr_font = Font(name=fn, size=10, bold=True, color="FFFFFF")
    zone_font = Font(name=fn, size=10, bold=True, color="0F172A")
    data_font = Font(name=fn, size=10, color="0F172A")
    tot_font = Font(name=fn, size=10, bold=True, color="EF4444")
    tot_fill = PatternFill("solid", fgColor="F1F5F9")

    # Header Row 1
    ws.cell(1, 1, "Zone / Post Office").font = hdr_font
    ws.cell(1, 1).fill = _HDR_FILL
    ws.cell(1, 1).border = _BORDER
    ws.cell(1, 1).alignment = _LEFT

    col_idx_map = {}
    for idx, dk in enumerate(day_keys):
        col_num = 2 + idx
        col_idx_map[dk] = col_num
        cell = ws.cell(1, col_num, f"{dk[1]:02d}")
        cell.font = hdr_font
        cell.fill = _HDR_FILL
        cell.border = _BORDER
        cell.alignment = _CENTER

    tot_col = 2 + len(day_keys)
    ws.cell(1, tot_col, "Grand Total").font = hdr_font
    ws.cell(1, tot_col).fill = _HDR_FILL
    ws.cell(1, tot_col).border = _BORDER
    ws.cell(1, tot_col).alignment = _CENTER

    ws.row_dimensions[1].height = 24

    r = 2
    grand_col_totals = defaultdict(int)
    grand_overall = 0

    for zone in sorted(tree.keys()):
        # Zone Header Row
        ws.cell(r, 1, zone).font = zone_font
        ws.cell(r, 1).fill = _ZONE_FILL
        ws.cell(r, 1).border = _BORDER
        for c in range(2, tot_col + 1):
            cell = ws.cell(r, c)
            cell.fill = _ZONE_FILL
            cell.border = _BORDER
        ws.row_dimensions[r].height = 20
        r += 1

        zone_col_totals = defaultdict(int)
        zone_total = 0

        for po in sorted(tree[zone].keys()):
            ws.cell(r, 1, f"   {po}").font = data_font
            ws.cell(r, 1).border = _BORDER
            ws.cell(r, 1).alignment = _LEFT

            po_sum = 0
            for oid, dk in tree[zone][po].items():
                if dk in col_idx_map:
                    cn = col_idx_map[dk]
                    cur = ws.cell(r, cn).value or 0
                    ws.cell(r, cn, cur + 1)
                    zone_col_totals[dk] += 1
                    grand_col_totals[dk] += 1
                    po_sum += 1

            for c in range(2, tot_col + 1):
                cell = ws.cell(r, c)
                cell.border = _BORDER
                cell.alignment = _CENTER
                cell.font = data_font

            ws.cell(r, tot_col, po_sum).font = data_font
            ws.cell(r, tot_col).border = _BORDER
            ws.cell(r, tot_col).alignment = _CENTER

            zone_total += po_sum
            ws.row_dimensions[r].height = 20
            r += 1

        grand_overall += zone_total

    # Grand Total Row
    ws.cell(r, 1, "Grand Total").font = tot_font
    ws.cell(r, 1).fill = tot_fill
    ws.cell(r, 1).border = _BORDER
    ws.cell(r, 1).alignment = _LEFT

    for dk, cn in col_idx_map.items():
        v = grand_col_totals.get(dk, 0)
        c = ws.cell(r, cn, v if v > 0 else "")
        c.font = tot_font
        c.fill = tot_fill
        c.border = _BORDER
        c.alignment = _CENTER

    c_tot = ws.cell(r, tot_col, grand_overall)
    c_tot.font = tot_font
    c_tot.fill = tot_fill
    c_tot.border = _BORDER
    c_tot.alignment = _CENTER

    ws.row_dimensions[r].height = 22

    # Column widths
    ws.column_dimensions["A"].width = 28
    for c in range(2, tot_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 7

    wb.save(out_path)
    return out_path, grand_overall


def _merge_pivot_cfg(global_pivot, report):
    cfg = dict(global_pivot)
    if "pending_status_codes" in report:
        cfg["pending_status_codes"] = report["pending_status_codes"]
    cfg["is_label"] = report.get("is_label", report.get("title", "Pending"))
    return cfg


def run_report(rows, out_path, config, report):
    """Dung 1 report tu rows da doc san. Tra ve (out_path, total)."""
    pivot_cfg = _merge_pivot_cfg(config.get("pivot", {}), report)
    tree, days, month = build_pivot(rows, pivot_cfg, config["zone_mapping"])
    return export_pivot(tree, days, month, pivot_cfg, out_path)


def build_mega_pivot(rows, pivot_cfg, zone_cfg):
    """
    Builds a 2-level pivot tree for MEGA check matching Metfone OPS format:
      Level 1: CURRENT POST OFFICE (MEGA1 / DVCMEGA1)
      Level 2: DELIVERY PROVINCE (MON, BAT, PNP, SIE, etc.)
    Tracks: order counts, total fee, total cod
    """
    exclude_test     = pivot_cfg.get("exclude_test", False)
    test_keywords    = pivot_cfg.get("test_keywords", ["test"])
    exclude_statuses = {"410", "201", "520", "99", "100", "-99"}

    tree = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
    fee_tree = defaultdict(lambda: defaultdict(float))
    cod_tree = defaultdict(lambda: defaultdict(float))
    urgent_tree = defaultdict(lambda: defaultdict(int))
    day_keys_seen = set()
    today = datetime.now().date()
    day_keys_seen.add((today.month, today.day))

    for row in rows:
        if not row or len(row) <= COL_CURRENT_PO:
            continue
        if row[COL_ORDER_ID] in (None, ""):
            continue
        status_code = _status_code(row[COL_CURRENT_STATUS])
        if status_code in exclude_statuses:
            continue
        if exclude_test and _is_test_row(row, test_keywords):
            continue

        po = str(row[COL_CURRENT_PO] or "").strip().upper()
        col35 = str(row[COL_ACTION_PO_HUB] if len(row) > COL_ACTION_PO_HUB and row[COL_ACTION_PO_HUB] else "").strip().upper()

        # Filter matching Metfone Web Order Status Report (Select Branch = MEGA HUB)
        # CURRENT POST OFFICE (Col 15) must be physically at the Hub!
        hub_statuses = {"306", "309", "302", "311", "310"}
        if status_code not in hub_statuses:
            continue

        if po == "MEGA1":
            hub_label = "MEGA1"
        elif "DVC" in po or "MEGA" in po or "HUB" in po:
            hub_label = "DVCMEGA1"
        else:
            continue






        prov = str(row[COL_DELIVERY_PROV] if len(row) > COL_DELIVERY_PROV and row[COL_DELIVERY_PROV] else "").strip().upper() or "KHAC"

        # Use Action Time (Col 24 - CURRENT TIME) matching boss Action day pivot
        action_time_val = row[24] if len(row) > 24 and row[24] else row[COL_CREATED_DATE]
        month, day = _parse_day(action_time_val)
        if day is None:
            month, day = _parse_day(row[COL_CREATED_DATE])
        if day is None:
            continue


        key = (month, day)
        tree[hub_label][prov][key] += 1
        day_keys_seen.add(key)

        try:
            fee_tree[hub_label][prov] += float(row[COL_TOTAL_FEE] or 0)
        except (ValueError, TypeError):
            pass
        try:
            cod_tree[hub_label][prov] += float(row[COL_COD] or 0)
        except (ValueError, TypeError):
            pass

        act_val = row[24] if len(row) > 24 and row[24] else row[COL_CREATED_DATE]
        act_date = None
        if isinstance(act_val, datetime):
            act_date = act_val.date()
        elif act_val:
            s = str(act_val).strip().split(" ")[0]
            for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y", "%d-%m-%Y"):
                try:
                    act_date = datetime.strptime(s, fmt).date()
                    break
                except ValueError:
                    continue
        if not act_date and row[COL_CREATED_DATE]:
            c_val = row[COL_CREATED_DATE]
            if isinstance(c_val, datetime):
                act_date = c_val.date()
            elif c_val:
                s = str(c_val).strip().split(" ")[0]
                for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y", "%d-%m-%Y"):
                    try:
                        act_date = datetime.strptime(s, fmt).date()
                        break
                    except ValueError:
                        continue

        if act_date and (today - act_date).days >= 1:
            urgent_tree[hub_label][prov] += 1

    day_keys = sorted(day_keys_seen)
    return tree, day_keys, (fee_tree, cod_tree, urgent_tree)



def export_mega_pivot(tree, day_keys, out_path, extra_data=None):
    """
    Renders an Executive CEO-level Mega Hub Dashboard:
      - Premium Dark Obsidian (#0F172A) Top Banner & Grand Total Footer
      - Sleek Slate Header Bar (#1E293B)
      - Smart Date Handling with crisp typography
      - Emerald Green financial metrics (Fee/COD)
      - Coral Red Urgent indicators
      - High-contrast executive summary layout
    """
    fee_tree    = extra_data[0] if extra_data and len(extra_data) > 0 else defaultdict(lambda: defaultdict(float))
    cod_tree    = extra_data[1] if extra_data and len(extra_data) > 1 else defaultdict(lambda: defaultdict(float))
    urgent_tree = extra_data[2] if extra_data and len(extra_data) > 2 else defaultdict(lambda: defaultdict(int))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "EXECUTIVE MEGA PIVOT"

    fn = "Segoe UI"
    banner_font = Font(name=fn, size=11, bold=True, color="FFFFFF")
    hdr_font    = Font(name=fn, size=10, bold=True, color="FFFFFF")
    handle_font = Font(name=fn, size=10, bold=True, color="0F172A")
    data_font   = Font(name=fn, size=10, color="0F172A")
    blue_font   = Font(name=fn, size=10, color="1E40AF", bold=True)
    fee_font    = Font(name=fn, size=10, color="047857", bold=True)
    urg_font    = Font(name=fn, size=10, color="DC2626", bold=True)
    tot_font    = Font(name=fn, size=10, bold=True, color="0F172A")
    gt_font     = Font(name=fn, size=10, bold=True, color="FFFFFF")
    gt_urg_font = Font(name=fn, size=10, bold=True, color="FCA5A5")

    # Executive Palette Fills
    banner_fill   = PatternFill("solid", fgColor="0F172A") # Dark Obsidian
    hdr_slate     = PatternFill("solid", fgColor="1E293B") # Slate Header
    tot_hdr_fill  = PatternFill("solid", fgColor="1E40AF") # Royal Blue
    fee_hdr_fill  = PatternFill("solid", fgColor="065F46") # Deep Emerald
    urg_hdr_fill  = PatternFill("solid", fgColor="991B1B") # Deep Crimson

    fee_cell_fill = PatternFill("solid", fgColor="ECFDF5") # Soft Emerald Tint
    urg_cell_fill = PatternFill("solid", fgColor="FEE2E2") # Soft Coral Tint
    tot_cell_fill = PatternFill("solid", fgColor="EFF6FF") # Soft Blue Tint
    gt_row_fill   = PatternFill("solid", fgColor="0F172A") # Executive Dark Footer

    # Subtle Border
    thin_border = Side(style="thin", color="CBD5E1")
    cell_border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)

    # Detect current hub label & status code
    hub_keys = [h for h in ["MEGA1", "DVCMEGA1"] if h in tree and tree[h]]
    if len(hub_keys) == 1:
        h_name = hub_keys[0]
        status_info = "(Status 309, 306)" if h_name == "MEGA1" else "(Status 306)"
        hub_title = f"{h_name} {status_info}"
    else:
        hub_title = "MEGA (Status 306/309)"

    # Filter day_keys to only dates that have orders in this tree (eliminates 25 empty columns)
    active_days = [
        dk for dk in day_keys
        if any(tree[h][p].get(dk, 0) > 0 for h in tree for p in tree[h])
    ]
    if active_days:
        day_keys = active_days

    tot_col = 2 + len(day_keys)
    fee_col = tot_col + 1
    cod_col = fee_col + 1
    urg_col = cod_col + 1

    # 1. Top Banner Row 1
    stamp_str = datetime.now().strftime("%d/%m/%Y %H:%M")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=urg_col)
    top_cell = ws.cell(1, 1, f"📊 EXECUTIVE MEGA HUB REPORT — {hub_title}  |  {stamp_str}")
    top_cell.font = banner_font
    top_cell.fill = banner_fill
    top_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 28

    # 2. Row 2: Headers
    h_cell = ws.cell(2, 1, "PROVINCE / HANDLE")
    h_cell.font = hdr_font
    h_cell.fill = hdr_slate
    h_cell.border = cell_border
    h_cell.alignment = _CENTER

    col_idx_map = {}
    for idx, dk in enumerate(day_keys):
        col_num = 2 + idx
        col_idx_map[dk] = col_num
        cell = ws.cell(2, col_num, f"{dk[1]:02d}/{dk[0]:02d}")
        cell.font = hdr_font
        cell.fill = hdr_slate
        cell.border = cell_border
        cell.alignment = _CENTER

    c_tot_hdr = ws.cell(2, tot_col, "TOTAL")
    c_tot_hdr.font = hdr_font
    c_tot_hdr.fill = tot_hdr_fill
    c_tot_hdr.border = cell_border
    c_tot_hdr.alignment = _CENTER

    c_fee_hdr = ws.cell(2, fee_col, "Fee ($)")
    c_fee_hdr.font = hdr_font
    c_fee_hdr.fill = fee_hdr_fill
    c_fee_hdr.border = cell_border
    c_fee_hdr.alignment = _CENTER

    c_cod_hdr = ws.cell(2, cod_col, "COD ($)")
    c_cod_hdr.font = hdr_font
    c_cod_hdr.fill = fee_hdr_fill
    c_cod_hdr.border = cell_border
    c_cod_hdr.alignment = _CENTER

    c_urg_hdr = ws.cell(2, urg_col, "URGENT")
    c_urg_hdr.font = hdr_font
    c_urg_hdr.fill = urg_hdr_fill
    c_urg_hdr.border = cell_border
    c_urg_hdr.alignment = _CENTER

    for c in range(1, urg_col + 1):
        ws.cell(1, c).fill = banner_fill

    ws.row_dimensions[2].height = 26

    r = 3
    grand_col_totals = defaultdict(int)
    grand_overall_orders = 0
    grand_overall_fee = 0.0
    grand_overall_cod = 0.0
    grand_overall_urgent = 0

    row_bg_even = PatternFill("solid", fgColor="F8FAFC") # Crisp Ice Tint
    row_bg_odd  = PatternFill("solid", fgColor="FFFFFF") # Pure White

    for hub in hub_keys:
        prov_dict = tree.get(hub, {})
        if not prov_dict:
            continue

        hub_col_totals = defaultdict(int)
        hub_total_orders = 0
        hub_total_fee = 0.0
        hub_total_cod = 0.0
        hub_total_urgent = 0

        # Province Rows with Alternating Striping
        for prov_idx, prov in enumerate(sorted(prov_dict.keys())):
            row_bg = row_bg_even if prov_idx % 2 == 0 else row_bg_odd
            ws.row_dimensions[r].height = 22

            c_prov = ws.cell(r, 1, prov)
            c_prov.font = handle_font
            c_prov.fill = row_bg
            c_prov.border = cell_border
            c_prov.alignment = _CENTER

            row_sum = 0
            for idx, dk in enumerate(day_keys):
                col_num = col_idx_map[dk]
                val = prov_dict[prov].get(dk, 0)
                cell = ws.cell(r, col_num)
                cell.fill = row_bg
                cell.border = cell_border
                cell.font = blue_font if val > 0 else data_font
                cell.alignment = _CENTER
                if val > 0:
                    cell.value = val
                    row_sum += val
                    hub_col_totals[dk] += val
                    grand_col_totals[dk] += val

            p_fee = fee_tree[hub].get(prov, 0.0)
            p_cod = cod_tree[hub].get(prov, 0.0)
            p_urg = urgent_tree[hub].get(prov, 0)

            # TOTAL cell
            t_cell = ws.cell(r, tot_col, row_sum if row_sum > 0 else "")
            t_cell.font = tot_font
            t_cell.fill = tot_cell_fill if row_sum > 0 else row_bg
            t_cell.alignment = _CENTER
            t_cell.border = cell_border

            # Fee($) cell
            f_cell = ws.cell(r, fee_col, round(p_fee, 2) if p_fee > 0 else "")
            f_cell.font = fee_font
            f_cell.fill = fee_cell_fill if p_fee > 0 else row_bg
            f_cell.alignment = _CENTER
            f_cell.border = cell_border
            if p_fee > 0:
                f_cell.number_format = "$#,##0.00"

            # COD($) cell
            c_cell = ws.cell(r, cod_col, round(p_cod, 2) if p_cod > 0 else "")
            c_cell.font = fee_font
            c_cell.fill = fee_cell_fill if p_cod > 0 else row_bg
            c_cell.alignment = _CENTER
            c_cell.border = cell_border
            if p_cod > 0:
                c_cell.number_format = "$#,##0.00"

            # URGENT cell
            u_cell = ws.cell(r, urg_col, p_urg if p_urg > 0 else "")
            u_cell.font = urg_font
            u_cell.fill = urg_cell_fill if p_urg > 0 else row_bg
            u_cell.alignment = _CENTER
            u_cell.border = cell_border

            hub_total_orders += row_sum
            hub_total_fee += p_fee
            hub_total_cod += p_cod
            hub_total_urgent += p_urg
            r += 1

        grand_overall_orders += hub_total_orders
        grand_overall_fee += hub_total_fee
        grand_overall_cod += hub_total_cod
        grand_overall_urgent += hub_total_urgent

    # GRAND TOTAL Row (Executive Dark Obsidian Bar)
    ws.row_dimensions[r].height = 26
    gt_title = ws.cell(r, 1, "GRAND TOTAL")
    gt_title.font = gt_font
    gt_title.fill = gt_row_fill
    gt_title.border = cell_border
    gt_title.alignment = _CENTER

    for idx, dk in enumerate(day_keys):
        col_num = col_idx_map[dk]
        val = grand_col_totals.get(dk, 0)
        cell = ws.cell(r, col_num)
        cell.fill = gt_row_fill
        cell.font = gt_font
        cell.border = cell_border
        cell.alignment = _CENTER
        if val > 0:
            cell.value = val

    gt_c = ws.cell(r, tot_col, grand_overall_orders)
    gt_c.fill = tot_hdr_fill
    gt_c.font = Font(name=fn, size=10, bold=True, color="FFFFFF")
    gt_c.border = cell_border
    gt_c.alignment = _CENTER

    gf_c = ws.cell(r, fee_col, round(grand_overall_fee, 2))
    gf_c.fill = fee_hdr_fill
    gf_c.font = Font(name=fn, size=10, bold=True, color="FFFFFF")
    gf_c.alignment = _CENTER
    gf_c.border = cell_border
    gf_c.number_format = "$#,##0.00"

    gc_c = ws.cell(r, cod_col, round(grand_overall_cod, 2))
    gc_c.fill = fee_hdr_fill
    gc_c.font = Font(name=fn, size=10, bold=True, color="FFFFFF")
    gc_c.alignment = _CENTER
    gc_c.border = cell_border
    gc_c.number_format = "$#,##0.00"

    gu_c = ws.cell(r, urg_col, grand_overall_urgent)
    gu_c.fill = urg_hdr_fill
    gu_c.font = Font(name=fn, size=10, bold=True, color="FFFFFF")
    gu_c.alignment = _CENTER
    gu_c.border = cell_border

    # Column Widths
    ws.column_dimensions["A"].width = 22
    for c in range(2, tot_col):
        ws.column_dimensions[get_column_letter(c)].width = 8
    ws.column_dimensions[get_column_letter(tot_col)].width = 11
    ws.column_dimensions[get_column_letter(fee_col)].width = 15
    ws.column_dimensions[get_column_letter(cod_col)].width = 15
    ws.column_dimensions[get_column_letter(urg_col)].width = 11

    wb.save(out_path)
    return out_path, grand_overall_orders



def run_mega(source_path, out_path, config):
    rows = read_source(source_path)
    tree, day_keys, extra_data = build_mega_pivot(rows, config.get("pivot", {}), config.get("zone_mapping", {}))
    return export_mega_pivot(tree, day_keys, out_path, extra_data=extra_data)


def run_mega_combined(source_path, out_path, config):
    """Build combined Excel."""
    return run_mega(source_path, out_path, config)
