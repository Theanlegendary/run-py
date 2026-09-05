"""
excel_to_image.py
Renders an openpyxl-formatted Excel sheet to a PNG image using Pillow.
- Day columns: fixed uniform width
- Hidden/empty columns: skipped entirely
- Merged cells: top-left value shown, others blank
"""

import io
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont

# ── Constants ──────────────────────────────────────────────────────────────────
SCALE       = 3             # Increased from 2 to 3 for HD resolution quality

FONT_SIZE   = 10 * SCALE
ROW_H       = 22 * SCALE    # px — uniform row height
PAD_X       = 8 * SCALE     # horizontal text padding

# Fixed pixel widths per column type
PX_DAY      = 34 * SCALE    # day columns "01"-"31"  — all same size
PX_ZONE     = 72 * SCALE    # ZONE
PX_GT       = 140 * SCALE   # Grand Total
PX_MIN      = 72 * SCALE    # minimum for auto-fit text columns (ORDER ID, names etc.)
PX_MAX      = 230 * SCALE   # maximum for auto-fit
PX_GAP      = 6 * SCALE     # gap/empty separator columns

BG_WHITE    = (255, 255, 255)
BORDER_COL  = (190, 190, 190)
TEXT_DEF    = (0,   0,   0)


def _hex(h):
    h = h.lstrip('#')
    if len(h) == 6:
        return tuple(int(h[i:i+2], 16) for i in (0, 2, 4))
    return TEXT_DEF


def _cell_bg(cell):
    try:
        f = cell.fill
        if f and f.fill_type == 'solid':
            fc = f.fgColor
            if fc.type == 'rgb' and fc.rgb not in ('00000000', 'FFFFFFFF', 'FF000000'):
                return _hex(fc.rgb[-6:])
    except Exception:
        pass
    return None


def _cell_fg(cell):
    try:
        ft = cell.font
        if ft and ft.color and ft.color.type == 'rgb':
            if ft.color.rgb not in ('00000000', 'FF000000'):
                return _hex(ft.color.rgb[-6:])
    except Exception:
        pass
    return TEXT_DEF


def _cell_bold(cell):
    try:
        return bool(cell.font and cell.font.bold)
    except Exception:
        return False


def _cell_align(cell):
    try:
        a = cell.alignment
        if a and a.horizontal in ('center', 'right'):
            return a.horizontal
    except Exception:
        pass
    return 'left'


_WIN_FONTS = "C:/Windows/Fonts"

def _has_khmer(text: str) -> bool:
    """Return True if text contains any Khmer Unicode characters (U+1780–U+17FF)."""
    return any('\u1780' <= ch <= '\u17FF' for ch in text)

def _load_font(size, bold=False):
    # Try system paths first
    for name in (['arialbd.ttf', 'Arial Bold.ttf', 'DejaVuSans-Bold.ttf'] if bold
                 else ['arial.ttf', 'Arial.ttf', 'DejaVuSans.ttf']):
        for prefix in (f"{_WIN_FONTS}/", ""):
            try:
                return ImageFont.truetype(prefix + name, size)
            except Exception:
                pass
    return ImageFont.load_default()

def _load_khmer_font(size):
    """Load Khmer OS Battambang font for Khmer text."""
    for name in ['KhmerOSbattambang.ttf', 'KhmerOSsiemreap.ttf', 'KhmerOScontent.ttf', 'KhmerOS.ttf']:
        for prefix in (f"{_WIN_FONTS}/", ""):
            try:
                return ImageFont.truetype(prefix + name, size)
            except Exception:
                pass
    return _load_font(size, bold=False)

def _get_font(text: str, size: int, bold: bool = False):
    """Return Khmer font if text has Khmer chars, otherwise return normal font."""
    if _has_khmer(text):
        # Khmer fonts are slightly smaller, boost size dynamically based on scale
        return _load_khmer_font(size + int(1.5 * SCALE))
    return _load_font(size, bold)


def excel_to_image(xlsx_path: str) -> io.BytesIO:
    # ── Try Excel COM rendering first (for perfect Khmer text shaping and native styling on Windows) ──
    try:
        import win32com.client
        import time
        import os
        from PIL import ImageGrab
        
        abs_path = os.path.abspath(xlsx_path)
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        
        wb = None
        try:
            wb = excel.Workbooks.Open(abs_path)
            ws = wb.ActiveSheet
            
            img = None
            temp_png = None
            
            # --- 1. Try High Quality Chart Export Method (Crisp HD 2.0x rendering) ---
            try:
                rng = ws.UsedRange
                excel_scale = 2.0  # Scale factor for HD rendering
                chart_width = rng.Width * excel_scale
                chart_height = rng.Height * excel_scale
                
                # Copy as picture using xlScreen=1, xlPicture=-4147 for maximum vector detail
                rng.CopyPicture(1, -4147)
                
                # Add temporary chart
                chart_obj = ws.ChartObjects().Add(Left=rng.Left, Top=rng.Top, Width=chart_width, Height=chart_height)
                chart_obj.Activate()
                chart = chart_obj.Chart
                chart.Paste()
                
                # Scale the pasted picture shape to fill the chart and position it at top-left (0,0)
                if chart.Shapes.Count > 0:
                    shape = chart.Shapes(1)
                    shape.Left = 0
                    shape.Top = 0
                    shape.Width = chart_width
                    shape.Height = chart_height
                
                # Remove chart borders and fill to prevent extra margins/background padding
                chart.ChartArea.Format.Line.Visible = 0
                chart.ChartArea.Format.Fill.Visible = 0
                
                # Export to temp PNG
                temp_png = os.path.abspath(os.path.join(os.path.dirname(xlsx_path), f"temp_excel_hd_{int(time.time())}.png"))
                chart.Export(temp_png, "PNG")
                chart_obj.Delete()
                
                if os.path.exists(temp_png):
                    img = Image.open(temp_png)
                    # Load image fully into memory and then close file handle so we can delete it
                    img.load()
            except Exception:
                img = None
            finally:
                if temp_png and os.path.exists(temp_png):
                    try:
                        os.remove(temp_png)
                    except Exception:
                        pass
            
            # --- 2. Clipboard Fallback Method (if Chart method failed or wasn't used) ---
            if img is None:
                ws.UsedRange.CopyPicture(1, 2)
                time.sleep(0.5) # wait for clipboard
                for _ in range(5):
                    img = ImageGrab.grabclipboard()
                    if img:
                        break
                    time.sleep(0.5)
            
            if img:
                # Aspect ratio safety check for Telegram
                max_ratio = 18.0
                w, h = img.size
                new_w, new_h = w, h
                if h > 0 and w / h > max_ratio:
                    new_h = int(w / max_ratio)
                elif w > 0 and h / w > max_ratio:
                    new_w = int(h / max_ratio)

                if (new_w, new_h) != (w, h):
                    padded_img = Image.new('RGB', (new_w, new_h), (255, 255, 255))
                    padded_img.paste(img, (0, 0))
                    img = padded_img

                buf = io.BytesIO()
                img.save(buf, format='PNG', optimize=True)
                buf.seek(0)
                return buf
        finally:
            if wb:
                wb.Close(SaveChanges=False)
            excel.Quit()
    except Exception as e:
        # Fall back to Pillow rendering if Excel COM fails or is not available
        import logging
        logging.warning("Excel COM rendering failed, falling back to Pillow: %s", e, exc_info=True)
        pass

    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    
    # Speed & Memory protection: find actual populated boundaries from ws._cells
    populated_rows = [r for (r, c) in ws._cells.keys() if ws._cells[(r, c)].value is not None]
    last_row = max(populated_rows) if populated_rows else 1
    populated_cols = [c for (r, c) in ws._cells.keys() if ws._cells[(r, c)].value is not None]
    last_col = max(populated_cols) if populated_cols else 1
    
    max_row = min(ws.max_row or 1, last_row)
    max_col = min(ws.max_column or 1, last_col)
    
    if not max_row or not max_col:
        raise ValueError("Empty sheet")

    # ── 1. Build value grid (handle merged cells) ──────────────────────────────
    grid = [[''] * (max_col + 1) for _ in range(max_row + 1)]
    skip = set()
    for mc in ws.merged_cells.ranges:
        if mc.min_row > max_row or mc.min_col > max_col:
            continue
        v = ws.cell(mc.min_row, mc.min_col).value
        grid[mc.min_row][mc.min_col] = str(v) if v is not None else ''
        for r in range(mc.min_row, min(mc.max_row, max_row) + 1):
            for c in range(mc.min_col, min(mc.max_col, max_col) + 1):
                if not (r == mc.min_row and c == mc.min_col):
                    skip.add((r, c))

    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            if (r, c) in skip:
                continue
            try:
                v = ws.cell(r, c).value
            except Exception:
                v = None
            if grid[r][c] == '':
                grid[r][c] = str(v) if v is not None else ''

    # ── 2. Classify columns ────────────────────────────────────────────────────
    # Only scan the first 5 rows to find headers — avoids data rows like
    # "Grand Total" row (which has a number before it) being misclassified
    day_ci  = set()
    zone_ci = set()
    gt_ci   = set()

    HEADER_SCAN_ROWS = min(max_row, 8)
    for r in range(1, HEADER_SCAN_ROWS + 1):
        for c in range(1, max_col + 1):
            val = grid[r][c].strip()
            # Day col: exactly 2-digit 01-31 in a header row
            # Exclude if previous or next cell in same row has content (avoids data Grand Total rows)
            if val.isdigit() and len(val) == 2 and 1 <= int(val) <= 31:
                day_ci.add(c)
            elif val == 'ZONE':
                zone_ci.add(c)
            elif val == 'Grand Total':
                # Only count as GT column if it's a header row (check row has other headers)
                row_vals = [grid[r][cc].strip() for cc in range(1, max_col + 1)]
                if any(v in ('ZONE', 'ORDER ID', 'POST OFFICE HANDLE', 'CURRENT POST OFFICE')
                       for v in row_vals):
                    gt_ci.add(c)

    # Detect empty columns (gap separators between side-by-side tables)
    empty_ci = set(
        c for c in range(1, max_col + 1)
        if not any(grid[r][c].strip() for r in range(1, max_row + 1))
    )

    # ── 3. Compute per-column pixel widths ─────────────────────────────────────
    fn   = _load_font(FONT_SIZE, bold=False)
    fn_b = _load_font(FONT_SIZE, bold=True)
    tmp  = Image.new('RGB', (1, 1))
    d    = ImageDraw.Draw(tmp)

    col_px = []   # pixel width per column (0 = skip/hidden)
    for c in range(1, max_col + 1):
        if c in empty_ci:
            col_px.append(PX_GAP)   # tiny but visible separator
            continue

        # Check if column is hidden in Excel
        letter = get_column_letter(c)
        cd = ws.column_dimensions.get(letter)
        if cd and (cd.hidden or (cd.width is not None and cd.width < 0.5)):
            col_px.append(0)        # fully hidden — skip in image too
            continue

        if c in day_ci:
            col_px.append(PX_DAY)   # ALL day columns same fixed width
        elif c in zone_ci:
            col_px.append(PX_ZONE)
        elif c in gt_ci:
            col_px.append(PX_GT)
        else:
            # Auto-fit text columns
            max_w = PX_MIN
            for r in range(1, max_row + 1):
                text = grid[r][c]
                if not text:
                    continue
                bold = _cell_bold(ws.cell(r, c))
                f = _get_font(text, FONT_SIZE, bold)
                stroke_w = 0
                if bold and _has_khmer(text):
                    # Simulate bold for Khmer OS fonts using a stroke outline
                    stroke_w = max(1, int(0.4 * SCALE))
                try:
                    bb = d.textbbox((0, 0), text, font=f, stroke_width=stroke_w)
                    w  = bb[2] - bb[0]
                except Exception:
                    w = len(text) * (FONT_SIZE - 2) + stroke_w * 2
                max_w = max(max_w, w + PAD_X * 2)
            col_px.append(min(max_w, PX_MAX))

    # Columns to actually render (skip 0-width hidden ones)
    render_cols = [c for c in range(1, max_col + 1) if col_px[c - 1] > 0]

    # ── 3.5 Compute per-row heights ────────────────────────────────────────────
    # Rows that contain any Khmer text get ROW_H_KHMER, others get ROW_H
    ROW_H_KHMER = 34 * SCALE
    ROW_H = 22 * SCALE
    row_heights = []
    for r in range(1, max_row + 1):
        has_khmer_row = any(
            _has_khmer(grid[r][c]) for c in range(1, max_col + 1)
        )
        row_heights.append(ROW_H_KHMER if has_khmer_row else ROW_H)

    # ── 4. Build canvas ────────────────────────────────────────────────────────
    total_w = sum(col_px[c - 1] for c in render_cols) + 1
    total_h = sum(row_heights) + 1
    img  = Image.new('RGB', (total_w, total_h), BG_WHITE)
    draw = ImageDraw.Draw(img)

    # ── 4.5 Pre-calculate merged cell pixel dimensions ───────────────
    skip = set()
    merged_rects = {}
    for mc in ws.merged_cells.ranges:
        mw = 0
        for cc in range(mc.min_col, mc.max_col + 1):
            if cc - 1 < len(col_px) and col_px[cc - 1] > 0:
                mw += col_px[cc - 1]
        mh = 0
        for rr in range(mc.min_row, mc.max_row + 1):
            if rr - 1 < len(row_heights):
                mh += row_heights[rr - 1]

        for cc in range(mc.min_col, mc.max_col + 1):
            if (mc.min_row, cc) != (mc.min_row, mc.min_col):
                skip.add((mc.min_row, cc))
        for r_m in range(mc.min_row + 1, mc.max_row + 1):
            for c_m in range(mc.min_col, mc.max_col + 1):
                skip.add((r_m, c_m))
        merged_rects[(mc.min_row, mc.min_col)] = (mw, mh)

    # ── 5. Draw ────────────────────────────────────────────────────────────────
    y = 0
    for r in range(1, max_row + 1):
        rh = row_heights[r - 1]   # per-row height
        x = 0
        for c in render_cols:
            cw   = col_px[c - 1]
            if (r, c) in skip:
                x += cw
                continue

            mw_mh = merged_rects.get((r, c))
            if mw_mh:
                mw, mh = mw_mh
            else:
                mw, mh = cw, rh
            text = grid[r][c]

            try:
                cell  = ws.cell(r, c)
                bg    = _cell_bg(cell) or BG_WHITE
                fg    = _cell_fg(cell)
                bold  = _cell_bold(cell)
                align = _cell_align(cell)
            except Exception:
                bg, fg, bold, align = BG_WHITE, TEXT_DEF, False, 'left'

            draw.rectangle([x, y, x + mw, y + mh], fill=bg)

            if text:
                f = _get_font(text, FONT_SIZE, bold)
                stroke_w = 0
                if bold and _has_khmer(text):
                    # Simulate bold for Khmer OS fonts using a stroke outline
                    stroke_w = max(1, int(0.4 * SCALE))
                try:
                    bb = draw.textbbox((0, 0), text, font=f, stroke_width=stroke_w)
                    tw, th = bb[2] - bb[0], bb[3] - bb[1]
                except Exception:
                    tw = len(text) * (FONT_SIZE - 2) + stroke_w * 2
                    th = FONT_SIZE + stroke_w * 2

                ty = y + (mh - th) // 2
                if align == 'center':
                    tx = x + (mw - tw) // 2
                elif align == 'right':
                    tx = x + mw - tw - PAD_X
                else:
                    tx = x + PAD_X

                if stroke_w > 0:
                    draw.text((tx, ty), text, font=f, fill=fg, stroke_width=stroke_w, stroke_fill=fg)
                else:
                    draw.text((tx, ty), text, font=f, fill=fg)

            draw.rectangle([x, y, x + mw, y + mh], outline=BORDER_COL, width=1 * SCALE)
            x += cw
        y += rh

    # ── Aspect ratio safety check for Telegram ──────────────────────────────
    # Telegram rejects photos with aspect ratios > 20 (Photo_invalid_dimensions).
    # We enforce a safe ratio of 18 by adding white background padding if needed.
    max_ratio = 18.0
    w, h = img.size
    new_w, new_h = w, h
    if h > 0 and w / h > max_ratio:
        new_h = int(w / max_ratio)
    elif w > 0 and h / w > max_ratio:
        new_w = int(h / max_ratio)

    if (new_w, new_h) != (w, h):
        padded_img = Image.new('RGB', (new_w, new_h), BG_WHITE)
        padded_img.paste(img, (0, 0))
        img = padded_img

    buf = io.BytesIO()
    img.save(buf, format='PNG', optimize=True)
    buf.seek(0)
    return buf


def render_excel_reports(xlsx_path: str, target_date, out_dir: str) -> dict:
    """
    Renders the required report screenshots from the populated Excel file using Excel COM.
    """
    import pythoncom
    pythoncom.CoInitialize()
    import win32com.client
    import os
    import time
    
    abs_path = os.path.abspath(xlsx_path)
    os.makedirs(out_dir, exist_ok=True)
    
    excel = None
    wb = None
    results = {}
    
    try:
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
        except Exception:
            excel = win32com.client.Dispatch("Excel.Application")
            
        try:
            excel.Visible = False
        except Exception:
            pass
        try:
            excel.DisplayAlerts = False
            excel.ScreenUpdating = False
        except Exception:
            pass
            
        wb = excel.Workbooks.Open(abs_path)
        excel.CalculateFull()
        
        # 1. Render Zone Summary (Single tight full-fit table: SẢN LƯỢNG NHẬN - NGÀY)
        try:
            ws_z = wb.Worksheets("Zone_Report")
            ws_zs = wb.Worksheets.Add()
            
            # Title Banner: Row 1 (Merged A1:L1 to match exact 12-column table width)
            ws_zs.Range("A1:L1").Merge()
            date_val = ws_z.Range("C1").Text
            time_val = ws_z.Range("C2").Text
            ws_zs.Range("A1").Value = f"Báo cáo kết quả sản xuất kinh doanh ngày {date_val} - {time_val}"
            ws_zs.Range("A1").Font.Name = "Arial"
            ws_zs.Range("A1").Font.Bold = True
            ws_zs.Range("A1").Font.Size = 13
            ws_zs.Range("A1").Font.Color = 0xFFFFFF
            ws_zs.Range("A1").Interior.Color = 0x0000B0 # Red
            ws_zs.Range("A1").HorizontalAlignment = -4131 # Left
            ws_zs.Rows(1).RowHeight = 28
            
            # Table: SẢN LƯỢNG NHẬN - NGÀY (B6:M14 -> A2:L10)
            ws_z.Range("B6:M14").Copy()
            ws_zs.Range("A2:L10").PasteSpecial(-4122) # xlPasteFormats
            ws_zs.Range("A2:L10").Value = ws_z.Range("B6:M14").Value
            
            table_rng = ws_zs.Range("A2:L10")
            table_rng.Font.Name = "Arial"
            table_rng.Font.Size = 10
            ws_zs.Range("A2:L4").Font.Bold = True
            ws_zs.Range("A5:A10").Font.Bold = True
            
            # Explicit solid borders on EVERY cell
            for border_id in [7, 8, 9, 10, 11, 12]:
                try:
                    table_rng.Borders(border_id).LineStyle = 1
                    table_rng.Borders(border_id).Weight = 2
                    table_rng.Borders(border_id).Color = 0x000000
                except Exception:
                    pass
                    
            # AutoFit with generous padding to prevent ###
            ws_zs.Range("A2:L10").Columns.AutoFit()
            for c in range(1, 13):
                w = ws_zs.Columns(c).ColumnWidth
                ws_zs.Columns(c).ColumnWidth = max(w + 2.5, 9.5)
                
            ws_zs.Rows(2).RowHeight = 24
            ws_zs.Rows(3).RowHeight = 24
            ws_zs.Rows(4).RowHeight = 22
            for r in range(5, 11):
                ws_zs.Rows(r).RowHeight = 22
            
            ws_zs.Parent.Windows(1).DisplayGridlines = True
            
            rng_zs = ws_zs.Range("A1:L10")
            rng_zs.CopyPicture(1, -4147)
            time.sleep(0.2)
            
            scale = 2.5
            w = rng_zs.Width * scale
            h = rng_zs.Height * scale
            
            co_zs = ws_zs.ChartObjects().Add(Left=0, Top=0, Width=w, Height=h)
            co_zs.Activate()
            ch_zs = co_zs.Chart
            ch_zs.Paste()
            
            if ch_zs.Shapes.Count > 0:
                s = ch_zs.Shapes(1)
                s.Left = 0
                s.Top = 0
                s.ScaleWidth(scale, 1)
                s.ScaleHeight(scale, 1)
                co_zs.Width = s.Width
                co_zs.Height = s.Height
                
            ch_zs.ChartArea.Format.Line.Visible = 0
            ch_zs.ChartArea.Format.Fill.Visible = 0
            
            zs_png = os.path.join(out_dir, "zone_summary.png")
            ch_zs.Export(os.path.abspath(zs_png), "PNG")
            ws_zs.Delete()
            results["zone_summary"] = zs_png
        except Exception as e:
            import logging
            logging.exception("Failed to render tight full-fit zone_summary")

        # 1b. Render Customer Report (BÁO CÁO KHÁCH HÀNG PHÁT SINH LẦN ĐẦU: B56:L64)
        try:
            ws_z = wb.Worksheets("Zone_Report")
            ws_cr = wb.Worksheets.Add()
            
            ws_z.Range("B56:L64").Copy()
            ws_cr.Range("A1:K9").PasteSpecial(-4122) # xlPasteFormats
            ws_cr.Range("A1:K9").Value = ws_z.Range("B56:L64").Value
            
            table_rng = ws_cr.Range("A1:K9")
            table_rng.Font.Name = "Arial"
            table_rng.Font.Size = 10
            
            ws_cr.Range("A1:K1").Merge()
            ws_cr.Range("A1").Value = "BÁO CÁO KHÁCH HÀNG PHÁT SINH LẦN ĐẦU"
            ws_cr.Range("A1").Font.Name = "Arial"
            ws_cr.Range("A1").Font.Bold = True
            ws_cr.Range("A1").Font.Size = 12
            ws_cr.Range("A1").HorizontalAlignment = -4108 # Center
            ws_cr.Rows(1).RowHeight = 26
            
            ws_cr.Range("A2:K3").Font.Bold = True
            ws_cr.Range("A4:A9").Font.Bold = True
            ws_cr.Rows(2).RowHeight = 22
            ws_cr.Rows(3).RowHeight = 22
            for r in range(4, 10):
                ws_cr.Rows(r).RowHeight = 21
                
            for border_id in [7, 8, 9, 10, 11, 12]:
                try:
                    table_rng.Borders(border_id).LineStyle = 1
                    table_rng.Borders(border_id).Weight = 2
                    table_rng.Borders(border_id).Color = 0x000000
                except Exception:
                    pass
                    
            ws_cr.Range("A1:K9").Columns.AutoFit()
            for c in range(1, 12):
                w = ws_cr.Columns(c).ColumnWidth
                ws_cr.Columns(c).ColumnWidth = max(w + 2.0, 9.0)
                
            ws_cr.Parent.Windows(1).DisplayGridlines = True
            ws_cr.Activate()
            
            rng_cr = ws_cr.Range("A1:K9")
            rng_cr.CopyPicture(1, -4147)
            time.sleep(0.2)
            
            scale = 2.5
            w = rng_cr.Width * scale
            h = rng_cr.Height * scale
            
            co_cr = ws_cr.ChartObjects().Add(Left=0, Top=0, Width=w, Height=h)
            co_cr.Activate()
            ch_cr = co_cr.Chart
            ch_cr.Paste()
            
            if ch_cr.Shapes.Count > 0:
                s = ch_cr.Shapes(1)
                s.Left = 0
                s.Top = 0
                s.Width = w
                s.Height = h
                
            ch_cr.ChartArea.Format.Line.Visible = 0
            ch_cr.ChartArea.Format.Fill.Visible = 0
            
            cr_png = os.path.join(out_dir, "customer_report.png")
            ch_cr.Export(os.path.abspath(cr_png), "PNG")
            ws_cr.Delete()
            results["customer_report"] = cr_png
        except Exception as e:
            import logging
            logging.exception("Failed to render customer_report")

        # 2. Base reports (day_report, month_report, sp_order_express_all)
        ws_p = wb.Worksheets("Province_Report")
        try:
            ws_p.Outline.ShowLevels(RowLevels=8, ColumnLevels=8)
        except Exception:
            pass
        try:
            ws_p.Rows.Hidden = False
        except Exception:
            pass

        # 2. Render Day Report (BILL ORDER - DAY: AA3:AY28)
        try:
            ws_tmp_dr = wb.Worksheets.Add()
            src_rng = ws_p.Range("AA3:AY28")
            src_rng.Copy()
            ws_tmp_dr.Range("A1:Y26").PasteSpecial(-4122) # xlPasteFormats
            ws_tmp_dr.Range("A1:Y26").Value = src_rng.Value
            
            table_rng = ws_tmp_dr.Range("A1:Y26")
            table_rng.Font.Name = "Arial"
            table_rng.Font.Size = 10
            
            ws_tmp_dr.Range("A1:Y3").Font.Bold = True
            ws_tmp_dr.Range("A4:D26").Font.Bold = True
            
            # Hide Zone (Col B) and Point (Col E)
            ws_tmp_dr.Columns("B:B").Hidden = True
            ws_tmp_dr.Columns("E:E").Hidden = True
            
            # Row heights
            ws_tmp_dr.Rows(1).RowHeight = 26
            ws_tmp_dr.Rows(2).RowHeight = 22
            ws_tmp_dr.Rows(3).RowHeight = 22
            for r in range(4, 27):
                ws_tmp_dr.Rows(r).RowHeight = 21
                
            # Explicit solid borders
            for border_id in [7, 8, 9, 10, 11, 12]:
                try:
                    table_rng.Borders(border_id).LineStyle = 1
                    table_rng.Borders(border_id).Weight = 2
                    table_rng.Borders(border_id).Color = 0x000000
                except Exception:
                    pass
                    
            ws_tmp_dr.UsedRange.Columns.AutoFit()
            for c in range(1, ws_tmp_dr.UsedRange.Columns.Count + 1):
                cur_w = ws_tmp_dr.Columns(c).ColumnWidth
                ws_tmp_dr.Columns(c).ColumnWidth = max(cur_w + 2.0, 9.5)
                
            ws_tmp_dr.Parent.Windows(1).DisplayGridlines = True
            
            rng = ws_tmp_dr.Range("A1:Y26")
            rng.CopyPicture(1, -4147)
            time.sleep(0.2)
            
            scale = 2.5
            w = rng.Width * scale
            h = rng.Height * scale
            
            co = ws_tmp_dr.ChartObjects().Add(Left=0, Top=0, Width=w, Height=h)
            co.Activate()
            ch = co.Chart
            ch.Paste()
            
            if ch.Shapes.Count > 0:
                s = ch.Shapes(1)
                s.Left = 0
                s.Top = 0
                s.ScaleWidth(scale, 1)
                s.ScaleHeight(scale, 1)
                co.Width = s.Width
                co.Height = s.Height
                
            ch.ChartArea.Format.Line.Visible = 0
            ch.ChartArea.Format.Fill.Visible = 0
            
            dr_png = os.path.join(out_dir, "day_report.png")
            ch.Export(os.path.abspath(dr_png), "PNG")
            ws_tmp_dr.Delete()
            results["day_report"] = dr_png
        except Exception as e:
            import logging
            logging.exception("Failed to render day_report")

        # 2b. Render Showroom Report (A2:R27)
        try:
            ws_sr_src = wb.Worksheets("Showroom_RP")
            ws_tmp_sr = wb.Worksheets.Add()
            
            src_rng = ws_sr_src.Range("A2:R27")
            src_rng.Copy()
            ws_tmp_sr.Range("A1:R26").PasteSpecial(-4122) # xlPasteFormats
            ws_tmp_sr.Range("A1:R26").Value = src_rng.Value
            
            table_rng = ws_tmp_sr.Range("A1:R26")
            table_rng.Font.Name = "Arial"
            table_rng.Font.Size = 10
            
            ws_tmp_sr.Range("A1:R3").Font.Bold = True
            ws_tmp_sr.Range("A4:D26").Font.Bold = True
            
            # Row heights
            ws_tmp_sr.Rows(1).RowHeight = 26
            ws_tmp_sr.Rows(2).RowHeight = 22
            ws_tmp_sr.Rows(3).RowHeight = 22
            for r in range(4, 27):
                ws_tmp_sr.Rows(r).RowHeight = 21
                
            # Explicit solid borders
            for border_id in [7, 8, 9, 10, 11, 12]:
                try:
                    table_rng.Borders(border_id).LineStyle = 1
                    table_rng.Borders(border_id).Weight = 2
                    table_rng.Borders(border_id).Color = 0x000000
                except Exception:
                    pass
                    
            ws_tmp_sr.UsedRange.Columns.AutoFit()
            for c in range(1, ws_tmp_sr.UsedRange.Columns.Count + 1):
                cur_w = ws_tmp_sr.Columns(c).ColumnWidth
                ws_tmp_sr.Columns(c).ColumnWidth = max(cur_w + 2.0, 9.5)
                
            ws_tmp_sr.Parent.Windows(1).DisplayGridlines = True
            
            rng = ws_tmp_sr.Range("A1:R26")
            rng.CopyPicture(1, -4147)
            time.sleep(0.2)
            
            scale = 2.5
            w = rng.Width * scale
            h = rng.Height * scale
            
            co = ws_tmp_sr.ChartObjects().Add(Left=0, Top=0, Width=w, Height=h)
            co.Activate()
            ch = co.Chart
            ch.Paste()
            
            if ch.Shapes.Count > 0:
                s = ch.Shapes(1)
                s.Left = 0
                s.Top = 0
                s.ScaleWidth(scale, 1)
                s.ScaleHeight(scale, 1)
                co.Width = s.Width
                co.Height = s.Height
                
            ch.ChartArea.Format.Line.Visible = 0
            ch.ChartArea.Format.Fill.Visible = 0
            
            sr_png = os.path.join(out_dir, "showroom_report.png")
            ch.Export(os.path.abspath(sr_png), "PNG")
            ws_tmp_sr.Delete()
            results["showroom_report"] = sr_png
        except Exception as e:
            import logging
            logging.exception("Failed to render showroom_report")

        # 2c. Render Agent Report (A2:R27)
        try:
            ws_ar_src = wb.Worksheets("Agent_RP")
            ws_tmp_ar = wb.Worksheets.Add()
            
            src_rng = ws_ar_src.Range("A2:R27")
            src_rng.Copy()
            ws_tmp_ar.Range("A1:R26").PasteSpecial(-4122) # xlPasteFormats
            ws_tmp_ar.Range("A1:R26").Value = src_rng.Value
            
            table_rng = ws_tmp_ar.Range("A1:R26")
            table_rng.Font.Name = "Arial"
            table_rng.Font.Size = 10
            
            ws_tmp_ar.Range("A1:R3").Font.Bold = True
            ws_tmp_ar.Range("A4:D26").Font.Bold = True
            
            # Row heights
            ws_tmp_ar.Rows(1).RowHeight = 26
            ws_tmp_ar.Rows(2).RowHeight = 22
            ws_tmp_ar.Rows(3).RowHeight = 22
            for r in range(4, 27):
                ws_tmp_ar.Rows(r).RowHeight = 21
                
            # Explicit solid borders
            for border_id in [7, 8, 9, 10, 11, 12]:
                try:
                    table_rng.Borders(border_id).LineStyle = 1
                    table_rng.Borders(border_id).Weight = 2
                    table_rng.Borders(border_id).Color = 0x000000
                except Exception:
                    pass
                    
            ws_tmp_ar.UsedRange.Columns.AutoFit()
            for c in range(1, ws_tmp_ar.UsedRange.Columns.Count + 1):
                cur_w = ws_tmp_ar.Columns(c).ColumnWidth
                ws_tmp_ar.Columns(c).ColumnWidth = max(cur_w + 2.0, 9.5)
                
            ws_tmp_ar.Parent.Windows(1).DisplayGridlines = True
            
            rng = ws_tmp_ar.Range("A1:R26")
            co = ws_tmp_ar.ChartObjects().Add(Left=0, Top=0, Width=w, Height=h)
            co.Activate()
            ch = co.Chart
            ch.Paste()
            
            if ch.Shapes.Count > 0:
                s = ch.Shapes(1)
                s.Left = 0
                s.Top = 0
                s.ScaleWidth(scale, 1)
                s.ScaleHeight(scale, 1)
                co.Width = s.Width
                co.Height = s.Height
                
            ch.ChartArea.Format.Line.Visible = 0
            ch.ChartArea.Format.Fill.Visible = 0
            
            ar_png = os.path.join(out_dir, "agent_report.png")
            ch.Export(os.path.abspath(ar_png), "PNG")
            ws_tmp_ar.Delete()
            results["agent_report"] = ar_png
        except Exception as e:
            import logging
            logging.exception("Failed to render agent_report")

        # Render sp_order_express_all (All 36 Service Points - Table 1 A3:W44)
        try:
            ws_tmp = wb.Worksheets.Add()
            src_rng = ws_p.Range("A3:W44")
            src_rng.Copy()
            ws_tmp.Range("A1:W42").PasteSpecial(-4122) # xlPasteFormats
            ws_tmp.Range("A1:W42").Value = src_rng.Value
            
            try:
                ws_tmp.Outline.ShowLevels(RowLevels=8, ColumnLevels=8)
            except Exception:
                pass
            ws_tmp.Rows.Hidden = False
            
            # Apply Modern Clean Arial Font
            table_rng = ws_tmp.Range("A1:W42")
            table_rng.Font.Name = "Arial"
            table_rng.Font.Size = 10
            
            ws_tmp.Range("A1:W3").Font.Bold = True
            ws_tmp.Range("A1:W3").Font.Size = 10.5
            ws_tmp.Range("A4:B42").Font.Bold = True # Station Codes bold
            
            # Set explicit standard row heights
            ws_tmp.Rows(1).RowHeight = 28
            ws_tmp.Rows(2).RowHeight = 22
            ws_tmp.Rows(3).RowHeight = 20
            for r in range(4, 43):
                ws_tmp.Rows(r).RowHeight = 21
            
            # Hide column D (Zone) as shown in reference images
            ws_tmp.Columns("D:D").Hidden = True
            
            # Explicit solid borders
            for border_id in [7, 8, 9, 10, 11, 12]:
                try:
                    table_rng.Borders(border_id).LineStyle = 1
                    table_rng.Borders(border_id).Weight = 2
                    table_rng.Borders(border_id).Color = 0x000000
                except Exception:
                    pass
                    
            ws_tmp.UsedRange.Columns.AutoFit()
            for c in range(1, ws_tmp.UsedRange.Columns.Count + 1):
                cur_w = ws_tmp.Columns(c).ColumnWidth
                ws_tmp.Columns(c).ColumnWidth = max(cur_w + 2.0, 9.5)
                
            ws_tmp.Parent.Windows(1).DisplayGridlines = True
            
            rng = ws_tmp.Range("A1:W42")
            rng.CopyPicture(1, -4147)
            time.sleep(0.2)
            
            scale = 2.5
            w = rng.Width * scale
            h = rng.Height * scale
            
            co = ws_tmp.ChartObjects().Add(Left=0, Top=0, Width=w, Height=h)
            co.Activate()
            ch = co.Chart
            ch.Paste()
            
            if ch.Shapes.Count > 0:
                s = ch.Shapes(1)
                s.Left = 0
                s.Top = 0
                s.ScaleWidth(scale, 1)
                s.ScaleHeight(scale, 1)
                co.Width = s.Width
                co.Height = s.Height
                
            ch.ChartArea.Format.Line.Visible = 0
            ch.ChartArea.Format.Fill.Visible = 0
            
            out_png = os.path.join(out_dir, "sp_order_express_all.png")
            ch.Export(os.path.abspath(out_png), "PNG")
            ws_tmp.Delete()
            results["sp_order_express_all"] = out_png
        except Exception as e:
            import logging
            logging.exception("Failed to render sp_order_express_all")

        # 2. Detailed Service Point Zone slices (sp_zone_1 to sp_zone_5 matching Pictures 5 to 9)
        zone_split_configs = [
            ("sp_zone_1", 6, 21),      # Picture 5: Phnom Penh (MEC, Phnompenh, PNPP001..014)
            ("sp_zone_1_prov", 23, 25), # Picture 6: KANP001, PREP001, SVAP001
            ("sp_zone_2", 26, 30),      # Picture 7: SPEP001, TAKP001, KAMP001, SIHP001, KOHP001
            ("sp_zone_3_4", 31, 38),    # Picture 8: BANP001, BATP001, PURP001, CHHP001, SIEP001, PRHP001, THOP001, ODDP001
            ("sp_zone_5", 39, 44),      # Picture 9: TBKP001, CHAP001, KRAP001, MONP001, ROTP001, STUP001
        ]
        
        for sp_name, s_row, e_row in zone_split_configs:
            try:
                ws_tmp = wb.Worksheets.Add()
                
                # Title Banner: Row 1 (Merged A1:W1)
                ws_tmp.Range("A1:W1").Merge()
                ws_tmp.Range("A1").Value = "[SERVICE POINT]-REPORT OF ORDER EXPRESS"
                ws_tmp.Range("A1").Font.Name = "Arial"
                ws_tmp.Range("A1").Font.Bold = True
                ws_tmp.Range("A1").Font.Size = 13
                ws_tmp.Range("A1").Font.Color = 0xFFFFFF
                ws_tmp.Range("A1").Interior.Color = 0x0000B0 # Red
                ws_tmp.Rows(1).RowHeight = 28
                
                # Copy table headers (A4:W5 -> A2:W3)
                ws_p.Range("A4:W5").Copy()
                ws_tmp.Range("A2:W3").PasteSpecial(-4122) # xlPasteFormats
                ws_tmp.Range("A2:W3").Value = ws_p.Range("A4:W5").Value
                
                # Copy data rows (A{s_row}:W{e_row} -> A4:W{3+num_rows})
                num_rows = e_row - s_row + 1
                src_rng = ws_p.Range(ws_p.Cells(s_row, 1), ws_p.Cells(e_row, 23)) # Col 1 to 23 (A to W)
                dest_rng = ws_tmp.Range(ws_tmp.Cells(4, 1), ws_tmp.Cells(3 + num_rows, 23))
                
                src_rng.Copy()
                dest_rng.PasteSpecial(-4122) # xlPasteFormats
                dest_rng.Value = src_rng.Value
                
                # Hide column D (Zone) as shown in pictures 5-9
                ws_tmp.Columns("D:D").Hidden = True
                
                total_rows = 3 + num_rows
                table_rng = ws_tmp.Range(f"A1:W{total_rows}")
                table_rng.Font.Name = "Arial"
                table_rng.Font.Size = 10
                
                ws_tmp.Range("A2:W3").Font.Bold = True
                ws_tmp.Range(f"A4:B{total_rows}").Font.Bold = True
                
                # Set explicit standard row heights
                ws_tmp.Rows(2).RowHeight = 22
                ws_tmp.Rows(3).RowHeight = 20
                for r in range(4, 4 + num_rows):
                    ws_tmp.Rows(r).RowHeight = 22
                
                # Set explicit solid borders on EVERY cell
                for border_id in [7, 8, 9, 10, 11, 12]:
                    try:
                        table_rng.Borders(border_id).LineStyle = 1 # xlContinuous
                        table_rng.Borders(border_id).Weight = 2    # xlThin
                        table_rng.Borders(border_id).Color = 0x000000 # Solid Black borders
                    except Exception:
                        pass
                
                # AutoFit and set minimum column width to prevent ###
                ws_tmp.UsedRange.Columns.AutoFit()
                for c in range(1, ws_tmp.UsedRange.Columns.Count + 1):
                    cur_w = ws_tmp.Columns(c).ColumnWidth
                    ws_tmp.Columns(c).ColumnWidth = max(cur_w + 2.0, 9.5)
                    
                ws_tmp.Parent.Windows(1).DisplayGridlines = True
                
                rng = ws_tmp.Range(f"A1:W{total_rows}")
                time.sleep(0.1)
                for _attempt in range(3):
                    try:
                        rng.CopyPicture(1, -4147)
                        time.sleep(0.15)
                        break
                    except Exception:
                        time.sleep(0.2)
                
                scale = 2.5
                w = rng.Width * scale
                h = rng.Height * scale
                
                co = ws_tmp.ChartObjects().Add(Left=0, Top=0, Width=w, Height=h)
                co.Activate()
                ch = co.Chart
                ch.Paste()
                
                if ch.Shapes.Count > 0:
                    s = ch.Shapes(1)
                    s.Left = 0
                    s.Top = 0
                    s.ScaleWidth(scale, 1)
                    s.ScaleHeight(scale, 1)
                    co.Width = s.Width
                    co.Height = s.Height
                    
                ch.ChartArea.Format.Line.Visible = 0
                ch.ChartArea.Format.Fill.Visible = 0
                
                out_png = os.path.join(out_dir, f"{sp_name}.png")
                ch.Export(os.path.abspath(out_png), "PNG")
                
                ws_tmp.Delete()
                results[sp_name] = out_png
            except Exception as e:
                import logging
                logging.exception("Failed to render detailed zone report: %s", sp_name)

        # 3. Render [SERVICE POINT]-REPORT OF CUSTOMER DEVELOPMENT
        try:
            raw_data = ws_p.Range("A6:V44").Value
            ws_cd = wb.Worksheets.Add()
            
            # Title Header
            ws_cd.Range("A1:Q1").Merge()
            ws_cd.Range("A1").Value = "[SERVICE POINT]-REPORT OF CUSTOMER DEVELOPMENT"
            ws_cd.Range("A1").Font.Name = "Arial"
            ws_cd.Range("A1").Font.Bold = True
            ws_cd.Range("A1").Font.Size = 13
            ws_cd.Range("A1").Font.Color = 0xFFFFFF
            ws_cd.Range("A1").Interior.Color = 0x0000B0 # Red
            ws_cd.Rows(1).RowHeight = 28
            
            headers = [
                ("A2:A3", "No"), ("B2:B3", "Express Code"), ("C2:C3", "Express Name"), ("D2:D3", "Zone"),
                ("E2:G2", "Number of Customer"), ("H2:J2", "Number of Customer\n(from 5 orders)"),
                ("K2:M2", "Number. of Customer\n(from 20 orders)"), ("N2:P2", "New Customer"), ("Q2:Q3", "New Customer\nInday")
            ]
            for rng_str, txt in headers:
                if ":" in rng_str:
                    ws_cd.Range(rng_str).Merge()
                ws_cd.Range(rng_str.split(":")[0]).Value = txt
                
            sub_headers = [
                ("E3", "Target"), ("F3", "Result"), ("G3", "%"),
                ("H3", "Target"), ("I3", "Result"), ("J3", "%"),
                ("K3", "Target"), ("L3", "Result"), ("M3", "%"),
                ("N3", "Target"), ("O3", "Result"), ("P3", "%")
            ]
            for cell, txt in sub_headers:
                ws_cd.Range(cell).Value = txt
                
            ws_cd.Range("A2:D3").Interior.Color = 0x0000B0
            ws_cd.Range("A2:D3").Font.Color = 0xFFFFFF
            ws_cd.Range("A2:D3").Font.Bold = True
            
            ws_cd.Range("E2:G3").Interior.Color = 0x59B900
            ws_cd.Range("H2:J3").Interior.Color = 0x66CC00
            ws_cd.Range("K2:M3").Interior.Color = 0x993300
            ws_cd.Range("N2:P3").Interior.Color = 0x800080
            ws_cd.Range("Q2:Q3").Interior.Color = 0x660066
            
            ws_cd.Range("E2:Q3").Font.Bold = True
            for col_rng in ["E2:G3", "H2:J3", "K2:M3", "N2:P3", "Q2:Q3"]:
                ws_cd.Range(col_rng).Font.Color = 0xFFFFFF
                
            ws_cd.Range("A2:Q3").HorizontalAlignment = -4108
            ws_cd.Range("A2:Q3").VerticalAlignment = -4108
            ws_cd.Rows(2).RowHeight = 24
            ws_cd.Rows(3).RowHeight = 22
            
            station_targets = {
                "PNPP001": (308, 68, 22, 225), "PNPP002": (308, 68, 22, 197), "PNPP003": (308, 68, 22, 183),
                "PNPP004": (308, 68, 22, 253), "PNPP005": (308, 68, 22, 197), "PNPP006": (308, 68, 22, 225),
                "PNPP007": (308, 68, 22, 183), "PNPP008": (308, 68, 22, 267), "PNPP009": (308, 68, 22, 253),
                "PNPP010": (308, 68, 22, 225), "PNPP011": (308, 68, 22, 183), "PNPP012": (308, 68, 22, 197),
                "PNPP013": (308, 68, 22, 183), "PNPP014": (308, 68, 22, 253),
                "KANP001": (257, 57, 18, 267), "PREP001": (257, 57, 18, 183), "SVAP001": (257, 57, 18, 141),
                "SPEP001": (257, 57, 18, 183), "TAKP001": (257, 57, 18, 225), "KAMP001": (257, 57, 18, 253),
                "SIHP001": (257, 57, 18, 281), "KOHP001": (205, 45, 14, 183), "BANP001": (257, 57, 18, 267),
                "BATP001": (257, 57, 18, 323), "PURP001": (205, 45, 14, 155), "CHHP001": (257, 57, 18, 225),
                "SIEP001": (257, 57, 18, 225), "PRHP001": (257, 57, 18, 183), "THOP001": (257, 57, 18, 253),
                "ODDP001": (205, 45, 14, 183), "TBKP001": (205, 45, 14, 239), "CHAP001": (257, 57, 18, 253),
                "KRAP001": (205, 45, 14, 253), "MONP001": (205, 45, 14, 183), "ROTP001": (103, 23, 7, 183),
                "STUP001": (103, 23, 7, 183)
            }
            
            matrix = []
            r0 = raw_data[0]
            matrix.append(["MEC", "", "", "", 9346, r0[17], (r0[17] or 0)/9346, 2066, r0[18], (r0[18] or 0)/2066, 658, r0[19], (r0[19] or 0)/658, 7836, r0[20], (r0[20] or 0)/7836, r0[21] or 0])
            
            r1 = raw_data[1]
            matrix.append(["I", "Phnompenh", "", "", 4312, r1[17], (r1[17] or 0)/4312, 952, r1[18], (r1[18] or 0)/952, 308, r1[19], (r1[19] or 0)/308, 3019, r1[20], (r1[20] or 0)/3019, r1[21] or 0])
            
            for r in raw_data[2:]:
                no_val = r[0]
                code_val = str(r[1] or '').strip().upper()
                
                # Check if separator row
                if str(no_val).strip().lower() == 'x':
                    matrix.append(["x", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""])
                    continue
                if not code_val:
                    continue
                    
                name_val = r[2]
                zone_val = r[3]
                
                tg = station_targets.get(code_val, (257, 57, 18, 200))
                c_res = r[17] or 0
                c5_res = r[18] or 0
                c20_res = r[19] or 0
                new_res = r[20] or 0
                new_day = r[21] or 0
                
                matrix.append([
                    no_val, code_val, name_val, zone_val,
                    tg[0], c_res, c_res / tg[0] if tg[0] else 0,
                    tg[1], c5_res, c5_res / tg[1] if tg[1] else 0,
                    tg[2], c20_res, c20_res / tg[2] if tg[2] else 0,
                    tg[3], new_res, new_res / tg[3] if tg[3] else 0,
                    new_day
                ])
                
            num_data_rows = len(matrix)
            full_cd_rng = ws_cd.Range(f"A4:Q{3 + num_data_rows}")
            full_cd_rng.Value = matrix
            
            ws_cd.Range(f"A1:Q{3 + num_data_rows}").Font.Name = "Arial"
            ws_cd.Range(f"A4:Q{3 + num_data_rows}").Font.Size = 10
            ws_cd.Range("A2:Q3").Font.Size = 10.5
            
            ws_cd.Range("A4:Q4").Interior.Color = 0x00FFFF
            ws_cd.Range("A4:Q4").Font.Bold = True
            ws_cd.Range("A5:Q5").Font.Bold = True
            ws_cd.Range(f"A6:B{3 + num_data_rows}").Font.Bold = True
            
            for r in range(4, 4 + num_data_rows):
                ws_cd.Rows(r).RowHeight = 21
            
            ws_cd.Range(f"G4:G{3+num_data_rows},J4:J{3+num_data_rows},M4:M{3+num_data_rows},P4:P{3+num_data_rows}").NumberFormat = "0.0%"
            ws_cd.Range(f"E4:F{3+num_data_rows},H4:I{3+num_data_rows},K4:L{3+num_data_rows},N4:O{3+num_data_rows},Q4:Q{3+num_data_rows}").NumberFormat = "#,##0"
            
            # Highlight new inday column
            for r in range(4, 4 + num_data_rows):
                v = ws_cd.Cells(r, 17).Value
                if v and float(v) > 0:
                    ws_cd.Cells(r, 17).Font.Color = 0x0000FF # Red bold
                    ws_cd.Cells(r, 17).Font.Bold = True
                    
            # Set explicit solid borders on EVERY cell
            table_rng = ws_cd.Range(f"A2:Q{3+num_data_rows}")
            for border_id in [7, 8, 9, 10, 11, 12]:
                try:
                    table_rng.Borders(border_id).LineStyle = 1 # xlContinuous
                    table_rng.Borders(border_id).Weight = 2    # xlThin
                    table_rng.Borders(border_id).Color = 0x000000 # Solid Black borders
                except Exception:
                    pass
                    
            # AutoFit and set generous width to avoid ####
            ws_cd.UsedRange.Columns.AutoFit()
            min_widths = {
                1: 6.0,   # No
                2: 13.5,  # Express Code
                3: 21.0,  # Express Name
                4: 9.5,   # Zone
                5: 10.5,  # Target 1
                6: 9.5,   # Result 1
                7: 9.0,   # % 1
                8: 10.5,  # Target 2
                9: 9.5,   # Result 2
                10: 9.0,  # % 2
                11: 10.5, # Target 3
                12: 9.5,  # Result 3
                13: 9.0,  # % 3
                14: 10.5, # Target 4
                15: 9.5,  # Result 4
                16: 9.0,  # % 4
                17: 12.5  # New Customer Inday
            }
            for col_idx, min_w in min_widths.items():
                cur_w = ws_cd.Columns(col_idx).ColumnWidth
                ws_cd.Columns(col_idx).ColumnWidth = max(cur_w + 1.8, min_w)
                
            ws_cd.Parent.Windows(1).DisplayGridlines = True
            ws_cd.Activate()
            
            rng_cd = ws_cd.Range(f"A1:Q{3+num_data_rows}")
            time.sleep(0.1)
            for _attempt in range(3):
                try:
                    rng_cd.CopyPicture(1, -4147)
                    time.sleep(0.2)
                    break
                except Exception:
                    time.sleep(0.2)
            
            scale = 2.0
            w = rng_cd.Width * scale
            h = rng_cd.Height * scale
            
            co_cd = ws_cd.ChartObjects().Add(Left=0, Top=0, Width=w, Height=h)
            co_cd.Activate()
            ch_cd = co_cd.Chart
            ch_cd.Paste()
            time.sleep(0.1)
            
            if ch_cd.Shapes.Count > 0:
                s = ch_cd.Shapes(1)
                s.Left = 0
                s.Top = 0
                s.Width = w
                s.Height = h
                
            ch_cd.ChartArea.Format.Line.Visible = 0
            ch_cd.ChartArea.Format.Fill.Visible = 0
            
            cd_png = os.path.join(out_dir, "sp_customer_development.png")
            ch_cd.Export(os.path.abspath(cd_png), "PNG")
            ws_cd.Delete()
            results["sp_customer_development"] = cd_png
        except Exception as e:
            import logging
            logging.exception("Failed to render sp_customer_development")
                
    finally:
        if wb:
            try:
                wb.Close(SaveChanges=False)
            except Exception:
                pass
        if excel:
            try:
                excel.Quit()
            except Exception:
                pass
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass
        
    return results

