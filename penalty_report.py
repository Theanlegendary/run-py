import os
import sys
import copy
import tempfile
from datetime import datetime, date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import excel_to_image

MAIN_36_BRANCHES = [
    'BANP001', 'BATP001', 'CHAP001', 'CHHP001', 'KAMP001', 'KANP001', 'KOHP001', 'KRAP001',
    'MONP001', 'ODDP001', 'PNPP001', 'PNPP002', 'PNPP003', 'PNPP004', 'PNPP005', 'PNPP006',
    'PNPP007', 'PNPP008', 'PNPP009', 'PNPP010', 'PNPP011', 'PNPP012', 'PNPP013', 'PNPP014',
    'PREP001', 'PRHP001', 'PURP001', 'ROTP001', 'SIEP001', 'SIHP001', 'SPEP001', 'STUP001',
    'SVAP001', 'TAKP001', 'TBKP001', 'THOP001'
]

def parse_date(val):
    if not val or pd.isna(val):
        return None
    if isinstance(val, (datetime, date)):
        return val.date() if isinstance(val, datetime) else val
    s = str(val).strip().split(" ")[0]
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None

def parse_time(val):
    if not val or pd.isna(val):
        return None
    if isinstance(val, datetime):
        return val
    s = str(val).strip()
    for fmt in ("%d/%m/%Y %H:%M:%S", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M", "%Y-%m-%d %H:%M"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None

def build_penalty_report(src_xlsx, out_xlsx, target_label="ALL", report_date=None):
    """
    CEO Executive Penalty Dashboard:
      - Sorted: From WORST performing branch (% On-Time) to BEST
      - Dynamic Font Coloring on % Columns:
          * < 75.0%: Bold Red (#DC2626)
          * 75.0% - 89.9%: Bold Amber (#D97706)
          * >= 90.0%: Bold Green (#16A34A)
      - Exact 9 Columns: No | Post Office | RIGHT Handover | RIGHT Delivery | Total Handover | Total Delivery | % RIGHT Handover | % RIGHT Delivery | Total Penalty ($)
    """
    os.makedirs(os.path.dirname(os.path.abspath(out_xlsx)), exist_ok=True)
    df = pd.read_excel(src_xlsx)
    df.columns = [str(c).strip().upper() for c in df.columns]

    col_order = next((c for c in df.columns if 'ORDER ID' in c or 'ORDER' in c), 'ORDER ID')
    col_dest_prov = next((c for c in df.columns if 'DELIVERY PROVINCE' in c or 'DESTINATION_BRANCH' in c), 'DELIVERY PROVINCE')
    col_dest_po = next((c for c in df.columns if 'DELIVERY POST' in c or 'DESTINATION_POST' in c), 'DELIVERY POST OFFICE')
    col_orig_br = next((c for c in df.columns if 'ACTION POST OFFICE' in c or 'ORIGIN_BRANCH' in c), 'ACTION POST OFFICE')
    col_orig_po = next((c for c in df.columns if 'CURRENT POST OFFICE' in c or 'ORIGIN_POST' in c), 'CURRENT POST OFFICE')
    col_status = next((c for c in df.columns if 'CURRENT STATUS' in c or 'STATUS' in c), 'CURRENT STATUS')
    col_created = next((c for c in df.columns if 'CREATED DATE' in c), 'CREATED DATE')
    col_receiver = next((c for c in df.columns if 'RECEIVER' in c), 'RECEIVER')
    col_action_time = next((c for c in df.columns if 'ACTION TIME' in c or 'CURRENT TIME' in c), 'CURRENT TIME')

    today = report_date or datetime.now().date()
    tgt = "".join(c for c in str(target_label).upper() if c.isalnum() or c in ("-", "_")).strip()
    if not tgt:
        tgt = "ALL"

    df['sc'] = df[col_status].astype(str).str.extract(r'^(\d{3})')[0]
    df['curr_po_clean'] = df[col_orig_po].astype(str).str.strip().str.upper()
    df['deliv_po_clean'] = df[col_dest_po].astype(str).str.strip().str.upper()

    handover_statuses = {'210', '230', '300', '302', '306', '309', '310', '311'}
    delivery_statuses = {'400', '401', '402', '420', '430', '460', '470', '471', '472', '480', '500', '510', '511', '512'}
    excused_statuses = {'420', '472'}

    active_df = df[df['sc'].isin(handover_statuses | delivery_statuses)].copy()

    if tgt not in ("ALL", "TOTAL"):
        if tgt.startswith("ZONE"):
            zone_by_prefix = {
                "KAN": "ZONE1", "PNP": "ZONE1", "PRE": "ZONE1", "SVA": "ZONE1",
                "KAM": "ZONE2", "KOH": "ZONE2", "SIH": "ZONE2", "SPE": "ZONE2", "TAK": "ZONE2", "KEP": "ZONE2",
                "BAN": "ZONE3", "BAT": "ZONE3", "CHH": "ZONE3", "PUR": "ZONE3", "PAI": "ZONE3",
                "ODD": "ZONE4", "PRH": "ZONE4", "SIE": "ZONE4", "THO": "ZONE4",
                "CHA": "ZONE5", "KRA": "ZONE5", "TBK": "ZONE5", "ROT": "ZONE5", "MON": "ZONE5", "STU": "ZONE5"
            }
            active_df['zone_h'] = active_df['curr_po_clean'].str[:3].map(zone_by_prefix).fillna("ZONE1")
            active_df['zone_d'] = active_df['deliv_po_clean'].str[:3].map(zone_by_prefix).fillna("ZONE1")
            active_df = active_df[
                ((active_df['sc'].isin(handover_statuses)) & (active_df['zone_h'] == tgt)) |
                ((active_df['sc'].isin(delivery_statuses)) & (active_df['zone_d'] == tgt))
            ].copy()
        elif len(tgt) == 3:
            active_df = active_df[
                ((active_df['sc'].isin(handover_statuses)) & (active_df['curr_po_clean'].str.startswith(tgt))) |
                ((active_df['sc'].isin(delivery_statuses)) & (active_df['deliv_po_clean'].str.startswith(tgt)))
            ].copy()
        else:
            active_df = active_df[
                ((active_df['sc'].isin(handover_statuses)) & (active_df['curr_po_clean'] == tgt)) |
                ((active_df['sc'].isin(delivery_statuses)) & (active_df['deliv_po_clean'] == tgt))
            ].copy()

    summary_data = {}
    if tgt in ("ALL", "TOTAL"):
        for b in MAIN_36_BRANCHES:
            summary_data[b] = {
                "po": b,
                "total_handover": 0,
                "total_delivery": 0,
                "penalty_handover": 0,
                "penalty_delivery": 0,
                "excused_count": 0,
                "total_fine": 0.0
            }

    base_rows = []
    r_idx = 1

    for idx, row in active_df.iterrows():
        sc = str(row['sc'])
        curr_po = str(row.get('curr_po_clean', '')).strip()
        deliv_po = str(row.get('deliv_po_clean', '')).strip()

        is_handover = sc in handover_statuses
        is_delivery = sc in delivery_statuses

        if is_handover:
            raw_po = curr_po
        elif is_delivery:
            raw_po = deliv_po
        else:
            continue

        if not raw_po or raw_po == 'NAN':
            continue

        # STRICT FILTER: 36 Main Post Offices only (Exclude agents & showrooms)
        if tgt in ("ALL", "TOTAL"):
            if raw_po not in MAIN_36_BRANCHES:
                continue
            po = raw_po
        elif len(tgt) >= 7:
            po = tgt
            if raw_po != tgt:
                continue
        else:
            po = raw_po

        if po not in summary_data:
            summary_data[po] = {
                "po": po,
                "total_handover": 0,
                "total_delivery": 0,
                "penalty_handover": 0,
                "penalty_delivery": 0,
                "excused_count": 0,
                "total_fine": 0.0
            }

        order_id = str(row.get(col_order, '')).strip()
        status_raw = str(row.get(col_status, '')).strip()

        act_val = row.get(col_action_time) if pd.notna(row.get(col_action_time)) else row.get(col_created)
        act_date = parse_date(act_val) or parse_date(row.get(col_created))
        age_days = (today - act_date).days if act_date else 0

        fine = 0.0
        risk_level = "Normal"
        is_excused = False

        if is_handover:
            summary_data[po]["total_handover"] += 1
            if age_days >= 3:
                fine = 0.40
                risk_level = "Urgent (> 3 days)"
                summary_data[po]["penalty_handover"] += 1
            elif age_days >= 1:
                fine = 0.10
                risk_level = "Backlog (1-2 days)"
                summary_data[po]["penalty_handover"] += 1
            else:
                risk_level = "Safe (< 1 day)"

        elif is_delivery:
            summary_data[po]["total_delivery"] += 1
            if sc in excused_statuses:
                is_excused = True
                risk_level = f"Excused ({sc})"
                summary_data[po]["excused_count"] += 1
                fine = 0.0
            else:
                if age_days >= 3:
                    fine = 0.40
                    risk_level = "Critical (> 3 days)"
                    summary_data[po]["penalty_delivery"] += 1
                elif age_days >= 1:
                    fine = 0.10
                    risk_level = "Stagnant (1-2 days)"
                    summary_data[po]["penalty_delivery"] += 1
                else:
                    risk_level = "Safe (< 1 day)"

        summary_data[po]["total_fine"] += fine

        base_rows.append({
            "no": r_idx,
            "order_number": order_id,
            "customer": str(row.get(col_receiver, ''))[:28],
            "origin_branch": str(row.get(col_orig_br, '')),
            "origin_post": curr_po,
            "destination_branch": str(row.get(col_dest_prov, '')),
            "destination_post": deliv_po,
            "assigned_branch": po,
            "created_at": str(row.get(col_created, '')),
            "last_action_time": str(act_val or ""),
            "status_code": sc,
            "status_name": status_raw,
            "type": "Handover" if is_handover else "Delivery",
            "age_days": age_days,
            "penalty_fine": fine,
            "risk_level": risk_level,
            "is_excused": is_excused
        })
        r_idx += 1

    if tgt not in ("ALL", "TOTAL") and tgt not in summary_data:
        summary_data[tgt] = {
            "po": tgt,
            "total_handover": 0,
            "total_delivery": 0,
            "penalty_handover": 0,
            "penalty_delivery": 0,
            "excused_count": 0,
            "total_fine": 0.0
        }

    # Build Excel Workbook
    wb = openpyxl.Workbook()

    # Sheet 1: INVENTORY PENALTY REPORT
    ws1 = wb.active
    ws1.title = "INVENTORY PENALTY REPORT"
    ws1.views.sheetView[0].showGridLines = True

    # ── CEO MIDNIGHT & ROYAL SAPPHIRE BLUE PALETTE ──
    fill_title_left   = PatternFill("solid", fgColor="0B132B") # Deepest Midnight Navy
    fill_hdr_left     = PatternFill("solid", fgColor="1C2541") # Midnight Slate
    fill_title_right  = PatternFill("solid", fgColor="0B132B") # Deepest Midnight Navy
    fill_sub_right    = PatternFill("solid", fgColor="1C2541") # Subtitle Midnight Slate
    fill_hdr_right    = PatternFill("solid", fgColor="1C3D82") # Royal Sapphire Blue Header
    fill_row_alt      = PatternFill("solid", fgColor="F8FAFC") # Clean Soft Zebra
    fill_left_tot     = PatternFill("solid", fgColor="E2E8F0") # Soft Slate Total
    fill_sum_tot      = PatternFill("solid", fgColor="E2E8F0") # Executive Accounting Total

    border_clean = Border(
        left=Side(style="thin", color="CBD5E1"), right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"), bottom=Side(style="thin", color="CBD5E1")
    )
    tot_border_accounting = Border(
        left=Side(style="thin", color="CBD5E1"), right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="64748B"), bottom=Side(style="double", color="0B132B")
    )

    font_banner = Font(name="Segoe UI", size=10.5, bold=True, color="FFFFFF")
    font_sub = Font(name="Segoe UI", size=8.0, italic=True, color="93C5FD")
    font_hdr = Font(name="Segoe UI", size=8.5, bold=True, color="FFFFFF")
    font_data = Font(name="Segoe UI", size=8.5, color="0F172A")
    font_bold_data = Font(name="Segoe UI", size=8.5, bold=True, color="0F172A")
    font_tot = Font(name="Segoe UI", size=9.5, bold=True, color="0F172A")

    # Font colors for % on-time metrics
    font_pct_green = Font(name="Segoe UI", size=8.5, bold=True, color="16A34A") # >= 90%
    font_pct_amber = Font(name="Segoe UI", size=8.5, bold=True, color="D97706") # 75% - 89.9%
    font_pct_red   = Font(name="Segoe UI", size=8.5, bold=True, color="DC2626") # < 75%

    font_tot_pct_green = Font(name="Segoe UI", size=9.5, bold=True, color="16A34A")
    font_tot_pct_amber = Font(name="Segoe UI", size=9.5, bold=True, color="D97706")
    font_tot_pct_red   = Font(name="Segoe UI", size=9.5, bold=True, color="DC2626")

    def get_pct_font(pct_val, is_tot=False):
        if pct_val >= 90.0:
            return font_tot_pct_green if is_tot else font_pct_green
        elif pct_val >= 75.0:
            return font_tot_pct_amber if is_tot else font_pct_amber
        else:
            return font_tot_pct_red if is_tot else font_pct_red

    date_str = today.strftime('%d/%m/%Y')

    # 1. Left Title Banner
    ws1.merge_cells("A1:H1")
    ws1.cell(1, 1, f"METFONE EXPRESS — INVENTORY PENALTY & STAGNANT GOODS ({tgt}) — {date_str}").font = font_banner
    ws1.cell(1, 1).alignment = Alignment(horizontal="left", vertical="center")
    for c in range(1, 9):
        ws1.cell(1, c).fill = fill_title_left
    ws1.row_dimensions[1].height = 26.0

    # 2. Right Title Banner (Cols J to R)
    ws1.merge_cells("J1:R1")
    ws1.cell(1, 10, f"EXECUTIVE PENALTY DASHBOARD ({tgt})").font = font_banner
    ws1.cell(1, 10).alignment = Alignment(horizontal="center", vertical="center")
    for c in range(10, 19):
        ws1.cell(1, c).fill = fill_title_right

    ws1.merge_cells("J2:R2")
    ws1.cell(2, 10, "SLA Penalty: 1-2 Days (-$0.10) | ≥ 3 Days (-$0.40) • Excused 420/472 ($0.00)").font = font_sub
    ws1.cell(2, 10).alignment = Alignment(horizontal="center", vertical="center")
    for c in range(10, 19):
        ws1.cell(2, c).fill = fill_sub_right
    ws1.row_dimensions[2].height = 18.0

    # Row 3: Headers
    headers_left = [
        "No", "Order Number", "Customer", "Post Office",
        "Status", "Type", "Age (Days)", "Penalty Fine ($)"
    ]
    headers_right = [
        "No", "Post Office", "RIGHT Handover", "RIGHT Delivery",
        "Total Handover", "Total Delivery", "% RIGHT Handover", "% RIGHT Delivery", "Total Penalty ($)"
    ]

    ws1.row_dimensions[3].height = 25.0
    for ci, h in enumerate(headers_left, 1):
        cell = ws1.cell(3, ci, h)
        cell.font = font_hdr
        cell.fill = fill_hdr_left
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_clean

    for ci, h in enumerate(headers_right, 10):
        cell = ws1.cell(3, ci, h)
        cell.font = font_hdr
        cell.fill = fill_hdr_right
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_clean

    # Populate Left Detail Order Rows
    r_curr = 4
    tot_fine_left = 0.0

    for idx, item in enumerate(base_rows, 1):
        ws1.row_dimensions[r_curr].height = 18.0
        fine_text = f"-${item['penalty_fine']:.2f}" if item['penalty_fine'] > 0 else "$0.00"
        row_vals = [
            idx,
            item["order_number"],
            item["customer"],
            item["assigned_branch"],
            item["status_name"][:20],
            item["type"],
            f"{item['age_days']} d",
            fine_text
        ]
        for col_idx, val in enumerate(row_vals, 1):
            c = ws1.cell(row=r_curr, column=col_idx, value=val)
            c.font = font_data
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = border_clean
            if r_curr % 2 == 0:
                c.fill = fill_row_alt

        tot_fine_left += item["penalty_fine"]
        r_curr += 1

    # Left Grand Total
    ws1.row_dimensions[r_curr].height = 24.0
    ws1.merge_cells(start_row=r_curr, start_column=1, end_row=r_curr, end_column=2)
    gt_left = ws1.cell(r_curr, 1, f"Total Orders: {len(base_rows)}")
    gt_left.font = font_tot
    gt_left.alignment = Alignment(horizontal="left", vertical="center")

    for c in range(1, 9):
        cell = ws1.cell(r_curr, c)
        cell.fill = fill_left_tot
        cell.border = tot_border_accounting
        if c == 8:
            cell.value = f"-${tot_fine_left:.2f}" if tot_fine_left > 0 else "$0.00"
            cell.font = font_tot
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Populate Right Executive Summary Table
    # SORT FROM WORSE TO BEST (% ON-TIME ASCENDING)
    def calc_sort_key(stats):
        tot_ops = stats["total_handover"] + stats["total_delivery"]
        r_ho = max(0, stats["total_handover"] - stats["penalty_handover"])
        r_del = max(0, stats["total_delivery"] - stats["penalty_delivery"])
        tot_r = r_ho + r_del
        overall_pct = (tot_r / tot_ops * 100.0) if tot_ops > 0 else 100.0
        # Worse first: lowest overall_pct, then largest fine (descending), then volume
        return (overall_pct, -stats["total_fine"], -tot_ops)

    if tgt in ("ALL", "TOTAL"):
        all_branches = [summary_data[b] for b in MAIN_36_BRANCHES if b in summary_data]
    else:
        all_branches = list(summary_data.values())

    sorted_branches = sorted(all_branches, key=calc_sort_key)

    r_sum = 4
    n_idx = 1
    tot_ho = 0
    tot_del = 0
    tot_pen_ho = 0
    tot_pen_del = 0
    tot_fine = 0.0

    for stats in sorted_branches:
        ws1.row_dimensions[r_sum].height = 18.0
        
        # Calculate RIGHT (On-Time / Excused)
        r_ho = max(0, stats["total_handover"] - stats["penalty_handover"])
        r_del = max(0, stats["total_delivery"] - stats["penalty_delivery"])
        
        pct_r_ho = (r_ho / stats["total_handover"] * 100) if stats["total_handover"] > 0 else 100.0
        pct_r_del = (r_del / stats["total_delivery"] * 100) if stats["total_delivery"] > 0 else 100.0
        fine_str = f"-${stats['total_fine']:.2f}" if stats['total_fine'] > 0 else "$0.00"

        s_vals = [
            n_idx,
            stats["po"],
            r_ho,
            r_del,
            stats["total_handover"],
            stats["total_delivery"],
            f"{pct_r_ho:.1f}%",
            f"{pct_r_del:.1f}%",
            fine_str
        ]
        for ci, val in enumerate(s_vals, 10):
            cell = ws1.cell(r_sum, ci, val)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border_clean
            
            # Dynamic Font Coloring on % Columns
            if ci == 16: # % RIGHT Handover
                cell.font = get_pct_font(pct_r_ho)
            elif ci == 17: # % RIGHT Delivery
                cell.font = get_pct_font(pct_r_del)
            elif ci in (11, 18):
                cell.font = font_bold_data
            else:
                cell.font = font_data

            if r_sum % 2 == 0:
                cell.fill = fill_row_alt

        tot_ho += stats["total_handover"]
        tot_del += stats["total_delivery"]
        tot_pen_ho += stats["penalty_handover"]
        tot_pen_del += stats["penalty_delivery"]
        tot_fine += stats["total_fine"]
        r_sum += 1
        n_idx += 1

    # Right Grand Total Row
    ws1.row_dimensions[r_sum].height = 24.0
    ws1.merge_cells(start_row=r_sum, start_column=10, end_row=r_sum, end_column=11)
    rt_tot = ws1.cell(r_sum, 10, "Grand Total")
    rt_tot.font = font_tot
    rt_tot.alignment = Alignment(horizontal="left", vertical="center")

    tot_r_ho = max(0, tot_ho - tot_pen_ho)
    tot_r_del = max(0, tot_del - tot_pen_del)
    tot_pct_r_ho = (tot_r_ho / tot_ho * 100) if tot_ho > 0 else 100.0
    tot_pct_r_del = (tot_r_del / tot_del * 100) if tot_del > 0 else 100.0
    tot_fine_str = f"-${tot_fine:.2f}" if tot_fine > 0 else "$0.00"

    tot_vals_right = [
        "", "",
        tot_r_ho,
        tot_r_del,
        tot_ho,
        tot_del,
        f"{tot_pct_r_ho:.1f}%",
        f"{tot_pct_r_del:.1f}%",
        tot_fine_str
    ]
    for c in range(10, 19):
        cell = ws1.cell(r_sum, c)
        if c >= 12:
            cell.value = tot_vals_right[c-10]
        
        # Color Grand Total %
        if c == 16:
            cell.font = get_pct_font(tot_pct_r_ho, is_tot=True)
        elif c == 17:
            cell.font = get_pct_font(tot_pct_r_del, is_tot=True)
        else:
            cell.font = font_tot

        cell.fill = fill_sum_tot
        cell.border = tot_border_accounting
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Generous Column Widths
    col_widths = {
        1: 5, 2: 15, 3: 20, 4: 12, 5: 18, 6: 10, 7: 12, 8: 15,
        9: 4,
        10: 6, 11: 14, 12: 16, 13: 16, 14: 16, 15: 16, 16: 18, 17: 18, 18: 18
    }
    for c, w in col_widths.items():
        ws1.column_dimensions[get_column_letter(c)].width = w

    # Sheet 2: base
    ws2 = wb.create_sheet(title="base")
    ws2.views.sheetView[0].showGridLines = True

    base_headers = [
        "No", "Order Number", "Customer", "Origin Branch", "Origin Post",
        "Destination Branch", "Destination Post", "Assigned Branch", "Created At",
        "Last Action Time", "Status Code", "Status Name", "Type", "Age (Days)",
        "Penalty Fine ($)", "Risk Level", "Is Excused"
    ]

    ws2.row_dimensions[1].height = 22
    for col_idx, h in enumerate(base_headers, 1):
        c = ws2.cell(row=1, column=col_idx, value=h)
        c.font = font_hdr
        c.fill = fill_hdr_left
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border_clean

    for idx, item in enumerate(base_rows, 1):
        r_num = idx + 1
        ws2.row_dimensions[r_num].height = 18
        fine_text = f"-${item['penalty_fine']:.2f}" if item['penalty_fine'] > 0 else "$0.00"
        row_data = [
            idx,
            item["order_number"],
            item["customer"],
            item["origin_branch"],
            item["origin_post"],
            item["destination_branch"],
            item["destination_post"],
            item["assigned_branch"],
            item["created_at"],
            item["last_action_time"],
            item["status_code"],
            item["status_name"],
            item["type"],
            item["age_days"],
            fine_text,
            item["risk_level"],
            "YES" if item["is_excused"] else "NO"
        ]
        for col_idx, val in enumerate(row_data, 1):
            c = ws2.cell(row=r_num, column=col_idx, value=val)
            c.font = font_data
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = border_clean

    for col in ws2.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws2.column_dimensions[col_letter].width = max(max_len + 3, 11)

    wb.save(out_xlsx)
    return tot_ho, tot_del, (tot_pen_ho + tot_pen_del), tot_fine


def render_penalty_summary_image(out_xlsx):
    wb = openpyxl.load_workbook(out_xlsx)
    ws = wb['INVENTORY PENALTY REPORT']

    wb_sum = openpyxl.Workbook()
    ws_sum = wb_sum.active
    ws_sum.title = 'Executive Summary'
    ws_sum.views.sheetView[0].showGridLines = True

    max_r = 1
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, 10).value is not None or ws.cell(r, 18).value is not None:
            max_r = r

    for r in range(1, max_r + 1):
        if ws.row_dimensions[r].height:
            ws_sum.row_dimensions[r].height = ws.row_dimensions[r].height
        for c_idx in range(9):
            orig_c = 10 + c_idx
            tgt_c = 1 + c_idx
            cell_orig = ws.cell(r, orig_c)
            cell_tgt = ws_sum.cell(r, tgt_c, cell_orig.value)
            if cell_orig.has_style:
                cell_tgt.font = copy.copy(cell_orig.font)
                cell_tgt.fill = copy.copy(cell_orig.fill)
                cell_tgt.border = copy.copy(cell_orig.border)
                cell_tgt.alignment = copy.copy(cell_orig.alignment)

    # Wide column dimensions for crisp unclipped header text
    col_widths = {1: 6, 2: 14, 3: 16, 4: 16, 5: 16, 6: 16, 7: 18, 8: 18, 9: 18}
    for c, w in col_widths.items():
        ws_sum.column_dimensions[get_column_letter(c)].width = w

    ws_sum.merge_cells("A1:I1")
    ws_sum.merge_cells("A2:I2")
    ws_sum.merge_cells(start_row=max_r, start_column=1, end_row=max_r, end_column=2)

    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_f:
        tmp_path = tmp_f.name
    wb_sum.save(tmp_path)

    try:
        buf = excel_to_image.excel_to_image(tmp_path)
        return buf
    finally:
        try:
            os.remove(tmp_path)
        except Exception:
            pass
