import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime

def build_shipments_tomorrow_report(src_xlsx, out_xlsx, target_label="Zone 1"):
    """
    Builds CEO-Level Executive SHIPMENTS TOMORROW REPORT Excel file:
      - Sheet 1: SHIPMENTS TOMORROW REPORT (Left main table + Right Executive Summary table)
      - Sheet 2: base (Raw order dataset)
    Reports all active transit orders (Status 306, 309, 302, 310, 311) destined for the target zone/branch.
    """
    import pandas as pd
    df = pd.read_excel(src_xlsx)
    df.columns = [str(c).strip().upper() for c in df.columns]

    # Find columns safely
    col_order = next((c for c in df.columns if 'ORDER ID' in c or 'ORDER_NUMBER' in c), 'ORDER ID')
    col_dest_prov = next((c for c in df.columns if 'DELIVERY PROVINCE' in c or 'DESTINATION_BRANCH' in c), 'DELIVERY PROVINCE')
    col_dest_po = next((c for c in df.columns if 'DELIVERY POST' in c or 'DESTINATION_POST' in c), 'DELIVERY POST OFFICE')
    col_orig_br = next((c for c in df.columns if 'ACTION POST OFFICE' in c or 'ORIGIN_BRANCH' in c), 'ACTION POST OFFICE')
    col_orig_po = next((c for c in df.columns if 'CURRENT POST OFFICE' in c or 'ORIGIN_POST' in c), 'CURRENT POST OFFICE')
    col_created = next((c for c in df.columns if 'CREATED DATE' in c or 'CREATED_AT' in c), 'CREATED DATE')
    col_status = next((c for c in df.columns if 'CURRENT STATUS' in c or 'STATUS' in c), 'CURRENT STATUS')
    col_weight = next((c for c in df.columns if 'WEIGHT' in c), 'WEIGHT (G)')
    col_fee = next((c for c in df.columns if 'TOTAL FEE' in c or 'TOTAL_AMOUNT' in c), 'TOTAL FEE (USD) (4)=(1)+(2)-(3)')
    col_vas = next((c for c in df.columns if 'VAS FEE' in c), 'VAS FEE (USD) (2)')
    col_cod = next((c for c in df.columns if 'COD' in c), 'COD (USD)')
    col_receiver = next((c for c in df.columns if 'RECEIVER' in c), 'RECEIVER')
    col_service = next((c for c in df.columns if 'SERVICE' in c), 'SERVICE')
    col_pay_meth = next((c for c in df.columns if 'PAYMENT METHOD' in c), 'PAYMENT METHOD')

    # Status Code extraction
    df['sc'] = df[col_status].astype(str).str.extract(r'^(\d{3})')[0]
    
    # Active transit shipments only (Status 306, 309, 302, 310, 311)
    df_active = df[df['sc'].isin(['306', '309', '302', '310', '311'])].copy()

    # Zone / Post office filtering
    tgt = target_label.upper().replace(" ", "")
    zone_by_prefix = {
        "KAN": "ZONE1", "PNP": "ZONE1", "PRE": "ZONE1", "SVA": "ZONE1",
        "KAM": "ZONE2", "KOH": "ZONE2", "SIH": "ZONE2", "SPE": "ZONE2", "TAK": "ZONE2", "KEP": "ZONE2",
        "BAN": "ZONE3", "BAT": "ZONE3", "CHH": "ZONE3", "PUR": "ZONE3", "PAI": "ZONE3",
        "ODD": "ZONE4", "PRH": "ZONE4", "SIE": "ZONE4", "THO": "ZONE4",
        "CHA": "ZONE5", "KRA": "ZONE5", "TBK": "ZONE5", "ROT": "ZONE5", "MON": "ZONE5", "STU": "ZONE5"
    }

    PROVINCIAL_BRANCH_CODES = {
        'BANP001', 'BATP001', 'CHHP001', 'PURP001', 'SIEP001', 'PRHP001',
        'ODDP001', 'THOP001', 'SIHP001', 'KOHP001', 'KAMP001', 'SPEP001',
        'TAKP001', 'TBKP001', 'CHAP001', 'KRAP001', 'STUP001', 'ROTP001',
        'MONP001', 'PREP001', 'SVAP001', 'PAIP001', 'KEPP001'
    }

    df_active['dest_prov_clean'] = df_active[col_dest_prov].astype(str).str.strip().str.upper()
    df_active['dest_po_clean'] = df_active[col_dest_po].astype(str).str.strip().str.upper()

    if tgt.startswith("ZONE"):
        target_zone_name = tgt if len(tgt) > 4 else "ZONE1"
        df_active['zone'] = df_active['dest_prov_clean'].map(zone_by_prefix).fillna("ZONE1")
        df_matched = df_active[df_active['zone'] == target_zone_name].copy()
    elif tgt in ("ALL", "TOTAL", "MEGA", "BRANCH", "BRANCHES"):
        df_matched = df_active.copy()
    elif tgt in PROVINCIAL_BRANCH_CODES or (len(tgt) == 3 and tgt in zone_by_prefix and tgt not in ("PNP", "KAN")):
        df_matched = df_active[
            (df_active['dest_prov_clean'] == tgt[:3]) |
            (df_active['dest_po_clean'].str.startswith(tgt[:3]))
        ].copy()
    elif len(tgt) >= 7 or (len(tgt) > 3 and tgt[3:4] in ("P", "A", "S")):
        df_matched = df_active[df_active['dest_po_clean'] == tgt].copy()
    else:
        df_matched = df_active[df_active['dest_prov_clean'].str.startswith(tgt[:3])].copy()

    base_rows = []
    r_idx = 1
    for _, row in df_matched.iterrows():
        order_id = str(row.get(col_order, '') or '').strip()
        if not order_id or order_id == 'nan':
            continue

        dest_prov = str(row.get(col_dest_prov, '') or '').strip().upper()
        dest_po = str(row.get(col_dest_po, '') or '').strip().upper()
        orig_br = str(row.get(col_orig_br, '') or '').strip().upper()
        orig_po = str(row.get(col_orig_po, '') or '').strip().upper()
        created = str(row.get(col_created, '') or '').strip()
        status = str(row.get(col_status, '') or '').strip()
        receiver = str(row.get(col_receiver, '') or '').strip()
        service = str(row.get(col_service, '') or '').strip().upper()
        pay_meth = str(row.get(col_pay_meth, '') or '').strip()

        try:
            weight = float(row.get(col_weight, 0) or 0)
        except (ValueError, TypeError):
            weight = 0.0

        try:
            fee = float(row.get(col_fee, 0) or 0)
        except (ValueError, TypeError):
            fee = 0.0

        try:
            vas_fee = float(row.get(col_vas, 0) or 0)
        except (ValueError, TypeError):
            vas_fee = 0.0

        try:
            cod = float(row.get(col_cod, 0) or 0)
        except (ValueError, TypeError):
            cod = 0.0


        # Determine VAS Code & VAS Khmer Description
        vas_codes = []
        vas_khmer_list = []

        if "NGƯỜI GỬI" in pay_meth.upper() or "SENDER" in pay_meth.upper():
            vas_codes.append("NTN")
            vas_khmer_list.append("អ្នកផ្ញើបង់")
        if vas_fee > 0:
            vas_codes.append("VBH")
            vas_khmer_list.append("ធានារ៉ាប់រង")
        if cod > 0:
            vas_codes.append("VBP")
            vas_khmer_list.append("ទារប្រាក់")
        if not vas_codes or service in ("CCN", "CLT"):
            vas_codes.append("VTT")
            vas_khmer_list.append("ដឹកដល់ផ្ទះ")

        vas_code_str = ", ".join(vas_codes)
        vas_khmer_str = ", ".join(vas_khmer_list)

        # Full district lookup — 3-level: exact PO code → 5-char prefix → province fallback
        DISTRICT_BY_PO = {
            # ── Phnom Penh (PNP) ──
            'PNPP001': 'Chamkar Mon',     'PNPP007': 'Boeng Keng Kang',
            'PNPA002': 'Boeng Keng Kang', 'PNPS007': 'Boeng Keng Kang',
            'PNPP005': 'Chbar Ampov',     'PNPA005': 'Chbar Ampov',
            'PNPP010': 'Chraoy Chongvar', 'PNPA010': 'Chraoy Chongvar',
            'PNPP011': 'Dangkao',         'PNPA011': 'Dangkao',
            'PNPP014': 'Doun Penh',       'PNPA014': 'Doun Penh',
            'PNPP012': 'Kamboul',         'PNPA016': 'Kamboul', 'PNPA012': 'Kamboul',
            'PNPP002': 'Mean Chey',       'PNPP003': 'Mean Chey',
            'PNPA029': 'Mean Chey',       'PNPA055': 'Mean Chey',
            'PNPA003': 'Mean Chey',
            'PNPP008': 'Pou Saen Chey',   'PNPP009': 'Pou Saen Chey',
            'PNPA028': 'Pou Saen Chey',   'PNPA008': 'Pou Saen Chey',
            'PNPP004': 'Preaek Pnov',     'PNPA004': 'Preaek Pnov',
            'PNPP013': 'Saen Sokh',       'PNPA040': 'Saen Sokh', 'PNPA013': 'Saen Sokh',
            'PNPP006': 'Tuol Kouk',       'PNPA036': 'Tuol Kouk', 'PNPA006': 'Tuol Kouk',
            'PNPP015': 'Russey Keo',      'PNPA015': 'Russey Keo',
            'PNPP016': 'Por Senchey',

            # ── Kandal (KAN) ──
            'KANP001': 'Ta Khmau',
            'KANA001': 'Ta Khmau',         'KANA023': 'Ta Khmau',
            'KANA002': 'Kandal Stueng',    'KANA024': 'Kandal Stueng',
            'KANA028': 'Kandal Stueng',    'KANA049': 'Kandal Stueng',
            'KANA003': 'Kaoh Thum',        'KANS003': 'Kaoh Thum',
            'KANA004': 'Khsach Kandal',    'KANA031': 'Khsach Kandal',
            'KANA005': 'Kien Svay',        'KANA012': 'Kien Svay',   'KANA013': 'Kien Svay',
            'KANA006': 'Leuk Daek',        'KANA008': 'Leuk Daek',
            'KANA007': "S'ang",            'KANA020': "S'ang",       'KANA026': "S'ang",
            'KANA009': 'Mukh Kampul',      'KANS004': 'Mukh Kampul',
            'KANA010': 'Ponhea Lueu',      'KANA019': 'Ponhea Lueu',
            'KANA017': 'Lvea Em',          'KANA018': 'Ang Snuol',
            'KANA040': 'Sampov Pun',

            # ── Prey Veng (PRE) ──
            'PREP001': 'Prey Veng Town',   'PRES001': 'Prey Veng Town',
            'PREA001': 'Prey Veng Town',   'PREA014': 'Prey Veng Town',
            'PREA036': 'Prey Veng Town',
            'PREA002': 'Pur Rieng',        'PREA016': 'Pur Rieng',
            'PREA003': 'Ba Phnum',         'PREA024': 'Ba Phnum',
            'PREA004': 'Kang Meas',        'PREA022': 'Kang Meas',
            'PREA005': 'Kampong Trabaek',
            'PREA006': 'Mesang',           'PREA021': 'Mesang',
            'PREA007': 'Peam Chor',        'PREA026': 'Peam Chor',
            'PREA008': 'Peam Ro',          'PREA023': 'Peam Ro',   'PREA032': 'Peam Ro',
            'PREA009': 'Pea Reang',        'PREA039': 'Pea Reang',
            'PREA010': 'Preah Sdach',      'PREA020': 'Preah Sdach',
            'PREA029': 'Preah Sdach',      'PREA035': 'Preah Sdach',
            'PREA030': 'Preah Sdach',      'PREA031': 'Preah Sdach',
            'PREA011': 'Sithor Kandal',    'PREA028': 'Sithor Kandal',
            'PREA012': 'Svay Antor',       'PREA025': 'Svay Antor',
            'PREA013': 'Svay Teab',        'PREA038': 'Svay Teab',
            'PREA017': 'Kamchay Mear',     'PREA037': 'Kamchay Mear',
            'PREA033': 'Ba Phnum East',    'PREA034': 'Mesang North',

            # ── Svay Rieng (SVA) ──
            'SVAP001': 'Svay Rieng Town',
            'SVAA001': 'Bavet',            'SVAA002': 'Bavet',   'SVAS002': 'Bavet',
            'SVAA003': 'Romeas Haek',
            'SVAA004': 'Rumduol',
            'SVAA005': 'Svay Chrum',
            'SVAA006': 'Chantrea',
            'SVAA007': 'Kampong Rou',
            'SVAA008': 'Svay Teab (SVA)',
            'SVAA009': 'Kong Pisei (SVA)',

            # ── Kampot (KAM) ──
            'KAMP001': 'Kampot Town',
            'KAMA001': 'Kampot Town',      'KAMA002': 'Angkor Chey',
            'KAMA003': 'Banteay Meas',     'KAMA004': 'Chhouk',
            'KAMA005': 'Dang Tong',        'KAMA006': 'Kampong Trach',
            'KAMA007': 'Prey Nob (KAM)',   'KAMA008': 'Tuek Chhou',

            # ── Koh Kong (KOH) ──
            'KOHP001': 'Koh Kong Town',
            'KOHA001': 'Koh Kong Town',    'KOHA002': 'Botum Sakor',
            'KOHA003': 'Kaoh Sdach',       'KOHA004': 'Mondol Seima',
            'KOHA005': 'Smach Mean Chey',  'KOHA006': 'Sre Ambel',
            'KOHA007': 'Thma Bang',        'KOHA008': 'Veal Veng (KOH)',
            'KOHA009': 'Khiri Sakor',

            # ── Takeo (TAK) ──
            'TAKP001': 'Takeo Town',
            'TAKA001': 'Takeo Town',       'TAKA002': 'Angkor Borei',
            'TAKA003': 'Bati',             'TAKA004': 'Bokor',
            'TAKA005': 'Daun Keo',         'TAKA006': 'Kirivong',
            'TAKA007': 'Prey Kabbas',      'TAKA008': 'Samraong (TAK)',
            'TAKA009': 'Treang',           'TAKA010': 'Tram Kak',

            # ── Battambang (BAT) ──
            'BATP001': 'Battambang Town',
            'BATA001': 'Battambang Town',  'BATA009': 'Battambang Town',
            'BATA040': 'Battambang Town',  'BATA042': 'Battambang Town',
            'BATA003': 'Banan',            'BATA016': 'Banan',
            'BATA011': 'Kamrieng',
            'BATA017': 'Samlout',
            'BATA010': 'Sampov Lun',       'BATA028': 'Sampov Lun',
            'BATA004': 'Sangkae',          'BATA008': 'Sangkae',
            'BATA023': 'Thma Koul',        'BATA025': 'Thma Koul',
            'BATS007': 'Moung Ruessei',
            'BATA002': 'Ek Phnom',         'BATA005': 'Phnom Proek',
            'BATA006': 'Rotanak Mondol',   'BATA012': 'Maung',
            'BATA013': 'Ou Reang Ov',      'BATA014': 'Pa Oy',

            # ── Banteay Meanchey (BAN) ──
            'BANP001': 'Serei Saophoan',
            'BANA001': 'Serei Saophoan',   'BANA002': 'Mongkol Borei',
            'BANA003': 'Malai',            'BANA004': 'Ou Chrov',
            'BANA005': 'Phnom Srok',       'BANA006': 'Poipet',
            'BANA007': 'Svay Chek',        'BANA008': 'Thma Puok',

            # ── Kampong Cham (CHA) ──
            'CHAP001': 'Kampong Cham Town',
            'CHAA001': 'Kampong Cham Town', 'CHAA002': 'Batheay',
            'CHAA003': 'Chamkar Leu',      'CHAA004': 'Cheung Prey',
            'CHAA005': 'Dambae',           'CHAA006': 'Kampong Siem',
            'CHAA007': 'Kang Meas (CHA)',  'CHAA008': 'Koh Sotin',
            'CHAA009': 'Memot',            'CHAA010': 'Mou Kinnh',
            'CHAA011': 'Prey Chhor',       'CHAA012': 'Srei Santhor',
            'CHAA013': 'Stueng Trang',     'CHAA015': 'Tbong Khmum Town',
            'CHAA017': 'Krouch Chhmar',    'CHAA018': 'Ponhea Kraek',
            'CHAA019': 'Suong',            'CHAA020': 'Tboung Khmum',
            'CHAA021': 'Ponhea Kraek West','CHAA022': 'Koh Sotin East',
            'CHAA023': 'Cheung Prey North','CHAA024': 'Kampong Siem South',
            'CHAA025': 'Batheay East',

            # ── Siem Reap (SIE) ──
            'SIEP001': 'Siem Reap Town',
            'SIEA001': 'Angkor Chum',      'SIEA002': 'Angkor Thon',
            'SIEA003': 'Banteay Srei',     'SIEA004': 'Chi Kraeng',
            'SIEA005': 'Kralanh',          'SIEA006': 'Prasat Bakong',
            'SIEA007': 'Puok',             'SIEA008': 'Soutr Nikom',
            'SIEA009': 'Srei Snam',        'SIEA010': 'Svay Leu',
            'SIEA011': 'Varin',

            # ── Sihanoukville (SIH) ──
            'SIHP001': 'Sihanoukville Town',
            'SIHA001': 'Preah Sihanouk',   'SIHA002': 'Stung Hav',
            'SIHA003': 'Kampong Seila',    'SIHA004': 'Prey Nob (SIH)',

            # ── Kampong Chhnang (CHH) ──
            'CHHP001': 'Kampong Chhnang Town',
            'CHHA001': 'Kampong Chhnang Town', 'CHHA002': 'Baribour',
            'CHHA003': 'Chol Kiri',        'CHHA004': 'Kampong Tralach',
            'CHHA005': 'Kirivong (CHH)',   'CHHA006': 'Oral',
            'CHHA007': 'Rolea Bier (CHH)', 'CHHA008': 'Sameakki Mean Chey',
            'CHHA009': 'Tuek Phos',

            # ── Kampong Thom (THO) ──
            'THOP001': 'Kampong Thom Town',
            'THOA001': 'Kampong Thom Town', 'THOA002': 'Baray',
            'THOA003': 'Kampong Svay',     'THOA004': 'Prasat Ballangk',
            'THOA005': 'Prasat Sambour',   'THOA006': 'Sandan',
            'THOA007': 'Santuk',           'THOA008': 'Stoung',

            # ── Pursat (PUR) ──
            'PURP001': 'Pursat Town',
            'PURA001': 'Pursat Town',      'PURA002': 'Bakan',
            'PURA003': 'Kandieng',         'PURA004': 'Krakor',
            'PURA005': 'Phnom Kravanh',    'PURA006': 'Veal Veng (PUR)',

            # ── Kratie (KRA) ──
            'KRAP001': 'Kratie Town',
            'KRAA001': 'Kratie Town',      'KRAA002': 'Chhloung',
            'KRAA003': 'Koh Nhek',         'KRAA004': 'Prek Prasab',
            'KRAA005': 'Sambour',          'KRAA006': 'Snuol',

            # ── Tbong Khmum (TBK) ──
            'TBKP001': 'Suong Town',
            'TBKA001': 'Suong Town',       'TBKA002': 'Krouch Chhmar (TBK)',
            'TBKA003': 'Memot (TBK)',      'TBKA004': 'Ponhea Kraek (TBK)',
            'TBKA005': 'Tbaeng Meanchey',  'TBKA006': 'Dambae (TBK)',

            # ── Kampong Speu (SPE) ──
            'SPEP001': 'Chbar Mon',
            'SPEA001': 'Chbar Mon',        'SPEA002': 'Basedth',
            'SPEA003': 'Kong Pisei',       'SPEA004': 'Odong',
            'SPEA005': 'Phnom Sruoch',     'SPEA006': 'Samraong Tong',
            'SPEA007': 'Thpong',

            # ── Preah Vihear (PRH) ──
            'PRHP001': 'Tbeng Meanchey',
            'PRHA001': 'Tbeng Meanchey',   'PRHA002': 'Chey Saen',
            'PRHA003': 'Choam Ksan',       'PRHA004': 'Kulen',
            'PRHA005': 'Rovieng',          'PRHA006': 'Sangkum Thmei',
            'PRHA007': 'Tbaeng',

            # ── Oddar Meanchey (ODD) ──
            'ODDP001': 'Samraong (ODD)',
            'ODDA001': 'Samraong (ODD)',   'ODDA002': 'Anlong Veng',
            'ODDA003': 'Banteay Ampil',    'ODDA004': 'Chong Kal',
            'ODDA005': 'Trapeang Prasat',

            # ── Kep (KEP) ──
            'KEPP001': 'Kep Town',
            'KEPA001': 'Kep Town',         'KEPA002': 'Damnak Chang Aeur',

            # ── Pailin (PAI) ──
            'PAIP001': 'Pailin Town',
            'PAIA001': 'Pailin Town',      'PAIA002': 'Sala Krau',

            # ── Mondulkiri (MON) ──
            'MONP001': 'Sen Monorom',
            'MONA001': 'Sen Monorom',      'MONA002': 'Kaoh Nheak',
            'MONA003': 'Keo Seima',        'MONA004': 'Pech Chreada',

            # ── Ratanakiri (ROT) ──
            'ROTP001': 'Banlung',
            'ROTA001': 'Banlung',          'ROTA002': 'Andong Meas',
            'ROTA003': 'Bar Kaev',         'ROTA004': 'Koun Mom',
            'ROTA005': 'Lumphat',          'ROTA006': 'O Chum',
            'ROTA007': 'Ou Ya Dav',        'ROTA008': 'Ta Veng',

            # ── Stung Treng (STU) ──
            'STUP001': 'Stung Treng Town',
            'STUA001': 'Stung Treng Town', 'STUA002': 'Sesan',
            'STUA003': 'Siem Bouk',        'STUA004': 'Siem Pang',
            'STUA005': 'Thala Barivat',    'STUA006': 'Voeun Sai',


            # ── Transit Hubs (not end-delivery) ──
            'DVCMEGA1': 'MEGA Hub (Transit)',
            'DVCMEGA2': 'MEGA Hub (Transit)',
            'PNPMEGA1': 'MEGA Hub PNP (Transit)',

            # ── Banteay Meanchey (BAN) – Agents / Showrooms ──
            'BANA014': 'Koub',           'BANA015': 'Changha',
            'BANS002': 'Paoy Paet',

            # ── Battambang (BAT) – Agents / Showrooms ──
            'BATA007': 'Snoeng',         'BATA015': 'Sdok Pravoek',
            'BATA018': 'Kampong Lpov',   'BATA019': 'Traeng',
            'BATA022': 'Bavel',          'BATA027': 'Ou Ta Ki',
            'BATA037': 'Peam Aek',       'BATA041': 'Moung Ruessei',

            # ── Kampong Cham (CHA) – Agents / Showrooms ──
            'CHAS002': 'Preak Kak',      'CHAS005': 'Svay Teab',

            # ── Kampong Chhnang (CHH) – Agents / Showrooms ──
            'CHHA023': 'Rolea Bier',

            # ── Kampot (KAM) – Agents / Showrooms ──
            'KAMA013': 'Trapeang Reang', 'KAMA020': 'Kampong Trach Khang Kaeut',
            'KAMA022': 'Praphnum',       'KAMS002': 'Tuk Meas Khang Lech',
            'KAMS003': 'Satv Pong',

            # ── Kandal (KAN) – Agents / Showrooms ──
            'KANS001': 'Ta Khmau',       'KANS005': 'Baek Chan',

            # ── Kratie (KRA) – Agents / Showrooms ──
            'KRAS002': 'Snuol',          'KRAS004': 'Sambour',

            # ── Mondulkiri (MON) – Agents / Showrooms ──
            'MONA006': 'Srae Ampum',     'MONS001': 'Spean Mean Chey',

            # ── Oddar Meanchey (ODD) – Agents / Showrooms ──
            'ODDA006': 'Samraong (ODD)', 'ODDA008': 'Kouk Mon',

            # ── Pailin (PAI) – Showrooms ──
            'PAIS001': 'Pailin Town',

            # ── Phnom Penh (PNP) – Agents / Showrooms ──
            'PNPA031': 'Dangkao',        'PNPA051': 'Tuek L\'ak',
            'PNPA088': 'Chaom Chau',
            'PNPS002': 'Boeng Kak',      'PNPS003': 'Nirouth',
            'PNPS004': 'Tuek Thla',      'PNPS005': 'Chaom Chau',
            'PNPS006': 'Srah Chak',      'PNPS008': 'Chaom Chau',

            # ── Prey Veng (PRE) – Agents / Showrooms ──
            'PREA015': 'Banteay Chakrei','PREA027': 'Chheang Seangkhin',
            'PRES004': 'Chheu Kach',

            # ── Preah Vihear (PRH) – Agents / Showrooms ──
            'PRHA008': 'Chamraeun',      'PRHA009': 'S\'ang',
            'PRHA015': 'Phnum Tbaeng',

            # ── Pursat (PUR) – Agents / Showrooms ──
            'PURA013': 'Ou Saom',        'PURA015': 'Ou Sandan',
            'PURS002': 'Boeng Khnar',

            # ── Ratanakiri (ROT) – Agents / Showrooms ──
            'ROTA009': 'Kachanh',        'ROTA011': 'Chey Otdam',
            'ROTS001': 'Labansiek',      'ROTS002': 'La Minh',

            # ── Siem Reap (SIE) – Agents / Showrooms ──
            'SIEA014': 'Chrouy Neang Nguon', 'SIEA022': 'Puok',
            'SIEA034': 'Anlung Samnar',  'SIEA036': 'Kaev Poar',
            'SIES001': 'Kouk Chak',      'SIES002': 'Siem Reab',
            'SIES003': 'Dam Daek',

            # ── Sihanoukville (SIH) – Agents / Showrooms ──
            'SIHA006': 'Cheung Kou',     'SIHA008': 'Lekh Muoy',
            'SIHS001': 'Lekh Buon',

            # ── Kampong Speu (SPE) – Agents / Showrooms ──
            'SPEA009': 'Traeng Trayueng','SPEA010': 'Pneay',
            'SPEA017': 'Khsem Khsan',    'SPEA020': 'Prey Sralaeng',
            'SPEA023': 'Trapeang Chou',  'SPES002': 'Snam Krapeu',

            # ── Stung Treng (STU) – Showrooms ──
            'STUS001': 'Stung Treng Town',

            # ── Svay Rieng (SVA) – Agents ──
            'SVAA019': 'Svay Chrum',

            # ── Takeo (TAK) – Agents / Showrooms ──
            'TAKA022': 'Kandoeng',       'TAKA033': 'Prey Kabbas',
            'TAKS002': 'Ang Ta Saom',

            # ── Tbong Khmum (TBK) – Agents / Showrooms ──
            'TBKA010': 'Dambae',         'TBKS003': 'Memut',
            'TBKS004': 'Ampil Ta Pok',

            # ── Kampong Thom (THO) – Agents / Showrooms ──
            'THOA012': 'Steung Saen',    'THOA015': 'Kampong Chen',
            'THOA017': 'Steung Saen',    'THOS002': 'Ballang',
            'THOA011': 'Sala Visai',     'THOA013': 'Chhuk',
            'THOA014': 'Krava',          'THOA016': 'Chamna Kraom',
            'THOA018': 'Steung Saen',    'THOS001': 'Prey Ta Hu',
            'THOS003': 'Kampong Chen',   'THOS004': 'Sandan',

            # ── Kampong Speu (SPE) – Agents / Showrooms (batch 2) ──
            'SPEA011': 'Rung Roeang',    'SPEA012': 'Amleang',
            'SPEA014': 'Chambak',        'SPEA015': 'Roleang Kreul',
            'SPEA019': 'Prey Kmeng',     'SPEA028': 'Veang Chas',
            'SPEA030': 'Khtum Krang',    'SPEA032': 'Roleang Kreul',
            'SPES001': 'Rokar Thum',

            # ── Battambang (BAT) – Agents (batch 2) ──
            'BATA021': 'Preaek Chik',    'BATA026': 'Svay Por',
            'BATA029': 'Thipakdei',      'BATA030': 'Prey Svay',
            'BATA035': 'Anlung Vil',     'BATA036': 'Vaot Ta Muem',
            'BATA038': 'Prey Khpos',     'BATA039': 'Mukh Rea',

            # ── Banteay Meanchey (BAN) – Agents (batch 2) ──
            'BANA009': 'Chub Vari',      'BANA011': 'Kampong Svay',
            'BANA016': 'Phnum Lieb',     'BANA020': 'Banteay Neang',
            'BANA021': 'Banteay Meanchey',
            'BANS001': 'Ou Ambel',       'BANS003': 'Thma Puok',
            'BANS004': 'Srah Chik',

            # ── Prey Veng (PRE) – Agents (batch 2) ──
            'PREA019': 'Kampong Trabaek','PREA041': 'Prey Veng',
            'PREA043': 'Prey Veng',      'PREA046': 'Prey Veng',
            'PRES004': 'Chheu Kach',     'PRES005': 'Roka',

            # ── Preah Vihear (PRH) – Agents (batch 2) ──
            'PRHA013': 'Rummeakney',     'PRHA016': 'Srayang',
            'PRHS001': 'Kampong Pranak',

            # ── Pursat (PUR) – Agents (batch 2) ──
            'PURA008': 'Anlung Tnaot',   'PURA009': 'Me Tuek',
            'PURA014': 'Snam Preah',     'PURA016': 'Ta Lou',
            'PURA017': 'Boeng Khnar',

            # ── Ratanakiri (ROT) – Agents (batch 2) ──
            'ROTA014': 'Trapeang Kraham','ROTA016': 'Trapeang Kraham',
            'ROTA017': 'Srae Angkrong',  'ROTS003': 'Boeng Kansaeng',

            # ── Siem Reap (SIE) – Agents (batch 2) ──
            'SIEA030': 'Char Chhuk',     'SIEA033': 'Svay Sa',
            'SIEA035': 'Sla Kram',       'SIEA037': 'Ruessei Lok',

            # ── Sihanoukville (SIH) – Agents (batch 2) ──
            'SIHA005': 'Andoung Thma',   'SIHA009': 'Stueng Chhay',

            # ── Mondulkiri (MON) – Agents (batch 2) ──
            'MONA005': 'Pu Chrey',       'MONA007': 'Srae Khtum',
            'MONA008': 'Srae Chhuk',

            # ── Oddar Meanchey (ODD) – Agents (batch 2) ──
            'ODDA007': 'Chong Kal',      'ODDS002': 'Anlung Veaeng',

            # ── Svay Rieng (SVA) – Agents (batch 2) ──
            'SVAA013': 'Chres',          'SVAA020': 'Pouthi',
            'SVAA022': 'Bos Svay',       'SVAA025': 'Chher Teal',
            'SVAA028': 'Sambour',        'SVAA029': 'Pong Tuek',
            'SVAA030': 'Chantrei',       'SVAS001': 'Svay Rieng Town',
            'SVAS004': 'Kampong Trach',

            # ── Takeo (TAK) – Agents (batch 2) ──
            'TAKA020': 'Prey Khla',      'TAKA023': 'Rovieng',
            'TAKA028': 'Borei Chulsar',  'TAKS003': 'Trapeang Sab',

            # ── Tbong Khmum (TBK) – Agents (batch 2) ──
            'TBKA009': 'Trapeang Phlong','TBKA011': 'Chong Cheach',
            'TBKA016': 'Kraek',          'TBKA021': 'Sralab',
            'TBKS001': 'Suong',          'TBKS002': 'Kraek',

            # ── Kampong Cham (CHA) – Agents (batch 2) ──
            'CHAS001': 'Veal Vung',      'CHAS003': 'Soutib',

            # ── Kampong Chhnang (CHH) – Agents (batch 2) ──
            'CHHA012': 'Trangel',        'CHHA015': 'Chrey Bak',
            'CHHA016': 'Chieb',          'CHHA020': 'Krang Lvea',
            'CHHS001': 'Kampong Chhnang Town',

            # ── Kampot (KAM) – Agents (batch 2) ──
            'KAMA017': 'Trapeang Sangkae','KAMA021': 'Trapeang Sala',
            'KAMA028': 'Kampot',         'KAMA029': 'Dambouk Khpuos',
            'KAMA031': 'Kampong Trach',  'KAMS001': 'Krang Ampil',

            # ── Kandal (KAN) – Agents (batch 2) ──
            'KANA011': 'Cheung Kaeub',   'KANA041': 'Leuk Daek',
            'KANA043': 'Puk Ruessei',    'KANS002': 'Kokir',

            # ── Kratie (KRA) – Agents (batch 2) ──
            'KRAA007': 'Ou Krieng',      'KRAA009': 'Sandan',
            'KRAS003': 'Chhloung',       'KRAS005': 'Preaek Prasab',

            # ── Phnom Penh (PNP) – Agents (batch 2) ──
            'PNPA017': 'Toul Svay Prey', 'PNPA020': 'Toul Svay Prey',
            'PNPA034': 'Dangkao',        'PNPA035': 'Ruessei Kaev',
            'PNPA052': 'Chak Angrae',    'PNPA074': 'Tuek Thla',
            'PNPA080': 'Phleung Chheh',  'PNPA086': 'Phnom Penh',
            'PNPA096': 'Tuek L\'ak',

            # ── Stung Treng (STU) – Agents (batch 2) ──
            'STUA009': 'Anlung Chrey',

            # ── Koh Kong (KOH) – Showrooms ──
            'KOHS001': 'Smach Mean Chey','DVCZ5':   'Koh Kong (Transit)',
        }

        PROVINCE_NAME_MAP = {
            'PNP': 'Phnom Penh',    'KAN': 'Kandal',
            'PRE': 'Prey Veng',     'SVA': 'Svay Rieng',
            'BAT': 'Battambang',    'SIE': 'Siem Reap',
            'SIH': 'Sihanoukville', 'KOH': 'Koh Kong',
            'KAM': 'Kampot',        'TAK': 'Takeo',
            'PUR': 'Pursat',        'PRH': 'Preah Vihear',
            'TBK': 'Tbong Khmum',   'THO': 'Kampong Thom',
            'CHA': 'Kampong Cham',  'KRA': 'Kratie',
            'BAN': 'Banteay Meanchey', 'CHH': 'Kampong Chhnang',
            'MON': 'Mondulkiri',    'ROT': 'Ratanakiri',
            'STU': 'Stung Treng',   'ODD': 'Oddar Meanchey',
            'KEP': 'Kep',           'PAI': 'Pailin',
            'SPE': 'Kampong Speu',
        }

        # Level 1: exact 7-char post office code
        if dest_po in DISTRICT_BY_PO:
            dist_name = DISTRICT_BY_PO[dest_po]
        # Level 2: 5-char prefix (e.g. PREA0 → covers PREA036, PREA037 etc.)
        elif len(dest_po) >= 5 and dest_po[:5] in DISTRICT_BY_PO:
            dist_name = DISTRICT_BY_PO[dest_po[:5]]
        # Level 3: province-code fallback
        else:
            br_prefix = dest_prov.upper()[:3] if dest_prov else (dest_po[:3] if dest_po else '')
            dist_name = PROVINCE_NAME_MAP.get(
                br_prefix,
                dest_prov if dest_prov and dest_prov.upper() not in ('NONE', 'NAN', '') else 'Unknown District'
            )




        base_rows.append({
            "no": r_idx,
            "order_number": order_id,
            "customer": receiver[:30],
            "origin_branch": orig_br,
            "origin_post": orig_po,
            "destination_branch": dest_prov,
            "destination_post": dest_po,
            "created_at": created,
            "fee": fee,
            "cod": cod,
            "weight_g": weight,
            "status": status,
            "receiver": receiver,
            "vas_code": vas_code_str,
            "vas_khmer": vas_khmer_str,
            "district": dist_name,
            "zone": {
                "KAN": "Zone 1", "PNP": "Zone 1", "PRE": "Zone 1", "SVA": "Zone 1",
                "KAM": "Zone 2", "KOH": "Zone 2", "SIH": "Zone 2", "SPE": "Zone 2", "TAK": "Zone 2", "KEP": "Zone 2",
                "BAN": "Zone 3", "BAT": "Zone 3", "CHH": "Zone 3", "PUR": "Zone 3", "PAI": "Zone 3",
                "ODD": "Zone 4", "PRH": "Zone 4", "SIE": "Zone 4", "THO": "Zone 4",
                "CHA": "Zone 5", "KRA": "Zone 5", "TBK": "Zone 5", "ROT": "Zone 5", "MON": "Zone 5", "STU": "Zone 5"
            }.get(dest_prov, "Zone 1")
        })
        r_idx += 1

    wb = openpyxl.Workbook()

    # Sheet 1: SHIPMENTS TOMORROW REPORT
    ws1 = wb.active
    ws1.title = "SHIPMENTS TOMORROW REPORT"
    ws1.views.sheetView[0].showGridLines = True

    # Executive CEO Color Palette (Subtle, High-Contrast, Professional)
    fill_title_left  = PatternFill("solid", fgColor="0F172A") # Deep Slate Navy
    fill_title_right = PatternFill("solid", fgColor="0F766E") # Deep Teal Slate
    fill_hdr_left    = PatternFill("solid", fgColor="1E293B") # Executive Navy Slate
    fill_hdr_right   = PatternFill("solid", fgColor="0F766E") # Deep Teal Slate
    fill_row_alt     = PatternFill("solid", fgColor="F8FAFC") # Subtle Zebra Tint
    fill_left_tot    = PatternFill("solid", fgColor="CBD5E1") # Refined Slate Grey Total
    fill_sum_tot     = PatternFill("solid", fgColor="CCFBF1") # Refined Soft Teal Total

    border_clean = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0")
    )

    # Double-line Accounting Bottom Border for CEO Grand Total
    tot_border_accounting = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="64748B"),
        bottom=Side(style="double", color="0F172A") # Executive Double Line
    )

    font_banner = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_hdr    = Font(name="Segoe UI", size=9,  bold=True, color="FFFFFF")
    font_data   = Font(name="Segoe UI", size=9,  color="0F172A")
    font_data_b = Font(name="Segoe UI", size=9,  bold=True, color="0F172A")
    font_tot    = Font(name="Segoe UI", size=10, bold=True, color="0F172A")
    font_tot_red= Font(name="Segoe UI", size=10, bold=True, color="991B1B")

    # Row 1: Title Banners (Height 36)
    stamp_date = datetime.now().strftime("%d.%m")
    target_clean = target_label.upper()
    title_left_txt = f"SHIPMENTS TOMORROW REPORT {stamp_date} (Báo cáo hàng đến {target_clean})"
    title_right_txt= f"EXECUTIVE SUMMARY ({target_clean} / {target_clean[:3]})"

    ws1.merge_cells("A1:H1")
    ws1.cell(1, 1, title_left_txt).font = font_banner
    ws1.cell(1, 1).alignment = Alignment(horizontal="left", vertical="center")
    for c in range(1, 9):
        ws1.cell(1, c).fill = fill_title_left

    ws1.merge_cells("J1:N1")
    ws1.cell(1, 10, title_right_txt).font = font_banner
    ws1.cell(1, 10).alignment = Alignment(horizontal="center", vertical="center")
    for c in range(10, 15):
        ws1.cell(1, c).fill = fill_title_right

    ws1.row_dimensions[1].height = 36.0

    # Row 2: Header Rows (Height 32)
    headers_left = [
        "DESTINATION\n(សាខា)",
        "District\n(ស្រុក/ខណ្ឌ)",
        "DESTINATION_POS\n(បូស្តិ៍គោលដៅ)",
        "ORDER_NUMBER\n(លេខវិក្កយបត្រ)",
        "Receiver\n(អ្នកទទួល)",
        "SUM ACTUAL_WEIGHT (G)\n(ទម្ងន់សរុប g)",
        "VAS\n(សេវា)",
        "VAS Description\n(ឈ្មោះសេវាបន្ថែម)"
    ]
    headers_right = [
        "ZONE\n(តំបន់)",
        "DESTINATION_BRANCH\n(សាខា)",
        "District\n(ស្រុក/ខណ្ឌ)",
        "Bill\n(ចំនួនប័ណ្ណ)",
        "SUM ACTUAL_WEIGHT (G)\n(ទម្ងន់សរុប g)"
    ]

    ws1.row_dimensions[2].height = 32.0
    for ci, h in enumerate(headers_left, 1):
        cell = ws1.cell(2, ci, h)
        cell.font = font_hdr
        cell.fill = fill_hdr_left
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_clean

    for ci, h in enumerate(headers_right, 10):
        cell = ws1.cell(2, ci, h)
        cell.font = font_hdr
        cell.fill = fill_hdr_right
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_clean

    # Populate Left Data Rows (Single pure white background for all data rows)
    summary_data = {}
    total_bills = 0
    total_weight = 0.0

    r_curr = 3
    for idx_row, item in enumerate(base_rows):
        ws1.row_dimensions[r_curr].height = 20.0

        vals = [
            item["destination_branch"],
            item["district"],
            item["destination_post"],
            item["order_number"],
            item["receiver"],
            item["weight_g"],
            item["vas_code"],
            item["vas_khmer"]
        ]
        for ci, val in enumerate(vals, 1):
            cell = ws1.cell(r_curr, ci, val)
            cell.font = font_data_b if ci == 1 else font_data
            cell.border = border_clean


            if ci in (1, 2, 3, 4):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif ci in (5, 7, 8):
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif ci == 6:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = "#,##0"

        key = (item["zone"], item["destination_branch"], item["district"])
        summary_data.setdefault(key, {"bills": 0, "weight": 0.0})
        summary_data[key]["bills"] += 1
        summary_data[key]["weight"] += item["weight_g"]

        total_bills += 1
        total_weight += item["weight_g"]
        r_curr += 1

    # Left Grand Total Row (Refined CEO Double-Line Accounting Finish)
    ws1.row_dimensions[r_curr].height = 25.0
    ws1.merge_cells(start_row=r_curr, start_column=1, end_row=r_curr, end_column=5)
    gt_left = ws1.cell(r_curr, 1, "Grand Total / សរុប")
    gt_left.font = font_tot
    gt_left.alignment = Alignment(horizontal="left", vertical="center")
    for c in range(1, 6):
        cell = ws1.cell(r_curr, c)
        cell.fill = fill_left_tot
        cell.border = tot_border_accounting

    gt_w_cell = ws1.cell(r_curr, 6, total_weight)
    gt_w_cell.font = font_tot_red
    gt_w_cell.fill = fill_left_tot
    gt_w_cell.border = tot_border_accounting
    gt_w_cell.alignment = Alignment(horizontal="right", vertical="center")
    gt_w_cell.number_format = "#,##0"

    for c in (7, 8):
        cell = ws1.cell(r_curr, c)
        cell.fill = fill_left_tot
        cell.border = tot_border_accounting

    # Populate Executive Summary Table on Right with Province/Branch Subtotals
    r_sum = 3
    branch_groups = {}
    for (zone_str, br, dist), stats in sorted(summary_data.items()):
        if br not in branch_groups:
            branch_groups[br] = []
        branch_groups[br].append((zone_str, br, dist, stats))

    sub_fill = PatternFill("solid", fgColor="E0F2FE")
    sub_border = Border(
        top=Side(style="thin", color="94A3B8"),
        bottom=Side(style="thin", color="94A3B8"),
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0")
    )

    for br in sorted(branch_groups.keys()):
        br_items = branch_groups[br]
        br_bills = 0
        br_weight = 0

        for zone_str, b_code, dist, stats in br_items:
            ws1.row_dimensions[r_sum].height = 20.0
            s_vals = [zone_str, b_code, dist, stats["bills"], stats["weight"]]
            for ci, val in enumerate(s_vals, 10):
                cell = ws1.cell(r_sum, ci, val)
                cell.font = font_data
                cell.border = border_clean
                if ci in (10, 11, 12):
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif ci in (13, 14):
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    if ci == 14:
                        cell.number_format = "#,##0"
            br_bills += stats["bills"]
            br_weight += stats["weight"]
            r_sum += 1

        # Branch Subtotal Row (e.g. KAN Total, PNP Total, PRE Total, SVA Total)
        ws1.row_dimensions[r_sum].height = 22.0
        ws1.merge_cells(start_row=r_sum, start_column=10, end_row=r_sum, end_column=12)
        sub_lbl = ws1.cell(r_sum, 10, f"{br} Total")
        sub_lbl.font = Font(name="Segoe UI", size=10, bold=True, color="0F172A")
        sub_lbl.alignment = Alignment(horizontal="left", vertical="center")

        for c in range(10, 13):
            cell = ws1.cell(r_sum, c)
            cell.fill = sub_fill
            cell.border = sub_border

        sub_b_cell = ws1.cell(r_sum, 13, br_bills)
        sub_b_cell.font = Font(name="Segoe UI", size=10, bold=True, color="0F172A")
        sub_b_cell.fill = sub_fill
        sub_b_cell.border = sub_border
        sub_b_cell.alignment = Alignment(horizontal="right", vertical="center")
        sub_b_cell.number_format = "#,##0"

        sub_w_cell = ws1.cell(r_sum, 14, br_weight)
        sub_w_cell.font = Font(name="Segoe UI", size=10, bold=True, color="991B1B")
        sub_w_cell.fill = sub_fill
        sub_w_cell.border = sub_border
        sub_w_cell.alignment = Alignment(horizontal="right", vertical="center")
        sub_w_cell.number_format = "#,##0"
        r_sum += 1

    # Right Summary Total Row (CEO Double-Line Accounting Finish)
    ws1.row_dimensions[r_sum].height = 25.0
    ws1.merge_cells(start_row=r_sum, start_column=10, end_row=r_sum, end_column=12)
    tot_label_cell = ws1.cell(r_sum, 10, f"{target_clean[:3]} Total")
    tot_label_cell.font = font_tot
    tot_label_cell.alignment = Alignment(horizontal="left", vertical="center")
    for c in range(10, 13):
        cell = ws1.cell(r_sum, c)
        cell.fill = fill_sum_tot
        cell.border = tot_border_accounting

    tot_b_cell = ws1.cell(r_sum, 13, total_bills)
    tot_b_cell.font = font_tot
    tot_b_cell.fill = fill_sum_tot
    tot_b_cell.border = tot_border_accounting
    tot_b_cell.alignment = Alignment(horizontal="right", vertical="center")

    tot_w_cell = ws1.cell(r_sum, 14, total_weight)
    tot_w_cell.font = font_tot_red
    tot_w_cell.fill = fill_sum_tot
    tot_w_cell.border = tot_border_accounting
    tot_w_cell.alignment = Alignment(horizontal="right", vertical="center")
    tot_w_cell.number_format = "#,##0"

    # Exact Column Widths matching example file
    exact_widths = {
        'A': 16.0, 'B': 18.0, 'C': 20.0, 'D': 18.0, 'E': 35.0, 'F': 24.0,
        'G': 14.0, 'H': 26.0, 'I': 4.0,  'J': 12.0, 'K': 22.0, 'L': 18.0,
        'M': 12.0, 'N': 24.0
    }
    for col_let, w in exact_widths.items():
        ws1.column_dimensions[col_let].width = w

    # Sheet 2: base (Raw order dataset)
    ws2 = wb.create_sheet(title="base")
    ws2.views.sheetView[0].showGridLines = True

    base_headers = [
        "No", "ORDER_NUMBER", "CUSTOMER", "ORIGIN_BRANCH", "ORIGIN_POST",
        "DESTINATION_BRANCH", "DESTINATION_POST", "CREATED_BY", "CREATED_AT",
        "PAYMENT_METHOD", "SHIPPING_FEE", "DISCOUNT", "SERVICE_FEE",
        "TOTAL_FEE", "COD", "WEIGHT_G", "LENGTH", "WIDTH",
        "VAS_CODE", "VAS_DESCRIPTION", "DESTINATION_PROVINCE", "DESTINATION_DISTRICT",
        "STATUS", "RECEIVER_NAME"
    ]
    ws2.append(base_headers)
    ws2.row_dimensions[1].height = 24.0
    for c in range(1, len(base_headers) + 1):
        cell = ws2.cell(1, c)
        cell.font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0F172A")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for item in base_rows:
        row_data = [
            item["no"],
            item["order_number"],
            item["customer"],
            item["origin_branch"],
            item["origin_post"],
            item["destination_branch"],
            item["destination_post"],
            item["customer"],
            item["created_at"],
            "Sender",
            item["fee"],
            0.0,
            0.0,
            item["fee"],
            item["cod"],
            item["weight_g"],
            "",
            "",
            item["vas_code"],
            item["vas_khmer"],
            item.get("province", item.get("destination_branch", "")),
            item["district"],
            item["status"],
            item["receiver"]
        ]
        ws2.append(row_data)


    wb.save(out_xlsx)
    return total_bills, total_weight


def render_executive_summary_image(out_xlsx):
    """Renders ONLY the small right Executive Summary table to a pixel-perfect PNG image."""
    import tempfile, copy, openpyxl, excel_to_image
    from openpyxl.utils import get_column_letter

    wb = openpyxl.load_workbook(out_xlsx)
    ws = wb['SHIPMENTS TOMORROW REPORT']

    # Create 1-table Executive Summary workbook
    wb_sum = openpyxl.Workbook()
    ws_sum = wb_sum.active
    ws_sum.title = 'Executive Summary'
    ws_sum.views.sheetView[0].showGridLines = True

    # Find max row in summary table (Col J=10, Col M=13)
    max_r = 1
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, 10).value is not None or ws.cell(r, 13).value is not None:
            max_r = r

    # Copy Cols J to N (10 to 14) into Cols A to E (1 to 5)
    for r in range(1, max_r + 1):
        if ws.row_dimensions[r].height:
            ws_sum.row_dimensions[r].height = ws.row_dimensions[r].height
        for c_idx in range(5):
            orig_c = 10 + c_idx
            target_c = 1 + c_idx
            cell_orig = ws.cell(r, orig_c)
            cell_tgt  = ws_sum.cell(r, target_c, cell_orig.value)
            
            if cell_orig.has_style:
                cell_tgt.font = copy.copy(cell_orig.font)
                cell_tgt.fill = copy.copy(cell_orig.fill)
                cell_tgt.border = copy.copy(cell_orig.border)
                cell_tgt.alignment = copy.copy(cell_orig.alignment)
                cell_tgt.number_format = cell_orig.number_format

    # Copy all merged ranges for Cols J..N (10..14) -> A..E (1..5)
    for m_range in ws.merged_cells.ranges:
        if m_range.min_col >= 10 and m_range.max_col <= 14:
            new_min_c = m_range.min_col - 9
            new_max_c = m_range.max_col - 9
            ws_sum.merge_cells(
                start_row=m_range.min_row,
                end_row=m_range.max_row,
                start_column=new_min_c,
                end_column=new_max_c
            )

    # Column Widths
    col_widths = [14, 22, 18, 12, 24]
    for ci, w in enumerate(col_widths, 1):
        ws_sum.column_dimensions[get_column_letter(ci)].width = w

    tmp_sum_dir = tempfile.mkdtemp()
    tmp_sum_xlsx = os.path.join(tmp_sum_dir, 'exec_summary_only.xlsx')
    wb_sum.save(tmp_sum_xlsx)

    # Render to pixel-perfect cropped image using excel_to_image
    return excel_to_image.excel_to_image(tmp_sum_xlsx)
