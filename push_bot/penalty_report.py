import os
import sys
import copy
import tempfile
import re
from datetime import datetime, date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import excel_to_image

MAIN_36_BRANCHES = [
    'BANP001', 'BATP001', 'CHAP001', 'CHHP001', 'KAMP001', 'KANP001', 'KOHP001', 'KRAP001',
    'MONP001', 'ODDP001', 'PNPP001', 'PNPP002', 'PNPP003', 'PNPP004', 'PNPP005', 'PNPP006',
    'PNPP007', 'PNPP008', 'PNPP009', 'PNPP010', 'PNPP011', 'PNPP012', 'PNPP013', 'PNPP014',
    'PREP001', 'PRHP001', 'PURP001', 'ROTP001', 'SIEP001', 'SIHP001', 'SPEP001', 'STUP001',
    'SVAP001', 'TAKP001', 'TBKP001', 'THOP001'
]

ZONE_BRANCHES_MAP = {
    "ZONE1": ['PNPP001', 'PNPP002', 'PNPP003', 'PNPP004', 'PNPP005', 'PNPP006', 'PNPP007', 'PNPP008', 'PNPP009', 'PNPP010', 'PNPP011', 'PNPP012', 'PNPP013', 'PNPP014', 'KANP001', 'PREP001', 'SVAP001'],
    "ZONE2": ['KAMP001', 'KOHP001', 'SIHP001', 'SPEP001', 'TAKP001'],
    "ZONE3": ['BANP001', 'BATP001', 'CHHP001', 'PURP001'],
    "ZONE4": ['ODDP001', 'PRHP001', 'SIEP001', 'THOP001'],
    "ZONE5": ['CHAP001', 'KRAP001', 'MONP001', 'ROTP001', 'STUP001', 'TBKP001'],
}

STATUS_NAME_EN = {
    '110': 'Order Created (Not Collected)',
    '120': 'Assigned to Pickup Staff',
    '200': 'Pending Pickup / Collecting',
    '210': 'Pickup Collected',
    '230': 'Pickup Failed',
    '300': 'Assigned to Bag / Transit',
    '302': 'Bag / Transit Completed',
    '306': 'In Storage / Handover',
    '309': 'Received at Post Office',
    '310': 'Packing / Sorting',
    '311': 'In Transit to Hub',
    '400': 'Assigned to Rider (Pending)',
    '401': 'Delivery Dispatched (Completed)',
    '402': 'Confirmed Dispatch',
    '410': 'Delivered Successfully',
    '420': 'Rescheduled / Customer Appointment',
    '430': 'Delivery Failed (Undelivered)',
    '460': 'Return Notice Created',
    '470': 'Return in Progress',
    '471': 'Checking Customer Info',
    '472': 'Resolving Delivery Issue',
    '480': 'Confirming New Address',
    '500': 'Out for Return to Sender',
    '510': 'Return Received at Post',
    '511': 'Return in Transit',
    '512': 'Return Dispatch Confirmed',
    '520': 'Returned to Hub',
    '540': 'Return Completed to Merchant',
}

def parse_date(val):
    if not val or pd.isna(val):
        return None
    if isinstance(val, (datetime, date)):
        return val.date() if isinstance(val, datetime) else val
    s = str(val).strip().split(" ")[0]
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None

def parse_time(val):
    if not val or pd.isna(val):
        return None
    if isinstance(val, datetime):
        return val
    s = str(val).strip()
    for fmt in ("%d/%m/%Y %H:%M:%S", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M", "%Y-%m-%d %H:%M"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None

def load_test_bills(cfg=None):
    """Load ignored/test bill IDs from test_bills.txt, delayed_bills.json, and test_receipts Excel file."""
    test_ids = set()
    base_dirs = [
        os.getcwd(),
        os.path.dirname(os.path.abspath(__file__)),
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    ]
    for d in base_dirs:
        txt_path = os.path.join(d, "test_bills.txt")
        if os.path.exists(txt_path):
            try:
                with open(txt_path, "r", encoding="utf-8") as f:
                    for line in f:
                        val = line.strip()
                        if val and not val.startswith("#"):
                            clean_val = str(val).strip().upper()
                            clean_val = re.sub(r'\.0$', '', clean_val)
                            if clean_val:
                                test_ids.add(clean_val)
            except Exception:
                pass

    for d in base_dirs:
        delay_path = os.path.join(d, "delayed_bills.json")
        if os.path.exists(delay_path):
            try:
                import json
                with open(delay_path, "r", encoding="utf-8") as f:
                    delayed = json.load(f)
                today_d = datetime.now().date()
                for bill_id, exp_date_str in delayed.items():
                    try:
                        exp_date = datetime.strptime(exp_date_str, "%Y-%m-%d").date()
                        if today_d < exp_date:
                            clean_val = str(bill_id).strip().upper()
                            clean_val = re.sub(r'\.0$', '', clean_val)
                            if clean_val:
                                test_ids.add(clean_val)
                    except Exception:
                        pass
            except Exception:
                pass

    excel_paths = []
    if cfg and isinstance(cfg, dict):
        tr_cfg = cfg.get("test_receipts", {})
        if tr_cfg.get("enabled") and tr_cfg.get("path"):
            excel_paths.append(tr_cfg.get("path"))
    else:
        for d in base_dirs:
            cfg_path = os.path.join(d, "config.json")
            if os.path.exists(cfg_path):
                try:
                    import json
                    with open(cfg_path, "r", encoding="utf-8") as f:
                        c = json.load(f)
                    tr_cfg = c.get("test_receipts", {})
                    if tr_cfg.get("enabled") and tr_cfg.get("path"):
                        excel_paths.append(tr_cfg.get("path"))
                        break
                except Exception:
                    pass

    for d in base_dirs:
        tx_path = os.path.join(d, "test.xlsx")
        if os.path.exists(tx_path) and tx_path not in excel_paths:
            excel_paths.append(tx_path)

    for ep in excel_paths:
        if os.path.exists(ep):
            try:
                import pandas as pd
                if ep.lower().endswith((".xlsx", ".xls")):
                    df_test = pd.read_excel(ep, dtype=str)
                else:
                    df_test = pd.read_csv(ep, dtype=str, keep_default_na=False)
                df_test = df_test.fillna("")
                if not df_test.empty:
                    order_col = next(
                        (
                            c for c in df_test.columns
                            if any(k in str(c).lower() for k in ("order", "code", "bill", "phi", "shipment"))
                        ),
                        df_test.columns[0]
                    )
                    for val in df_test[order_col].tolist():
                        clean_val = str(val).strip().upper()
                        clean_val = re.sub(r'\.0$', '', clean_val)
                        if clean_val and clean_val != "NAN":
                            test_ids.add(clean_val)
            except Exception:
                pass

    return test_ids

def build_penalty_report(src_xlsx, out_xlsx, target_label="ALL", report_date=None):
    """
    CEO Executive Penalty Dashboard:
      - Sorted: From WORST performing branch (% On-Time) to BEST
      - Dynamic Font Coloring on % Columns:
          * < 75.0%: Bold Red (#DC2626)
          * 75.0% - 89.9%: Bold Amber (#D97706)
          * >= 90.0%: Bold Green (#16A34A)
      - Exact 9 Columns: No | Post Office | RIGHT Handover | RIGHT Delivery | Total Handover | Total Delivery | % RIGHT Handover | % RIGHT Delivery | Total Penalty ($)
    """
    os.makedirs(os.path.dirname(os.path.abspath(out_xlsx)), exist_ok=True)
    df = pd.read_excel(src_xlsx)
    df.columns = [str(c).strip().upper() for c in df.columns]

    col_order = next((c for c in df.columns if 'ORDER ID' in c or 'ORDER' in c), 'ORDER ID')
    col_dest_prov = next((c for c in df.columns if 'DELIVERY PROVINCE' in c or 'DESTINATION_BRANCH' in c), 'DELIVERY PROVINCE')
    col_dest_po = next((c for c in df.columns if 'DELIVERY POST' in c or 'DESTINATION_POST' in c), 'DELIVERY POST OFFICE')
    col_orig_br = next((c for c in df.columns if 'ACTION POST OFFICE' in c or 'ORIGIN_BRANCH' in c), 'ACTION POST OFFICE')
    col_orig_po = next((c for c in df.columns if 'CURRENT POST OFFICE' in c or 'ORIGIN_POST' in c), 'CURRENT POST OFFICE')
    col_status = next((c for c in df.columns if 'CURRENT STATUS' in c or 'STATUS' in c), 'CURRENT STATUS')
    col_created = next((c for c in df.columns if 'CREATED DATE' in c), 'CREATED DATE')
    col_receiver = next((c for c in df.columns if 'RECEIVER' in c), 'RECEIVER')
    col_action_time = next((c for c in df.columns if 'ACTION TIME' in c or 'CURRENT TIME' in c), 'CURRENT TIME')

    # Exclude all testing bills from test_bills.txt and delayed_bills.json
    test_bills = load_test_bills()
    if test_bills and col_order in df.columns:
        order_series = df[col_order].astype(str).str.strip().str.upper().str.replace(r'\.0$', '', regex=True)
        df = df[~order_series.isin(test_bills)].copy()

    # Exclude orders with test keywords in Order ID, Sender, Receiver, Remark, Note, or Description
    test_keywords = ['test', 'kiểm thử', 'kiem thu', 'demo', 'trial', 'sample', 'dummy', 'thử nghiệm', 'thu nghiem']
    test_mask = pd.Series(False, index=df.index)

    if col_order in df.columns:
        ord_lower = df[col_order].astype(str).str.lower()
        for kw in test_keywords:
            test_mask |= ord_lower.str.contains(kw, na=False)

    check_cols = [c for c in df.columns if any(k in str(c).upper() for k in ('SENDER', 'RECEIVER', 'NOTE', 'REMARK', 'GOODS', 'DESC', 'COMMODITY', 'ITEM', 'CUSTOMER'))]
    for c in check_cols:
        if c in df.columns:
            s_lower = df[c].astype(str).str.lower()
            for kw in test_keywords:
                test_mask |= s_lower.str.contains(kw, na=False)

    if test_mask.any():
        df = df[~test_mask].copy()

    today = report_date or datetime.now().date()
    tgt = "".join(c for c in str(target_label).upper() if c.isalnum() or c in ("-", "_")).strip()
    if not tgt:
        tgt = "ALL"

    df['sc'] = df[col_status].astype(str).str.extract(r'^(\d{3})')[0]

    def get_actual_handover_po(r):
        sc_val = str(r.get('sc', ''))
        # Return statuses (500, 510, 511, 512, 540) -> last post office that scanned it (CURRENT POST OFFICE / latest action PO)
        if sc_val in ('500', '510', '511', '512', '540'):
            for col in ['ACTION POST OFFICE.4', 'ACTION POST OFFICE.3', 'ACTION POST OFFICE.2', 'ACTION POST OFFICE.1', 'ACTION POST OFFICE', col_orig_po]:
                act_po = str(r.get(col, '') or '').strip().upper()
                if act_po and act_po not in ('NAN', 'MEGA1', 'DVCMEGA1') and 'HUB' not in act_po:
                    return act_po
            return str(r.get(col_orig_po, '') or '').strip().upper()
        # 1. If status is 110, 120, 200, check RECEIVE POST OFFICE or ORIGIN_POST
        if sc_val in ('110', '120', '200'):
            for col in ['RECEIVE POST OFFICE', 'ORIGIN_POST', col_orig_po]:
                act_po = str(r.get(col, '') or '').strip().upper()
                if act_po and act_po not in ('NAN', 'MEGA1', 'DVCMEGA1') and 'HUB' not in act_po:
                    return act_po
        # 2. If status is 210, check ACTION POST OFFICE (under STATUS 210 TIME)
        elif sc_val == '210':
            act_po = str(r.get('ACTION POST OFFICE', '') or '').strip().upper()
            if act_po and act_po != 'NAN':
                return act_po
        # 3. If status is 302 or 310, check ACTION POST OFFICE.1 then ACTION POST OFFICE
        elif sc_val in ('302', '310'):
            for col in ['ACTION POST OFFICE.1', 'ACTION POST OFFICE']:
                act_po = str(r.get(col, '') or '').strip().upper()
                if act_po and act_po != 'NAN':
                    return act_po
        # 4. If status is 306 or 311, check ACTION POST OFFICE.2, .1, etc.
        elif sc_val in ('306', '311'):
            for col in ['ACTION POST OFFICE.2', 'ACTION POST OFFICE.1', 'ACTION POST OFFICE']:
                act_po = str(r.get(col, '') or '').strip().upper()
                if act_po and act_po != 'NAN':
                    return act_po
        # Fallback to CURRENT POST OFFICE
        cur = str(r.get(col_orig_po, '') or '').strip().upper()
        if cur in ('MEGA1', 'DVCMEGA1') or 'HUB' in cur:
            for col in ['RECEIVE POST OFFICE', 'ORIGIN_POST']:
                cand = str(r.get(col, '') or '').strip().upper()
                if cand and cand not in ('NAN', 'MEGA1', 'DVCMEGA1') and 'HUB' not in cand:
                    return cand
        return cur

    def get_action_user(r):
        sc_val = str(r.get('sc', ''))
        if sc_val in ('500', '510', '511', '512', '540'):
            for col in ['ACTION USER.4', 'ACTION USER.3', 'ACTION USER.2', 'ACTION USER.1', 'ACTION USER']:
                u = str(r.get(col, '') or '').strip()
                if u and u.lower() != 'nan':
                    return u
        # 1. If status is 210, check ACTION USER.1
        if sc_val == '210':
            for col in ['ACTION USER.1', 'ACTION USER']:
                u = str(r.get(col, '') or '').strip()
                if u and u.lower() != 'nan':
                    return u
        # 2. If status is 302 or 310, check ACTION USER.2
        elif sc_val in ('302', '310'):
            for col in ['ACTION USER.2', 'ACTION USER.1', 'ACTION USER']:
                u = str(r.get(col, '') or '').strip()
                if u and u.lower() != 'nan':
                    return u
        # 3. If status is 306 or 311, check ACTION USER.3, .2, .1
        elif sc_val in ('306', '311'):
            for col in ['ACTION USER.3', 'ACTION USER.2', 'ACTION USER.1', 'ACTION USER']:
                u = str(r.get(col, '') or '').strip()
                if u and u.lower() != 'nan':
                    return u
        # Default / Delivery: ACTION USER
        u = str(r.get('ACTION USER', '') or '').strip()
        return u if u.lower() != 'nan' else ""

    df['curr_po_clean'] = df.apply(get_actual_handover_po, axis=1)
    df['deliv_po_clean'] = df[col_dest_po].astype(str).str.strip().str.upper()

    # Load Post Office Handle mapping (to group agents under their 36 main branches)
    here = os.path.dirname(os.path.abspath(__file__))
    lookup_path = os.path.join(here, "post_office_lookup.csv")
    po_handle_map = {}
    if os.path.exists(lookup_path):
        import csv
        try:
            with open(lookup_path, encoding="utf-8", errors="ignore") as f:
                for r_csv in csv.reader(f):
                    if len(r_csv) >= 2 and r_csv[0].strip() and r_csv[1].strip():
                        po_handle_map[r_csv[0].strip().upper()] = r_csv[1].strip().upper()
        except Exception:
            pass

    def map_po_to_main(raw_code):
        c = str(raw_code).strip().upper()
        if not c or c == 'NAN':
            return None
        if c in MAIN_36_BRANCHES:
            return c
        mapped = po_handle_map.get(c)
        if mapped and mapped in MAIN_36_BRANCHES:
            return mapped
        if len(c) >= 3:
            prov = f"{c[:3]}P001"
            if prov in MAIN_36_BRANCHES:
                return prov
            if c.startswith("PNP"):
                return mapped or "PNPP001"
            if c.startswith("PAI"):
                return "BATP001"
            if c.startswith("KEP"):
                return "KAMP001"
            if c.startswith("TBK"):
                return "CHAP001"
        return mapped or c

    # EXCLUDE ONLY TERMINAL / COMPLETED / CANCELLED STATUSES (matching old report)
    excluded_statuses = {'410', '520', '201', '99', '100', '-99'}
    active_df = df[~df['sc'].isin(excluded_statuses)].copy()

    # Also exclude delivered / returned keywords in status text
    if col_status in active_df.columns:
        for kw in ['GIAO THÀNH CÔNG', 'DELIVERED', 'COMPLETED', 'ĐÃ GIAO', 'DA GIAO', 'RETURN COMPLETED']:
            active_df = active_df[~active_df[col_status].astype(str).str.upper().str.contains(kw, na=False)].copy()

    customer_delay_statuses = {'420', '471', '472', '480'}
    return_statuses = {'500', '510', '511', '512', '540'}
    excused_statuses = customer_delay_statuses

    summary_data = {}
    if tgt in ("ALL", "TOTAL"):
        for b in MAIN_36_BRANCHES:
            summary_data[b] = {
                "po": b,
                "total_handover": 0,
                "total_delivery": 0,
                "penalty_handover": 0,
                "penalty_delivery": 0,
                "excused_count": 0,
                "total_fine": 0.0
            }
    elif tgt.startswith("ZONE") and tgt in ZONE_BRANCHES_MAP:
        for b in ZONE_BRANCHES_MAP[tgt]:
            summary_data[b] = {
                "po": b,
                "total_handover": 0,
                "total_delivery": 0,
                "penalty_handover": 0,
                "penalty_delivery": 0,
                "excused_count": 0,
                "total_fine": 0.0
            }

    base_rows = []
    r_idx = 1

    for idx, row in active_df.iterrows():
        sc = str(row['sc'])
        curr_po = str(row.get('curr_po_clean', '')).strip()
        deliv_po = str(row.get('deliv_po_clean', '')).strip()

        is_return = sc in return_statuses
        is_delivery = sc.startswith('4') and not is_return
        is_handover = not (is_delivery or is_return)

        if is_return:
            # Return penalty MUST NOT go to sender! It goes to the last guy who scan (curr_po / CURRENT POST OFFICE)
            raw_po = curr_po
            if not raw_po or raw_po in ('MEGA1', 'DVCMEGA1') or 'HUB' in raw_po:
                for col in ['ACTION POST OFFICE.4', 'ACTION POST OFFICE.3', 'ACTION POST OFFICE.2', 'ACTION POST OFFICE.1', 'ACTION POST OFFICE']:
                    cand = str(row.get(col, '') or '').strip().upper()
                    if cand and cand not in ('NAN', 'MEGA1', 'DVCMEGA1') and 'HUB' not in cand:
                        raw_po = cand
                        break
        elif is_delivery:
            raw_po = deliv_po
        else:
            raw_po = curr_po
            if raw_po in ('MEGA1', 'DVCMEGA1') or 'HUB' in raw_po:
                for col in ['RECEIVE POST OFFICE', 'ORIGIN_POST', 'ACTION POST OFFICE']:
                    cand = str(row.get(col, '') or '').strip().upper()
                    if cand and cand not in ('NAN', 'MEGA1', 'DVCMEGA1') and 'HUB' not in cand:
                        raw_po = cand
                        break

        if not raw_po or raw_po == 'NAN':
            continue

        po = map_po_to_main(raw_po)
        if not po or po == 'NAN':
            continue

        # Target filtering
        if tgt in ("ALL", "TOTAL"):
            if po not in MAIN_36_BRANCHES:
                continue
        elif tgt.startswith("ZONE") and tgt in ZONE_BRANCHES_MAP:
            if po not in ZONE_BRANCHES_MAP[tgt]:
                continue
        elif len(tgt) >= 7:
            if po != tgt and raw_po != tgt:
                continue
            po = tgt
        elif len(tgt) == 3:
            if not (po.startswith(tgt) or raw_po.startswith(tgt)):
                continue

        if po not in summary_data:
            summary_data[po] = {
                "po": po,
                "total_handover": 0,
                "total_delivery": 0,
                "penalty_handover": 0,
                "penalty_delivery": 0,
                "excused_count": 0,
                "total_fine": 0.0
            }

        order_id = str(row.get(col_order, '')).strip()
        status_raw = str(row.get(col_status, '')).strip()

        # Smarter SLA Timing:
        # For delivery: measure age from physical arrival at delivery branch / dispatch
        # For handover: measure age from physical pickup scan time
        if is_delivery or is_return:
            arr_val = None
            for cand_col in [
                'STATUS 306 AT STORE / AGENT FROM HUB (FIRST TIME)',
                'STATUS 306 AT STORE / AGENT (LAST TIME)',
                col_action_time
            ]:
                if cand_col in row and pd.notna(row.get(cand_col)):
                    v = str(row.get(cand_col)).strip()
                    if v and v.lower() != 'nan':
                        arr_val = v
                        break
            act_val = arr_val or row.get(col_action_time) or row.get(col_created)
        else:
            p_val = None
            if sc in ('110', '120', '200'):
                p_val = row.get(col_created)
            elif sc == '210' and 'STATUS 210 TIME' in row and pd.notna(row.get('STATUS 210 TIME')):
                v = str(row.get('STATUS 210 TIME')).strip()
                if v and v.lower() != 'nan':
                    p_val = v
            act_val = p_val or row.get(col_action_time) or row.get(col_created)

        act_date = parse_date(act_val) or parse_date(row.get(col_created))
        age_days = (today - act_date).days if act_date else 0

        fine = 0.0
        risk_level = "Normal"
        is_excused = False

        if is_handover:
            summary_data[po]["total_handover"] += 1
            if age_days >= 3:
                fine = 0.40
                risk_level = "Urgent (> 3 days)"
                summary_data[po]["penalty_handover"] += 1
            elif age_days >= 1:
                fine = 0.10
                risk_level = "Backlog (1-2 days)"
                summary_data[po]["penalty_handover"] += 1
            else:
                risk_level = "Safe (< 1 day)"

        elif is_delivery or is_return:
            summary_data[po]["total_delivery"] += 1
            if sc in customer_delay_statuses:
                is_excused = True
                risk_level = f"Customer Delay ({sc})"
                summary_data[po]["excused_count"] += 1
                fine = 0.0
            else:
                if age_days >= 3:
                    fine = 0.40
                    risk_level = "Critical (> 3 days)"
                    summary_data[po]["penalty_delivery"] += 1
                elif age_days >= 1:
                    fine = 0.10
                    risk_level = "Stagnant (1-2 days)"
                    summary_data[po]["penalty_delivery"] += 1
                else:
                    risk_level = "Safe (< 1 day)"

        summary_data[po]["total_fine"] += fine

        action_user = get_action_user(row)
        staff_display = action_user
        if " - " in staff_display:
            staff_display = staff_display.split(" - ", 1)[1].strip()
        display_branch = f"{po} ({staff_display})" if staff_display else po

        base_rows.append({
            "no": r_idx,
            "order_number": order_id,
            "customer": str(row.get(col_receiver, ''))[:28],
            "origin_branch": str(row.get(col_orig_br, '')),
            "origin_post": curr_po,
            "destination_branch": str(row.get(col_dest_prov, '')),
            "destination_post": deliv_po,
            "assigned_branch": po,
            "display_branch": display_branch,
            "staff_user": action_user,
            "created_at": str(row.get(col_created, '')),
            "last_action_time": str(act_val or ""),
            "status_code": sc,
            "status_name": STATUS_NAME_EN.get(sc, "Processing"),
            "type": "Handover" if is_handover else "Delivery",
            "age_days": age_days,
            "penalty_fine": fine,
            "risk_level": risk_level,
            "is_excused": is_excused
        })
        r_idx += 1

    if tgt not in ("ALL", "TOTAL") and tgt not in summary_data:
        summary_data[tgt] = {
            "po": tgt,
            "total_handover": 0,
            "total_delivery": 0,
            "penalty_handover": 0,
            "penalty_delivery": 0,
            "excused_count": 0,
            "total_fine": 0.0
        }

    # Build Excel Workbook
    wb = openpyxl.Workbook()

    # Sheet 1: INVENTORY PENALTY REPORT
    ws1 = wb.active
    ws1.title = "INVENTORY PENALTY REPORT"
    ws1.views.sheetView[0].showGridLines = True

    # ── CEO MIDNIGHT & ROYAL SAPPHIRE BLUE PALETTE ──
    fill_title_left   = PatternFill("solid", fgColor="0F172A") # Midnight Navy
    fill_hdr_left     = PatternFill("solid", fgColor="1E293B") # Midnight Slate
    fill_title_right  = PatternFill("solid", fgColor="0B132B") # Deepest Midnight Navy
    fill_sub_right    = PatternFill("solid", fgColor="1C2541") # Subtitle Midnight Slate
    fill_hdr_right    = PatternFill("solid", fgColor="1C3D82") # Royal Sapphire Blue Header
    fill_row_white    = PatternFill("solid", fgColor="FFFFFF") # 100% Pure White (No 2 alternating colors)
    fill_left_tot     = PatternFill("solid", fgColor="E2E8F0") # Soft Slate Total
    fill_sum_tot      = PatternFill("solid", fgColor="E2E8F0") # Executive Accounting Total
    fill_penalty_pink = PatternFill("solid", fgColor="FFEAEA") # Light Pink / Soft Red for penalty cells

    border_clean = Border(
        left=Side(style="thin", color="CBD5E1"), right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"), bottom=Side(style="thin", color="CBD5E1")
    )
    tot_border_accounting = Border(
        left=Side(style="thin", color="CBD5E1"), right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="94A3B8"), bottom=Side(style="double", color="0B132B")
    )

    font_banner = Font(name="Segoe UI", size=10.5, bold=True, color="FFFFFF")
    font_sub = Font(name="Segoe UI", size=8.0, italic=True, color="93C5FD")
    font_hdr = Font(name="Segoe UI", size=8.5, bold=True, color="FFFFFF")
    font_data = Font(name="Segoe UI", size=8.5, color="0F172A")
    font_bold_data = Font(name="Segoe UI", size=8.5, bold=True, color="0F172A")
    font_tot = Font(name="Segoe UI", size=9.5, bold=True, color="0F172A")

    # High-contrast bold font colors for % on-time metrics
    font_pct_green = Font(name="Segoe UI", size=8.5, bold=True, color="16A34A") # >= 90%
    font_pct_amber = Font(name="Segoe UI", size=8.5, bold=True, color="D97706") # 75% - 89.9%
    font_pct_red   = Font(name="Segoe UI", size=8.5, bold=True, color="DC2626") # < 75%

    font_tot_pct_green = Font(name="Segoe UI", size=9.5, bold=True, color="16A34A")
    font_tot_pct_amber = Font(name="Segoe UI", size=9.5, bold=True, color="D97706")
    font_tot_pct_red   = Font(name="Segoe UI", size=9.5, bold=True, color="DC2626")

    def get_pct_font(pct_val, is_tot=False):
        if pct_val >= 90.0:
            return font_tot_pct_green if is_tot else font_pct_green
        elif pct_val >= 75.0:
            return font_tot_pct_amber if is_tot else font_pct_amber
        else:
            return font_tot_pct_red if is_tot else font_pct_red

    date_str = today.strftime('%d/%m/%Y')

    # 1. Left Title Banner
    ws1.merge_cells("A1:H1")
    ws1.cell(1, 1, f"METFONE EXPRESS — INVENTORY PENALTY & STAGNANT GOODS ({tgt}) — {date_str}").font = font_banner
    ws1.cell(1, 1).alignment = Alignment(horizontal="left", vertical="center")
    for c in range(1, 9):
        ws1.cell(1, c).fill = fill_title_left
    ws1.row_dimensions[1].height = 26.0

    # 2. Right Title Banner (Cols J to R)
    ws1.merge_cells("J1:R1")
    ws1.cell(1, 10, f"EXECUTIVE PENALTY DASHBOARD ({tgt})").font = font_banner
    ws1.cell(1, 10).alignment = Alignment(horizontal="center", vertical="center")
    for c in range(10, 19):
        ws1.cell(1, c).fill = fill_title_right

    ws1.merge_cells("J2:R2")
    ws1.cell(2, 10, "SLA Penalty: 1-2 Days (-$0.10) | ≥ 3 Days (-$0.40) • Excused 420/472 ($0.00)").font = font_sub
    ws1.cell(2, 10).alignment = Alignment(horizontal="center", vertical="center")
    for c in range(10, 19):
        ws1.cell(2, c).fill = fill_sub_right
    ws1.row_dimensions[2].height = 18.0

    # Row 3: Headers
    headers_left = [
        "No", "Order Number", "Customer", "Post Office (Staff)",
        "Status", "Type", "Age (Days)", "Penalty Fine ($)"
    ]
    headers_right = [
        "No", "Post Office", "RIGHT Handover", "RIGHT Delivery",
        "Total Handover", "Total Delivery", "% RIGHT Handover", "% RIGHT Delivery", "Total Penalty ($)"
    ]

    ws1.row_dimensions[3].height = 26.0
    for ci, h in enumerate(headers_left, 1):
        cell = ws1.cell(3, ci, h)
        cell.font = font_hdr
        cell.fill = fill_hdr_left
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_clean

    for ci, h in enumerate(headers_right, 10):
        cell = ws1.cell(3, ci, h)
        cell.font = font_hdr
        cell.fill = fill_hdr_right
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_clean

    # Populate Left Detail Order Rows (Only overdue / penalized bills)
    r_curr = 4
    tot_fine_left = 0.0

    overdue_rows = [item for item in base_rows if item["penalty_fine"] > 0]

    if overdue_rows:
        for idx, item in enumerate(overdue_rows, 1):
            ws1.row_dimensions[r_curr].height = 18.0
            fine_text = f"-${item['penalty_fine']:.2f}"
            row_vals = [
                idx,
                item["order_number"],
                item["customer"],
                item.get("display_branch", item["assigned_branch"]),
                item["status_code"],
                item["type"],
                item["age_days"],
                fine_text
            ]
            for col_idx, val in enumerate(row_vals, 1):
                c = ws1.cell(row=r_curr, column=col_idx, value=val)
                c.font = font_data
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = border_clean
                c.fill = fill_row_white

            tot_fine_left += item["penalty_fine"]
            r_curr += 1
    else:
        ws1.row_dimensions[r_curr].height = 22.0
        ws1.merge_cells(start_row=r_curr, start_column=1, end_row=r_curr, end_column=8)
        no_pen_cell = ws1.cell(r_curr, 1, "✓ No overdue or penalized bills")
        no_pen_cell.font = Font(name="Segoe UI", size=9.5, bold=True, color="16A34A")
        no_pen_cell.alignment = Alignment(horizontal="center", vertical="center")
        for c in range(1, 9):
            ws1.cell(r_curr, c).fill = fill_row_white
            ws1.cell(r_curr, c).border = border_clean
        r_curr += 1

    # Left Grand Total
    ws1.row_dimensions[r_curr].height = 24.0
    ws1.merge_cells(start_row=r_curr, start_column=1, end_row=r_curr, end_column=2)
    gt_left = ws1.cell(r_curr, 1, f"Total Overdue: {len(overdue_rows)}")
    gt_left.font = font_tot
    gt_left.alignment = Alignment(horizontal="left", vertical="center")

    for c in range(1, 9):
        cell = ws1.cell(r_curr, c)
        cell.fill = fill_left_tot
        cell.border = tot_border_accounting
        if c == 8:
            cell.value = f"-${tot_fine_left:.2f}" if tot_fine_left > 0 else "$0.00"
            cell.font = font_tot
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Populate Right Executive Summary Table
    # SORT: % RIGHT Handover ascending (worst handover → top), then % RIGHT Delivery, then fine
    def calc_sort_key(stats):
        r_ho = max(0, stats["total_handover"] - stats["penalty_handover"])
        pct_ho = (r_ho / stats["total_handover"] * 100.0) if stats["total_handover"] > 0 else 100.0
        r_del = max(0, stats["total_delivery"] - stats["penalty_delivery"])
        pct_del = (r_del / stats["total_delivery"] * 100.0) if stats["total_delivery"] > 0 else 100.0
        return (pct_ho, pct_del, -stats["total_fine"])

    if tgt in ("ALL", "TOTAL"):
        all_branches = [summary_data[b] for b in MAIN_36_BRANCHES if b in summary_data]
    elif tgt.startswith("ZONE") and tgt in ZONE_BRANCHES_MAP:
        all_branches = [summary_data[b] for b in ZONE_BRANCHES_MAP[tgt] if b in summary_data]
    else:
        all_branches = list(summary_data.values())

    sorted_branches = sorted(all_branches, key=calc_sort_key)

    r_sum = 4
    n_idx = 1
    tot_ho = 0
    tot_del = 0
    tot_pen_ho = 0
    tot_pen_del = 0
    tot_fine = 0.0

    fill_pen_pink = PatternFill("solid", fgColor="FFEAEA") # Soft Light Pink for penalty cells
    font_pen_red  = Font(name="Segoe UI", size=8.5, bold=True, color="DC2626") # Bold Red

    for stats in sorted_branches:
        ws1.row_dimensions[r_sum].height = 18.0
        
        # Calculate RIGHT (On-Time)
        r_ho = max(0, stats["total_handover"] - stats["penalty_handover"])
        r_del = max(0, stats["total_delivery"] - stats["penalty_delivery"])
        
        # Show N/A for % Handover when branch has zero handover records
        if stats["total_handover"] > 0:
            pct_r_ho = r_ho / stats["total_handover"] * 100
            pct_ho_str = f"{pct_r_ho:.1f}%"
        else:
            pct_r_ho = None
            pct_ho_str = "N/A"

        pct_r_del = (r_del / stats["total_delivery"] * 100) if stats["total_delivery"] > 0 else 100.0
        fine_str = f"-${stats['total_fine']:.2f}" if stats['total_fine'] > 0 else "$0.00"

        s_vals = [
            n_idx,
            stats["po"],
            r_ho,
            r_del,
            stats["total_handover"],
            stats["total_delivery"],
            pct_ho_str,
            f"{pct_r_del:.1f}%",
            fine_str
        ]
        for ci, val in enumerate(s_vals, 10):
            cell = ws1.cell(r_sum, ci, val)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border_clean
            
            # Dynamic Font Coloring on % Columns
            if ci == 16:  # % RIGHT Handover
                if pct_r_ho is not None:
                    cell.font = get_pct_font(pct_r_ho)
                else:
                    cell.font = Font(name="Segoe UI", size=8.5, italic=True, color="94A3B8")
            elif ci == 17:  # % RIGHT Delivery
                cell.font = get_pct_font(pct_r_del)
            elif ci == 18:  # Total Penalty ($) → Light Pink + Bold Red text
                if stats["total_fine"] > 0:
                    cell.fill = fill_pen_pink
                    cell.font = font_pen_red
                else:
                    cell.font = font_bold_data
            elif ci == 11:
                cell.font = font_bold_data
            else:
                cell.font = font_data

            if ci != 18:
                cell.fill = fill_row_white

        tot_ho += stats["total_handover"]
        tot_del += stats["total_delivery"]
        tot_pen_ho += stats["penalty_handover"]
        tot_pen_del += stats["penalty_delivery"]
        tot_fine += stats["total_fine"]
        r_sum += 1
        n_idx += 1

    # Right Grand Total Row
    ws1.row_dimensions[r_sum].height = 24.0
    ws1.merge_cells(start_row=r_sum, start_column=10, end_row=r_sum, end_column=11)
    rt_tot = ws1.cell(r_sum, 10, "Grand Total")
    rt_tot.font = font_tot
    rt_tot.alignment = Alignment(horizontal="left", vertical="center")

    tot_r_ho = max(0, tot_ho - tot_pen_ho)
    tot_r_del = max(0, tot_del - tot_pen_del)
    tot_pct_r_ho = (tot_r_ho / tot_ho * 100) if tot_ho > 0 else None
    tot_pct_ho_str = f"{tot_pct_r_ho:.1f}%" if tot_pct_r_ho is not None else "N/A"
    tot_pct_r_del = (tot_r_del / tot_del * 100) if tot_del > 0 else 100.0
    tot_fine_str = f"-${tot_fine:.2f}" if tot_fine > 0 else "$0.00"

    tot_vals_right = [
        "", "",
        tot_r_ho,
        tot_r_del,
        tot_ho,
        tot_del,
        tot_pct_ho_str,
        f"{tot_pct_r_del:.1f}%",
        tot_fine_str
    ]
    for c in range(10, 19):
        cell = ws1.cell(r_sum, c)
        if c >= 12:
            cell.value = tot_vals_right[c-10]
        
        # Color Grand Total %
        if c == 16:  # % RIGHT Handover
            if tot_pct_r_ho is not None:
                cell.font = get_pct_font(tot_pct_r_ho, is_tot=True)
            else:
                cell.font = Font(name="Segoe UI", size=9.5, italic=True, color="94A3B8")
        elif c == 17:
            cell.font = get_pct_font(tot_pct_r_del, is_tot=True)
        elif c == 18:  # Total penalty grand total → light pink
            cell.font = Font(name="Segoe UI", size=9.5, bold=True, color="DC2626")
            cell.fill = fill_pen_pink
            cell.border = tot_border_accounting
            cell.alignment = Alignment(horizontal="center", vertical="center")
            continue
        else:
            cell.font = font_tot

        cell.fill = fill_sum_tot
        cell.border = tot_border_accounting
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Generous Column Widths
    col_widths = {
        1: 5, 2: 15, 3: 20, 4: 22, 5: 18, 6: 10, 7: 12, 8: 15,
        9: 4,
        10: 5, 11: 14, 12: 15, 13: 15, 14: 15, 15: 15, 16: 16, 17: 16, 18: 16
    }
    for c, w in col_widths.items():
        ws1.column_dimensions[get_column_letter(c)].width = w

    # Sheet 2: base
    ws2 = wb.create_sheet(title="base")
    ws2.views.sheetView[0].showGridLines = True

    base_headers = [
        "No", "Order Number", "Customer", "Origin Branch", "Origin Post",
        "Destination Branch", "Destination Post", "Assigned Branch", "Staff / User", "Created At",
        "Last Action Time", "Status Code", "Status Name", "Type", "Age (Days)",
        "Penalty Fine ($)", "Risk Level", "Is Excused"
    ]

    ws2.row_dimensions[1].height = 22
    for col_idx, h in enumerate(base_headers, 1):
        c = ws2.cell(row=1, column=col_idx, value=h)
        c.font = font_hdr
        c.fill = fill_hdr_left
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border_clean

    for idx, item in enumerate(base_rows, 1):
        r_num = idx + 1
        ws2.row_dimensions[r_num].height = 18
        fine_text = f"-${item['penalty_fine']:.2f}" if item['penalty_fine'] > 0 else "$0.00"
        row_data = [
            idx,
            item["order_number"],
            item["customer"],
            item["origin_branch"],
            item["origin_post"],
            item["destination_branch"],
            item["destination_post"],
            item["assigned_branch"],
            item.get("staff_user", ""),
            item["created_at"],
            item["last_action_time"],
            item["status_code"],
            item["status_name"],
            item["type"],
            item["age_days"],
            fine_text,
            item["risk_level"],
            "YES" if item["is_excused"] else "NO"
        ]
        for col_idx, val in enumerate(row_data, 1):
            c = ws2.cell(row=r_num, column=col_idx, value=val)
            c.font = font_data
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = border_clean

    for col in ws2.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws2.column_dimensions[col_letter].width = max(max_len + 3, 11)

    wb.save(out_xlsx)
    return tot_ho, tot_del, (tot_pen_ho + tot_pen_del), tot_fine


def render_penalty_summary_image(out_xlsx):
    wb = openpyxl.load_workbook(out_xlsx)
    ws = wb['INVENTORY PENALTY REPORT']

    wb_sum = openpyxl.Workbook()
    ws_sum = wb_sum.active
    ws_sum.title = 'Executive Summary'
    ws_sum.views.sheetView[0].showGridLines = True

    max_r = 1
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, 10).value is not None or ws.cell(r, 18).value is not None:
            max_r = r

    for r in range(1, max_r + 1):
        if ws.row_dimensions[r].height:
            ws_sum.row_dimensions[r].height = ws.row_dimensions[r].height
        for c_idx in range(9):
            orig_c = 10 + c_idx
            tgt_c = 1 + c_idx
            cell_orig = ws.cell(r, orig_c)
            cell_tgt = ws_sum.cell(r, tgt_c, cell_orig.value)
            if cell_orig.has_style:
                cell_tgt.font = copy.copy(cell_orig.font)
                cell_tgt.fill = copy.copy(cell_orig.fill)
                cell_tgt.border = copy.copy(cell_orig.border)
                cell_tgt.alignment = copy.copy(cell_orig.alignment)

    # Wide column dimensions for crisp unclipped header text
    col_widths = {1: 5, 2: 14, 3: 15, 4: 15, 5: 15, 6: 15, 7: 16, 8: 16, 9: 16}
    for c, w in col_widths.items():
        ws_sum.column_dimensions[get_column_letter(c)].width = w

    ws_sum.merge_cells("A1:I1")
    ws_sum.merge_cells("A2:I2")
    ws_sum.merge_cells(start_row=max_r, start_column=1, end_row=max_r, end_column=2)

    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_f:
        tmp_path = tmp_f.name
    wb_sum.save(tmp_path)

    try:
        buf = excel_to_image.excel_to_image(tmp_path)
        return buf
    finally:
        try:
            os.remove(tmp_path)
        except Exception:
            pass
