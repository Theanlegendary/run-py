"""
excel_to_image.py
Renders an openpyxl-formatted Excel sheet to a PNG image using Pillow.
- Day columns: fixed uniform width
- Hidden/empty columns: skipped entirely
- Merged cells: top-left value shown, others blank
"""

import io
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont

# ── Constants ──────────────────────────────────────────────────────────────────
SCALE       = 3             # HD resolution quality

FONT_SIZE   = 11 * SCALE
ROW_H       = 26 * SCALE    # px — spacious row height
PAD_X       = 10 * SCALE    # horizontal text padding

# Fixed pixel widths per column type
PX_DAY      = 36 * SCALE    # day columns "01"-"31"
PX_ZONE     = 85 * SCALE    # ZONE
PX_GT       = 150 * SCALE   # Grand Total
PX_MIN      = 80 * SCALE    # minimum for auto-fit text columns
PX_MAX      = 260 * SCALE   # maximum for auto-fit
PX_GAP      = 8 * SCALE     # gap/empty separator columns

BG_WHITE    = (255, 255, 255)
BORDER_COL  = (218, 222, 229) # Clean, subtle modern grid border
TEXT_DEF    = (15,  23,  42)  # Modern dark slate


def _hex(h):
    h = h.lstrip('#')
    if len(h) == 6:
        return tuple(int(h[i:i+2], 16) for i in (0, 2, 4))
    return TEXT_DEF


def _cell_bg(cell):
    try:
        f = cell.fill
        if f and f.fill_type == 'solid':
            fc = f.fgColor
            if fc.type == 'rgb' and fc.rgb not in ('00000000', 'FFFFFFFF', 'FF000000'):
                return _hex(fc.rgb[-6:])
    except Exception:
        pass
    return None


def _cell_fg(cell):
    try:
        ft = cell.font
        if ft and ft.color and ft.color.type == 'rgb':
            if ft.color.rgb not in ('00000000', 'FF000000'):
                return _hex(ft.color.rgb[-6:])
    except Exception:
        pass
    return TEXT_DEF


def _cell_bold(cell):
    try:
        return bool(cell.font and cell.font.bold)
    except Exception:
        return False


def _cell_align(cell):
    try:
        a = cell.alignment
        if a and a.horizontal in ('center', 'right'):
            return a.horizontal
    except Exception:
        pass
    return 'left'


_WIN_FONTS = "C:/Windows/Fonts"
_LINUX_FONT_DIRS = [
    "/usr/share/fonts/truetype/dejavu",
    "/usr/share/fonts/truetype/liberation",
    "/usr/share/fonts/truetype/freefont",
    "/usr/share/fonts/truetype/ubuntu",
    "/usr/share/fonts/truetype",
    "/usr/share/fonts",
]

def _has_khmer(text: str) -> bool:
    """Return True if text contains any Khmer Unicode characters (U+1780–U+17FF)."""
    return any('\u1780' <= ch <= '\u17FF' for ch in text)

def _load_font(size, bold=False):
    import os
    here = os.path.dirname(os.path.abspath(__file__))
    bundled_fonts = [
        os.path.join(here, "fonts", "arialbd.ttf" if bold else "arial.ttf"),
        os.path.join(here, "fonts", "calibrib.ttf" if bold else "calibri.ttf"),
    ]
    for b_font in bundled_fonts:
        if os.path.exists(b_font):
            try:
                return ImageFont.truetype(b_font, size)
            except Exception:
                pass

    candidates = (
        ['arialbd.ttf', 'Arial Bold.ttf', 'DejaVuSans-Bold.ttf', 'LiberationSans-Bold.ttf', 'FreeSansBold.ttf', 'Ubuntu-B.ttf']
        if bold else
        ['arial.ttf', 'Arial.ttf', 'DejaVuSans.ttf', 'LiberationSans-Regular.ttf', 'FreeSans.ttf', 'Ubuntu-R.ttf']
    )
    prefixes = [f"{_WIN_FONTS}/", ""] + [f"{d}/" for d in _LINUX_FONT_DIRS] + ["fonts/"]
    for name in candidates:
        for prefix in prefixes:
            try:
                return ImageFont.truetype(prefix + name, size)
            except Exception:
                pass
    return ImageFont.load_default()

def _load_khmer_font(size):
    """Load Khmer OS Battambang font for Khmer text."""
    import os
    here = os.path.dirname(os.path.abspath(__file__))
    for k_name in ['KhmerOSbattambang.ttf', 'KhmerOSsiemreap.ttf', 'KhmerOScontent.ttf', 'KhmerOS.ttf']:
        k_path = os.path.join(here, "fonts", k_name)
        if os.path.exists(k_path):
            try:
                return ImageFont.truetype(k_path, size)
            except Exception:
                pass
        for prefix in (f"{_WIN_FONTS}/", "") + tuple(f"{d}/" for d in _LINUX_FONT_DIRS):
            try:
                return ImageFont.truetype(prefix + k_name, size)
            except Exception:
                pass
    return _load_font(size, bold=False)

def _get_font(text: str, size: int, bold: bool = False):
    """Return Khmer font if text has Khmer chars, otherwise return normal font."""
    if _has_khmer(text):
        return _load_khmer_font(size + int(1.5 * SCALE))
    return _load_font(size, bold)


def _kill_zombie_excel():
    """Auto-terminate stuck/zombie EXCEL.EXE background processes on Windows."""
    try:
        import subprocess
        subprocess.run(["taskkill", "/F", "/IM", "EXCEL.EXE"], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    except Exception:
        pass

def _try_excel_com(xlsx_path: str):
    import win32com.client
    import time
    import os
    from PIL import ImageGrab
    
    try:
        import pythoncom
        pythoncom.CoInitialize()
    except Exception:
        pass
    
    abs_path = os.path.abspath(xlsx_path)
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
    except Exception:
        excel = win32com.client.Dispatch("Excel.Application")
        
    try:
        excel.Visible = False
    except Exception:
        pass
        
    try:
        excel.DisplayAlerts = False
    except Exception:
        pass
    
    wb = None
    try:
        wb = excel.Workbooks.Open(abs_path)
        ws = wb.ActiveSheet
        
        img = None
        temp_png = None
        
        try:
            rng = ws.UsedRange
            chart_width = rng.Width
            chart_height = rng.Height
            
            rng.CopyPicture(1, -4147)
            
            chart_obj = ws.ChartObjects().Add(Left=rng.Left, Top=rng.Top, Width=chart_width, Height=chart_height)
            chart_obj.Activate()
            chart = chart_obj.Chart
            chart.Paste()
            
            chart.ChartArea.Format.Line.Visible = 0
            chart.ChartArea.Format.Fill.Visible = 0
            
            temp_png = os.path.abspath(os.path.join(os.path.dirname(xlsx_path), f"temp_excel_hd_{int(time.time())}.png"))
            chart.Export(temp_png, "PNG")
            chart_obj.Delete()
            
            if os.path.exists(temp_png):
                img = Image.open(temp_png)
                img.load()
        except Exception:
            img = None
        finally:
            if temp_png and os.path.exists(temp_png):
                try:
                    os.remove(temp_png)
                except Exception:
                    pass
        
        if img is None:
            ws.UsedRange.CopyPicture(1, 2)
            time.sleep(0.5)
            for _ in range(5):
                img = ImageGrab.grabclipboard()
                if img:
                    break
                time.sleep(0.5)
        
        if img:
            w, h = img.size
            max_dim = 2400
            if w > max_dim or h > max_dim:
                scale_factor = min(max_dim / float(w), max_dim / float(h))
                new_w = max(1, int(w * scale_factor))
                new_h = max(1, int(h * scale_factor))
                img = img.resize((new_w, new_h), Image.Resampling.LANCZOS)

            w, h = img.size
            max_ratio = 18.0
            new_w, new_h = w, h
            if h > 0 and w / h > max_ratio:
                new_h = int(w / max_ratio)
            elif w > 0 and h / w > max_ratio:
                new_w = int(h / max_ratio)

            if (new_w, new_h) != (w, h):
                padded_img = Image.new('RGB', (new_w, new_h), (255, 255, 255))
                padded_img.paste(img, (0, 0))
                img = padded_img

            buf = io.BytesIO()
            img.save(buf, format='PNG', optimize=True)
            buf.seek(0)
            return buf
    finally:
        if wb:
            try:
                wb.Close(SaveChanges=False)
            except Exception:
                pass
        try:
            excel.Quit()
        except Exception:
            pass
        try:
            import pythoncom
            pythoncom.CoUninitialize()
        except Exception:
            pass
    return None

def excel_to_image(xlsx_path: str) -> io.BytesIO:
    # ── 1. Try Excel COM rendering ──
    try:
        buf = _try_excel_com(xlsx_path)
        if buf:
            return buf
    except Exception:
        pass

    # ── 2. Auto-kill frozen/zombie EXCEL.EXE background processes and retry COM ──
    _kill_zombie_excel()
    import time
    time.sleep(0.5)

    try:
        buf = _try_excel_com(xlsx_path)
        if buf:
            return buf
    except Exception:
        pass

    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    max_row = ws.max_row
    max_col = ws.max_column
    if not max_row or not max_col:
        raise ValueError("Empty sheet")

    # ── 1. Build value grid (handle merged cells) ──────────────────────────────
    grid = [[''] * (max_col + 1) for _ in range(max_row + 1)]
    skip = set()
    multi_col_merged = set()
    for mc in ws.merged_cells.ranges:
        v = ws.cell(mc.min_row, mc.min_col).value
        grid[mc.min_row][mc.min_col] = str(v) if v is not None else ''
        if mc.max_col > mc.min_col:
            for r in range(mc.min_row, mc.max_row + 1):
                multi_col_merged.add((r, mc.min_col))
        for r in range(mc.min_row, mc.max_row + 1):
            for c in range(mc.min_col, mc.max_col + 1):
                if not (r == mc.min_row and c == mc.min_col):
                    skip.add((r, c))

    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            if (r, c) in skip:
                continue
            try:
                v = ws.cell(r, c).value
            except Exception:
                v = None
            if grid[r][c] == '':
                grid[r][c] = str(v) if v is not None else ''

    # ── 2. Classify columns ────────────────────────────────────────────────────
    # Day columns are only classified if a header row has multiple (>=3) day numbers (01-31)
    day_ci  = set()
    zone_ci = set()
    gt_ci   = set()

    HEADER_SCAN_ROWS = min(max_row, 5)
    for r in range(1, HEADER_SCAN_ROWS + 1):
        row_days = []
        for c in range(1, max_col + 1):
            val = grid[r][c].strip()
            if val.isdigit() and len(val) == 2 and 1 <= int(val) <= 31:
                row_days.append(c)
            elif val == 'ZONE':
                zone_ci.add(c)
            elif val == 'Grand Total':
                row_vals = [grid[r][cc].strip() for cc in range(1, max_col + 1)]
                if any(v in ('ZONE', 'ORDER ID', 'POST OFFICE HANDLE', 'CURRENT POST OFFICE')
                       for v in row_vals):
                    gt_ci.add(c)
        if len(row_days) >= 3:
            day_ci.update(row_days)

    # Detect empty columns (gap separators between side-by-side tables)
    empty_ci = set(
        c for c in range(1, max_col + 1)
        if not any(grid[r][c].strip() for r in range(1, max_row + 1))
    )

    # ── 3. Compute per-column pixel widths ─────────────────────────────────────
    fn   = _load_font(FONT_SIZE, bold=False)
    fn_b = _load_font(FONT_SIZE, bold=True)
    tmp  = Image.new('RGB', (1, 1))
    d    = ImageDraw.Draw(tmp)

    col_px = []   # pixel width per column (0 = skip/hidden)
    for c in range(1, max_col + 1):
        if c in empty_ci:
            col_px.append(PX_GAP)   # tiny but visible separator
            continue

        # Check if column is hidden in Excel
        letter = get_column_letter(c)
        cd = ws.column_dimensions.get(letter)
        if cd and (cd.hidden or (cd.width is not None and cd.width < 0.5)):
            col_px.append(0)        # fully hidden — skip in image too
            continue

        if c in day_ci:
            col_px.append(PX_DAY)   # ALL day columns same fixed width
        elif c in zone_ci:
            col_px.append(PX_ZONE)
        elif c in gt_ci:
            col_px.append(PX_GT)
        else:
            # Auto-fit text columns (ignore banner text merged across multiple columns)
            max_w = PX_MIN
            for r in range(1, max_row + 1):
                if (r, c) in multi_col_merged:
                    continue
                text = grid[r][c]
                if not text:
                    continue
                bold = _cell_bold(ws.cell(r, c))
                f = _get_font(text, FONT_SIZE, bold)
                stroke_w = 0
                if bold and _has_khmer(text):
                    # Simulate bold for Khmer OS fonts using a stroke outline
                    stroke_w = max(1, int(0.4 * SCALE))
                try:
                    bb = d.textbbox((0, 0), text, font=f, stroke_width=stroke_w)
                    w  = bb[2] - bb[0]
                except Exception:
                    w = len(text) * (FONT_SIZE - 2) + stroke_w * 2
                max_w = max(max_w, w + PAD_X * 2)

            # Respect explicit Excel column width if defined
            if cd and cd.width and cd.width >= 2:
                excel_px = int(cd.width * 7.5 * SCALE)
                max_w = max(max_w, excel_px)

            col_px.append(min(max_w, PX_MAX))

    # Columns to actually render (skip 0-width hidden ones)
    render_cols = [c for c in range(1, max_col + 1) if col_px[c - 1] > 0]

    # ── 3.5 Compute per-row heights ────────────────────────────────────────────
    # Rows that contain any Khmer text get ROW_H_KHMER, others get ROW_H
    ROW_H_KHMER = 34 * SCALE
    ROW_H = 22 * SCALE
    row_heights = []
    for r in range(1, max_row + 1):
        has_khmer_row = any(
            _has_khmer(grid[r][c]) for c in range(1, max_col + 1)
        )
        row_heights.append(ROW_H_KHMER if has_khmer_row else ROW_H)

    # ── 4. Build canvas ────────────────────────────────────────────────────────
    total_w = sum(col_px[c - 1] for c in render_cols) + 1
    total_h = sum(row_heights) + 1
    img  = Image.new('RGB', (total_w, total_h), BG_WHITE)
    draw = ImageDraw.Draw(img)

    # ── 4.5 Pre-calculate merged cell pixel dimensions ───────────────
    skip = set()
    merged_rects = {}
    for mc in ws.merged_cells.ranges:
        mw = 0
        for cc in range(mc.min_col, mc.max_col + 1):
            if cc - 1 < len(col_px) and col_px[cc - 1] > 0:
                mw += col_px[cc - 1]
        mh = 0
        for rr in range(mc.min_row, mc.max_row + 1):
            if rr - 1 < len(row_heights):
                mh += row_heights[rr - 1]

        for cc in range(mc.min_col, mc.max_col + 1):
            if (mc.min_row, cc) != (mc.min_row, mc.min_col):
                skip.add((mc.min_row, cc))
        for r_m in range(mc.min_row + 1, mc.max_row + 1):
            for c_m in range(mc.min_col, mc.max_col + 1):
                skip.add((r_m, c_m))
        merged_rects[(mc.min_row, mc.min_col)] = (mw, mh)

    # ── 5. Draw ────────────────────────────────────────────────────────────────
    y = 0
    for r in range(1, max_row + 1):
        rh = row_heights[r - 1]   # per-row height
        x = 0
        for c in render_cols:
            cw   = col_px[c - 1]
            if (r, c) in skip:
                x += cw
                continue

            mw_mh = merged_rects.get((r, c))
            if mw_mh:
                mw, mh = mw_mh
            else:
                mw, mh = cw, rh
            text = grid[r][c]

            try:
                cell  = ws.cell(r, c)
                bg    = _cell_bg(cell) or BG_WHITE
                fg    = _cell_fg(cell)
                bold  = _cell_bold(cell)
                align = _cell_align(cell)
            except Exception:
                bg, fg, bold, align = BG_WHITE, TEXT_DEF, False, 'left'

            draw.rectangle([x, y, x + mw, y + mh], fill=bg)

            if text:
                text_str = str(text)
                dot_color = None
                clean_text = text_str
                
                if text_str.startswith("🟢"):
                    dot_color = "#10B981"  # Bright Vivid Green
                    clean_text = text_str[1:].strip()
                elif text_str.startswith("🟡"):
                    dot_color = "#EF4444"  # Map yellow to Red (no yellow)
                    clean_text = text_str[1:].strip()
                elif text_str.startswith("🔴"):
                    dot_color = "#EF4444"  # Bright Vivid Red
                    clean_text = text_str[1:].strip()

                if dot_color:
                    f = _get_font(clean_text, FONT_SIZE, True)
                    try:
                        bb = draw.textbbox((0, 0), clean_text, font=f)
                        tw_t, th_t = bb[2] - bb[0], bb[3] - bb[1]
                    except Exception:
                        tw_t = len(clean_text) * (FONT_SIZE - 2)
                        th_t = FONT_SIZE

                    dot_r = max(4 * SCALE, th_t // 3)
                    dot_w = dot_r * 2
                    gap = 5 * SCALE
                    total_w = dot_w + gap + tw_t

                    if align == 'center':
                        start_x = x + (mw - total_w) // 2
                    elif align == 'right':
                        start_x = x + mw - total_w - PAD_X
                    else:
                        start_x = x + PAD_X

                    cy = y + mh // 2
                    # Draw solid filled color circle
                    draw.ellipse([start_x, cy - dot_r, start_x + dot_w, cy + dot_r], fill=dot_color)

                    # Draw text in matching bold color font
                    tx = start_x + dot_w + gap
                    ty = y + (mh - th_t) // 2
                    draw.text((tx, ty), clean_text, font=f, fill=dot_color)
                else:
                    f = _get_font(text, FONT_SIZE, bold)
                    stroke_w = 0
                    if bold and _has_khmer(text):
                        stroke_w = max(1, int(0.4 * SCALE))
                    try:
                        bb = draw.textbbox((0, 0), text, font=f, stroke_width=stroke_w)
                        tw, th = bb[2] - bb[0], bb[3] - bb[1]
                    except Exception:
                        tw = len(text) * (FONT_SIZE - 2) + stroke_w * 2
                        th = FONT_SIZE + stroke_w * 2

                    ty = y + (mh - th) // 2
                    if align == 'center':
                        tx = x + (mw - tw) // 2
                    elif align == 'right':
                        tx = x + mw - tw - PAD_X
                    else:
                        tx = x + PAD_X

                    if stroke_w > 0:
                        draw.text((tx, ty), text, font=f, fill=fg, stroke_width=stroke_w, stroke_fill=fg)
                    else:
                        draw.text((tx, ty), text, font=f, fill=fg)

            draw.rectangle([x, y, x + mw, y + mh], outline=BORDER_COL, width=1 * SCALE)
            x += cw
        y += rh

    # ── Aspect ratio & max dimension safety check for Telegram ─────────────────
    # Telegram rejects photos with dimensions > 2500px or aspect ratio > 20 (Photo_invalid_dimensions).
    w, h = img.size
    max_dim = 2400
    if w > max_dim or h > max_dim:
        scale_factor = min(max_dim / float(w), max_dim / float(h))
        new_w = max(1, int(w * scale_factor))
        new_h = max(1, int(h * scale_factor))
        img = img.resize((new_w, new_h), Image.Resampling.LANCZOS)

    # Check aspect ratio safety
    w, h = img.size
    max_ratio = 18.0
    new_w, new_h = w, h
    if h > 0 and w / h > max_ratio:
        new_h = int(w / max_ratio)
    elif w > 0 and h / w > max_ratio:
        new_w = int(h / max_ratio)

    if (new_w, new_h) != (w, h):
        padded_img = Image.new('RGB', (new_w, new_h), BG_WHITE)
        padded_img.paste(img, (0, 0))
        img = padded_img

    buf = io.BytesIO()
    img.save(buf, format='PNG', optimize=True)
    buf.seek(0)
    return buf
