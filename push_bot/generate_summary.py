"""
generate_summary.py
Builds a zone push summary image with the layout:
  Row 1  : Title bar  — "DAILY REPORT — ZONE3  06/07/2026  09:30"
  Row 2  : Month sub-header spanning date columns
  Row 3  : Column headers — HANDLE | Pickup | Delivery | Pending | <DD> … | TOTAL | URGENT
  Row 4+ : Data — branch row with counts; URGENT column red bold
  Last   : Grand Total — all red bold
"""

import io
import re
import calendar as _calendar
from datetime import datetime, date as _date
from PIL import Image, ImageDraw, ImageFont

SUMMARY_HEADER_KHMER = {
    "HANDLE": "ប៉ុស្តិ៍ / ហាង",
    "Pickup": "ត្រូវយក",
    "Delivery": "ត្រូវដឹក",
    "Pending": "កំពុងរង់ចាំ",
    "Transit": "ប្រគល់ទៅ MEGA",
    "Branch": "ចាត់តាំងដឹក",
    "Send Mega": "ប្រគល់ទៅ MEGA",
    "Not Assign": "ចាត់តាំងដឹក",
    "Handover to Mega": "ប្រគល់ទៅ MEGA",
    "Assign Deliver": "ចាត់តាំងដឹក",
    "TOTAL": "សរុប",
    "> 1 Day": "> 1 ថ្ងៃ",
    "> 3 Days": "> 3 ថ្ងៃ",
    "U.Pickup": "ប្រញាប់.យក",
    "U.Delivery": "ប្រញាប់.ដឹក",
    "U.Pending": "ប្រញាប់.រង់ចាំ",
    "U.Transit": "ប្រញាប់.MEGA",
    "U.Branch": "ប្រញាប់.ចាត់ដឹក",
    "GRAND TOTAL": "សរុបទាំងអស់"
}

# ── Palette ────────────────────────────────────────────────────────────────────
C_TITLE_BG    = ( 10,  15,  35)   # near-black navy
C_TITLE_FG    = (255, 255, 255)
C_MONTH_BG    = ( 22,  34,  64)   # deep indigo for month row
C_MONTH_FG    = (180, 200, 255)   # soft blue-white
C_HEADER_BG   = ( 30,  45,  80)   # dark slate-blue
C_HEADER_FG   = (255, 255, 255)
C_URGENT_HDR  = (180,  20,  20)   # dark red header for URGENT col
C_ROW_BG      = (255, 255, 255)
C_ROW_ALT     = (245, 248, 255)   # very light blue stripe
C_TOTAL_BG    = (232, 238, 250)   # soft blue-grey footer
C_TOTAL_FG    = (200,  30,  30)   # strong red
C_NUM_FG      = ( 30,  30, 160)   # deep blue for Pickup/Delivery/Pending counts
C_PENDING_FG  = (180,  80,   0)   # amber for Pending counts
C_DATE_FG     = ( 30,  30, 160)   # blue for date counts
C_URGENT_FG   = (210,  30,  30)   # red for urgent counts
C_HANDLE_FG   = ( 10,  15,  40)   # near-black for branch name
C_BORDER      = (180, 195, 220)
C_BORDER_DARK = ( 80, 100, 140)   # darker border for section separators

_WIN_FONTS = "C:/Windows/Fonts"


def _load_font(size, bold=False):
    candidates = (
        [
            f"{_WIN_FONTS}/calibrib.ttf",
            f"{_WIN_FONTS}/arialbd.ttf",
            f"{_WIN_FONTS}/verdanab.ttf",
            f"{_WIN_FONTS}/DejaVuSans-Bold.ttf",
            "arialbd.ttf", "DejaVuSans-Bold.ttf",
        ]
        if bold else
        [
            f"{_WIN_FONTS}/calibri.ttf",
            f"{_WIN_FONTS}/arial.ttf",
            f"{_WIN_FONTS}/verdana.ttf",
            f"{_WIN_FONTS}/DejaVuSans.ttf",
            "arial.ttf", "DejaVuSans.ttf",
        ]
    )
    for name in candidates:
        try:
            return ImageFont.truetype(name, size)
        except Exception:
            pass
    return ImageFont.load_default()


def _tw(draw, text, font):
    try:
        bb = draw.textbbox((0, 0), text, font=font)
        return bb[2] - bb[0]
    except Exception:
        return len(text) * max(8, font.size - 2)


def _th(draw, text, font):
    try:
        bb = draw.textbbox((0, 0), text, font=font)
        return bb[3] - bb[1]
    except Exception:
        return font.size


def _draw_cell(draw, x, y, w, h, bg, text, font, fg, align="center", pad=8,
               border=True, border_col=None, border_w=1):
    draw.rectangle([x, y, x + w - 1, y + h - 1], fill=bg)
    if text:
        tw = _tw(draw, text, font)
        th = _th(draw, text, font)
        ty = y + (h - th) // 2
        if align == "center":
            tx = x + (w - tw) // 2
        elif align == "right":
            tx = x + w - tw - pad
        else:
            tx = x + pad
        draw.text((tx, ty), text, font=font, fill=fg)
    if border:
        bc = border_col or C_BORDER
        draw.rectangle([x, y, x + w - 1, y + h - 1], outline=bc, width=border_w)


def build_summary_image(
    handle_results: list,
    overall: dict,
    today: datetime = None,
    zone_label: str = "",
    day_date_counts: dict = None,   # {handle: {date_obj: count}} for date columns
    urgent_counts: dict = None,     # {handle: urgent_count}
    fee_counts: dict = None,        # {handle: float} total fee per handle
    cod_counts: dict = None,        # {handle: float} total COD per handle
    vip_counts: dict = None,        # {handle: int} total VIP count per handle
) -> io.BytesIO:
    """
    New layout:
      Row 1  : Title   — "DAILY REPORT — ZONE3  06/07/2026 09:30"
      Row 2  : Month sub-header spanning date columns (blank over fixed cols)
      Row 3  : Headers — HANDLE | Pickup | Delivery | Pending | VIP | DD DD DD … | TOTAL | URGENT
      Row 4+ : Data rows
      Last   : Grand Total (all red)
    """
    now = today or datetime.now()
    n_data_rows = len(handle_results)

    # ── Scale factor ───────────────────────────────────────────────────────────
    if n_data_rows > 75:
        sc = 1
    elif n_data_rows > 35:
        sc = 2
    else:
        sc = 3

    FS       = 11 * sc   # base font size
    FS_TITLE = 13 * sc   # title font size
    FS_SM    = 9  * sc   # small (month row)
    ROW_H    = 26 * sc
    TITLE_H  = 34 * sc
    MONTH_H  = 18 * sc
    PAD      =  8 * sc

    fn       = _load_font(FS,       bold=False)
    fn_b     = _load_font(FS,       bold=True)
    fn_title = _load_font(FS_TITLE, bold=True)
    fn_sm    = _load_font(FS_SM,    bold=True)

    # ── Collect all unique sorted dates ───────────────────────────────────────
    all_dates: list[_date] = []
    if day_date_counts:
        date_set = set()
        for dc in day_date_counts.values():
            date_set.update(dc.keys())
        all_dates = sorted(date_set)

    # ── Zone column detection & sorting ───────────────────────────────────────
    import json
    handle_to_zone = {}
    try:
        cfg_p = os.path.join(os.path.dirname(__file__), "config.json")
        if not os.path.exists(cfg_p):
            cfg_p = os.path.join(os.path.dirname(os.path.dirname(__file__)), "config.json")
        if os.path.exists(cfg_p):
            with open(cfg_p, encoding="utf-8") as _f:
                _cfg = json.load(_f)
            for zk, h_list in _cfg.get("total_zones", {}).items():
                z_name = f"Zone {zk[-1]}" if zk.lower().startswith("zone") and zk[-1].isdigit() else zk.upper()
                for _h in h_list:
                    handle_to_zone[_h.upper()] = z_name
            for _h, _z in _cfg.get("zone_mapping", {}).get("by_post_office", {}).items():
                if _h.upper() not in handle_to_zone:
                    handle_to_zone[_h.upper()] = _z
    except Exception:
        pass

    def resolve_zone(h_str):
        if not h_str:
            return "Zone ?"
        hu = str(h_str).strip().upper()
        if hu in handle_to_zone:
            return handle_to_zone[hu]
        for k, v in handle_to_zone.items():
            if k in hu or hu in k:
                return v
        p3 = hu[:3]
        if p3 in ["PNP", "KAN", "PRE", "SVA"]: return "Zone 1"
        elif p3 in ["KAM", "KOH", "SIH", "SPE", "TAK"]: return "Zone 2"
        elif p3 in ["BAN", "BAT", "CHH", "PUR"]: return "Zone 3"
        elif p3 in ["ODD", "PRH", "SIE", "THO"]: return "Zone 4"
        elif p3 in ["CHA", "KRA", "TBK", "ROT", "MON", "STU"]: return "Zone 5"
        return "Zone ?"

    show_zone_col = bool(zone_label and ("ZONE" in zone_label.upper() or "ALL" in zone_label.upper()))

    if show_zone_col:
        def _sort_key(hr):
            h_u = hr["handle"].upper()
            z_val = resolve_zone(h_u)
            return (z_val, h_u)
        handle_results = sorted(handle_results, key=_sort_key)

    tmp  = Image.new("RGB", (1, 1))
    draw = ImageDraw.Draw(tmp)

    handle_strs = [hr["handle"] for hr in handle_results] + ["GRAND TOTAL"]
    W_HANDLE = max(_tw(draw, s, fn_b) for s in handle_strs) + PAD * 2
    W_HANDLE = max(W_HANDLE, 80 * sc)
    W_ZONE   = max(_tw(draw, "Zone 5", fn_b) + PAD * 2, 54 * sc) if show_zone_col else 0

    W_NUM    = max(_tw(draw, h, fn_b) for h in ["Delivery", "Assign Deliver", "Pickup", "Handover to Mega", "VIP", "TOTAL"]) + PAD * 2
    W_NUM    = max(W_NUM, 56 * sc)

    W_DATE   = max(_tw(draw, "00", fn_b) + PAD * 2, 32 * sc)
    W_URGENT = max(_tw(draw, "> 1 Day", fn_sm) + PAD * 2, 52 * sc)
    W_URGENT_3 = max(_tw(draw, "> 3 Days", fn_sm) + PAD * 2, 56 * sc)
    W_U_COL  = max(_tw(draw, "U.Delivery", fn_sm) + PAD * 2, 48 * sc)

    # Column order: [ZONE] | Handle | Delivery | Assign Deliver | Pickup | Handover to Mega | [VIP] | [dates…] | TOTAL | [Fee | COD] | > 1 Day | > 3 Days
    fixed_cols  = (["ZONE"] if show_zone_col else []) + ["HANDLE", "Delivery", "Assign Deliver", "Pickup", "Handover to Mega"]
    if vip_counts is not None:
        fixed_cols.append("VIP")
    date_labels = [f"{d.day:02d}" for d in all_dates]
    tail_cols   = ["TOTAL"]

    has_fee_cod = fee_counts is not None or cod_counts is not None
    W_FEE = max(_tw(draw, "$9999.99", fn_b) + PAD * 2, 64 * sc)

    has_split_urgent = False
    if urgent_counts:
        first_val = next(iter(urgent_counts.values()))
        if isinstance(first_val, dict):
            has_split_urgent = True

    if has_fee_cod:
        if fee_counts is not None:
            tail_cols.append("Fee($)")
        if cod_counts is not None:
            tail_cols.append("COD($)")

    # Add > 1 Day and > 3 Days columns
    if urgent_counts is not None:
        tail_cols.append("> 1 Day")
        tail_cols.append("> 3 Days")

    col_widths = (
        ([W_ZONE] if show_zone_col else [])
        + [W_HANDLE] + [W_NUM] * (len(fixed_cols) - (2 if show_zone_col else 1))
        + [W_DATE] * len(all_dates)
        + [W_NUM]                          # TOTAL
    )
    if has_fee_cod:
        if fee_counts is not None:
            col_widths.append(W_FEE)
        if cod_counts is not None:
            col_widths.append(W_FEE)
    if urgent_counts is not None:
        col_widths.append(W_URGENT)      # > 1 Day
        col_widths.append(W_URGENT_3)    # > 3 Days
    col_labels = fixed_cols + date_labels + tail_cols

    n_cols  = len(col_widths)
    total_w = sum(col_widths) + 1

    # ── Canvas height ─────────────────────────────────────────────────────────
    # Title + month row + header row + data rows + grand total
    total_h = TITLE_H + MONTH_H + ROW_H + n_data_rows * ROW_H + ROW_H + 1

    img  = Image.new("RGB", (total_w, total_h), C_ROW_BG)
    draw = ImageDraw.Draw(img)

    # ── Helper: draw a full row of cells ──────────────────────────────────────
    def _row(y, h, cells, bgs, fgs, fonts, aligns, border_col=C_BORDER):
        x = 0
        for ci, (text, cw) in enumerate(zip(cells, col_widths)):
            _draw_cell(draw, x, y, cw, h,
                       bg=bgs[ci] if isinstance(bgs, list) else bgs,
                       text=text,
                       font=fonts[ci] if isinstance(fonts, list) else fonts,
                       fg=fgs[ci] if isinstance(fgs, list) else fgs,
                       align=aligns[ci] if isinstance(aligns, list) else aligns,
                       pad=PAD, border=True, border_col=border_col)
            x += cw

    # ── Row 1: Title ──────────────────────────────────────────────────────────
    y = 0
    zone_part = f" — {zone_label.upper()}" if zone_label else ""
    title_text = f"DAILY REPORT{zone_part}    {now.strftime('%d/%m/%Y  %H:%M')}"
    draw.rectangle([0, 0, total_w - 1, TITLE_H - 1], fill=C_TITLE_BG)
    tw = _tw(draw, title_text, fn_title)
    th = _th(draw, title_text, fn_title)
    draw.text((PAD * 2, (TITLE_H - th) // 2), title_text, font=fn_title, fill=C_TITLE_FG)
    # right-side date accent bar
    accent_w = 6 * sc
    draw.rectangle([total_w - accent_w - 1, 0, total_w - 1, TITLE_H - 1], fill=(60, 100, 200))
    draw.rectangle([0, 0, total_w - 1, TITLE_H - 1], outline=C_BORDER_DARK, width=sc)
    y += TITLE_H

    # ── Row 2: Month sub-header ───────────────────────────────────────────────
    # Fixed cols + tail cols are blank (same dark bg); date cols show month name merged
    month_cells  = [""] * n_cols
    month_bgs    = [C_MONTH_BG] * n_cols
    month_fgs    = [C_MONTH_FG] * n_cols
    month_fonts  = [fn_sm] * n_cols
    month_aligns = ["center"] * n_cols

    # Group consecutive dates by month and write month name into first cell of group
    if all_dates:
        n_fixed = 4   # HANDLE + 3 type cols
        groups = []
        cur_mo, g_start = None, None
        for i, d in enumerate(all_dates):
            mo = (d.year, d.month)
            if mo != cur_mo:
                if cur_mo is not None:
                    groups.append((cur_mo, g_start, i - 1))
                cur_mo, g_start = mo, i
        if cur_mo is not None:
            groups.append((cur_mo, g_start, len(all_dates) - 1))

        for (yr, mo), gi_start, gi_end in groups:
            col_idx = n_fixed + gi_start   # 0-based column index
            month_cells[col_idx] = _calendar.month_abbr[mo].upper()

    # Draw month row manually (need to visually merge month spans)
    x = 0
    for ci, cw in enumerate(col_widths):
        is_date_col = (4 <= ci < 4 + len(all_dates))
        bg = C_MONTH_BG
        # lighter bg for non-date cols
        cell_bg = C_MONTH_BG if is_date_col else C_HEADER_BG
        _draw_cell(draw, x, y, cw, MONTH_H,
                   bg=cell_bg,
                   text=month_cells[ci],
                   font=fn_sm, fg=C_MONTH_FG,
                   align="center", pad=PAD,
                   border=True, border_col=C_BORDER_DARK)
        x += cw
    y += MONTH_H

    # ── Row 3: Column headers ─────────────────────────────────────────────────
    hdr_bgs = []
    hdr_fgs = []
    for ci, label in enumerate(col_labels):
        if label in ("> 1 Day", "> 3 Days"):
            hdr_bgs.append(C_URGENT_HDR)
            hdr_fgs.append((255, 230, 230))
        elif label in ("TOTAL",):
            hdr_bgs.append((20, 60, 120))
            hdr_fgs.append(C_HEADER_FG)
        elif label in ("Fee($)", "COD($)"):
            hdr_bgs.append((15, 100, 60))   # dark green header for fee/cod
            hdr_fgs.append((220, 255, 220))
        else:
            hdr_bgs.append(C_HEADER_BG)
            hdr_fgs.append(C_HEADER_FG)

    _row(y, ROW_H,
         cells=col_labels,
         bgs=hdr_bgs,
         fgs=hdr_fgs,
         fonts=fn_b,
         aligns=["left"] + ["center"] * (n_cols - 1),
         border_col=C_BORDER_DARK)
    y += ROW_H

    # ── Data rows ─────────────────────────────────────────────────────────────
    for i, hr in enumerate(handle_results):
        counts   = hr["handle_counts"]
        handle   = hr["handle"]
        pickup   = counts.get("Pickup",   0)
        delivery = counts.get("Delivery", 0)
        transit  = counts.get("Transit",  0)
        branch   = counts.get("Branch",   0)
        total    = pickup + delivery + transit + branch
        urgent   = (urgent_counts or {}).get(handle, 0)

        row_bg = C_ROW_ALT if i % 2 else C_ROW_BG

        z_str  = resolve_zone(handle) if show_zone_col else None
        cells  = (([z_str] if show_zone_col else []) +
                  [handle,
                   str(delivery) if delivery else "",
                   str(branch)   if branch   else "",
                   str(pickup)   if pickup   else "",
                   str(transit)  if transit  else ""])

        fgs    = (([C_HANDLE_FG] if show_zone_col else []) +
                  [C_HANDLE_FG,
                   C_NUM_FG,
                   C_NUM_FG,
                   C_NUM_FG,
                   C_NUM_FG])

        fonts  = (([fn_b] if show_zone_col else []) +
                  [fn_b, fn, fn, fn, fn])
        aligns = (["center"] if show_zone_col else []) + ["left", "center", "center", "center", "center"]

        if vip_counts is not None:
            v_cnt = vip_counts.get(handle, 0)
            cells.append(str(v_cnt) if v_cnt else "")
            fgs.append((220, 38, 38) if v_cnt else C_NUM_FG)
            fonts.append(fn_b if v_cnt else fn)
            aligns.append("center")

        # Date columns
        day_dc = (day_date_counts or {}).get(handle, {})
        for d in all_dates:
            cnt = day_dc.get(d, 0)
            cells.append(str(cnt) if cnt else "")
            fgs.append(C_DATE_FG)
            fonts.append(fn)
            aligns.append("center")

        # TOTAL
        cells.append(str(total) if total else "")
        fgs.append(C_TOTAL_FG)
        fonts.append(fn_b)
        aligns.append("center")

        # Fee / COD columns
        C_FEE_FG = (20, 140, 60)   # dark green text
        C_FEE_BG = (230, 255, 235) # light green tint
        if has_fee_cod:
            if fee_counts is not None:
                fee_v = (fee_counts or {}).get(handle, 0.0)
                cells.append(f"${fee_v:.2f}" if fee_v else "")
                fgs.append(C_FEE_FG)
                fonts.append(fn_b)
                aligns.append("center")
            if cod_counts is not None:
                cod_v = (cod_counts or {}).get(handle, 0.0)
                cells.append(f"${cod_v:.2f}" if cod_v else "")
                fgs.append(C_FEE_FG)
                fonts.append(fn_b)
                aligns.append("center")

        # > 1 Day and > 3 Days columns
        if urgent_counts is not None:
            h_urgent = urgent_counts.get(hr["handle"], {})
            
            if isinstance(h_urgent, dict):
                urgent_1day = h_urgent.get("1day", 0)
                urgent_3days = h_urgent.get("3days", 0)
            else:
                urgent_1day = h_urgent
                urgent_3days = 0
            
            cells.append(str(urgent_1day) if urgent_1day else "")
            cells.append(str(urgent_3days) if urgent_3days else "")
            
            fgs.extend([C_URGENT_FG, C_URGENT_FG])
            fonts.extend([fn_b, fn_b])
            aligns.extend(["center", "center"])

        bgs = [row_bg] * n_cols
        if has_fee_cod:
            n_fee_cod = (1 if fee_counts is not None else 0) + (1 if cod_counts is not None else 0)
            total_idx = (6 if show_zone_col else 5) + len(all_dates) + (1 if vip_counts is not None else 0)
            for fi in range(n_fee_cod):
                if total_idx + 1 + fi < n_cols:
                    bgs[total_idx + 1 + fi] = (230, 255, 235)
        
        if urgent_counts is not None:
            h_urgent = urgent_counts.get(hr["handle"], {})
            if isinstance(h_urgent, dict):
                urgent_1day = h_urgent.get("1day", 0)
                urgent_3days = h_urgent.get("3days", 0)
            else:
                urgent_1day = h_urgent
                urgent_3days = 0
            
            if urgent_1day > 0 or urgent_3days > 0:
                bgs[-2:] = [(255, 235, 235), (255, 235, 235)]

        _row(y, ROW_H,
             cells=cells, bgs=bgs, fgs=fgs, fonts=fonts, aligns=aligns,
             border_col=C_BORDER)
        y += ROW_H

    # ── Grand Total row ───────────────────────────────────────────────────────
    g_pickup   = overall.get("Pickup",   0)
    g_delivery = overall.get("Delivery", 0)
    g_transit  = overall.get("Transit",  0)
    g_branch   = overall.get("Branch",   0)
    g_total    = g_pickup + g_delivery + g_transit + g_branch

    gt_cells  = ([""] if show_zone_col else []) + [
                 "GRAND TOTAL",
                 str(g_delivery) if g_delivery else "",
                 str(g_branch)   if g_branch   else "",
                 str(g_pickup)   if g_pickup   else "",
                 str(g_transit)  if g_transit  else ""]

    if vip_counts is not None:
        g_vip = sum((vip_counts or {}).values())
        gt_cells.append(str(g_vip) if g_vip else "")

    # Date totals
    for d in all_dates:
        day_total = sum(
            (day_date_counts or {}).get(hr["handle"], {}).get(d, 0)
            for hr in handle_results
        )
        gt_cells.append(str(day_total) if day_total else "")

    gt_cells.append(str(g_total) if g_total else "")

    # Grand total Fee / COD
    if has_fee_cod:
        if fee_counts is not None:
            g_fee = sum((fee_counts or {}).values())
            gt_cells.append(f"${g_fee:.2f}" if g_fee else "")
        if cod_counts is not None:
            g_cod = sum((cod_counts or {}).values())
            gt_cells.append(f"${g_cod:.2f}" if g_cod else "")

    # Add > 1 Day and > 3 Days to Grand Total
    if urgent_counts is not None:
        g_1day = 0
        g_3days = 0
        for h_urgent in urgent_counts.values():
            if isinstance(h_urgent, dict):
                g_1day += h_urgent.get("1day", 0)
                g_3days += h_urgent.get("3days", 0)
            else:
                # Backward compatibility
                g_1day += h_urgent
        
        gt_cells.append(str(g_1day) if g_1day else "")
        gt_cells.append(str(g_3days) if g_3days else "")

    gt_bgs = [C_TOTAL_BG] * n_cols
    if urgent_counts is not None:
        g_1day = sum(u.get("1day", 0) if isinstance(u, dict) else u for u in urgent_counts.values())
        g_3days = sum(u.get("3days", 0) if isinstance(u, dict) else 0 for u in urgent_counts.values())
        if g_1day > 0 or g_3days > 0:
            gt_bgs[-2:] = [(255, 200, 200), (255, 200, 200)]  # Last 2 cells

    _row(y, ROW_H,
         cells=gt_cells,
         bgs=gt_bgs,
         fgs=C_TOTAL_FG,
         fonts=fn_b,
         aligns=["left"] + ["center"] * (n_cols - 1),
         border_col=C_BORDER_DARK)

    # ── Telegram aspect-ratio guard (max 20:1) ────────────────────────────────
    w, h = img.size
    max_ratio = 18.0
    new_w, new_h = w, h
    if h > 0 and w / h > max_ratio:
        new_h = int(w / max_ratio)
    elif w > 0 and h / w > max_ratio:
        new_w = int(h / max_ratio)
    if (new_w, new_h) != (w, h):
        padded = Image.new("RGB", (new_w, new_h), C_ROW_BG)
        padded.paste(img, (0, 0))
        img = padded

    buf = io.BytesIO()
    img.save(buf, format="PNG", optimize=True)
    buf.seek(0)
    return buf


def compute_kpi_info(row, age_adjust_hours=0):
    """
    Compute age and KPI info for a row.
    
    Args:
        row: DataFrame row with order data
        age_adjust_hours: Hours to subtract from age (e.g., 12 for morning reports to exclude overnight hold)
    """
    import pandas as pd
    from datetime import datetime

    status_code = str(row.get('STATUS_CODE', '') or row.get('CURRENT STATUS', '') or '').strip()
    if status_code and ' ' in status_code:
        status_code = status_code.split()[0].strip()

    scan_time = None
    for col in [
        'CURRENT TIME',
        'STATUS 306 AT STORE / AGENT (LAST TIME)',
        'STATUS 306 AT STORE / AGENT FROM HUB (FIRST TIME)',
        'STATUS 302/310 AT RECEIVING STORE / RECEIVING AGENT (FIRST TIME)',
        'STATUS 306  AT ORIGIN HUB (FIRST TIME)',
        'CREATED DATE'
    ]:
        val = row.get(col)
        if pd.notna(val) and str(val).strip() and str(val).strip().lower() != 'nan':
            parsed_dt = pd.to_datetime(val, dayfirst=True, format='mixed', errors='coerce')
            if pd.notna(parsed_dt):
                scan_time = parsed_dt
                break

    if scan_time is None:
        return '', '10h'

    now = datetime.now()
    diff = now - scan_time
    total_seconds = max(0, diff.total_seconds())
    total_minutes = int(total_seconds // 60)
    
    # Apply age adjustment (subtract hours for morning reports)
    adjusted_minutes = max(0, total_minutes - (age_adjust_hours * 60))
    
    hours = int(adjusted_minutes // 60)
    minutes = int(adjusted_minutes % 60)

    raw_age = f"{hours}h {minutes:02d}m"
    kpi_target = "10h"

    # Status 420 (Store Waiting) and 472 (Resolving Issue) are Green status rows -> ALWAYS GREEN 🟢!
    if status_code in ('420', '472'):
        dot = "🟢"
    elif adjusted_minutes <= 600:  # Use adjusted minutes for color
        dot = "🟢"
    else:  # >10h = Red (no yellow)
        dot = "🔴"

    age_with_dot = f"{dot} {raw_age}"
    return age_with_dot, kpi_target


# ── Total Excel builder ────────────────────────────────────────────────────────

def build_total_excel(result, out_path, lang='kh', age_adjust_hours=0):
    """
    Build total Excel report with optional age adjustment.
    
    Args:
        result: Report data dictionary
        out_path: Output file path
        lang: Language ('kh' or 'en')
        age_adjust_hours: Hours to subtract from age (e.g., 12 for morning reports to exclude overnight)
    """
    import calendar
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter

    REPORT_ORDER = ['Delivery', 'Assign Deliver', 'Pickup', 'Handover to Mega']
    REPORT_COLS = {
        'Pickup':   ['ZONE', 'POST OFFICE HANDLE', 'CURRENT POST OFFICE', 'ORDER ID', 'Cus name', 'Phone'],
        'Delivery': ['ZONE', 'POST OFFICE HANDLE', 'CURRENT POST OFFICE', 'ORDER ID', 'RECEIVER', 'VIP', 'STATUS_CODE', 'NEXT_STEP', 'TOTAL FEE (USD)', 'COD (USD)', 'Age', '10H KPI'],
        'Handover to Mega': ['ZONE', 'POST OFFICE HANDLE', 'CURRENT POST OFFICE', 'ORDER ID', 'STATUS_CODE', 'NEXT_STEP'],
        'Send Mega':  ['ZONE', 'POST OFFICE HANDLE', 'CURRENT POST OFFICE', 'ORDER ID', 'STATUS_CODE', 'NEXT_STEP'],
        'Assign Deliver': ['ZONE', 'POST OFFICE HANDLE', 'CURRENT POST OFFICE', 'ORDER ID', 'RECEIVER', 'VIP', 'STATUS_CODE', 'NEXT_STEP', 'TOTAL FEE (USD)', 'COD (USD)', 'Age', '10H KPI'],
        'Not Assign': ['ZONE', 'POST OFFICE HANDLE', 'CURRENT POST OFFICE', 'ORDER ID', 'RECEIVER', 'VIP', 'STATUS_CODE', 'NEXT_STEP', 'TOTAL FEE (USD)', 'COD (USD)', 'Age', '10H KPI'],
    }

    ACTION_TRANSLATIONS_EN = {
        'ដឹកជញ្ជូន': 'Deliver',
        'ត្រួតពិនិត្យ': 'Check',
        'ពិនិត្យ': 'Check',
        'ត្រឡប់': 'Return',
        'ផ្ញើត្រឡប់': 'Return',
        'ចាត់អ្នកដឹក': 'Assign Driver',
        'ដឹកជូនថ្ងៃនេះ': 'Deliver Today',
        'ដឹកជូនឡើងវិញ': 'Redeliver',
        'ជូនដំណឹងភ្ញៀវមកទទួល': 'Notify Customer to Collect',
        'ទាក់ទងអ្នកទទួល': 'Contact Receiver',
        'បញ្ជូនត្រឡប់ទៅហាងផ្ញើ': 'Return to Sender Store',
        'បន្តដំណើរការត្រឡប់': 'Continue Return Process',
        'ពិនិត្យព័ត៌មាន': 'Verify Info',
        'ដោះស្រាយបញ្ហា': 'Resolve Issue',
        'បញ្ជាក់អាសយដ្ឋានថ្មី': 'Confirm New Address',
        'ផ្ទេរទៅឡានដឹក': 'Transfer to Truck',
        'ចាកចេញពីស្ថានីយ៍រង': 'Depart Sub-station',
        'ទទួលការផ្ទេរ': 'Receive Transfer',
        'បញ្ជូនទៅឃ្លាំងធំ': 'Dispatch to Main Warehouse',
        'ទទួលការបញ្ជូន': 'Receive Dispatch',
    }

    def _translate_val(v):
        if lang != 'en' or not v or not isinstance(v, str):
            return v
        res = str(v)
        for kh, en in ACTION_TRANSLATIONS_EN.items():
            if kh in res:
                res = res.replace(kh, en)
        return res

    type_data = result.get('type_data', {})
    order_status_map = result.get('order_status_map', {})
    day_cols  = result.get('day_cols', [])
    date_col  = result.get('cur_time_col') or 'CURRENT TIME'
    now_str   = datetime.now().strftime('%d.%m_%Hh%M')

    fn    = 'Segoe UI'
    RED   = 'EF4444'
    NAVY  = '0F172A'
    SLATE = '1E293B'
    thin  = Side(style='thin', color='BFBFBF')
    bdr   = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = Workbook()

    for idx, rn in enumerate(REPORT_ORDER):
        if idx == 0:
            ws = wb.active
            ws.title = rn
        else:
            ws = wb.create_sheet(title=rn)
        ws.views.sheetView[0].showGridLines = True
        current_row = 1

        internal_key_map = {
            'Assign Deliver': 'Branch',
            'Not Assign': 'Branch',
            'Handover to Mega': 'Transit',
            'Send Mega': 'Transit',
            'Delivery': 'Delivery',
            'Pickup': 'Pickup',
        }
        internal_key = internal_key_map.get(rn, rn)
        df = type_data.get(internal_key)
        if df is None:
            df = type_data.get(rn)
        if df is None:
            df = pd.DataFrame(columns=REPORT_COLS.get(rn, REPORT_COLS.get(internal_key, [])))
        elif df.empty:
            df = df.copy()
            for col in REPORT_COLS.get(rn, REPORT_COLS.get(internal_key, [])):
                if col not in df.columns:
                    df[col] = ''

        idx_cols = REPORT_COLS.get(rn, REPORT_COLS.get(internal_key, []))
        for col in idx_cols:
            if col not in df.columns:
                df[col] = ''

        # Map dates to day columns
        if '_scan_date' in df.columns and df['_scan_date'].notna().any():
            df['_date'] = df['_scan_date']
        elif date_col in df.columns:
            parsed = pd.to_datetime(df[date_col], dayfirst=True, format='mixed', errors='coerce')
            df['_date'] = parsed.dt.date
        else:
            df['_date'] = None

        dates_present = set(df['_date'].dropna().unique())
        # Always include today's date so today's column is never missing
        dates_present.add(datetime.now().date())
        active_days = [d for d in day_cols if d in dates_present]

        if rn in ('Delivery', 'Assign Deliver', 'Not Assign'):
            kpi_res = df.apply(lambda row: compute_kpi_info(row, age_adjust_hours), axis=1)
            df['Age'] = [r[0] for r in kpi_res]
            df['10H KPI'] = [r[1] for r in kpi_res]

        threshold_hours = 48 if rn in ('Transit', 'Send Mega') else 24
        def calc_overdue_row(r):
            age_str = str(r.get('Age', '') or '')
            m = re.search(r'(\d+)\s*h(?:\s*(\d+)\s*m)?', age_str, re.IGNORECASE)
            if m:
                h = int(m.group(1))
                return (h >= threshold_hours, h >= 168)
            for col in [
                'STATUS 306 AT STORE / AGENT FROM HUB (FIRST TIME)',
                'STATUS 306 AT STORE / AGENT (LAST TIME)',
                'STATUS 302/310 AT RECEIVING STORE / RECEIVING AGENT (FIRST TIME)',
                'CURRENT TIME',
                'CREATED DATE'
            ]:
                val = r.get(col)
                if pd.notna(val) and str(val).strip() and str(val).strip().lower() != 'nan':
                    parsed_dt = pd.to_datetime(val, dayfirst=True, format='mixed', errors='coerce')
                    if pd.notna(parsed_dt):
                        diff_h = (datetime.now() - parsed_dt).total_seconds() / 3600
                        return (diff_h >= threshold_hours, diff_h >= 168)
            return (False, False)

        ov_flags = df.apply(calc_overdue_row, axis=1)
        df['_is_overdue'] = [f[0] for f in ov_flags]
        df['_is_overdue_7days'] = [f[1] for f in ov_flags]

        idx_cols_with_flags = idx_cols + ['_is_overdue', '_is_overdue_7days']

        for d in active_days:
            df[d] = (df['_date'] == d).astype(int)
        df['Grand Total'] = 1

        agg = df.groupby(idx_cols_with_flags, sort=False, dropna=False)[
            active_days + ['Grand Total']
        ].sum().reset_index()

        for d in active_days:
            agg[d] = agg[d].apply(lambda v: int(v) if v > 0 else '')

        if rn in ('Delivery', 'Send Mega', 'Not Assign'):
            def get_age_minutes(row):
                scan_time = None
                for col in [
                    'CURRENT TIME',
                    'STATUS 306 AT STORE / AGENT (LAST TIME)',
                    'STATUS 306 AT STORE / AGENT FROM HUB (FIRST TIME)',
                    'STATUS 302/310 AT RECEIVING STORE / RECEIVING AGENT (FIRST TIME)',
                    'STATUS 306  AT ORIGIN HUB (FIRST TIME)',
                    'CREATED DATE'
                ]:
                    val = row.get(col)
                    if pd.notna(val) and str(val).strip() and str(val).strip().lower() != 'nan':
                        parsed_dt = pd.to_datetime(val, dayfirst=True, format='mixed', errors='coerce')
                        if pd.notna(parsed_dt):
                            scan_time = parsed_dt
                            break
                if scan_time is None:
                    return 0
                now = datetime.now()
                diff = now - scan_time
                return max(0, int(diff.total_seconds() // 60))

            agg['_age_mins'] = agg.apply(get_age_minutes, axis=1)
            sort_by = []
            ascending = []
            if 'ZONE' in agg.columns:
                sort_by.append('ZONE')
                ascending.append(True)
            sort_by.append('_age_mins')
            ascending.append(False)  # Oldest age first!
            agg = agg.sort_values(by=sort_by, ascending=ascending).reset_index(drop=True)
            agg = agg.drop(columns=['_age_mins'])
        else:
            sort_cols = [c for c in ['POST OFFICE HANDLE', 'CURRENT POST OFFICE', 'ORDER ID']
                         if c in agg.columns]
            agg = agg.sort_values(sort_cols).reset_index(drop=True)

        all_cols = idx_cols + active_days + ['Grand Total']
        n_idx    = len(idx_cols)
        n_total  = len(all_cols)

        title_row = current_row
        ws.row_dimensions[title_row].height = 22
        today_str = datetime.now().strftime('%d/%m/%Y %H:%M')
        tc = ws.cell(title_row, 1, f"{rn.upper()} REPORT — {today_str}")
        tc.font      = Font(name=fn, color='FFFFFF', bold=True, size=12)
        tc.fill      = PatternFill(start_color=NAVY, end_color=NAVY, fill_type='solid')
        tc.alignment = Alignment(horizontal='center', vertical='center')
        tc.border    = bdr
        for ci in range(2, n_total + 1):
            c = ws.cell(title_row, ci, '')
            c.fill   = PatternFill(start_color=NAVY, end_color=NAVY, fill_type='solid')
            c.border = bdr
        if n_total > 1:
            ws.merge_cells(start_row=title_row, end_row=title_row,
                           start_column=1, end_column=n_total)

        month_row  = current_row + 1
        header_row = current_row + 2
        ws.row_dimensions[month_row].height  = 18
        ws.row_dimensions[header_row].height = 17

        for ci in range(1, n_total + 1):
            for ri in (month_row, header_row):
                cell = ws.cell(ri, ci, '')
                cell.fill      = PatternFill(start_color=SLATE, end_color=SLATE, fill_type='solid')
                cell.font      = Font(name=fn, color='FFFFFF', bold=True, size=10)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border    = bdr

        for ci, col_name in enumerate(all_cols, start=1):
            is_idx = (ci <= n_idx)
            is_gt  = (ci == n_total)
            if is_idx or is_gt:
                ws.cell(month_row, ci, col_name)
                ws.merge_cells(start_row=month_row, end_row=header_row,
                               start_column=ci, end_column=ci)

        for ci in range(n_idx + 1, n_total):
            d = all_cols[ci - 1]
            ws.cell(header_row, ci, f"{d.day:02d}")

        month_groups = []
        cur_month, grp_start = None, None
        for ci in range(n_idx + 1, n_total):
            d = all_cols[ci - 1]
            m_val = (d.year, d.month)
            if m_val != cur_month:
                if cur_month is not None:
                    month_groups.append((cur_month, grp_start, ci - 1))
                cur_month, grp_start = m_val, ci
        if cur_month is not None:
            month_groups.append((cur_month, grp_start, n_total - 1))

        for (yr, mo), start_c, end_c in month_groups:
            ws.cell(month_row, start_c, calendar.month_name[mo])
            if end_c > start_c:
                ws.merge_cells(start_row=month_row, end_row=month_row,
                               start_column=start_c, end_column=end_c)

        order_created_map = {}
        if 'ORDER ID' in df.columns and 'CREATED DATE' in df.columns:
            parsed_created = pd.to_datetime(df['CREATED DATE'], dayfirst=True, format='mixed', errors='coerce')
            for order_id, dt in zip(df['ORDER ID'].astype(str).str.strip(), parsed_created):
                if pd.notna(dt):
                    order_created_map[order_id] = dt.date()

        day_totals  = {d: 0 for d in active_days}
        grand_total = 0
        data_start  = current_row + 3

        for ri, row in agg.iterrows():
            r = data_start + ri
            ws.row_dimensions[r].height = 15
            gt_val = int(row.get('Grand Total', 0))
            grand_total += gt_val

            is_overdue = bool(row.get('_is_overdue', False))
            is_overdue_7days = bool(row.get('_is_overdue_7days', False))
            status_code = None
            if 'ORDER ID' in row:
                oid = str(row['ORDER ID']).strip()
                status_code = order_status_map.get(oid)

            if not is_overdue:
                age_str = str(row.get('Age', '') or '')
                match = re.search(r'(\d+)\s*h(?:\s*(\d+)\s*m)?', age_str, re.IGNORECASE)
                if match:
                    h_val = int(match.group(1))
                    if h_val >= 24:
                        is_overdue = True
                    if h_val >= 168:
                        is_overdue_7days = True

            row_fill = None
            age_val_str = str(row.get('Age', '') or '')
            is_green_kpi = age_val_str.startswith('🟢')

            if status_code in ('420', '472'):
                row_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
            elif not is_green_kpi and (is_overdue or status_code in ('500', '510', '511', '512', '520', '540')):
                row_fill = PatternFill(start_color='FFEBEB', end_color='FFEBEB', fill_type='solid')

            for ci, col in enumerate(all_cols, start=1):
                val  = _translate_val(row.get(col, ''))
                cell = ws.cell(r, ci, val if val != '' else None)
                cell.border = bdr
                
                cell_fill = row_fill
                cell_font = Font(name=fn, size=10)

                # AGE Column (2 Status Colors: 🟢 Green 0-10h, 🔴 Red >10h - no yellow)
                if col == 'Age':
                    val_str = str(val or '').strip()
                    match = re.search(r'(\d+)\s*h(?:\s*(\d+)\s*m)?', val_str, re.IGNORECASE)
                    if match:
                        h_val = int(match.group(1))
                        m_val = int(match.group(2)) if match.group(2) else 0
                        t_mins = h_val * 60 + m_val
                        if status_code in ('420', '472'):
                            cell_font = Font(name=fn, size=10, bold=True, color='065F46')
                        elif t_mins <= 600:  # 0-10h = Green
                            cell_font = Font(name=fn, size=10, bold=True, color='065F46')
                        else:  # >10h = Red (no yellow)
                            cell_font = Font(name=fn, size=10, bold=True, color='991B1B')
                    else:
                        cell_font = Font(name=fn, size=10, bold=True)

                if cell_fill:
                    cell.fill = cell_fill
                cell.font = cell_font

                if col in active_days:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    if isinstance(val, (int, float)) and val:
                        day_totals[col] = day_totals.get(col, 0) + int(val)
                elif col == 'Grand Total':
                    cell.font      = Font(name=fn, color=RED, bold=True, size=10)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')

        gt_row = data_start + len(agg)
        ws.row_dimensions[gt_row].height = 17
        for ci, col in enumerate(all_cols, start=1):
            cell = ws.cell(gt_row, ci)
            cell.font      = Font(name=fn, color=RED, bold=True, size=10)
            cell.fill      = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')
            cell.border    = bdr
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if ci == 1:
                cell.value = 'Grand Total'
            elif col in active_days:
                cell.value = day_totals.get(col) or None
            elif col == 'Grand Total':
                cell.value = grand_total or None

        for ci, col in enumerate(all_cols, start=1):
            letter = get_column_letter(ci)
            if col == 'Grand Total':
                ws.column_dimensions[letter].width = 20
            elif col == 'ZONE':
                ws.column_dimensions[letter].width = 9
            elif isinstance(col, __import__('datetime').date):
                ws.column_dimensions[letter].width = 7
            elif col in ('Cus name', 'RECEIVER'):
                max_len = max(
                    (len(str(ws.cell(r_iter, ci).value or ''))
                     for r_iter in range(1, gt_row + 1)), default=20)
                ws.column_dimensions[letter].width = min(max(max_len + 3, 22), 50)
            elif col == 'Phone':
                max_len = max(
                    (len(str(ws.cell(r_iter, ci).value or ''))
                     for r_iter in range(current_row, gt_row + 1)), default=14)
                existing = ws.column_dimensions[letter].width or 0
                ws.column_dimensions[letter].width = max(existing, min(max(max_len + 3, 16), 35))
            else:
                max_len = max(
                    (len(str(ws.cell(r_iter, ci).value or ''))
                     for r_iter in range(current_row, gt_row + 1)), default=8)
                existing = ws.column_dimensions[letter].width or 0
                ws.column_dimensions[letter].width = max(existing, min(max(max_len + 2, 10), 28))

        current_row = gt_row + 3

    wb.save(out_path)
    return out_path
