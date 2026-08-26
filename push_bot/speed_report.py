import os
import sys
import re
from datetime import datetime, date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd

def parse_date(val):
    if not val or pd.isna(val):
        return None
    if isinstance(val, (datetime, date)):
        return val.date() if isinstance(val, datetime) else val
    s = str(val).strip().split(" ")[0]
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None

def parse_time(val):
    if not val or pd.isna(val):
        return None
    if isinstance(val, datetime):
        return val
    s = str(val).strip()
    for fmt in ("%d/%m/%Y %H:%M:%S", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M", "%Y-%m-%d %H:%M"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None

def build_speed_report(src_xlsx, out_xlsx, target_label="ALL", report_date=None):
    os.makedirs(os.path.dirname(os.path.abspath(out_xlsx)), exist_ok=True)
    df = pd.read_excel(src_xlsx)
    df.columns = [str(c).strip().upper() for c in df.columns]

    col_order = next((c for c in df.columns if 'ORDER ID' in c or 'ORDER' in c), 'ORDER ID')
    col_curr_po = next((c for c in df.columns if 'CURRENT POST OFFICE' in c), 'CURRENT POST OFFICE')
    col_deliv_po = next((c for c in df.columns if 'DELIVERY POST OFFICE' in c), 'DELIVERY POST OFFICE')
    col_status = next((c for c in df.columns if 'CURRENT STATUS' in c or 'STATUS' in c), 'CURRENT STATUS')
    col_created = next((c for c in df.columns if 'CREATED DATE' in c), 'CREATED DATE')
    col_action_time = next((c for c in df.columns if 'ACTION TIME' in c or 'CURRENT TIME' in c), 'CURRENT TIME')
    col_400_time = next((c for c in df.columns if '400' in c or 'OUT' in c or 'ASSIGN' in c), None)

    today = report_date or datetime.now().date()
    tgt = "".join(c for c in str(target_label).upper() if c.isalnum() or c in ("-", "_")).strip()
    if not tgt:
        tgt = "ALL"

    df['sc'] = df[col_status].astype(str).str.extract(r'^(\d{3})')[0]
    df['curr_po_clean'] = df[col_curr_po].astype(str).str.strip().str.upper()
    df['deliv_po_clean'] = df[col_deliv_po].astype(str).str.strip().str.upper()

    df['branch'] = df.apply(
        lambda r: r['deliv_po_clean'] if r['deliv_po_clean'] != 'NAN' and r['deliv_po_clean'] else r['curr_po_clean'],
        axis=1
    )

    # Filter Delivered bills (410)
    delivered_df = df[df['sc'] == '410'].copy()

    # Filter target branch / zone / all first
    if tgt not in ("ALL", "TOTAL"):
        if tgt.startswith("ZONE"):
            zone_by_prefix = {
                "KAN": "ZONE1", "PNP": "ZONE1", "PRE": "ZONE1", "SVA": "ZONE1",
                "KAM": "ZONE2", "KOH": "ZONE2", "SIH": "ZONE2", "SPE": "ZONE2", "TAK": "ZONE2", "KEP": "ZONE2",
                "BAN": "ZONE3", "BAT": "ZONE3", "CHH": "ZONE3", "PUR": "ZONE3", "PAI": "ZONE3",
                "ODD": "ZONE4", "PRH": "ZONE4", "SIE": "ZONE4", "THO": "ZONE4",
                "CHA": "ZONE5", "KRA": "ZONE5", "TBK": "ZONE5", "ROT": "ZONE5", "MON": "ZONE5", "STU": "ZONE5"
            }
            delivered_df['zone'] = delivered_df['branch'].str[:3].map(zone_by_prefix).fillna("ZONE1")
            delivered_df = delivered_df[delivered_df['zone'] == tgt].copy()
        elif len(tgt) == 3:
            delivered_df = delivered_df[delivered_df['branch'].str.startswith(tgt)].copy()
        else:
            delivered_df = delivered_df[
                (delivered_df['branch'] == tgt) |
                (delivered_df['curr_po_clean'] == tgt) |
                (delivered_df['deliv_po_clean'] == tgt)
            ].copy()

    # EVERYDAY FILTER: Filter by delivery date
    def get_deliv_date(row):
        t410 = parse_time(row.get(col_action_time)) or parse_time(row.get(col_created))
        return t410.date() if t410 else None

    delivered_df['deliv_date'] = delivered_df.apply(get_deliv_date, axis=1)
    
    # Check if there are bills for today
    today_df = delivered_df[delivered_df['deliv_date'] == today].copy()
    if len(today_df) > 0:
        delivered_df = today_df
    elif len(delivered_df) > 0:
        # Take the most recent date available in the dataset
        latest_date = delivered_df['deliv_date'].dropna().max()
        if latest_date:
            today = latest_date
            delivered_df = delivered_df[delivered_df['deliv_date'] == latest_date].copy()

    branch_stats = {}
    detail_rows = []

    for idx, row in delivered_df.iterrows():
        po = str(row.get('branch', 'OTHER')).strip()
        if not po or po == 'NAN':
            continue

        if po not in branch_stats:
            branch_stats[po] = {
                "po": po,
                "total_delivered": 0,
                "under_2h_count": 0,
                "2_to_4h_count": 0,
                "4_to_8h_count": 0,
                "over_8h_count": 0,
                "total_commission_usd": 0.0
            }

        order_id = str(row.get(col_order, '')).strip()
        t410 = parse_time(row.get(col_action_time)) or parse_time(row.get(col_created))
        t400 = parse_time(row.get(col_400_time)) if col_400_time else None
        if not t400 and t410:
            t400 = parse_time(row.get(col_created))

        duration_hours = (t410 - t400).total_seconds() / 3600.0 if (t410 and t400 and t410 >= t400) else 1.5

        branch_stats[po]["total_delivered"] += 1

        if duration_hours < 2.0:
            tier = "< 2 Hours (+50%)"
            rate_usd = 0.30
            branch_stats[po]["under_2h_count"] += 1
            tag_color = "GREEN"
        elif duration_hours <= 4.0:
            tier = "2 - 4 Hours (+25%)"
            rate_usd = 0.25
            branch_stats[po]["2_to_4h_count"] += 1
            tag_color = "BLUE"
        elif duration_hours <= 8.0:
            tier = "4 - 8 Hours (Normal)"
            rate_usd = 0.20
            branch_stats[po]["4_to_8h_count"] += 1
            tag_color = "NORMAL"
        else:
            tier = "> 8 Hours (-25% Fine)"
            rate_usd = 0.15
            branch_stats[po]["over_8h_count"] += 1
            tag_color = "RED"

        branch_stats[po]["total_commission_usd"] += rate_usd
        dur_str = f"{int(duration_hours)}h {int((duration_hours%1)*60)}m" if duration_hours else "N/A"

        detail_rows.append({
            "order_id": order_id,
            "po": po,
            "t400": str(t400 or ""),
            "t410": str(t410 or ""),
            "duration_hours": round(duration_hours, 2),
            "dur_str": dur_str,
            "tier": tier,
            "rate_usd": rate_usd,
            "tag_color": tag_color
        })

    sorted_branches = sorted(branch_stats.values(), key=lambda x: x["po"])

    # Build Excel
    wb = openpyxl.Workbook()
    fn = "Segoe UI"
    banner_font = Font(name=fn, size=11, bold=True, color="FFFFFF")
    hdr_font = Font(name=fn, size=10, bold=True, color="FFFFFF")
    data_font = Font(name=fn, size=10, color="0F172A")
    bold_data_font = Font(name=fn, size=10, bold=True, color="0F172A")
    tot_font = Font(name=fn, size=10, bold=True, color="15803D")

    hdr_fill = PatternFill("solid", fgColor="0B132B")      # Dark Navy
    sub_hdr_fill = PatternFill("solid", fgColor="1E3A8A")  # Royal Blue
    tot_fill = PatternFill("solid", fgColor="DCFCE7")      # Light Green
    green_fill = PatternFill("solid", fgColor="DCFCE7")    # Light Green
    blue_fill = PatternFill("solid", fgColor="EFF6FF")     # Light Blue
    red_fill = PatternFill("solid", fgColor="FEE2E2")      # Light Red
    zebra_fill = PatternFill("solid", fgColor="F8FAFC")

    border_thin = Border(
        left=Side(style="thin", color="CBD5E1"), right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"), bottom=Side(style="thin", color="CBD5E1")
    )

    # 1. SUMMARY SHEET
    ws_sum = wb.active
    ws_sum.title = "SUMMARY_DAILY_SPEED"
    ws_sum.views.sheetView[0].showGridLines = True

    date_str = today.strftime('%d/%m/%Y')
    ws_sum.merge_cells("A1:H1")
    ws_sum["A1"].value = f"METFONE EXPRESS — DAILY DELIVERY SPEED BONUS REPORT ({tgt}) — {date_str}"
    ws_sum["A1"].font = banner_font
    ws_sum["A1"].fill = hdr_fill
    ws_sum["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.row_dimensions[1].height = 28

    ws_sum.merge_cells("A2:H2")
    ws_sum["A2"].value = f"Date: {date_str}  |  <2h (+50% / $0.30)  |  2-4h (+25% / $0.25)  |  4-8h ($0.20)  |  >8h (-25% / $0.15 Fine)"
    ws_sum["A2"].font = Font(name=fn, size=9, italic=True, color="FFFFFF")
    ws_sum["A2"].fill = sub_hdr_fill
    ws_sum["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.row_dimensions[2].height = 20

    sum_headers = [
        "Post Office",
        "Delivered Today (410)",
        "< 2 Hours (+50%)",
        "2 - 4 Hours (+25%)",
        "4 - 8 Hours (Normal)",
        "> 8 Hours (-25% Fine)",
        "Fast Rate (<4h %)",
        "Daily Commission ($)"
    ]

    ws_sum.row_dimensions[4].height = 26
    for col_idx, h in enumerate(sum_headers, 1):
        c = ws_sum.cell(row=4, column=col_idx, value=h)
        c.font = hdr_font
        c.fill = sub_hdr_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border_thin

    tot_del = 0
    tot_u2 = 0
    tot_24 = 0
    tot_48 = 0
    tot_o8 = 0
    tot_pay = 0.0

    r_cur = 5
    for b in sorted_branches:
        ws_sum.row_dimensions[r_cur].height = 22
        fast_pct = ((b["under_2h_count"] + b["2_to_4h_count"]) / b["total_delivered"] * 100) if b["total_delivered"] > 0 else 0
        vals = [
            b["po"],
            b["total_delivered"],
            b["under_2h_count"],
            b["2_to_4h_count"],
            b["4_to_8h_count"],
            b["over_8h_count"],
            f"{fast_pct:.1f}%",
            f"${b['total_commission_usd']:.2f}"
        ]
        for col_idx, val in enumerate(vals, 1):
            c = ws_sum.cell(row=r_cur, column=col_idx, value=val)
            c.font = bold_data_font if col_idx in (1, 8) else data_font
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = border_thin
            if r_cur % 2 == 0:
                c.fill = zebra_fill
            if col_idx == 8:
                c.font = Font(name=fn, size=10, bold=True, color="15803D")

        tot_del += b["total_delivered"]
        tot_u2 += b["under_2h_count"]
        tot_24 += b["2_to_4h_count"]
        tot_48 += b["4_to_8h_count"]
        tot_o8 += b["over_8h_count"]
        tot_pay += b["total_commission_usd"]
        r_cur += 1

    # If no data found for specific branch, still write a blank row
    if not sorted_branches and tgt not in ("ALL", "TOTAL"):
        ws_sum.row_dimensions[r_cur].height = 22
        vals = [tgt, 0, 0, 0, 0, 0, "0.0%", "$0.00"]
        for col_idx, val in enumerate(vals, 1):
            c = ws_sum.cell(row=r_cur, column=col_idx, value=val)
            c.font = data_font
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = border_thin
        r_cur += 1

    # TOTAL ROW
    ws_sum.row_dimensions[r_cur].height = 26
    tot_fast_pct = ((tot_u2 + tot_24) / tot_del * 100) if tot_del > 0 else 0
    tot_vals = [
        "TOTAL",
        tot_del,
        tot_u2,
        tot_24,
        tot_48,
        tot_o8,
        f"{tot_fast_pct:.1f}%",
        f"${tot_pay:.2f}"
    ]
    for col_idx, val in enumerate(tot_vals, 1):
        c = ws_sum.cell(row=r_cur, column=col_idx, value=val)
        c.font = tot_font
        c.fill = tot_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border_thin

    for col in ws_sum.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_sum.column_dimensions[col_letter].width = max(max_len + 4, 16)

    # 2. DETAIL SHEET
    ws_det = wb.create_sheet(title="DETAIL_DELIVERY_SPEED")
    ws_det.views.sheetView[0].showGridLines = True

    det_headers = [
        "No", "Order ID", "Post Office", "Out for Delivery (400)",
        "Delivered Time (410)", "Duration", "Hours (Dec)", "Speed Category", "Rate ($/Bill)"
    ]

    ws_det.row_dimensions[1].height = 24
    for col_idx, h in enumerate(det_headers, 1):
        c = ws_det.cell(row=1, column=col_idx, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border_thin

    for idx, item in enumerate(detail_rows, 1):
        r_num = idx + 1
        ws_det.row_dimensions[r_num].height = 20
        row_data = [
            idx,
            item["order_id"],
            item["po"],
            item["t400"],
            item["t410"],
            item["dur_str"],
            item["duration_hours"],
            item["tier"],
            f"${item['rate_usd']:.2f}"
        ]
        for col_idx, val in enumerate(row_data, 1):
            c = ws_det.cell(row=r_num, column=col_idx, value=val)
            c.font = data_font
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = border_thin
            if item["tag_color"] == "GREEN":
                c.fill = green_fill
            elif item["tag_color"] == "BLUE":
                c.fill = blue_fill
            elif item["tag_color"] == "RED":
                c.fill = red_fill

    for col in ws_det.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_det.column_dimensions[col_letter].width = max(max_len + 4, 14)

    wb.save(out_xlsx)
    return tot_del, tot_u2, tot_24, tot_o8, tot_pay
