# -*- coding: utf-8 -*-
"""
branch_today.py - Generate Branch / Post Office Today Performance Summary Report
Tracks 3 key metrics per post office / district:
1. From Mega (Status 306 - Received from Hub)
2. Pending (Pending delivery - status not 410, 520, 201)
3. Success Delivery (Status 410 - Delivered successfully)
"""

import os
import openpyxl
import copy
import pandas as pd
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

POST_OFFICE_DISTRICT_MAP = {
    # Phnom Penh (PNP)
    'PNPA002': 'Boeng Keng Kang', 'PNPP001': 'Boeng Keng Kang', 'PNPP007': 'Boeng Keng Kang', 'PNPS007': 'Boeng Keng Kang',
    'PNPP005': 'Chbar Ampov', 'PNPP010': 'Chraoy Chongvar', 'PNPP011': 'Dangkao', 'PNPP014': 'Doun Penh',
    'PNPA016': 'Kamboul', 'PNPP012': 'Kamboul', 'PNPA029': 'Mean Chey', 'PNPA055': 'Mean Chey', 'PNPP002': 'Mean Chey', 'PNPP003': 'Mean Chey',
    'PNPA028': 'Pou Saen Chey', 'PNPA074': 'Pou Saen Chey', 'PNPP008': 'Pou Saen Chey', 'PNPP009': 'Pou Saen Chey', 'PNPP004': 'Preaek Pnov',
    'PNPA040': 'Saen Sokh', 'PNPP013': 'Saen Sokh', 'PNPA036': 'Tuol Kouk', 'PNPP006': 'Tuol Kouk',

    # Kandal (KAN)
    'KANA024': 'Kandal Stueng', 'KANA028': 'Kandal Stueng', 'KANA049': 'Kandal Stueng',
    'KANS003': 'Kaoh Thum', 'KANA031': 'Khsach Kandal', 'KANA012': 'Kien Svay', 'KANA013': 'Kien Svay',
    'KANA008': 'Leuk Daek', 'KANS004': 'Mukh Kampul', 'KANA019': 'Ponhea Lueu',
    'KANA007': "S'ang", 'KANA020': "S'ang", 'KANA026': "S'ang", 'KANA040': 'Sampov Pun',
    'KANA023': 'Ta Khmau', 'KANP001': 'Ta Khmau',

    # Prey Veng (PRE)
    'PREA024': 'Ba Phnum', 'PREA023': 'Peam Ro', 'PREA020': 'Preah Sdach', 'PREA029': 'Preah Sdach', 'PREA035': 'Preah Sdach',
    'PREP001': 'Prey Veng', 'PRES001': 'Prey Veng', 'PREA002': 'Pur Rieng', 'PREA028': 'Sithor Kandal',

    # Svay Rieng (SVA)
    'SVAP001': 'Svay Rieng', 'SVAS002': 'Bavet', 'SVAA001': 'Bavet', 'SVAA002': 'Bavet', 'SVAA003': 'Romeas Haek', 'SVAA004': 'Rumduol', 'SVAA005': 'Svay Chrum',

    # Battambang (BAT)
    'BATA003': 'Banan', 'BATA016': 'Banan',
    'BATA001': 'Battambang', 'BATA009': 'Battambang', 'BATA040': 'Battambang', 'BATA042': 'Battambang', 'BATP001': 'Battambang',
    'BATA011': 'Kamrieng', 'BATS007': 'Moung Ruessei', 'BATA017': 'Samlout',
    'BATA010': 'Sampov Lun', 'BATA028': 'Sampov Lun', 'BATA004': 'Sangkae', 'BATA008': 'Sangkae',
    'BATA023': 'Thma Koul', 'BATA025': 'Thma Koul',

    # Siem Reap (SIE)
    'SIEP001': 'Siem Reap', 'SIEA001': 'Angkor Chum', 'SIEA002': 'Angkor Thon', 'SIEA003': 'Banteay Srei', 'SIEA004': 'Chi Kraeng',

    # Sihanoukville (SIH)
    'SIHP001': 'Sihanoukville', 'SIHA001': 'Preah Sihanouk', 'SIHA002': 'Stung Hav', 'SIHA003': 'Kampong Seila'
}


def build_branch_today_report(src_excel, out_xlsx, target_label="ALL"):
    """Builds Branch/Post Office Today Performance Report Excel file."""
    df = pd.read_excel(src_excel)

    # Standardize target label
    tgt = str(target_label).strip().upper()
    target_clean = tgt.replace(" ", "_")

    zone_by_prefix = {
        "KAN": "ZONE1", "PNP": "ZONE1", "PRE": "ZONE1", "SVA": "ZONE1",
        "KAM": "ZONE2", "KOH": "ZONE2", "SIH": "ZONE2", "SPE": "ZONE2", "TAK": "ZONE2",
        "BAN": "ZONE3", "BAT": "ZONE3", "CHH": "ZONE3", "PUR": "ZONE3",
        "ODD": "ZONE4", "PRH": "ZONE4", "SIE": "ZONE4", "THO": "ZONE4",
        "CHA": "ZONE5", "KRA": "ZONE5", "TBK": "ZONE5", "ROT": "ZONE5", "MON": "ZONE5", "STU": "ZONE5"
    }

    summary_data = {} # (zone_str, br, dist_or_po) -> {"from_mega": c1, "pending": c2, "success": c3}
    total_from_mega = 0
    total_pending = 0
    total_success = 0

    # Determine if target is a specific 7-char Post Office Code (e.g. PNPP012, SVAP001, BATP001)
    is_specific_po = len(tgt) >= 7 and (tgt.endswith("001") or tgt.endswith("012") or tgt.endswith("002") or tgt[3:4] in ("P", "A", "S"))

    for idx, row in df.iterrows():
        dest_po = str(row.get("DELIVERY POST OFFICE", "")).strip().upper()
        dest_prov = str(row.get("DELIVERY PROVINCE", "")).strip().upper()
        curr_status = str(row.get("CURRENT STATUS", "")).strip().upper()

        if not dest_po or dest_po == "NAN":
            continue

        dest_br = dest_po[:3] if len(dest_po) >= 3 else dest_prov

        # Filtering target logic
        tgt_norm = tgt.replace(" ", "")
        if is_specific_po:
            if dest_po != tgt:
                continue
        elif tgt_norm.startswith("ZONE"):
            target_zone_name = tgt_norm
            item_zone = zone_by_prefix.get(dest_prov, "ZONE1")
            if item_zone != target_zone_name:
                continue
        elif tgt not in ("ALL", "TOTAL", "MEGA"):
            branch_prefix = tgt[:3]
            if dest_prov != branch_prefix and dest_po != tgt:
                continue

        # District / PO Label
        if is_specific_po:
            dist_name = dest_po
        elif dest_po in POST_OFFICE_DISTRICT_MAP:
            dist_name = POST_OFFICE_DISTRICT_MAP[dest_po]
        else:
            if dest_br == 'SVA':
                dist_name = 'Svay Rieng'
            elif dest_br == 'PNP':
                dist_name = 'Phnom Penh'
            elif dest_br == 'KAN':
                dist_name = 'General District'
            elif dest_br == 'BAT':
                dist_name = 'Battambang'
            elif dest_br == 'PRE':
                dist_name = 'Prey Veng'
            elif dest_br == 'SIE':
                dist_name = 'Siem Reap'
            elif dest_br == 'SIH':
                dist_name = 'Sihanoukville'
            else:
                dist_name = dest_prov if dest_prov != dest_po else 'General District'

        zone_str = zone_by_prefix.get(dest_prov, "Zone 1").replace("ZONE", "Zone ")

        # Status check
        is_from_mega = "306" in curr_status or pd.notna(row.get("STATUS 306 AT STORE / AGENT FROM HUB (FIRST TIME)"))
        is_success   = "410" in curr_status or "GIAO THÀNH CÔNG" in curr_status
        is_done      = is_success or any(st in curr_status for st in ["520", "201", "CANCEL"])
        is_pending   = not is_done

        key = (zone_str, dest_br, dist_name)
        if key not in summary_data:
            summary_data[key] = {"from_mega": 0, "pending": 0, "success": 0}

        if is_from_mega:
            summary_data[key]["from_mega"] += 1
            total_from_mega += 1
        if is_pending:
            summary_data[key]["pending"] += 1
            total_pending += 1
        if is_success:
            summary_data[key]["success"] += 1
            total_success += 1

    # Create Workbook
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "BRANCH TODAY PERFORMANCE"
    ws1.views.sheetView[0].showGridLines = True

    # Styling Tokens
    font_title = Font(name="Segoe UI", size=14, bold=True, color="FFFFFF")
    font_hdr   = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_data  = Font(name="Segoe UI", size=10, color="0F172A")
    font_tot   = Font(name="Segoe UI", size=11, bold=True, color="0F172A")
    font_tot_green = Font(name="Segoe UI", size=11, bold=True, color="15803D")
    font_tot_amber = Font(name="Segoe UI", size=11, bold=True, color="B45309")

    fill_hdr      = PatternFill("solid", fgColor="0F766E")
    fill_sub      = PatternFill("solid", fgColor="E0F2FE")
    fill_tot      = PatternFill("solid", fgColor="CCFBF1")

    border_clean = Border(
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0")
    )
    border_tot_acc = Border(
        top=Side(style="thin", color="0F766E"),
        bottom=Side(style="double", color="0F766E"),
        left=Side(style="thin", color="0F766E"),
        right=Side(style="thin", color="0F766E")
    )

    # Title Banner (Row 1)
    ws1.row_dimensions[1].height = 35.0
    ws1.merge_cells("A1:F1")
    title_cell = ws1.cell(1, 1, f"TODAY PERFORMANCE SUMMARY ({target_label})")
    title_cell.font = font_title
    title_cell.fill = fill_hdr
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Table Headers (Row 2)
    headers = [
        "ZONE\n(តំបន់)",
        "DESTINATION_BRANCH\n(សាខា)",
        "Post Office / District\n(ស្រុក/ខណ្ឌ/បុស្តិ៍)",
        "From Mega\n(ទទួលពី MEGA)",
        "Pending\n(កំពុងរង់ចាំ)",
        "Success Delivery\n(ដឹកជោគជ័យ)"
    ]
    ws1.row_dimensions[2].height = 35.0
    for c_idx, h in enumerate(headers, 1):
        cell = ws1.cell(2, c_idx, h)
        cell.font = font_hdr
        cell.fill = fill_hdr
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data & Subtotal Rows
    r_curr = 3
    branch_groups = {}
    for (zone_str, br, dist), stats in sorted(summary_data.items()):
        if br not in branch_groups:
            branch_groups[br] = []
        branch_groups[br].append((zone_str, br, dist, stats))

    for br in sorted(branch_groups.keys()):
        br_items = branch_groups[br]
        br_mega    = 0
        br_pending = 0
        br_success = 0

        for zone_str, b_code, dist, stats in br_items:
            ws1.row_dimensions[r_curr].height = 20.0
            row_vals = [zone_str, b_code, dist, stats["from_mega"], stats["pending"], stats["success"]]
            for ci, val in enumerate(row_vals, 1):
                cell = ws1.cell(r_curr, ci, val)
                cell.font = font_data
                cell.border = border_clean
                if ci in (1, 2, 3):
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.number_format = "#,##0"

            br_mega    += stats["from_mega"]
            br_pending += stats["pending"]
            br_success += stats["success"]
            r_curr += 1

        # Branch Subtotal Row
        ws1.row_dimensions[r_curr].height = 22.0
        ws1.merge_cells(start_row=r_curr, start_column=1, end_row=r_curr, end_column=3)
        lbl_c = ws1.cell(r_curr, 1, f"{br} Total")
        lbl_c.font = font_tot
        lbl_c.alignment = Alignment(horizontal="left", vertical="center")

        for c in range(1, 4):
            cell = ws1.cell(r_curr, c)
            cell.fill = fill_sub
            cell.border = border_clean

        m_c = ws1.cell(r_curr, 4, br_mega)
        m_c.font = font_tot
        m_c.fill = fill_sub
        m_c.border = border_clean
        m_c.alignment = Alignment(horizontal="right", vertical="center")
        m_c.number_format = "#,##0"

        p_c = ws1.cell(r_curr, 5, br_pending)
        p_c.font = font_tot_amber
        p_c.fill = fill_sub
        p_c.border = border_clean
        p_c.alignment = Alignment(horizontal="right", vertical="center")
        p_c.number_format = "#,##0"

        s_c = ws1.cell(r_curr, 6, br_success)
        s_c.font = font_tot_green
        s_c.fill = fill_sub
        s_c.border = border_clean
        s_c.alignment = Alignment(horizontal="right", vertical="center")
        s_c.number_format = "#,##0"

        r_curr += 1

    # Grand Total Row
    ws1.row_dimensions[r_curr].height = 25.0
    ws1.merge_cells(start_row=r_curr, start_column=1, end_row=r_curr, end_column=3)
    gt_lbl = ws1.cell(r_curr, 1, f"{target_clean[:7]} Total")
    gt_lbl.font = font_tot
    gt_lbl.alignment = Alignment(horizontal="left", vertical="center")

    for c in range(1, 4):
        cell = ws1.cell(r_curr, c)
        cell.fill = fill_tot
        cell.border = border_tot_acc

    gt_m = ws1.cell(r_curr, 4, total_from_mega)
    gt_m.font = font_tot
    gt_m.fill = fill_tot
    gt_m.border = border_tot_acc
    gt_m.alignment = Alignment(horizontal="right", vertical="center")
    gt_m.number_format = "#,##0"

    gt_p = ws1.cell(r_curr, 5, total_pending)
    gt_p.font = font_tot_amber
    gt_p.fill = fill_tot
    gt_p.border = border_tot_acc
    gt_p.alignment = Alignment(horizontal="right", vertical="center")
    gt_p.number_format = "#,##0"

    gt_s = ws1.cell(r_curr, 6, total_success)
    gt_s.font = font_tot_green
    gt_s.fill = fill_tot
    gt_s.border = border_tot_acc
    gt_s.alignment = Alignment(horizontal="right", vertical="center")
    gt_s.number_format = "#,##0"

    # Column Widths
    col_widths = [14, 22, 24, 18, 18, 20]
    for ci, w in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(ci)].width = w

    wb.save(out_xlsx)
    return total_from_mega, total_pending, total_success


def render_today_summary_image(out_xlsx):
    """Renders pixel-perfect PNG image of Today Performance Summary table."""
    import excel_to_image
    return excel_to_image.excel_to_image(out_xlsx)
