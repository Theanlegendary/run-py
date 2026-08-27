import os
import json
import glob
import re
import pandas as pd
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

COMPARE_DIR = "compare"
os.makedirs(COMPARE_DIR, exist_ok=True)

SNAPSHOT_FILE = os.path.join(COMPARE_DIR, "compare_snapshots.json")
EXCLUDE_KEYWORDS = ["TRAINER", "GLOBAL", "EXTERNAL", "TEST", "CENTER02", "DVCZ"]

ALL_COMPANY_BRANCHES = [
    # Zone 1
    ("Zone 1", "KANP001"),
    ("Zone 1", "PNPP001"),
    ("Zone 1", "PNPP002"),
    ("Zone 1", "PNPP003"),
    ("Zone 1", "PNPP004"),
    ("Zone 1", "PNPP005"),
    ("Zone 1", "PNPP006"),
    ("Zone 1", "PNPP007"),
    ("Zone 1", "PNPP008"),
    ("Zone 1", "PNPP009"),
    ("Zone 1", "PNPP010"),
    ("Zone 1", "PNPP011"),
    ("Zone 1", "PNPP012"),
    ("Zone 1", "PNPP013"),
    ("Zone 1", "PNPP014"),
    ("Zone 1", "PREP001"),
    ("Zone 1", "SVAP001"),
    # Zone 2
    ("Zone 2", "SIHP001"),
    ("Zone 2", "KAMP001"),
    ("Zone 2", "SPEP001"),
    ("Zone 2", "TAKP001"),
    ("Zone 2", "KOHP001"),
    ("Zone 2", "KEPP001"),
    # Zone 3
    ("Zone 3", "BANP001"),
    ("Zone 3", "BATP001"),
    ("Zone 3", "CHHP001"),
    ("Zone 3", "PURP001"),
    # Zone 4
    ("Zone 4", "SIEP001"),
    ("Zone 4", "ODDP001"),
    ("Zone 4", "THOP001"),
    ("Zone 4", "PRHP001"),
    # Zone 5
    ("Zone 5", "KRAP001"),
    ("Zone 5", "TBKP001"),
    ("Zone 5", "STUP001"),
    ("Zone 5", "ROTP001"),
    ("Zone 5", "MONP001"),
    ("Zone 5", "CHAP001"),
    ("Zone 5", "PAIP001")
]

def clean_cache_3_times_daily():
    """Cleans old files in compare/ folder 3 times per day."""
    try:
        if os.path.exists(SNAPSHOT_FILE):
            with open(SNAPSHOT_FILE, "r", encoding="utf-8") as f:
                snapshots = json.load(f)
            
            cutoff_date = datetime.now() - timedelta(days=3)
            cleaned = {}
            for d_str, val in snapshots.items():
                try:
                    dt = datetime.strptime(d_str, "%d/%m/%Y")
                    if dt >= cutoff_date:
                        cleaned[d_str] = val
                except Exception:
                    cleaned[d_str] = val

            with open(SNAPSHOT_FILE, "w", encoding="utf-8") as f:
                json.dump(cleaned, f, ensure_ascii=False, indent=2)

        for fpath in glob.glob(os.path.join(COMPARE_DIR, "*.xlsx")):
            try:
                mtime = datetime.fromtimestamp(os.path.getmtime(fpath))
                if datetime.now() - mtime > timedelta(hours=24):
                    os.remove(fpath)
            except Exception:
                pass
    except Exception:
        pass

def load_post_office_lookup_map():
    po_map = {}
    if os.path.exists("post_office_lookup.csv"):
        try:
            df_po = pd.read_csv("post_office_lookup.csv", encoding="utf-8-sig")
            df_po.columns = [str(c).strip().lower() for c in df_po.columns]
            if "current_post_office" in df_po.columns and "post_office_handle" in df_po.columns:
                for _, r in df_po.iterrows():
                    cur = str(r["current_post_office"]).strip().upper()
                    hnd = str(r["post_office_handle"]).strip().upper()
                    if cur and hnd and cur != "NAN" and hnd != "NAN":
                        po_map[cur] = hnd
        except Exception:
            pass
    return po_map

PO_LOOKUP_MAP = load_post_office_lookup_map()

def resolve_post_office_handle(raw_po):
    po = str(raw_po).strip().upper()
    if not po or po == "NAN" or any(kw in po for kw in EXCLUDE_KEYWORDS):
        return None

    global PO_LOOKUP_MAP
    if not PO_LOOKUP_MAP:
        PO_LOOKUP_MAP = load_post_office_lookup_map()

    if po in PO_LOOKUP_MAP:
        return PO_LOOKUP_MAP[po]

    return po

def load_config_zone_mapping():
    try:
        if os.path.exists("config.json"):
            with open("config.json", "r", encoding="utf-8") as f:
                cfg = json.load(f)
            return cfg.get("zone_mapping", {})
    except Exception:
        pass
    return {}

def resolve_branch_zone(branch_code, df_detail=None):
    b_code = str(branch_code).strip().upper()
    for z_name, h_code in ALL_COMPANY_BRANCHES:
        if h_code == b_code:
            return z_name

    zm = load_config_zone_mapping()
    by_po = zm.get("by_post_office", {})
    if b_code in by_po:
        return by_po[b_code]
        
    by_prefix = zm.get("by_prefix", {})
    prefix3 = b_code[:3]
    if prefix3 in by_prefix:
        return by_prefix[prefix3]
        
    return zm.get("default_zone", "Zone 1")

def is_urgent_bill(r):
    sc = ""
    for sc_cand in ["STATUS_CODE", "STATUS CODE", "STATUS", "STATUS_NAME", "STATUS NAME", "STATE_CODE", "STATE", "CODE"]:
        if sc_cand in r and pd.notna(r[sc_cand]):
            sc_val = str(r[sc_cand]).lstrip("S").strip()
            if sc_val:
                sc = sc_val.split()[0].strip()
                break

    if sc in ("99", "100", "201", "410", "520"):
        return False

    r_text = ' '.join(str(v) for v in r.values() if pd.notna(v)).upper()
    if any(kw in r_text for kw in ['410', '520', 'GIAO THÀNH CÔNG', 'DELIVERED', 'COMPLETED', 'ĐÃ GIAO', 'DA GIAO']):
        return False

    if "_IS_OVERDUE" in r and pd.notna(r["_IS_OVERDUE"]):
        return bool(r["_IS_OVERDUE"])

    age_val = str(r.get("AGE", "") or r.get("CURRENT TIME", "") or r.get("SCAN TIME", "") or "")
    if "🔴" in age_val:
        return True

    cat_raw = str(r.get("REPORT TYPE", "") or r.get("TYPE", "") or r.get("_REPORT_CLASS", "") or "").upper()
    is_transit = ("TRANSIT" in cat_raw or "MEGA" in cat_raw or sc in ("306", "309"))
    threshold = 48 if is_transit else 24

    if age_val:
        match = re.search(r'(\d+)\s*h', age_val, re.IGNORECASE)
        if match:
            return float(match.group(1)) >= threshold

    return False

def get_row_post_office_handle(r):
    for col in ["POST OFFICE HANDLE", "CURRENT POST OFFICE", "DELIVERY POST OFFICE", "RECEIVE POST OFFICE"]:
        val = str(r.get(col, "") or "").strip()
        hnd = resolve_post_office_handle(val)
        if hnd and hnd != "NAN" and not any(kw in hnd for kw in EXCLUDE_KEYWORDS):
            if "MEGA" not in hnd and "DVC" not in hnd:
                return hnd

    for col in ["DELIVERY POST OFFICE", "RECEIVE POST OFFICE", "CURRENT POST OFFICE"]:
        val = str(r.get(col, "") or "").strip()
        hnd = resolve_post_office_handle(val)
        if hnd and hnd != "NAN" and not any(kw in hnd for kw in EXCLUDE_KEYWORDS):
            return hnd

    return None

def load_snapshots():
    clean_cache_3_times_daily()
    if os.path.exists(SNAPSHOT_FILE):
        try:
            with open(SNAPSHOT_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}
    return {}

def save_snapshots(data):
    with open(SNAPSHOT_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def extract_total_report_counts(df_detail):
    """Extracts exact overdue urgent counts per post office handle using generate_report engine."""
    import generate_report
    df = df_detail.copy()
    
    handle_map = {}
    try:
        tmp_out = os.path.join(COMPARE_DIR, "tmp_gen_report")
        os.makedirs(tmp_out, exist_ok=True)
        res = generate_report.generate_reports_from_data(df, target_handles=None, output_dir=tmp_out)
        h_results = res.get("handle_results", [])
        
        for hr in h_results:
            raw_h = str(hr.get("handle", "")).strip().upper()
            hnd = resolve_post_office_handle(raw_h)
            if not hnd:
                hnd = raw_h
            if any(kw in hnd for kw in EXCLUDE_KEYWORDS):
                continue
            
            urgent_cnt = 0
            for sec_name, sec_rows, sec_tot, sec_icols, active_days in hr.get("sections", []):
                urgent_cnt += len(sec_rows)
                
            handle_map[hnd] = {
                "urgent": urgent_cnt,
                "pickup": hr.get("handle_counts", {}).get("Pickup", 0),
                "delivery": hr.get("handle_counts", {}).get("Delivery", 0),
                "transit": hr.get("handle_counts", {}).get("Handover to Mega", hr.get("handle_counts", {}).get("Send Mega", hr.get("handle_counts", {}).get("Transit", 0))),
                "branch": hr.get("handle_counts", {}).get("Assign Deliver", hr.get("handle_counts", {}).get("Not Assign", hr.get("handle_counts", {}).get("Branch", 0)))
            }
    except Exception as e:
        pass

    if not handle_map:
        df.columns = [str(c).strip().upper() for c in df.columns]
        for _, r in df.iterrows():
            if not is_urgent_bill(r):
                continue
            hnd = get_row_post_office_handle(r)
            if not hnd or any(kw in hnd for kw in EXCLUDE_KEYWORDS):
                continue
            if hnd not in handle_map:
                handle_map[hnd] = {"urgent": 0, "pickup": 0, "delivery": 0, "transit": 0, "branch": 0}
            handle_map[hnd]["urgent"] += 1

    return handle_map

def determine_shift(now=None):
    """Determines shift based on exact clock hours: 9AM (7-11 AM), 2PM (12-3 PM), 5PM (4 PM+)."""
    if now is None:
        now = datetime.now()
    hour = now.hour
    if hour < 12:
        return "9AM"
    elif hour < 16:
        return "2PM"
    else:
        return "5PM"

def record_total_snapshot(date_str, df_detail):
    snapshots = load_snapshots()
    if date_str not in snapshots:
        snapshots[date_str] = {}

    curr_shift = determine_shift()
    handle_map = extract_total_report_counts(df_detail)
    now_str = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

    snapshots[date_str][curr_shift] = {
        "captured_at": now_str,
        "handles": handle_map
    }

    if "9AM" not in snapshots[date_str] or not snapshots[date_str]["9AM"].get("handles"):
        snapshots[date_str]["9AM"] = {
            "captured_at": now_str,
            "handles": handle_map
        }

    if curr_shift == "5PM" and ("2PM" not in snapshots[date_str] or not snapshots[date_str]["2PM"].get("handles")):
        snapshots[date_str]["2PM"] = {
            "captured_at": now_str,
            "handles": handle_map
        }

    save_snapshots(snapshots)
    return snapshots

def build_comparison_summary(date_str, df_detail=None, target_shift=None):
    if df_detail is not None:
        record_total_snapshot(date_str, df_detail)

    snapshots = load_snapshots()
    day_data = snapshots.get(date_str, {})

    h_9am = day_data.get("9AM", {}).get("handles", {})
    h_2pm = day_data.get("2PM", {}).get("handles", {})
    h_5pm = day_data.get("5PM", {}).get("handles", {})

    has_9am = bool(h_9am)
    has_2pm = bool(h_2pm)
    has_5pm = bool(h_5pm)

    rows = []
    tot_urg_9am = tot_urg_2pm = tot_urg_5pm = 0

    for z, h in ALL_COMPANY_BRANCHES:
        u9 = h_9am.get(h, {}).get("urgent", 0) if has_9am else 0
        u2 = h_2pm.get(h, {}).get("urgent", 0) if has_2pm else 0
        u5 = h_5pm.get(h, {}).get("urgent", 0) if has_5pm else 0

        if has_5pm and has_9am:
            urg_change = u5 - u9
            clear_pct = ((u9 - u5) / u9 * 100.0) if u9 > 0 else 100.0
        elif has_2pm and has_9am:
            urg_change = u2 - u9
            clear_pct = ((u9 - u2) / u9 * 100.0) if u9 > 0 else 100.0
        else:
            urg_change = 0
            clear_pct = 0.0

        urg_change_str = f"{urg_change}" if urg_change == 0 else (f"+{urg_change}" if urg_change > 0 else f"{urg_change}")

        rows.append({
            "ZONE": z,
            "POST OFFICE HANDLE": h,
            "URGENT_9AM": u9,
            "URGENT_2PM": u2,
            "URGENT_5PM": u5,
            "URGENT_CHANGE": urg_change_str,
            "CLEARANCE_PCT": clear_pct
        })

        tot_urg_9am += u9
        tot_urg_2pm += u2
        tot_urg_5pm += u5

    if has_5pm and has_9am:
        grand_change = tot_urg_5pm - tot_urg_9am
        grand_pct = ((tot_urg_9am - tot_urg_5pm) / tot_urg_9am * 100.0) if tot_urg_9am > 0 else 100.0
    elif has_2pm and has_9am:
        grand_change = tot_urg_2pm - tot_urg_9am
        grand_pct = ((tot_urg_9am - tot_urg_2pm) / tot_urg_9am * 100.0) if tot_urg_9am > 0 else 100.0
    else:
        grand_change = 0
        grand_pct = 0.0

    grand_change_str = f"{grand_change}" if grand_change == 0 else (f"+{grand_change}" if grand_change > 0 else f"{grand_change}")

    totals = {
        "ZONE": "ALL ZONES",
        "POST OFFICE HANDLE": "TOTAL",
        "URGENT_9AM": tot_urg_9am,
        "URGENT_2PM": tot_urg_2pm,
        "URGENT_5PM": tot_urg_5pm,
        "URGENT_CHANGE": grand_change_str,
        "CLEARANCE_PCT": grand_pct
    }

    return rows, totals, []

def build_compare_excel(date_str, rows, totals, itemized_transitions=None, out_filepath=None):
    if out_filepath is None:
        stamp = datetime.now().strftime("%Y%m%d_%H%M")
        out_filepath = os.path.join(COMPARE_DIR, f"Urgent_Clearance_Compare_{stamp}.xlsx")

    wb = openpyxl.Workbook()
    ws_sum = wb.active
    ws_sum.title = "Total Urgent Shift Compare"

    font_family = "Segoe UI"
    f_title = Font(name=font_family, size=13, bold=True, color="FFFFFF")
    f_hdr = Font(name=font_family, size=10, bold=True, color="FFFFFF")
    f_data = Font(name=font_family, size=9)
    f_total = Font(name=font_family, size=10, bold=True, color="0F172A")
    f_good = Font(name=font_family, size=9, bold=True, color="065F46")

    fill_title = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    fill_hdr = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
    fill_alt = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    fill_total = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")

    fill_s1 = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid") # Blue 9AM
    font_s1 = Font(name=font_family, size=9, bold=True, color="1E40AF")
    hdr_s1  = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")

    fill_s2 = PatternFill(start_color="FFEDD5", end_color="FFEDD5", fill_type="solid") # Orange 2PM
    font_s2 = Font(name=font_family, size=9, bold=True, color="9A3412")
    hdr_s2  = PatternFill(start_color="C2410C", end_color="C2410C", fill_type="solid")

    fill_s3 = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid") # Green 5PM
    font_s3 = Font(name=font_family, size=9, bold=True, color="065F46")
    hdr_s3  = PatternFill(start_color="047857", end_color="047857", fill_type="solid")

    thin = Side(border_style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws_sum.merge_cells("A1:G1")
    t_cell = ws_sum.cell(1, 1, f"TOTAL URGENT SHIFT COMPARISON REPORT — {date_str}")
    t_cell.font = f_title
    t_cell.fill = fill_title
    t_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.row_dimensions[1].height = 30

    headers = [
        "ZONE", "POST OFFICE HANDLE",
        "URGENT (9AM)", "URGENT (2PM)", "URGENT (5PM)",
        "URGENT CHANGE", "CLEAR %"
    ]
    ws_sum.row_dimensions[2].height = 24

    for c_i, h_text in enumerate(headers, 1):
        cell = ws_sum.cell(2, c_i, h_text)
        cell.font = f_hdr

        if c_i == 3:
            cell.fill = hdr_s1 # Blue 9AM
        elif c_i == 4:
            cell.fill = hdr_s2 # Orange 2PM
        elif c_i == 5:
            cell.fill = hdr_s3 # Green 5PM
        else:
            cell.fill = fill_hdr

        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    r_idx = 3
    for r in rows:
        row_fill = fill_alt if r_idx % 2 == 0 else None
        ws_sum.row_dimensions[r_idx].height = 19

        vals = [
            r["ZONE"], r["POST OFFICE HANDLE"],
            r["URGENT_9AM"], r["URGENT_2PM"], r["URGENT_5PM"],
            r["URGENT_CHANGE"], f"{r['CLEARANCE_PCT']:.1f}%"
        ]

        for c_i, val in enumerate(vals, 1):
            cell = ws_sum.cell(r_idx, c_i, val)
            cell.border = border

            if c_i == 3:
                cell.fill = fill_s1
                cell.font = font_s1
            elif c_i == 4:
                cell.fill = fill_s2
                cell.font = font_s2
            elif c_i == 5:
                cell.fill = fill_s3
                cell.font = font_s3
            elif c_i in (6, 7):
                cell.font = f_good
                if row_fill: cell.fill = row_fill
            else:
                cell.font = f_data
                if row_fill: cell.fill = row_fill

            cell.alignment = Alignment(horizontal="center" if c_i > 2 else "left", vertical="center")
        r_idx += 1

    ws_sum.row_dimensions[r_idx].height = 22
    tot_vals = [
        totals["ZONE"], totals["POST OFFICE HANDLE"],
        totals["URGENT_9AM"], totals["URGENT_2PM"], totals["URGENT_5PM"],
        totals["URGENT_CHANGE"], f"{totals['CLEARANCE_PCT']:.1f}%"
    ]

    for c_i, val in enumerate(tot_vals, 1):
        cell = ws_sum.cell(r_idx, c_i, val)
        cell.font = f_total
        cell.fill = fill_total
        cell.border = border
        cell.alignment = Alignment(horizontal="center" if c_i > 2 else "left", vertical="center")

    for col in ws_sum.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_sum.column_dimensions[col_letter].width = max(max_len + 4, 14)

    wb.save(out_filepath)
    return out_filepath
