# -*- coding: utf-8 -*-
"""
lag_report.py
Module for scanning and generating the API Lag / Status Mismatch Excel report.
Identifies bills where Web Tracking is already S410 (Delivered) / S520 (Returned),
but the Database / Report API is lagging behind in pending statuses.
"""

import os, time, json, requests, pandas as pd
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_REALTIME_DONE_STATUSES = {
    'S410', '410', 'S520', '520', 'S99', '99', 'S100', '100', 'S201', '201',
    'S500', '500', 'S510', '510'
}

def build_lag_report(src_xlsx: str, out_xlsx: str, api_cfg: dict) -> tuple[int, int, int]:
    """
    Scans pending bills from src_xlsx against real-time web tracking.
    Builds a styled Excel report of all lagging bills.
    Returns (total_lag_bills, delivered_count, returned_count).
    """
    df = pd.read_excel(src_xlsx)
    df.columns = [str(c).strip().upper() for c in df.columns]

    col_order = next((c for c in df.columns if 'ORDER ID' in c), 'ORDER ID')
    col_status = next((c for c in df.columns if 'CURRENT STATUS' in c), 'CURRENT STATUS')
    col_dest_po = next((c for c in df.columns if 'DELIVERY POST' in c), 'DELIVERY POST OFFICE')
    col_dest_prov = next((c for c in df.columns if 'DELIVERY PROVINCE' in c), 'DELIVERY PROVINCE')
    col_cur_po = next((c for c in df.columns if 'CURRENT POST' in c), 'CURRENT POST OFFICE')

    df['sc'] = df[col_status].astype(str).str.extract(r'^(\d{3})')[0]

    pending_mask = df['sc'].isin({'400', '401', '402', '420', '430', '440', '480', '306', '309', '310', '311', '210', '302', '110', '500', '512'})
    df_pending = df[pending_mask].copy()

    headers = {
        "Authorization": f"Bearer {api_cfg['bearer_token']}",
        "Referer": "https://opsexpress.metfone.com.kh/",
        "Accept-Language": "vi-VN",
        "Accept": "application/json, text/plain, */*",
        "Content-Type": "application/json",
        "x-client-id": "TMS_ANDROID",
        "User-Agent": "Mozilla/5.0 Chrome/148.0.0.0",
    }

    session = requests.Session()
    adapter = HTTPAdapter(pool_connections=60, pool_maxsize=60, max_retries=Retry(total=1, backoff_factor=0.1))
    session.mount("https://", adapter)
    session.headers.update(headers)

    TRACKING_URL = "https://gw-express.metfone.com.kh/tms-tracking/api/v1/order-tracking"

    def check_bill(row_data):
        oid = str(row_data[col_order]).strip()
        if not oid or oid == 'nan':
            return None
        try:
            r = session.get(TRACKING_URL, params={"order_id": oid}, timeout=5)
            if not r.ok:
                return None
            data = r.json()
            trips = data.get("trackingTrips", [])
            if not trips:
                return None
            
            t0 = trips[0]
            latest_status = str(t0.get("status", "")).upper().strip()
            status_name = str(t0.get("statusName", "")).upper().strip()
            item_type = str(t0.get("itemType", "")).upper().strip()
            desc = str(t0.get("desc", "")).strip()
            action_time = str(t0.get("updatedAt", "")).strip()

            is_done = (
                latest_status in _REALTIME_DONE_STATUSES
                or status_name in ('SHIPPED', 'DELIVERED', 'RETURNED')
                or item_type in ('SHIPPED', 'DELIVERED', 'RETURN_SUCCESS', 'CUSTOMER_RETURN')
                or (latest_status.startswith('S4') and 'giao thành công' in desc.lower() and 'bàn giao' not in desc.lower())
            )

            if not is_done:
                return None

            u_info = t0.get("postmanDelivery") or t0.get("updatedBy") or {}
            p_name = str(u_info.get("name", "")).strip()
            p_phone = str(u_info.get("phone", "")).strip()

            is_delivered = '410' in latest_status or 'SHIPPED' in status_name or 'giao thành công' in desc.lower()

            return {
                'order_id': oid,
                'export_status': str(row_data[col_status]).strip(),
                'live_status': latest_status,
                'live_status_name': status_name,
                'live_desc': desc,
                'action_time': action_time,
                'postman_name': p_name,
                'postman_phone': p_phone,
                'dest_po': str(row_data.get(col_dest_po, '')).strip(),
                'dest_prov': str(row_data.get(col_dest_prov, '')).strip(),
                'is_delivered': is_delivered,
                'web_link': f"https://opsexpress.metfone.com.kh/tra-cuu-hanh-trinh?orderCode={oid}",
            }
        except Exception:
            return None

    results = []
    with ThreadPoolExecutor(max_workers=60) as executor:
        futures = [executor.submit(check_bill, row) for _, row in df_pending.iterrows()]
        for f in as_completed(futures):
            res = f.result()
            if res:
                results.append(res)

    results.sort(key=lambda x: x['order_id'])

    delivered_count = sum(1 for r in results if r['is_delivered'])
    returned_count = len(results) - delivered_count

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "API Lag Bills"

    headers_list = [
        "No", "Order Code", "Web Tra Cứu Link (Clickable)", "Lagging Status in Database",
        "Real-Time Status on Web", "Tracking Description", "Completed Action Time",
        "Staff / Driver", "Phone Number", "Delivery Post Office", "Province"
    ]
    ws.append(headers_list)

    font_header = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    fill_header = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    font_data = Font(name="Segoe UI", size=9)
    font_link = Font(name="Segoe UI", size=9, color="2563EB", underline="single")
    border_thin = Border(
        left=Side(style='thin', color='CBD5E1'), right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'), bottom=Side(style='thin', color='CBD5E1')
    )

    fill_green = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    fill_red = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")

    for col_idx in range(1, len(headers_list) + 1):
        cell = ws.cell(1, col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 28

    for idx, r in enumerate(results, start=1):
        r_idx = idx + 1
        ws.row_dimensions[r_idx].height = 20

        ws.cell(r_idx, 1, idx)
        ws.cell(r_idx, 2, r['order_id'])
        
        link_cell = ws.cell(r_idx, 3, "View on Web 🌐")
        link_cell.hyperlink = r['web_link']
        link_cell.font = font_link

        ws.cell(r_idx, 4, r['export_status'])
        ws.cell(r_idx, 5, f"{r['live_status']} ({r['live_status_name']})")
        ws.cell(r_idx, 6, r['live_desc'])
        ws.cell(r_idx, 7, r['action_time'])
        ws.cell(r_idx, 8, r['postman_name'])
        ws.cell(r_idx, 9, r['postman_phone'])
        ws.cell(r_idx, 10, r['dest_po'])
        ws.cell(r_idx, 11, r['dest_prov'])

        row_fill = fill_green if r['is_delivered'] else fill_red

        for c_idx in range(1, len(headers_list) + 1):
            c = ws.cell(r_idx, c_idx)
            c.border = border_thin
            if c_idx != 3:
                c.font = font_data
            if c_idx == 5:
                c.fill = row_fill
                c.font = Font(name="Segoe UI", size=9, bold=True)
            if c_idx in (1, 2, 3, 5, 7, 9, 10, 11):
                c.alignment = Alignment(horizontal="center", vertical="center")
            else:
                c.alignment = Alignment(horizontal="left", vertical="center")

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 28
    ws.column_dimensions['E'].width = 22
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 24
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 16
    ws.column_dimensions['K'].width = 12

    wb.save(out_xlsx)
    return len(results), delivered_count, returned_count
