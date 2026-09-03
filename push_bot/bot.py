"""
bot.py — Push Bot with pause/resume/register/status commands.

Commands:
  push      — fetch data and send reports
  /test     — fetch data and send reports only to the requester
  /pause    — pause forwarding to groups (bot still works for you)
  /resume   — resume forwarding to groups
  /status   — show current state
  /mode     — toggle wide/long image mode
  /register — register current group to receive forwards
  /groups   — list registered groups
"""

import json
import logging
import os
import re
import sys
import tempfile
import asyncio
import io
from datetime import datetime, timedelta
import threading
import pandas as pd
from http.server import HTTPServer, BaseHTTPRequestHandler

from telegram import Update
from telegram.ext import (Application, MessageHandler, CommandHandler,
                          ContextTypes, filters)
from telegram.error import RetryAfter, NetworkError, Forbidden

class WebAppHandler(BaseHTTPRequestHandler):
    def do_GET(self):
        if self.path == '/' or self.path.startswith('/?'):
            self.send_response(200)
            self.send_header('Content-type', 'text/html; charset=utf-8')
            self.end_headers()
            try:
                # Read index.html from same directory
                with open(os.path.join(HERE, "index.html"), "rb") as f:
                    self.wfile.write(f.read())
            except Exception as e:
                self.wfile.write(f"Error loading index.html: {e}".encode('utf-8'))
        elif self.path == '/data':
            self.send_response(200)
            self.send_header('Content-type', 'application/json; charset=utf-8')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            data_path = os.path.join(HERE, "cache", "latest_data.json")
            if os.path.exists(data_path):
                with open(data_path, "rb") as f:
                    self.wfile.write(f.read())
            else:
                self.wfile.write(b"[]")
        else:
            self.send_response(404)
            self.end_headers()
            self.wfile.write(b"Not Found")

    def log_message(self, format, *args):
        # Suppress logging every single asset request to avoid cluttering the bot terminal logs
        pass

def start_webapp_server():
    try:
        server = HTTPServer(('0.0.0.0', 8080), WebAppHandler)
        log.info("Telegram WebApp Server listening on port 8080...")
        server.serve_forever()
    except Exception as e:
        log.error(f"Failed to start Telegram WebApp Server: {e}")

def update_webapp_cache(result):
    items = []
    try:
        type_data = result.get("type_data", {})
        for sheet_name, df in type_data.items():
            if df is None or df.empty:
                continue
            
            for _, row in df.iterrows():
                order_id = str(row.get("ORDER ID") or row.get("Order ID") or "").strip()
                if not order_id or order_id.lower() == "nan":
                    continue
                
                province = str(row.get("PROVINCE") or row.get("Province") or "").strip()
                district = str(row.get("DISTRICT") or row.get("District") or "").strip()
                store = str(row.get("POST OFFICE HANDLE") or row.get("Post Office") or row.get("CURRENT STATUS") or "").strip()
                address = str(row.get("DELIVERY ADDRESS") or row.get("Address") or row.get("RECEIVER ADDRESS") or "").strip()
                
                if province.lower() == "nan": province = ""
                if district.lower() == "nan": district = ""
                if store.lower() == "nan": store = ""
                if address.lower() == "nan": address = ""
                
                items.append({
                    "id": order_id,
                    "province": province,
                    "district": district,
                    "store": store,
                    "address": address
                })
        
        cache_dir = os.path.join(HERE, "cache")
        os.makedirs(cache_dir, exist_ok=True)
        cache_path = os.path.join(cache_dir, "latest_data.json")
        with open(cache_path, "w", encoding="utf-8") as f:
            json.dump(items, f, ensure_ascii=False, indent=2)
        log.info(f"Updated WebApp data cache with {len(items)} items.")
    except Exception as e:
        log.error(f"Failed to update WebApp data cache: {e}")

def save_highlight_history(result):
    try:
        today_date = datetime.now().date()
        today_str = today_date.strftime("%Y-%m-%d")
        timestamp_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        run_data = {"Pickup": {}, "Delivery": {}, "Pending": {}}
        type_data = result.get("type_data", {})
        filter_cols = {
            'Pickup': 'POST OFFICE HANDLE',
            'Delivery': 'POST OFFICE HANDLE',
            'Pending': 'POST OFFICE HANDLE'
        }

        for rn in ['Pickup', 'Delivery', 'Pending']:
            df = type_data.get(rn)
            if df is None or df.empty:
                continue

            fcol = filter_cols[rn]
            if fcol not in df.columns:
                continue

            for h, df_h in df.groupby(fcol):
                h_str = str(h).strip().upper()
                if not h_str:
                    continue
                
                oids = []
                for _, row in df_h.iterrows():
                    oid = str(row.get("ORDER ID") or "").strip()
                    if not oid or oid.lower() == "nan":
                        continue
                    
                    sc = str(row.get("STATUS_CODE") or "").strip()
                    cd_val = row.get("CREATED DATE")
                    
                    is_highlight = False
                    if sc in ('500', '520', '540'):
                        is_highlight = True
                    elif sc in ('420', '472'):
                        if cd_val:
                            cd = pd.to_datetime(cd_val, dayfirst=True, format="mixed", errors="coerce")
                            if not pd.isna(cd):
                                if (today_date - cd.date()).days > 7:
                                    is_highlight = True
                    else:
                        if cd_val:
                            cd = pd.to_datetime(cd_val, dayfirst=True, format="mixed", errors="coerce")
                            if not pd.isna(cd):
                                if (today_date - cd.date()).days > 1:
                                    is_highlight = True
                    
                    if is_highlight:
                        oids.append(oid)
                
                if oids:
                    run_data[rn][h_str] = oids

        history_path = os.path.join(HERE, "highlight_history.json")
        history = {}
        if os.path.exists(history_path):
            try:
                with open(history_path, encoding="utf-8") as f:
                    history = json.load(f)
            except Exception:
                pass

        # Keep last 7 days of history
        history = {k: v for k, v in history.items() if (datetime.now() - datetime.strptime(k, "%Y-%m-%d")).days < 7}

        if today_str not in history:
            history[today_str] = []

        history[today_str].append({
            "timestamp": timestamp_str,
            "data": run_data
        })

        with open(history_path, "w", encoding="utf-8") as f:
            json.dump(history, f, ensure_ascii=False, indent=2)

        log.info(f"Saved highlighted orders history to highlight_history.json at {timestamp_str}")
    except Exception as e:
        log.warning(f"Failed to save highlight history: {e}")

import warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

import downloader
import generate_report
import excel_to_image

HERE = os.path.dirname(os.path.abspath(__file__))
CONFIG_PATH        = os.path.join(HERE, "config.json")
REF_PATH           = os.path.join(HERE, "post_office_lookup.csv")
PICKUP_BRANCH_LOOKUP_PATH = os.path.join(HERE, "pickup_branch_lookup.csv")
REGISTERED_GROUPS_PATH = os.path.join(HERE, "registered_groups.json")
REPORTS_LOG_PATH   = os.path.join(HERE, "reports_today.json")

from logging.handlers import RotatingFileHandler
log_file_path = os.path.join(HERE, "bot.log")
file_handler = RotatingFileHandler(
    log_file_path,
    maxBytes=5 * 1024 * 1024,
    backupCount=3,
    encoding="utf-8"
)

logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO,
    handlers=[
        logging.StreamHandler(sys.stdout),
        file_handler
    ]
)
logging.getLogger("httpx").setLevel(logging.WARNING)
logging.getLogger("httpcore").setLevel(logging.WARNING)
log = logging.getLogger("push_bot")



# ── Today's report tracker (for /clean) ───────────────────────────────────────

def track_report_dir(tmpdir: str):
    """Record a tempdir so /clean can delete it later."""
    today_str = datetime.now().strftime("%Y-%m-%d")
    try:
        if os.path.exists(REPORTS_LOG_PATH):
            with open(REPORTS_LOG_PATH, encoding="utf-8") as f:
                data = json.load(f)
        else:
            data = {}
        # Prune old days (keep only today)
        data = {k: v for k, v in data.items() if k == today_str}
        if today_str not in data:
            data[today_str] = []
        if tmpdir not in data[today_str]:
            data[today_str].append(tmpdir)
        with open(REPORTS_LOG_PATH, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception as e:
        log.warning(f"track_report_dir failed: {e}")


from functools import wraps

class PrivateMessageRequired(Exception):
    """Raised when a command cannot reply privately because the user hasn't started the bot in PM."""
    pass

def pm_required_handler(func):
    @wraps(func)
    async def wrapper(update: Update, context: ContextTypes.DEFAULT_TYPE, *args, **kwargs):
        # ── User allowlist check ───────────────────────────────────────────────
        try:
            cfg = load_config()
            allowed_users = cfg["telegram"].get("allowed_user_ids") or []
            if allowed_users:
                user = update.effective_user
                if not user or user.id not in allowed_users:
                    log.info("Ignoring command from unauthorized user %s", user and user.id)
                    return
        except Exception:
            pass  # If config fails, allow through
        # ── Run handler ───────────────────────────────────────────────────────
        try:
            return await func(update, context, *args, **kwargs)
        except PrivateMessageRequired:
            log.info("Command aborted because user has not started PM with the bot.")
            return
    return wrapper


def user_guard(func):
    """Standalone decorator for commands NOT decorated with pm_required_handler.
    Silently ignores users not in the allowed_user_ids list."""
    @wraps(func)
    async def wrapper(update: Update, context: ContextTypes.DEFAULT_TYPE, *args, **kwargs):
        try:
            cfg = load_config()
            allowed_users = cfg["telegram"].get("allowed_user_ids") or []
            if allowed_users:
                user = update.effective_user
                if not user or user.id not in allowed_users:
                    log.info("Ignoring command from unauthorized user %s", user and user.id)
                    return
        except Exception:
            pass
        return await func(update, context, *args, **kwargs)
    return wrapper



# ── Config helpers ─────────────────────────────────────────────────────────────

def load_config():
    with open(CONFIG_PATH, encoding="utf-8") as f:
        return json.load(f)


def save_config(cfg):
    with open(CONFIG_PATH, "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)


def load_registered_groups():
    if not os.path.exists(REGISTERED_GROUPS_PATH):
        return []
    with open(REGISTERED_GROUPS_PATH, encoding="utf-8") as f:
        return json.load(f)


def save_registered_groups(groups):
    with open(REGISTERED_GROUPS_PATH, "w", encoding="utf-8") as f:
        json.dump(groups, f, ensure_ascii=False, indent=2)


_MIGRATION_PATTERN = re.compile(r"new chat id:\s*(-?\d+)", re.IGNORECASE)


def _extract_migrated_chat_id(error_msg: str):
    """Return new chat id from a 'Group migrated to supergroup' error, or None."""
    match = _MIGRATION_PATTERN.search(error_msg or "")
    if not match:
        return None
    try:
        return int(match.group(1))
    except (TypeError, ValueError):
        return None


def _replace_chat_id_in_args(args, kwargs, new_chat_id):
    """Return (args, kwargs) with chat_id replaced by new_chat_id."""
    new_args = tuple(new_chat_id if isinstance(a, int) else a for a in args)
    new_kwargs = dict(kwargs)
    if "chat_id" in new_kwargs:
        new_kwargs["chat_id"] = new_chat_id
    return new_args, new_kwargs


def _reset_io_buffers(args, kwargs):
    """Seek any file-like arguments back to position 0 before retrying.

    Telegram reads uploaded files during the request, which moves the
    BytesIO cursor to the end. Without resetting, retries would send
    empty data and trigger "File must be non-empty" errors.
    """
    candidates = list(args) + list(kwargs.values())

    def _maybe_reset(obj):
        if isinstance(obj, io.IOBase):
            try:
                obj.seek(0)
            except Exception:
                pass

    for obj in candidates:
        _maybe_reset(obj)


def _update_migrated_chat_id(old_chat_id, new_chat_id):
    """Persist chat_id migration to registered_groups.json and config.json."""
    old_str = str(old_chat_id)
    new_str = str(new_chat_id)

    # 1. registered_groups.json — replace old entry with new entry (dedup)
    try:
        groups = load_registered_groups()
        old_entry = None
        for g in groups:
            if str(g.get("chat_id")) == old_str:
                old_entry = dict(g)
                break
        if old_entry is not None:
            groups = [g for g in groups if str(g.get("chat_id")) != old_str]
            already_has_new = any(str(g.get("chat_id")) == new_str for g in groups)
            if not already_has_new:
                old_entry["chat_id"] = new_chat_id
                groups.append(old_entry)
            save_registered_groups(groups)
            log.info("Updated registered_groups.json: %s -> %s", old_str, new_str)
    except Exception as e:
        log.warning("Could not update registered_groups.json: %s", e)

    # 2. config.json — forward_mapping + zone_forward_mapping
    try:
        cfg = load_config()
        changed = False
        fwd = cfg.get("telegram", {}).get("forward_mapping", {})
        if old_str in fwd:
            if new_str not in fwd:
                fwd[new_str] = fwd[old_str]
            del fwd[old_str]
            cfg["telegram"]["forward_mapping"] = fwd
            changed = True
        zfwd = cfg.get("zone_forward_mapping", {})
        if old_str in zfwd:
            if new_str not in zfwd:
                zfwd[new_str] = zfwd[old_str]
            del zfwd[old_str]
            cfg["zone_forward_mapping"] = zfwd
            changed = True
        if changed:
            save_config(cfg)
            log.info("Updated config.json: %s -> %s", old_str, new_str)
    except Exception as e:
        log.warning("Could not update config.json: %s", e)


async def safe_api_call(func, *args, **kwargs):
    max_retries = 5
    for attempt in range(max_retries):
        try:
            return await func(*args, **kwargs)
        except RetryAfter as e:
            wait_time = e.retry_after
            log.warning(f"Flood control exceeded. Waiting for {wait_time} seconds before retrying (attempt {attempt + 1})...")
            await asyncio.sleep(wait_time + 1)
        except NetworkError as e:
            log.warning(f"Network error: {e}. Retrying in 3 seconds (attempt {attempt + 1})...")
            await asyncio.sleep(3)
            _reset_io_buffers(args, kwargs)
        except Forbidden as e:
            raise e
        except Exception as e:
            error_msg = str(e)
            new_chat_id = _extract_migrated_chat_id(error_msg)
            if new_chat_id is not None:
                old_chat_id = kwargs.get("chat_id")
                if old_chat_id is None:
                    for a in args:
                        if isinstance(a, int):
                            old_chat_id = a
                            break
                if old_chat_id is not None and new_chat_id != old_chat_id:
                    log.warning(
                        "Group %s migrated to supergroup %s. Updating storage and retrying...",
                        old_chat_id, new_chat_id,
                    )
                    _update_migrated_chat_id(old_chat_id, new_chat_id)
                    args, kwargs = _replace_chat_id_in_args(args, kwargs, new_chat_id)
                    _reset_io_buffers(args, kwargs)
                    # Don't count migration handling as a retry attempt
                    continue
            if attempt == max_retries - 1:
                raise e
            log.warning(f"Error calling Telegram API: {e}. Retrying in 2 seconds...")
            await asyncio.sleep(2)
            _reset_io_buffers(args, kwargs)


import re

def get_forward_mapping(cfg):
    # Groups configured manually in config.json
    mapping = dict(cfg["telegram"].get("forward_mapping", {}))
    
    # Auto-detect from registered group titles!
    for g in load_registered_groups():
        chat_id_str = str(g["chat_id"])
        title = g.get("title", "")
        # Look for handles like PNPP014, SVAP001, KANP001 in the title
        found_handles = re.findall(r'\b[A-Z]{3}P\d{3}\b', title.upper())
        
        # If no full handles found, try matching 3-letter branch codes (like SIH, KOH)
        if not found_handles:
            words = re.findall(r'\b[A-Z]{3}\b', title.upper())
            for w in words:
                known_prefixes = []
                for zb in cfg.get("zone_branches", {}).values():
                    known_prefixes.extend([p.strip().upper() for p in zb.split(",") if p.strip()])
                
                if w in known_prefixes:
                    found_handles.append(f"{w}P001")
        
        if found_handles:
            # If the group is already in mapping, don't overwrite manual settings
            if chat_id_str not in mapping:
                mapping[chat_id_str] = found_handles
                
    return mapping


def get_all_forward_groups(cfg):
    """Merge config groups + registered groups, excluding zone groups."""
    mapping = get_forward_mapping(cfg)
    configured  = list(mapping.keys())
    registered  = [str(g["chat_id"]) for g in load_registered_groups()]
    all_groups = list(dict.fromkeys(configured + registered))
    
    # Exclude zone groups from regular forwarding (they receive zone reports instead)
    zone_groups = set(str(k) for k in cfg.get("zone_forward_mapping", {}).keys())
    return [g for g in all_groups if g not in zone_groups]


def is_paused(cfg):
    return bool(cfg["telegram"].get("paused", False))


def get_mode(cfg):
    return cfg["telegram"].get("image_mode", "long")


def is_group_chat(update: Update):
    chat = update.effective_chat
    return bool(chat and chat.type in ("group", "supergroup"))


def is_user_allowed(update: Update, cfg: dict) -> bool:
    """Return True if the user is allowed to use the bot.
    If allowed_user_ids is empty, all users are allowed.
    """
    allowed = cfg["telegram"].get("allowed_user_ids") or []
    if not allowed:
        return True
    user = update.effective_user
    return bool(user and user.id in allowed)


def requester_chat_id(update: Update):
    if is_group_chat(update):
        user = update.effective_user
        return user.id if user else None

    chat = update.effective_chat
    return chat.id if chat else None


async def delete_group_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Best-effort removal of control commands typed in groups."""
    if not is_group_chat(update) or not update.message:
        return
    try:
        await context.bot.delete_message(
            chat_id=update.effective_chat.id,
            message_id=update.message.message_id,
        )
    except Exception as e:
        log.info("Could not delete group command message: %s", e)


async def private_or_current_reply(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
    text: str,
    parse_mode: str = None,
):
    """Send command feedback privately when invoked from a group."""
    msg = await send_requester_text(update, context, text, parse_mode=parse_mode)
    return bool(msg)


async def send_requester_text(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
    text: str,
    parse_mode: str = None,
):
    chat_id = requester_chat_id(update)
    if chat_id is None:
        log.warning("Cannot send requester text without a chat id.")
        return None

    try:
        # Try to send privately to user
        return await safe_api_call(
            context.bot.send_message, chat_id=chat_id, text=text, parse_mode=parse_mode
        )
    except Exception as e:
        log.warning("Could not send requester text privately to %s: %s", chat_id, e)
        # If in a group chat, fallback to sending directly in the group
        if is_group_chat(update) and update.effective_chat:
            try:
                return await safe_api_call(
                    context.bot.send_message,
                    chat_id=update.effective_chat.id,
                    text=text,
                    parse_mode=parse_mode,
                )
            except Exception as e2:
                log.warning("Failed to send fallback requester text to group: %s", e2)
        return None



async def edit_or_send_requester_text(
    message,
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
    text: str,
    parse_mode: str = None,
):
    if message:
        try:
            await safe_api_call(message.edit_text, text, parse_mode=parse_mode)
            return message
        except Exception as e:
            log.warning("Could not edit requester status message: %s", e)

    return await send_requester_text(update, context, text, parse_mode=parse_mode)


async def send_requester_photo(update: Update, context: ContextTypes.DEFAULT_TYPE, photo):
    chat_id = requester_chat_id(update)
    if chat_id is None:
        log.warning("Cannot send requester photo without a chat id.")
        return False

    try:
        await safe_api_call(context.bot.send_photo, chat_id=chat_id, photo=photo)
        return True
    except Exception as e:
        log.warning("Could not send requester photo to %s: %s", chat_id, e)
        if is_group_chat(update) and update.effective_chat:
            try:
                await safe_api_call(
                    context.bot.send_photo,
                    chat_id=update.effective_chat.id, 
                    photo=photo,
                )
                return True
            except Exception as e2:
                log.warning("Fallback send photo to group failed: %s", e2)
        return False


async def send_requester_media_group(update: Update, context: ContextTypes.DEFAULT_TYPE, media_list):
    chat_id = requester_chat_id(update)
    if chat_id is None:
        log.warning("Cannot send requester media group without a chat id.")
        return False

    try:
        await safe_api_call(context.bot.send_media_group, chat_id=chat_id, media=media_list)
        return True
    except Exception as e:
        log.warning("Could not send requester media group to %s: %s", chat_id, e)
        if is_group_chat(update) and update.effective_chat:
            try:
                await safe_api_call(
                    context.bot.send_media_group,
                    chat_id=update.effective_chat.id, 
                    media=media_list,
                )
                return True
            except Exception as e2:
                log.warning("Fallback send media group to group failed: %s", e2)
        return False


async def send_requester_document(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
    document,
    filename: str,
    caption: str = None,
):
    chat_id = requester_chat_id(update)
    if chat_id is None:
        log.warning("Cannot send requester document without a chat id.")
        return False

    try:
        await safe_api_call(
            context.bot.send_document,
            chat_id=chat_id,
            document=document,
            filename=filename,
            caption=caption,
        )
        return True
    except Exception as e:
        log.warning("Could not send requester document to %s: %s", chat_id, e)
        if is_group_chat(update) and update.effective_chat:
            try:
                document.seek(0)
                await safe_api_call(
                    context.bot.send_document,
                    chat_id=update.effective_chat.id,
                    document=document,
                    filename=filename,
                    caption=caption,
                )
                return True
            except Exception as e2:
                log.warning("Fallback send document to group failed: %s", e2)
        return False



async def forward_result_to_groups(context: ContextTypes.DEFAULT_TYPE, payload):
    result = payload["result"]
    forward_groups = payload["forward_groups"]
    forward_mapping = payload["forward_mapping"]
    sent_groups = 0

    for group_id_str in forward_groups:
        try:
            group_id = int(group_id_str)
        except ValueError:
            group_id = group_id_str  # fallback if it's a string like "@channel"

        allowed_handles = forward_mapping.get(group_id_str, ["*"])
        wants_all = "*" in allowed_handles
        sent_any = False

        for hr in result["handle_results"]:
            handle = hr["handle"]
            if not wants_all and handle not in allowed_handles:
                continue

            for hf in hr["handle_files"]:
                try:
                    img_buf = excel_to_image.excel_to_image(hf["path"])
                    img_buf.name = f"{handle}.png"
                    await safe_api_call(context.bot.send_photo, chat_id=group_id, photo=img_buf)
                    sent_any = True
                    await asyncio.sleep(0.5)
                except Exception as e:
                    log.error(f"Image to group {group_id}: {e}")

            # Auto-send Excel file if handle has 50+ pending rows
            pending_count = hr.get("handle_counts", {}).get("Pending", 0)
            threshold_count = pending_count if not wants_all else sum(hr.get("handle_counts", {}).values())
            if threshold_count > 50:
                for hf in hr["handle_files"]:
                    if not wants_all and "_Pending_" not in os.path.basename(hf["path"]):
                        continue
                    try:
                        with open(hf["path"], "rb") as ef:
                            await safe_api_call(
                                context.bot.send_document,
                                chat_id=group_id,
                                document=ef,
                                filename=os.path.basename(hf["path"]),
                            )
                            sent_any = True
                            await asyncio.sleep(0.5)
                    except Exception as e:
                        log.error(f"Excel file to group {group_id} for {handle}: {e}")

            try:
                await safe_api_call(context.bot.send_message, chat_id=group_id, text=hr["remark"])
                sent_any = True
                await asyncio.sleep(0.5)
            except Exception as e:
                log.error(f"Remark to group {group_id}: {e}")

        if wants_all:
            try:
                with open(result["final_xlsx"], "rb") as f:
                    await safe_api_call(
                        context.bot.send_document,
                        chat_id=group_id,
                        document=f,
                        filename=os.path.basename(result["final_xlsx"]),
                    )
                    sent_any = True
                    await asyncio.sleep(0.5)
            except Exception as e:
                log.error(f"Excel to group {group_id}: {e}")
            try:
                await safe_api_call(
                    context.bot.send_message,
                    chat_id=group_id,
                    text=result["summary_caption"],
                )
                sent_any = True
                await asyncio.sleep(0.5)
            except Exception as e:
                log.error(f"Summary to group {group_id}: {e}")

        if sent_any:
            sent_groups += 1
            await asyncio.sleep(1.0)

    return sent_groups


# ── Commands ───────────────────────────────────────────────────────────────────

@user_guard
async def cmd_pause(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Stop forwarding to groups. Bot still responds to push for the sender."""
    await delete_group_command(update, context)
    cfg = load_config()
    cfg["telegram"]["paused"] = True
    save_config(cfg)
    await private_or_current_reply(
        update,
        context,
        "⏸ Bot PAUSED.\n"
        "Forwarding to groups is disabled.\n"
        "You can still use 'push' to test — reports will only be sent to you.\n"
        "Use /resume to re-enable group forwarding."
    )


@user_guard
async def cmd_resume(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Re-enable forwarding to groups."""
    await delete_group_command(update, context)
    cfg = load_config()
    cfg["telegram"]["paused"] = False
    save_config(cfg)
    groups = get_all_forward_groups(cfg)
    await private_or_current_reply(
        update,
        context,
        f"▶️ Bot RESUMED.\n"
        f"Forwarding to {len(groups)} group(s) is enabled again."
    )


@user_guard
async def cmd_status(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Show current bot state."""
    await delete_group_command(update, context)
    cfg    = load_config()
    paused = is_paused(cfg)
    mode   = get_mode(cfg)
    groups = get_all_forward_groups(cfg)
    reg    = load_registered_groups()

    lines = [
        f"{'⏸ PAUSED' if paused else '▶️ ACTIVE'}",
        f"Image mode: {mode.upper()}",
        f"Forward groups: {len(groups)} total",
    ]
    if reg:
        lines.append("Registered groups:")
        for g in reg:
            lines.append(f"  • {g.get('title', '')} ({g['chat_id']})")
    else:
        lines.append("No registered groups (configure forward_groups in config.json)")

    await private_or_current_reply(update, context, "\n".join(lines))


@user_guard
async def cmd_statues(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/statues [code] - show status meanings, next flow, and app action mapping."""
    await delete_group_command(update, context)
    cfg = load_config()
    flows = cfg.get("ask_status_flow", {})
    actions = cfg.get("ask_app_actions", {})
    if not isinstance(flows, dict):
        flows = {}
    if not isinstance(actions, dict):
        actions = {}

    requested = [a.strip() for a in (context.args or []) if a.strip()]
    if requested:
        codes = [requested[0]]
    else:
        codes = sorted(
            {k for k in set(flows) | set(actions) if k != "default"},
            key=lambda v: int(v) if str(v).lstrip("-").isdigit() else 99999,
        )

    def one_status(code):
        flow = flows.get(code, {})
        action = actions.get(code, {})
        if not isinstance(flow, dict):
            flow = {}
        if not isinstance(action, dict):
            action = {}

        lines = [f"{code}"]
        meaning = flow.get("meaning")
        next_statuses = flow.get("next_statuses", [])
        if not isinstance(next_statuses, list):
            next_statuses = [next_statuses] if next_statuses else []
        next_statuses = [str(s) for s in next_statuses if str(s).strip()]

        if meaning:
            lines.append(f"Meaning: {meaning}")
        if next_statuses:
            lines.append(f"Next: {', '.join(next_statuses)}")
        elif flow:
            lines.append("Next: none")
        if flow.get("next_flow"):
            lines.append(f"Flow: {flow['next_flow']}")

        actor = action.get("actor")
        app = action.get("app")
        tab = action.get("tab")
        function = action.get("function")
        action_text = action.get("action")
        if actor:
            lines.append(f"Actor: {actor}")
        if app:
            lines.append(f"App: {app}")
        if tab:
            lines.append(f"Tab: {tab}")
        if function:
            lines.append(f"Function: {function}")
        if action_text:
            lines.append(f"Do: {action_text}")

        if len(lines) == 1:
            lines.append("No mapping found.")
        return "\n".join(lines)

    if not codes:
        await private_or_current_reply(update, context, "No status mappings configured.")
        return

    header = "Status details"
    if requested:
        header = f"Status details for {codes[0]}"
    chunks = []
    cur = [header]
    for code in codes:
        block = one_status(code)
        candidate = "\n\n".join(cur + [block])
        if len(candidate) > 3600 and len(cur) > 1:
            chunks.append("\n\n".join(cur))
            cur = [header + " (continued)", block]
        else:
            cur.append(block)
    if cur:
        chunks.append("\n\n".join(cur))

    for chunk in chunks:
        await private_or_current_reply(update, context, chunk)


@user_guard
async def cmd_test_mode(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await run_push(update, context, force_test=True)


@user_guard
async def cmd_clean(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Delete all report files generated today."""
    await delete_group_command(update, context)
    today_str = datetime.now().strftime("%Y-%m-%d")
    deleted_dirs = 0
    deleted_files = 0
    errors = []

    try:
        if not os.path.exists(REPORTS_LOG_PATH):
            await private_or_current_reply(update, context,
                "🗑 No reports found for today.")
            return

        with open(REPORTS_LOG_PATH, encoding="utf-8") as f:
            data = json.load(f)

        dirs_today = data.get(today_str, [])
        if not dirs_today:
            await private_or_current_reply(update, context,
                "🗑 No reports found for today.")
            return

        import shutil
        for d in dirs_today:
            if os.path.isdir(d):
                try:
                    # Count files before deleting
                    for root, _, files in os.walk(d):
                        deleted_files += len(files)
                    shutil.rmtree(d, ignore_errors=True)
                    deleted_dirs += 1
                except Exception as e:
                    errors.append(str(e))

        # Clear today's log
        data[today_str] = []
        with open(REPORTS_LOG_PATH, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

        msg_lines = [
            f"🗑 Cleaned today's reports ({today_str})",
            f"• Removed {deleted_dirs} report folder(s)",
            f"• Deleted {deleted_files} file(s)",
        ]
        if errors:
            msg_lines.append(f"⚠️ {len(errors)} error(s): {'; '.join(errors[:3])}")
        await private_or_current_reply(update, context, "\n".join(msg_lines))

    except Exception as e:
        log.exception("Error in cmd_clean")
        await private_or_current_reply(update, context, f"❌ Clean failed: {e}")



@user_guard
async def cmd_delete_report(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Delete a report message by replying to it with /deletereport."""
    user_id = update.effective_user.id if update.effective_user else "Unknown"
    chat_id = update.effective_chat.id if update.effective_chat else "Unknown"
    log.info("=== cmd_delete_report triggered by user %s in chat %s ===", user_id, chat_id)
    
    await delete_group_command(update, context)
    
    message = update.effective_message
    if not message:
        log.warning("cmd_delete_report: No effective message found in update.")
        return
        
    if not message.reply_to_message:
        log.info("cmd_delete_report: Message is not a reply to another message.")
        await private_or_current_reply(
            update, context, 
            "⚠️ Please reply to the report message you want to delete with `/deletereport`."
        )
        return

    target_msg = message.reply_to_message
    log.info("cmd_delete_report: Replying to message ID %s sent by user %s (is_bot: %s)", 
             target_msg.message_id, 
             target_msg.from_user.id if target_msg.from_user else "Unknown",
             target_msg.from_user.is_bot if target_msg.from_user else "Unknown")
    
    try:
        bot_user = await context.bot.get_me()
        log.info("cmd_delete_report: Bot user ID is %s", bot_user.id)
        if target_msg.from_user.id != bot_user.id:
            log.info("cmd_delete_report: Message was not sent by the bot (sender ID %s != bot ID %s). Ignoring delete request.",
                     target_msg.from_user.id, bot_user.id)
            await private_or_current_reply(
                update, context, 
                "⚠️ You can only delete messages sent by the bot."
            )
            return
    except Exception as e:
        log.warning("Could not verify bot identity: %s", e)

    try:
        log.info("cmd_delete_report: Attempting to delete message %s in chat %s", target_msg.message_id, chat_id)
        await context.bot.delete_message(
            chat_id=chat_id,
            message_id=target_msg.message_id
        )
        log.info("cmd_delete_report: Message %s deleted successfully.", target_msg.message_id)
    except Exception as e:
        log.error("Failed to delete report message: %s", e)
        await private_or_current_reply(
            update, context, 
            f"❌ Failed to delete message: {e}"
        )



@user_guard
async def cmd_mode(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Toggle wide/long image mode."""
    await delete_group_command(update, context)
    cfg = load_config()
    cur = get_mode(cfg)
    new = "long" if cur == "wide" else "wide"
    cfg["telegram"]["image_mode"] = new
    save_config(cfg)
    desc = "stacked vertically (tall)" if new == "long" else "side by side (wide)"
    await private_or_current_reply(update, context, f"Image mode → {new.upper()} ({desc})")


@user_guard
async def cmd_register(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Register the current group to receive report forwards."""
    await delete_group_command(update, context)
    chat = update.effective_chat
    if chat.type not in ("group", "supergroup"):
        await private_or_current_reply(
            update,
            context,
            "Use /register inside the Telegram group you want reports sent to."
        )
        return
    groups = load_registered_groups()
    if any(g["chat_id"] == chat.id for g in groups):
        await private_or_current_reply(
            update,
            context,
            f"This group is already registered.\nChat ID: {chat.id}"
        )
        return
    groups.append({"chat_id": chat.id, "title": chat.title or str(chat.id)})
    save_registered_groups(groups)
    await private_or_current_reply(
        update,
        context,
        f"✅ Group registered for report forwards.\n"
        f"Title: {chat.title}\n"
        f"Chat ID: {chat.id}\n\n"
        f"Use /pause to pause forwarding and /resume to re-enable."
    )


@user_guard
async def cmd_unregister(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Remove the current group from forwards."""
    await delete_group_command(update, context)
    chat = update.effective_chat
    groups = load_registered_groups()
    new_groups = [g for g in groups if g["chat_id"] != chat.id]
    if len(new_groups) == len(groups):
        await private_or_current_reply(update, context, "This group is not registered.")
        return
    save_registered_groups(new_groups)
    await private_or_current_reply(update, context, f"Group removed from forwards: {chat.title}")


@user_guard
async def cmd_help(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Show all available commands."""
    text = (
        "📋 *Bot Commands*\n"
        "━━━━━━━━━━━━━━━━━━━━━━\n"
        "\n"
        "📊 *Reports*\n"
        "`push` — Fetch data & send to 24 groups\n"
        "`push zone` — Fetch & send to 5 zone groups\n"
        "`push all` — Fetch & send to all groups + zones\n"
        "`push zone5` — Push to zone 5 only\n"
        "`/total` — Summary image + Excel (all data)\n"
        "`/total zone5` — Summary image + Excel (zone5 only)\n"
        "`/dailyreport` — Text daily report (today vs last week same time)\n"
        "`/deletereport` — Delete a report (reply to the bot's report message)\n"
        "\n"
        "📥 *Export*\n"
        "`/export KAM` — Export Kampot post office list\n"
        "`/export PNP` — Export Phnom Penh post office list\n"
        "`/export KAM,PNP` — Download multiple branches\n"
        "\n"
        "🔎 *Search & Explorer*\n"
        "`/app` — Open interactive Data Explorer app 📱\n"
        "`/find <phone/order>` — Search order tracking by phone number or ID\n"
        "`/ask <order_id>` — Show status, next steps & responsible scanning office\n"
        "`/statues [code]` — Show status flow/app action details\n"
        "`/qr <order_id>` — Generate a scanable QR code image for direct scanning\n"
        "\n"
        "🛠 *Test Bills*\n"
        "`/add <id>` — Ignore test bill(s) (e.g., `/add 12345`)\n"
        "`/remove <id>` — Stop ignoring test bill(s)\n"
        "`/list` — List all currently ignored test bills\n"
        "`/delay <id> <date/days>` — Delay a bill (ignore temporarily)\n"
        "`/undelay <id>` — Remove delay on a bill\n"
        "`/delaylist` — List all delayed bills and resume dates\n"
        "\n"
        "⚙️ *Control*\n"
        "`/pause` — Stop forwarding to groups (bot still works for YOU to test)\n"
        "`/resume` — Re-enable forwarding to groups\n"
        "`/status` — Show current state (paused/active, mode, groups)\n"
        "\n"
        "🖼 *Display*\n"
        "`/mode` — Toggle image layout: LONG (stacked) ↔ WIDE (side by side)\n"
        "\n"
        "👥 *Groups*\n"
        "`/register` — Register this group to receive report forwards (use inside the group)\n"
        "`/unregister` — Remove this group from forwards\n"
        "`/groups` — List all registered groups\n"
        "\n"
        "❓ *Help*\n"
        "`/help` — Show this message\n"
        "━━━━━━━━━━━━━━━━━━━━━━\n"
        "*Tip:* Use `/pause` before testing so groups don't get spammed."
    )
    await update.message.reply_text(text, parse_mode="Markdown")


@pm_required_handler
async def cmd_total(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/total [zone] — summary image + Excel sorted by report type.
    Examples: /total  (all data)  |  /total zone5  |  /total zone1 | /total mega
    """
    await delete_group_command(update, context)
    cfg = load_config()

    # Parse optional zone and force arguments
    args = [a.strip().lower() for a in (context.args or []) if a.strip()]
    force_refresh = "force" in args
    args = [a for a in args if a != "force"]

    zone_filter = None
    zone_label = "ALL"
    if args:
        zone_key = args[0]
        if zone_key == "mega":
            msg = await send_requester_text(update, context, "Fetching data for TỒN MEGA CHECK...")
            tmpdir = tempfile.mkdtemp(prefix="mega_")
            track_report_dir(tmpdir)
            stamp  = datetime.now().strftime("%d.%m_%HH%M")
            src    = os.path.join(tmpdir, f"export_{stamp}.xlsx")
            try:
                downloader.download_detail(cfg["api"], src, force_refresh=force_refresh)
                msg = await edit_or_send_requester_text(msg, update, context, "Building TỒN MEGA CHECK report...")
                import pivot

                # 1. Combined Excel (Zone + Mega tables)
                combined_xlsx = os.path.join(tmpdir, f"Report_MEGA_Combined_{stamp}.xlsx")
                pivot.run_mega_combined(src, combined_xlsx, cfg)
                img_combined = excel_to_image.excel_to_image(combined_xlsx)
                img_combined.name = "mega_zone_combined.png"
                await send_requester_photo(update, context, img_combined)

                # 2. Mega-only pivot Excel + image
                mega_xlsx = os.path.join(tmpdir, f"Report_MEGA_{stamp}.xlsx")
                _, grand_total = pivot.run_mega(src, mega_xlsx, cfg)
                img_buf = excel_to_image.excel_to_image(mega_xlsx)
                img_buf.name = "mega_check.png"
                await send_requester_photo(update, context, img_buf)

                caption = f"TỒN MEGA CHECK {datetime.now().strftime('%d/%m/%Y %H:%M')}\nGrand Total: {grand_total}"
                with open(combined_xlsx, "rb") as f:
                    await send_requester_document(
                        update,
                        context,
                        f,
                        os.path.basename(combined_xlsx),
                        caption=caption,
                    )

                # Build detail Excel with actual order data for MEGA/HUB/DVC
                try:
                    import mega_detail
                    detail_xlsx = os.path.join(tmpdir, f"MEGA_Detail_{stamp}.xlsx")
                    result_detail = mega_detail.build_mega_detail(src, detail_xlsx, cfg)
                    total_orders = result_detail[0] if result_detail else 0
                    urgent_orders = result_detail[1] if result_detail else 0
                    with open(detail_xlsx, "rb") as f:
                        await send_requester_document(
                            update, context, f,
                            os.path.basename(detail_xlsx),
                            caption=f"📋 ទិន្នន័យលម្អិត MEGA {datetime.now().strftime('%d/%m/%Y %H:%M')}\nTotal: {total_orders} | Urgent: {urgent_orders}",
                        )
                except Exception as e:
                    log.warning("Failed to build mega detail Excel: %s", e)

                await edit_or_send_requester_text(msg, update, context, f"Done. TỒN MEGA CHECK {datetime.now().strftime('%d.%m.%Y %H:%M')}")
            except Exception as e:
                log.exception("Error in /total mega")
                await edit_or_send_requester_text(msg, update, context, f"Error: {e}")
            return

        if zone_key == "penalty":
            target_label = " ".join(args[1:]) if len(args) > 1 else "ALL"
            msg = await send_requester_text(update, context, f"⏳ Generating INVENTORY PENALTY REPORT ({target_label.upper()})...")
            tmpdir = tempfile.mkdtemp(prefix="penalty_")
            track_report_dir(tmpdir)
            stamp  = datetime.now().strftime("%d.%m_%HH%M")
            src    = os.path.join(tmpdir, f"export_{stamp}.xlsx")
            try:
                downloader.download_detail(cfg["api"], src, force_refresh=force_refresh)
                import penalty_report
                out_xlsx = os.path.join(tmpdir, f"INVENTORY_PENALTY_REPORT_{stamp}_{target_label.replace(' ', '_')}.xlsx")
                tot_ho, tot_del, tot_pen_cnt, tot_fine = penalty_report.build_penalty_report(src, out_xlsx, target_label=target_label)
                
                # Render and send executive summary image
                try:
                    img_buf = penalty_report.render_penalty_summary_image(out_xlsx)
                    img_buf.name = f"penalty_summary_{stamp}.png"
                    await send_requester_photo(update, context, img_buf)
                except Exception as e:
                    log.warning("Failed to render penalty summary image: %s", e)

                with open(out_xlsx, "rb") as f:
                    await send_requester_document(
                        update, context, f,
                        os.path.basename(out_xlsx)
                    )
                await edit_or_send_requester_text(msg, update, context, f"✅ Done! Sent INVENTORY PENALTY REPORT ({target_label.upper()}).")
            except Exception as e:
                log.exception("Error in /total penalty")
                await edit_or_send_requester_text(msg, update, context, f"Error: {e}")
            return

        if zone_key == "speed":
            target_label = " ".join(args[1:]) if len(args) > 1 else "ALL"
            msg = await send_requester_text(update, context, f"⏳ Generating EXECUTIVE DELIVERY SPEED DASHBOARD ({target_label.upper()})...")
            tmpdir = tempfile.mkdtemp(prefix="speed_")
            track_report_dir(tmpdir)
            stamp  = datetime.now().strftime("%d.%m_%HH%M")
            src    = os.path.join(tmpdir, f"export_{stamp}.xlsx")
            try:
                downloader.download_detail(cfg["api"], src, force_refresh=force_refresh)
                import speed_report
                out_xlsx = os.path.join(tmpdir, f"DELIVERY_SPEED_REPORT_{stamp}_{target_label.replace(' ', '_')}.xlsx")
                tot_del, tot_u2, tot_24, tot_o8, tot_pay = speed_report.build_speed_report(src, out_xlsx, target_label=target_label)
                
                # Render and send executive summary image
                try:
                    img_buf = speed_report.render_speed_summary_image(out_xlsx)
                    img_buf.name = f"speed_summary_{stamp}.png"
                    await send_requester_photo(update, context, img_buf)
                except Exception as e:
                    log.warning("Failed to render speed summary image: %s", e)

                with open(out_xlsx, "rb") as f:
                    await send_requester_document(
                        update, context, f,
                        os.path.basename(out_xlsx)
                    )
                await edit_or_send_requester_text(msg, update, context, f"✅ Done! Sent EXECUTIVE DELIVERY SPEED DASHBOARD ({target_label.upper()}).")
            except Exception as e:
                log.exception("Error in /total speed")
                await edit_or_send_requester_text(msg, update, context, f"Error: {e}")
            return

        total_zones = cfg.get("total_zones", {})
        if zone_key in total_zones:
            zone_filter = [h.upper() for h in total_zones[zone_key]]
            zone_label = zone_key.upper()
        else:
            available = ", ".join(sorted(total_zones.keys())) if total_zones else "none configured"
            await private_or_current_reply(
                update,
                context,
                f"Unknown zone '{zone_key}'.\n"
                f"Available zones: {available}\n"
                f"Usage: /total zone5"
            )
            return

    msg = await send_requester_text(update, context, f"Fetching data for {zone_label} summary...")

    tmpdir = tempfile.mkdtemp(prefix="total_")
    track_report_dir(tmpdir)
    stamp  = datetime.now().strftime("%d.%m_%HH%M")
    src    = os.path.join(tmpdir, f"export_{stamp}.xlsx")

    try:
        downloader.download_detail(cfg["api"], src, force_refresh=force_refresh)
        msg = await edit_or_send_requester_text(msg, update, context, "Building summary...")

        mode = get_mode(cfg)
        result = generate_report.generate_reports_from_data(
            src, REF_PATH, tmpdir, return_metadata=True, mode=mode,
        )
        update_webapp_cache(result)
        save_highlight_history(result)

        # ── Zone filtering ────────────────────────────────────────────────
        if zone_filter:
            # Filter handle_results to only matching handles
            result["handle_results"] = [
                hr for hr in result["handle_results"]
                if hr["handle"] in zone_filter
            ]
            # Recalculate overall_counts from filtered handles
            overall = {"Pickup": 0, "Delivery": 0, "Pending": 0}
            for hr in result["handle_results"]:
                for k in overall:
                    overall[k] += hr["handle_counts"].get(k, 0)
            result["overall_counts"] = overall

            # Filter type_data DataFrames
            import pandas as pd
            for rn in ["Pickup", "Delivery", "Pending"]:
                df = result.get("type_data", {}).get(rn)
                if df is not None and not df.empty:
                    filter_col = "POST OFFICE HANDLE"
                    if filter_col in df.columns:
                        result["type_data"][rn] = df[df[filter_col].isin(zone_filter)].copy()

            # Re-fetch overall if zone_filter was processed
            overall = result["overall_counts"]

        # Calculate day_date_counts and urgent_counts for /total image
        total_day_date_counts = {}
        total_urgent_counts   = {}
        urgent_by_type        = {"Pickup": 0, "Delivery": 0, "Pending": 0}
        today_date = datetime.now().date()
        import pandas as pd

        for rn in ["Pickup", "Delivery", "Pending"]:
            df_z = result.get("type_data", {}).get(rn)
            if df_z is None or df_z.empty:
                continue
            date_col_z = result.get("date_col") or (
                "CREATED DATE" if "CREATED DATE" in df_z.columns else
                "CURRENT TIME"  if "CURRENT TIME"  in df_z.columns else None
            )
            if date_col_z and date_col_z in df_z.columns:
                parsed_z = pd.to_datetime(df_z[date_col_z], dayfirst=True,
                                          format="mixed", errors="coerce")
                df_z = df_z.copy()
                df_z["_zdate"] = parsed_z.dt.date

            handle_col = "POST OFFICE HANDLE"
            if handle_col not in df_z.columns:
                continue

            for _, row_z in df_z.iterrows():
                h = str(row_z.get(handle_col, "")).strip().upper()
                if not h:
                    continue
                d_val = row_z.get("_zdate") if "_zdate" in df_z.columns else None
                if d_val and not pd.isna(d_val):
                    total_day_date_counts.setdefault(h, {})
                    total_day_date_counts[h][d_val] = total_day_date_counts[h].get(d_val, 0) + 1
                created_d = None
                if "CREATED DATE" in df_z.columns:
                    cd = pd.to_datetime(row_z.get("CREATED DATE"), dayfirst=True,
                                        format="mixed", errors="coerce")
                    if not pd.isna(cd):
                        created_d = cd.date()
                if created_d and (today_date - created_d).days > 1:
                    if h not in total_urgent_counts:
                        total_urgent_counts[h] = {"Pickup": 0, "Delivery": 0, "Pending": 0}
                    total_urgent_counts[h][rn] = total_urgent_counts[h].get(rn, 0) + 1
                    urgent_by_type[rn] += 1

        overall = result["overall_counts"]
        grand_total = sum(overall.values())
        total_urgent_sum = sum(urgent_by_type.values())

        # Build final formatted caption
        result["summary_caption"] = "\n".join([
            f"📋 {zone_label} Report  {datetime.now().strftime('%d/%m/%Y %H:%M')}",
            f"Pickup: {overall['Pickup']} (Urgent: {urgent_by_type['Pickup']})  |  "
            f"Delivery: {overall['Delivery']} (Urgent: {urgent_by_type['Delivery']})  |  "
            f"Pending: {overall['Pending']} (Urgent: {urgent_by_type['Pending']})",
            f"Grand Total: {grand_total}  |  Total Urgent: {total_urgent_sum}",
        ])

        # 1. Summary image — totals per handle
        img_buf = generate_summary.build_summary_image(
            result["handle_results"],
            result["overall_counts"],
            zone_label=zone_label,
            day_date_counts=total_day_date_counts if total_day_date_counts else None,
            urgent_counts=total_urgent_counts if total_urgent_counts else None,
        )
        img_buf.name = "summary.png"
        await send_requester_photo(update, context, img_buf)

        # 2. Total Excel — 3 tables on one sheet (Pickup / Delivery / Pending)
        label = f"Total_{zone_label}_" if zone_filter else "Total_"
        total_xlsx = os.path.join(tmpdir, f"{label}{stamp}.xlsx")
        generate_summary.build_total_excel(result, total_xlsx)

        with open(total_xlsx, "rb") as f:
            await send_requester_document(
                update,
                context,
                f,
                os.path.basename(total_xlsx),
                caption=result["summary_caption"],
            )

        await edit_or_send_requester_text(msg, update, context, f"Done. {zone_label} {datetime.now().strftime('%d.%m.%Y %H:%M')}")

    except Exception as e:
        log.exception("Error in /total")
        await edit_or_send_requester_text(msg, update, context, f"Error: {e}")


@pm_required_handler
async def cmd_penalty(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Generates Stagnant Inventory & Handover Penalty Report and forwards to Zone and Branch groups."""
    await delete_group_command(update, context)
    cfg = load_config()

    args = [a.strip() for a in (context.args or []) if a.strip()]
    is_ytd = any(a.lower() in ("ytd", "yesterday", "homqua") for a in args)
    skip_zone = any(a.lower() in ("nozone", "nozones", "no_zone", "skipzone", "skipzones", "branch", "branches", "branch_only") for a in args)
    skip_branch = any(a.lower() in ("nobranch", "nobranches", "no_branch", "skipbranch", "skipbranches", "zone_only") for a in args)
    no_fwd = any(a.lower() in ("nofwd", "no_fwd", "onlyme", "me", "self", "quiet") for a in args)
    force_refresh = any(a.lower() in ("new", "refresh", "force") for a in args)

    ignore_tokens = {
        "ytd", "yesterday", "homqua", "new", "refresh", "force",
        "nozone", "nozones", "no_zone", "skipzone", "skipzones", "branch", "branches", "branch_only",
        "nobranch", "nobranches", "no_branch", "skipbranch", "skipbranches", "zone_only",
        "nofwd", "no_fwd", "onlyme", "me", "self", "quiet"
    }
    filtered_args = [a for a in args if a.lower() not in ignore_tokens]
    target_label = " ".join(filtered_args) if filtered_args else "ALL"
    target_date = (datetime.now().date() - timedelta(days=1)) if is_ytd else datetime.now().date()

    ytd_tag = f" — YESTERDAY {target_date.strftime('%d/%m/%Y')}" if is_ytd else ""
    msg = await send_requester_text(update, context, f"⏳ Generating INVENTORY PENALTY REPORT ({target_label.upper()}{ytd_tag})...")
    tmpdir = tempfile.mkdtemp(prefix="penalty_")
    track_report_dir(tmpdir)
    stamp = datetime.now().strftime("%d.%m_%HH%M")
    src = os.path.join(tmpdir, f"export_{stamp}.xlsx")

    try:
        downloader.download_detail(cfg["api"], src, force_refresh=force_refresh)
        import penalty_report
        safe_label = "".join(c for c in target_label if c.isalnum() or c in ("-", "_")).strip() or "ALL"
        suffix_file = "_YTD" if is_ytd else ""
        out_xlsx = os.path.join(tmpdir, f"INVENTORY_PENALTY_REPORT_{stamp}_{safe_label}{suffix_file}.xlsx")
        tot_ho, tot_del, tot_pen_cnt, tot_fine = penalty_report.build_penalty_report(src, out_xlsx, target_label=target_label, report_date=target_date)

        # 1. Render Summary Image & send to requester
        try:
            img_buf = penalty_report.render_penalty_summary_image(out_xlsx)
            img_buf.name = f"PENALTY_SUMMARY_{safe_label}{suffix_file}.png"
            await send_requester_photo(update, context, img_buf)
        except Exception as e_img:
            log.warning("Could not render penalty summary image: %s", e_img)

        # 2. Send Excel Document to requester
        with open(out_xlsx, "rb") as f:
            await send_requester_document(update, context, f, os.path.basename(out_xlsx))

        # 3. Group Forwarding (When target is ALL, MEGA, or TOTAL)
        tgt_upper = target_label.upper().replace(" ", "")
        total_sent_zones = 0
        total_sent_branches = 0

        if tgt_upper in ("ALL", "TOTAL", "MEGA") and not no_fwd:
            # A. Forward to 5 Zone Groups (Unless skip_zone)
            if not skip_zone:
                zone_fwd_map = cfg.get("zone_forward_mapping", {})
                for z_idx in range(1, 6):
                    z_name = f"Zone {z_idx}"
                    z_clean = f"zone{z_idx}"
                    z_xlsx = os.path.join(tmpdir, f"INVENTORY_PENALTY_REPORT_{stamp}_{z_name.replace(' ', '_')}{suffix_file}.xlsx")
                    z_ho, z_del, z_pen, z_fine = penalty_report.build_penalty_report(src, z_xlsx, target_label=z_name, report_date=target_date)
                    z_caption = (
                        f"📊 *INVENTORY & SLA PENALTY REPORT ({z_name}{ytd_tag})*\n"
                        f"Overdue Handover (> 4h): `{z_ho}`\n"
                        f"Overdue Delivery (> 10h): `{z_del}`\n"
                        f"Penalized Bills: `{z_pen}`\n"
                        f"Total Fine: `${z_fine:.2f}`"
                    )

                    for gid, zkey in zone_fwd_map.items():
                        if str(zkey).lower().strip() == z_clean:
                            try:
                                z_img = penalty_report.render_penalty_summary_image(z_xlsx)
                                z_img.name = f"PENALTY_SUMMARY_{z_name.replace(' ', '_')}{suffix_file}.png"
                                await safe_api_call(context.bot.send_photo, chat_id=int(gid), photo=z_img, caption=z_caption, parse_mode="Markdown")
                            except Exception as e_fwd_img:
                                log.warning("Failed forwarding penalty photo to zone group %s: %s", gid, e_fwd_img)

                            try:
                                with open(z_xlsx, "rb") as z_f_doc:
                                    await safe_api_call(
                                        context.bot.send_document,
                                        chat_id=int(gid),
                                        document=z_f_doc,
                                        filename=os.path.basename(z_xlsx)
                                    )
                                total_sent_zones += 1
                            except Exception as e_fwd_doc:
                                log.warning("Failed forwarding penalty doc to zone group %s: %s", gid, e_fwd_doc)

            # B. Forward to 36 Branch Groups in forward_mapping (Unless skip_branch)
            if not skip_branch:
                fwd_map = get_forward_mapping(cfg)
                for gid, handles in fwd_map.items():
                    if not handles or "*" in handles:
                        continue
                    br_code = handles[0].upper()
                    if br_code not in penalty_report.MAIN_36_BRANCHES:
                        continue

                    br_xlsx = os.path.join(tmpdir, f"INVENTORY_PENALTY_REPORT_{stamp}_{br_code}{suffix_file}.xlsx")
                    try:
                        b_ho, b_del, b_pen, b_fine = penalty_report.build_penalty_report(src, br_xlsx, target_label=br_code, report_date=target_date)
                        b_caption = (
                            f"📊 *INVENTORY & SLA PENALTY REPORT ({br_code}{ytd_tag})*\n"
                            f"Overdue Handover (> 4h): `{b_ho}`\n"
                            f"Overdue Delivery (> 10h): `{b_del}`\n"
                            f"Penalized Bills: `{b_pen}`\n"
                            f"Total Fine: `${b_fine:.2f}`"
                        )

                        try:
                            b_img = penalty_report.render_penalty_summary_image(br_xlsx)
                            b_img.name = f"PENALTY_SUMMARY_{br_code}{suffix_file}.png"
                            await safe_api_call(context.bot.send_photo, chat_id=int(gid), photo=b_img, caption=b_caption, parse_mode="Markdown")
                        except Exception as e_b_img:
                            log.warning("Failed sending penalty photo to branch group %s: %s", br_code, e_b_img)

                        try:
                            with open(br_xlsx, "rb") as b_f_doc:
                                await safe_api_call(
                                    context.bot.send_document,
                                    chat_id=int(gid),
                                    document=b_f_doc,
                                    filename=os.path.basename(br_xlsx)
                                )
                            total_sent_branches += 1
                        except Exception as e_b_doc:
                            log.warning("Failed sending penalty doc to branch group %s: %s", br_code, e_b_doc)
                    except Exception as e_br_gen:
                        log.warning("Failed building penalty report for branch %s: %s", br_code, e_br_gen)

        fwd_status_msg = ""
        if total_sent_zones > 0 or total_sent_branches > 0:
            fwd_status_msg = f" (Forwarded to {total_sent_zones} Zone groups & {total_sent_branches} Branch groups)"
        elif no_fwd or skip_branch and skip_zone:
            fwd_status_msg = " [Private / No-Forward Mode]"

        await edit_or_send_requester_text(msg, update, context, f"✅ Done! Sent INVENTORY PENALTY REPORT ({target_label.upper()}){fwd_status_msg}.")
    except Exception as e:
        log.exception("Error in /penalty command: %s", e)
        await edit_or_send_requester_text(msg, update, context, f"❌ Error generating penalty report: {e}")


@pm_required_handler
async def cmd_speed(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Generates Delivery Speed SLA Report."""
    await delete_group_command(update, context)
    cfg = load_config()

    args = [a.strip() for a in (context.args or []) if a.strip()]
    skip_zone = any(a.lower() in ("nozone", "nozones", "no_zone", "skipzone", "skipzones", "branch", "branches", "branch_only") for a in args)
    skip_branch = any(a.lower() in ("nobranch", "nobranches", "no_branch", "skipbranch", "skipbranches", "zone_only") for a in args)
    no_fwd = any(a.lower() in ("nofwd", "no_fwd", "onlyme", "me", "self", "quiet") for a in args)
    force_refresh = any(a.lower() in ("new", "refresh", "force") for a in args)

    ignore_tokens = {
        "new", "refresh", "force",
        "nozone", "nozones", "no_zone", "skipzone", "skipzones", "branch", "branches", "branch_only",
        "nobranch", "nobranches", "no_branch", "skipbranch", "skipbranches", "zone_only",
        "nofwd", "no_fwd", "onlyme", "me", "self", "quiet"
    }
    filtered_args = [a for a in args if a.lower() not in ignore_tokens]
    target_label = " ".join(filtered_args) if filtered_args else "ALL"

    msg = await send_requester_text(update, context, f"⏳ [1/3] Downloading TMS data ({target_label.upper()})...")
    tmpdir = tempfile.mkdtemp(prefix="speed_")
    track_report_dir(tmpdir)
    stamp = datetime.now().strftime("%d.%m_%HH%M")
    src = os.path.join(tmpdir, f"export_{stamp}.xlsx")

    try:
        await asyncio.to_thread(downloader.download_detail, cfg["api"], src, force_refresh=force_refresh)
        
        await edit_or_send_requester_text(msg, update, context, f"⏳ [2/3] Processing VTT speed SLA metrics ({target_label.upper()})...")
        cache_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "cache")
        rev_file = os.path.join(cache_dir, "latest_revenue.xlsx")
        try:
            await asyncio.to_thread(downloader.download_revenue_detail, cfg["api"], rev_file, force_refresh=force_refresh)
        except Exception as e_rev:
            log.warning("Could not refresh revenue detail for /speed: %s", e_rev)

        import speed_report
        safe_label = "".join(c for c in target_label if c.isalnum() or c in ("-", "_")).strip() or "ALL"
        out_xlsx = os.path.join(tmpdir, f"DELIVERY_SPEED_REPORT_{stamp}_{safe_label}.xlsx")
        tot_del, tot_u2, tot_24, tot_o8, tot_pay = await asyncio.to_thread(
            speed_report.build_speed_report, src, out_xlsx, target_label=target_label
        )

        # 1. Render Summary Image
        await edit_or_send_requester_text(msg, update, context, f"⏳ [3/3] Rendering clean dashboard image ({target_label.upper()})...")
        try:
            img_buf = await asyncio.to_thread(speed_report.render_speed_summary_image, out_xlsx)
            img_buf.name = f"speed_summary_{stamp}.png"
            await send_requester_photo(update, context, img_buf)
        except Exception as e_img:
            log.warning("Could not render speed summary image: %s", e_img)

        # 2. Send Excel Document
        with open(out_xlsx, "rb") as f:
            await send_requester_document(update, context, f, os.path.basename(out_xlsx))

        # 3. Group Forwarding (When target is ALL, MEGA, or TOTAL)
        tgt_upper = target_label.upper().replace(" ", "")
        total_sent_zones = 0
        total_sent_branches = 0

        if tgt_upper in ("ALL", "TOTAL", "MEGA") and not no_fwd:
            # A. Forward to 5 Zone Groups (Unless skip_zone)
            if not skip_zone:
                zone_fwd_map = cfg.get("zone_forward_mapping", {})
                for z_idx in range(1, 6):
                    z_name = f"Zone {z_idx}"
                    z_clean = f"zone{z_idx}"
                    z_xlsx = os.path.join(tmpdir, f"DELIVERY_SPEED_REPORT_{stamp}_{z_name.replace(' ', '_')}.xlsx")
                    try:
                        z_del, z_u2, z_24, z_o8, z_pay = await asyncio.to_thread(
                            speed_report.build_speed_report, src, z_xlsx, target_label=z_name
                        )
                        z_caption = (
                            f"⚡ *DELIVERY SPEED SLA REPORT ({z_name})*\n"
                            f"Total Delivered (410): `{z_del}`\n"
                            f"< 2 Hours (+50%): `{z_u2}`\n"
                            f"2 - 4 Hours (+25%): `{z_24}`\n"
                            f"> 8 Hours (-25%): `{z_o8}`\n"
                            f"Total Commission: `${z_pay:.2f}`"
                        )

                        for gid, zkey in zone_fwd_map.items():
                            if str(zkey).lower().strip() == z_clean:
                                try:
                                    z_img = await asyncio.to_thread(speed_report.render_speed_summary_image, z_xlsx)
                                    z_img.name = f"SPEED_SUMMARY_{z_name.replace(' ', '_')}.png"
                                    await safe_api_call(context.bot.send_photo, chat_id=int(gid), photo=z_img, caption=z_caption, parse_mode="Markdown")
                                except Exception as e_fwd_img:
                                    log.warning("Failed forwarding speed photo to zone group %s: %s", gid, e_fwd_img)

                                try:
                                    with open(z_xlsx, "rb") as z_f_doc:
                                        await safe_api_call(
                                            context.bot.send_document,
                                            chat_id=int(gid),
                                            document=z_f_doc,
                                            filename=os.path.basename(z_xlsx)
                                        )
                                    total_sent_zones += 1
                                except Exception as e_fwd_doc:
                                    log.warning("Failed forwarding speed doc to zone group %s: %s", gid, e_fwd_doc)
                    except Exception as e_z_gen:
                        log.warning("Failed building speed report for zone %s: %s", z_name, e_z_gen)

            # B. Forward to 36 Branch Groups in forward_mapping (Unless skip_branch)
            if not skip_branch:
                fwd_map = get_forward_mapping(cfg)
                for gid, handles in fwd_map.items():
                    if not handles or "*" in handles:
                        continue
                    br_code = handles[0].upper()
                    if br_code not in speed_report.MAIN_36_BRANCHES:
                        continue

                    br_xlsx = os.path.join(tmpdir, f"DELIVERY_SPEED_REPORT_{stamp}_{br_code}.xlsx")
                    try:
                        b_del, b_u2, b_24, b_o8, b_pay = await asyncio.to_thread(
                            speed_report.build_speed_report, src, br_xlsx, target_label=br_code
                        )
                        b_caption = (
                            f"⚡ *DELIVERY SPEED SLA REPORT ({br_code})*\n"
                            f"Total Delivered (410): `{b_del}`\n"
                            f"< 2 Hours (+50%): `{b_u2}`\n"
                            f"2 - 4 Hours (+25%): `{b_24}`\n"
                            f"> 8 Hours (-25%): `{b_o8}`\n"
                            f"Total Commission: `${b_pay:.2f}`"
                        )

                        try:
                            b_img = await asyncio.to_thread(speed_report.render_speed_summary_image, br_xlsx)
                            b_img.name = f"SPEED_SUMMARY_{br_code}.png"
                            await safe_api_call(context.bot.send_photo, chat_id=int(gid), photo=b_img, caption=b_caption, parse_mode="Markdown")
                        except Exception as e_b_img:
                            log.warning("Failed sending speed photo to branch group %s: %s", br_code, e_b_img)

                        try:
                            with open(br_xlsx, "rb") as b_f_doc:
                                await safe_api_call(
                                    context.bot.send_document,
                                    chat_id=int(gid),
                                    document=b_f_doc,
                                    filename=os.path.basename(br_xlsx)
                                )
                            total_sent_branches += 1
                        except Exception as e_b_doc:
                            log.warning("Failed sending speed doc to branch group %s: %s", br_code, e_b_doc)
                    except Exception as e_br_gen:
                        log.warning("Failed building speed report for branch %s: %s", br_code, e_br_gen)

        fwd_status_msg = ""
        if total_sent_zones > 0 or total_sent_branches > 0:
            fwd_status_msg = f" (Forwarded to {total_sent_zones} Zone groups & {total_sent_branches} Branch groups)"
        elif no_fwd or (skip_branch and skip_zone):
            fwd_status_msg = " [Private / No-Forward Mode]"

        await edit_or_send_requester_text(msg, update, context, f"✅ Done! Sent EXECUTIVE DELIVERY SPEED DASHBOARD ({target_label.upper()}){fwd_status_msg}.")
    except Exception as e:
        log.exception("Error in /speed command: %s", e)
        await edit_or_send_requester_text(msg, update, context, f"❌ Error generating speed report: {e}")


def get_highlighted_order_ids(df_t, today_date):
    """Return a set of order IDs that are highlighted in this DataFrame."""
    highlighted = set()
    if df_t.empty:
        return highlighted

    for _, row in df_t.iterrows():
        oid = str(row.get("ORDER ID") or "").strip()
        if not oid or oid.lower() == "nan":
            continue
        
        sc = str(row.get("STATUS_CODE") or "").strip()
        cd_val = row.get("CREATED DATE")
        
        is_highlight = False
        if sc in ('500', '520', '540'):
            is_highlight = True
        elif sc in ('420', '472'):
            if cd_val:
                cd = pd.to_datetime(cd_val, dayfirst=True, format="mixed", errors="coerce")
                if not pd.isna(cd):
                    if (today_date - cd.date()).days > 7:
                        is_highlight = True
        else:
            if cd_val:
                cd = pd.to_datetime(cd_val, dayfirst=True, format="mixed", errors="coerce")
                if not pd.isna(cd):
                    if (today_date - cd.date()).days > 1:
                        is_highlight = True
        
        if is_highlight:
            highlighted.add(oid)
            
    return highlighted


async def run_time_vs(update: Update, context: ContextTypes.DEFAULT_TYPE, start_hour: int, end_hour: int, command_label: str):
    await delete_group_command(update, context)

    cfg = load_config()
    allowed = cfg["telegram"].get("allowed_chat_ids") or []
    chat_id = update.effective_chat.id
    if allowed and chat_id not in allowed:
        await private_or_current_reply(update, context, "This chat is not allowed to use the bot.")
        return

    # Determine which handles to show based on arguments or chat context
    args = [a.strip().upper() for a in (context.args or []) if a.strip()]
    target_handles = None
    title_label = ""

    if args:
        arg = args[0]
        total_zones = cfg.get("total_zones", {})
        if arg.lower() in total_zones:
            target_handles = [h.upper() for h in total_zones[arg.lower()]]
            title_label = f"{arg} "
        else:
            target_handles = [arg]
            title_label = f"{arg} "
    else:
        # Check if in a registered group
        forward_mapping = get_forward_mapping(cfg)
        group_id_str = str(chat_id)
        if group_id_str in forward_mapping:
            group_handles = forward_mapping[group_id_str]
            if "*" not in group_handles:
                target_handles = [h.upper() for h in group_handles if h]
                title_label = f"{', '.join(target_handles)} "

        # Check if in a zone group
        zone_fwd_map = cfg.get("zone_forward_mapping", {})
        if group_id_str in zone_fwd_map:
            zone_key = zone_fwd_map[group_id_str]
            total_zones = cfg.get("total_zones", {})
            if zone_key in total_zones:
                target_handles = [h.upper() for h in total_zones[zone_key]]
                title_label = f"{zone_key.upper()} "

    # ── Try reading from JSON history first ─────────────────────────────────────
    history_path = os.path.join(HERE, "highlight_history.json")
    if os.path.exists(history_path):
        try:
            with open(history_path, encoding="utf-8") as f:
                history = json.load(f)

            today_str = datetime.now().strftime("%Y-%m-%d")
            runs = history.get(today_str, [])
            if runs:
                run_start = None
                run_end = None
                diff_start = None
                diff_end = None

                for r in runs:
                    dt = datetime.strptime(r["timestamp"], "%Y-%m-%d %H:%M:%S")
                    target_start = datetime.combine(dt.date(), datetime.min.time().replace(hour=start_hour))
                    target_end = datetime.combine(dt.date(), datetime.min.time().replace(hour=end_hour))

                    # start matching
                    is_valid_start = False
                    if start_hour == 8:
                        is_valid_start = (dt.hour < 12)
                    elif start_hour == 14:
                        is_valid_start = (dt.hour >= 12 and dt.hour < 16)

                    if is_valid_start:
                        d_start = abs((dt - target_start).total_seconds())
                        if diff_start is None or d_start < diff_start:
                            diff_start = d_start
                            run_start = r

                    # end matching
                    is_valid_end = False
                    if end_hour == 14:
                        is_valid_end = (dt.hour >= 12)
                    elif end_hour == 17:
                        is_valid_end = (dt.hour >= 16)

                    if is_valid_end:
                        d_end = abs((dt - target_end).total_seconds())
                        if diff_end is None or d_end < diff_end:
                            diff_end = d_end
                            run_end = r

                # Fallbacks
                if not run_start:
                    run_start = runs[0]
                if not run_end:
                    latest = runs[-1]
                    if latest["timestamp"] != run_start["timestamp"]:
                        run_end = latest

                if run_start and run_end and run_start["timestamp"] != run_end["timestamp"]:
                    dt_start = datetime.strptime(run_start["timestamp"], "%Y-%m-%d %H:%M:%S")
                    dt_end = datetime.strptime(run_end["timestamp"], "%Y-%m-%d %H:%M:%S")

                    t_start = dt_start.strftime('%I:%M %p')
                    t_end = dt_end.strftime('%I:%M %p')

                    all_h = set()
                    map_1 = run_start["data"]
                    map_2 = run_end["data"]

                    for rn in ['Pickup', 'Delivery', 'Pending']:
                        all_h.update(map_1.get(rn, {}).keys())
                        all_h.update(map_2.get(rn, {}).keys())

                    all_h = sorted(list(h for h in all_h if str(h).strip()))
                    if target_handles:
                        all_h = [h for h in all_h if h in target_handles]

                    if all_h:
                        text_lines = [
                            f"📊 {title_label}VS REPORT ({command_label}) — {t_start} vs {t_end}",
                            "=============================="
                        ]

                        grand_sets_1 = {"Pickup": set(), "Delivery": set(), "Pending": set()}
                        grand_sets_2 = {"Pickup": set(), "Delivery": set(), "Pending": set()}

                        for h in all_h:
                            h_lines = []
                            has_any_data = False
                            
                            for rn in ['Pickup', 'Delivery', 'Pending']:
                                set_1 = set(map_1.get(rn, {}).get(h, []))
                                set_2 = set(map_2.get(rn, {}).get(h, []))
                                
                                grand_sets_1[rn].update(set_1)
                                grand_sets_2[rn].update(set_2)
                                
                                n1 = len(set_1)
                                n2 = len(set_2)
                                cleared = len(set_1 - set_2)
                                
                                if n1 > 0 or n2 > 0:
                                    has_any_data = True
                                    h_lines.append(f"  • {rn}: {n1} vs {n2} (Cleared {cleared})")

                            if has_any_data and len(all_h) <= 10:
                                text_lines.append(f"\n🏢 {h}:")
                                text_lines.extend(h_lines)

                        # Grand Summary
                        text_lines.append("\n==============================")
                        text_lines.append("📈 GRAND SUMMARY:")
                        g_1 = 0
                        g_2 = 0
                        g_clear = 0
                        for rn in ['Pickup', 'Delivery', 'Pending']:
                            s1 = grand_sets_1[rn]
                            s2 = grand_sets_2[rn]
                            n1 = len(s1)
                            n2 = len(s2)
                            cleared = len(s1 - s2)
                            
                            g_1 += n1
                            g_2 += n2
                            g_clear += cleared
                            
                            text_lines.append(f"  • {rn}: {n1} vs {n2} (Cleared {cleared})")

                        text_lines.append(f"  • Total: {g_1} vs {g_2} (Cleared {g_clear})")

                        await send_requester_text(update, context, "\n".join(text_lines))
                        return
        except Exception as e:
            log.warning(f"Error checking highlight history JSON: {e}")

    if not os.path.exists(REPORTS_LOG_PATH):
        await private_or_current_reply(update, context, "No reports run today yet.")
        return

    try:
        with open(REPORTS_LOG_PATH, encoding="utf-8") as f:
            log_data = json.load(f)
    except Exception as e:
        await private_or_current_reply(update, context, f"Error reading reports log: {e}")
        return

    today_str = datetime.now().strftime("%Y-%m-%d")
    dirs = log_data.get(today_str, [])
    if not dirs:
        await private_or_current_reply(update, context, "No reports run today yet.")
        return

    # Find all export files today
    exports = []
    for d in dirs:
        if os.path.exists(d):
            for f in os.listdir(d):
                if f.startswith("export_") and f.endswith(".xlsx"):
                    fpath = os.path.join(d, f)
                    mtime = os.path.getmtime(fpath)
                    mtime_dt = datetime.fromtimestamp(mtime)
                    exports.append((mtime_dt, fpath))

    if not exports:
        await private_or_current_reply(update, context, "No raw exports found for today.")
        return

    exports = sorted(exports, key=lambda x: x[0])
    
    file_start = None
    file_end = None
    diff_start = None
    diff_end = None

    today_dt = exports[0][0].date()
    target_start = datetime.combine(today_dt, datetime.min.time().replace(hour=start_hour))
    target_end = datetime.combine(today_dt, datetime.min.time().replace(hour=end_hour))

    for dt, path in exports:
        # start file matching
        is_valid_start = False
        if start_hour == 8:
            is_valid_start = (dt.hour < 12)
        elif start_hour == 14:
            is_valid_start = (dt.hour >= 12 and dt.hour < 16)

        if is_valid_start:
            d_start = abs((dt - target_start).total_seconds())
            if diff_start is None or d_start < diff_start:
                diff_start = d_start
                file_start = (dt, path)

        # end file matching
        is_valid_end = False
        if end_hour == 14:
            is_valid_end = (dt.hour >= 12)
        elif end_hour == 17:
            is_valid_end = (dt.hour >= 16)

        if is_valid_end:
            d_end = abs((dt - target_end).total_seconds())
            if diff_end is None or d_end < diff_end:
                diff_end = d_end
                file_end = (dt, path)

    # Fallbacks
    if not file_start:
        file_start = exports[0]

    if not file_end:
        latest = exports[-1]
        if latest[1] != file_start[1]:
            file_end = latest

    if not file_end:
        await private_or_current_reply(
            update,
            context,
            f"Only one report has been run today (at {file_start[0].strftime('%H:%M')}). "
            "Please run another push first to compare."
        )
        return

    msg = await send_requester_text(update, context, "Calculating differences...")

    try:
        tmpdir = tempfile.mkdtemp(prefix="vs_")
        res_1 = generate_report.generate_reports_from_data(
            file_start[1], REF_PATH, tmpdir, return_metadata=True, mode="wide"
        )
        res_2 = generate_report.generate_reports_from_data(
            file_end[1], REF_PATH, tmpdir, return_metadata=True, mode="wide"
        )
        # Clean up
        for f in os.listdir(tmpdir):
            try:
                os.remove(os.path.join(tmpdir, f))
            except Exception:
                pass
        try:
            os.rmdir(tmpdir)
        except Exception:
            pass
    except Exception as e:
        await edit_or_send_requester_text(msg, update, context, f"Error processing reports: {e}")
        return

    today_date = datetime.now().date()
    filter_cols = {
        'Pickup': 'POST OFFICE HANDLE',
        'Delivery': 'POST OFFICE HANDLE',
        'Pending': 'POST OFFICE HANDLE'
    }

    # Aggregate by handle
    all_h = set()
    for rn in ['Pickup', 'Delivery', 'Pending']:
        df1 = res_1['type_data'].get(rn, pd.DataFrame())
        df2 = res_2['type_data'].get(rn, pd.DataFrame())
        fcol = filter_cols[rn]
        if fcol in df1.columns:
            all_h.update(df1[fcol].dropna().unique())
        if fcol in df2.columns:
            all_h.update(df2[fcol].dropna().unique())

    all_h = sorted(list(h for h in all_h if str(h).strip()))
    if target_handles:
        all_h = [h for h in all_h if h in target_handles]

    if not all_h:
        await edit_or_send_requester_text(msg, update, context, "No matching handles found in today's data.")
        return

    t_start = file_start[0].strftime('%I:%M %p')
    t_end = file_end[0].strftime('%I:%M %p')

    text_lines = [
        f"📊 {title_label}VS REPORT ({command_label}) — {t_start} vs {t_end}",
        "=============================="
    ]

    grand_sets_1 = {"Pickup": set(), "Delivery": set(), "Pending": set()}
    grand_sets_2 = {"Pickup": set(), "Delivery": set(), "Pending": set()}

    for h in all_h:
        h_lines = []
        has_any_data = False
        
        for rn in ['Pickup', 'Delivery', 'Pending']:
            df1 = res_1['type_data'].get(rn, pd.DataFrame())
            df2 = res_2['type_data'].get(rn, pd.DataFrame())
            fcol = filter_cols[rn]
            
            df1_h = df1[df1[fcol] == h] if fcol in df1.columns else pd.DataFrame()
            df2_h = df2[df2[fcol] == h] if fcol in df2.columns else pd.DataFrame()
            
            set_1 = get_highlighted_order_ids(df1_h, today_date)
            set_2 = get_highlighted_order_ids(df2_h, today_date)
            
            grand_sets_1[rn].update(set_1)
            grand_sets_2[rn].update(set_2)
            
            n1 = len(set_1)
            n2 = len(set_2)
            cleared = len(set_1 - set_2)
            
            if n1 > 0 or n2 > 0:
                has_any_data = True
                h_lines.append(f"  • {rn}: {n1} vs {n2} (Cleared {cleared})")

        if has_any_data and len(all_h) <= 10:
            text_lines.append(f"\n🏢 {h}:")
            text_lines.extend(h_lines)

    # Grand Summary
    text_lines.append("\n==============================")
    text_lines.append("📈 GRAND SUMMARY:")
    g_1 = 0
    g_2 = 0
    g_clear = 0
    for rn in ['Pickup', 'Delivery', 'Pending']:
        s1 = grand_sets_1[rn]
        s2 = grand_sets_2[rn]
        n1 = len(s1)
        n2 = len(s2)
        cleared = len(s1 - s2)
        
        g_1 += n1
        g_2 += n2
        g_clear += cleared
        
        text_lines.append(f"  • {rn}: {n1} vs {n2} (Cleared {cleared})")

    text_lines.append(f"  • Total: {g_1} vs {g_2} (Cleared {g_clear})")

    await edit_or_send_requester_text(msg, update, context, "\n".join(text_lines))


@pm_required_handler
async def cmd_vs(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/vs [handle/zone] — compare morning (8 AM) and afternoon (2 PM/current) reports."""
    await run_time_vs(update, context, start_hour=8, end_hour=14, command_label="8AM vs 2PM")


@pm_required_handler
async def cmd_vs2(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/vs2 [handle/zone] — compare afternoon (2 PM) and evening (5 PM/current) reports."""
    await run_time_vs(update, context, start_hour=14, end_hour=17, command_label="2PM vs 5PM")



@user_guard
async def cmd_groups(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """List all registered groups."""
    await delete_group_command(update, context)
    groups = load_registered_groups()
    if not groups:
        await private_or_current_reply(
            update,
            context,
            "No registered groups.\n"
            "Use /register inside a group to add it, or set forward_groups in config.json."
        )
        return
    lines = ["Registered groups:"]
    for g in groups:
        lines.append(f"  • {g.get('title', '')} — {g['chat_id']}")
    await private_or_current_reply(update, context, "\n".join(lines))


def _configured_export_branches(cfg):
    branches = []

    def add_branch(value):
        value = str(value or "").strip().upper()
        if value and value not in branches:
            branches.append(value)

    for raw in cfg.get("api", {}).get("branch_code", "").split(","):
        add_branch(raw)

    for raw_list in cfg.get("zone_branches", {}).values():
        for raw in str(raw_list or "").split(","):
            add_branch(raw)

    return branches


def _parse_export_branches(raw_args, cfg):
    branches = []
    for arg in raw_args:
        for piece in str(arg).replace(";", ",").split(","):
            piece = piece.strip().upper()
            if piece and piece not in branches:
                branches.append(piece)
    return branches or _configured_export_branches(cfg)


def _strip_department_code(value, code=""):
    text = str(value or "").strip()
    if not text or text.lower() == "nan":
        return ""
    if code:
        text = re.sub(rf"^{re.escape(str(code).strip())}\s*-\s*", "", text, flags=re.IGNORECASE)
    return re.sub(r"^[A-Z0-9]{3,10}\s*-\s*", "", text).strip()


def _clean_export_phone(value):
    phone = str(value or "").strip()
    if not phone or phone.lower() == "nan":
        return ""
    phone = re.sub(r"[^\d+]", "", phone)
    if phone.startswith("+855"):
        phone = "0" + phone[4:]
    if phone.isdigit() and not phone.startswith("0"):
        phone = "0" + phone
    return phone


async def fetch_lat_long(code: str, token: str, sem: asyncio.Semaphore):
    async with sem:
        url = f"https://gw-express.metfone.com.kh/vtp-user/api/v1/departments/{code}"
        headers = {
            "Authorization": f"Bearer {token}",
            "Referer": "https://opsexpress.metfone.com.kh/",
            "Accept": "application/json, text/plain, */*",
            "User-Agent": "Mozilla/5.0",
        }
        for attempt in range(3):
            try:
                import requests
                r = await asyncio.to_thread(
                    requests.get, url, headers=headers, timeout=10
                )
                if r.status_code == 200:
                    data = r.json()
                    addr = data.get("departmentAddress") or {}
                    return code, addr.get("latitude"), addr.get("longitude")
                elif r.status_code == 404:
                    return code, None, None
            except Exception as e:
                log.warning(f"Error fetching coordinates for {code} (attempt {attempt+1}): {e}")
                await asyncio.sleep(0.5)
        return code, None, None


PROVINCE_TO_MAIN_BRANCH = {
    "BAN": "BANP001", "BAT": "BATP001", "CHA": "CHAP001", "CHH": "CHHP001",
    "KAM": "KAMP001", "KAN": "KANP001", "KOH": "KOHP001", "KRA": "KRAP001",
    "MON": "MONP001", "ODD": "ODDP001", "PNP": "PNPP001", "PRE": "PREP001",
    "PRH": "PRHP001", "PUR": "PURP001", "ROT": "ROTP001", "SIE": "SIEP001",
    "SIH": "SIHP001", "SPE": "SPEP001", "STU": "STUP001", "SVA": "SVAP001",
    "TAK": "TAKP001", "TBK": "TBKP001", "THO": "THOP001", "KEP": "KAMP001",
    "PAI": "BATP001"
}

_location_phones_cache = None

def _get_fallback_location_phone(code: str, branch_code: str = "") -> str:
    global _location_phones_cache
    if _location_phones_cache is None:
        _location_phones_cache = {}
        here = os.path.dirname(os.path.abspath(__file__))
        
        # 1. Load from manager_contacts.json (main branch hubs)
        mc_path = os.path.join(here, "manager_contacts.json")
        if os.path.exists(mc_path):
            try:
                with open(mc_path, "r", encoding="utf-8") as f:
                    mc = json.load(f)
                    for k, v in mc.items():
                        if isinstance(v, dict) and v.get("phone"):
                            _location_phones_cache[k.upper()] = _clean_export_phone(v["phone"])
            except Exception:
                pass
                
        # 2. Load from post_office_lookup.csv
        po_path = os.path.join(here, "post_office_lookup.csv")
        if os.path.exists(po_path):
            try:
                import pandas as pd
                df_po = pd.read_csv(po_path, dtype=str)
                for _, r in df_po.iterrows():
                    po = str(r.get("current_post_office", "")).strip().upper()
                    ph = _clean_export_phone(r.get("phone", ""))
                    if po and ph:
                        _location_phones_cache[po] = ph
            except Exception:
                pass
                
        # 3. Load from pickup_branch_lookup.csv (Search Text / Phone)
        pk_path = os.path.join(here, "pickup_branch_lookup.csv")
        if os.path.exists(pk_path):
            try:
                import pandas as pd
                df_pk = pd.read_csv(pk_path, dtype=str)
                for _, r in df_pk.iterrows():
                    c = str(r.get("Pickup Branch", "") or r.get("Post code", "")).strip().upper()
                    st = str(r.get("Search Text", "")).strip()
                    if c and st and (c not in _location_phones_cache or not _location_phones_cache[c]):
                        parts = [p.strip() for p in st.split("|")]
                        for p in parts:
                            if re.match(r"^(0\d{7,10}|\d{8,10})$", p):
                                _location_phones_cache[c] = _clean_export_phone(p)
                                break
            except Exception:
                pass

    code_u = str(code or "").strip().upper()
    if code_u in _location_phones_cache and _location_phones_cache[code_u]:
        return _location_phones_cache[code_u]
    b_u = str(branch_code or "").strip().upper()
    if b_u in _location_phones_cache and _location_phones_cache[b_u]:
        return _location_phones_cache[b_u]
    return ""


def _classify_facility(code: str, type_label: str = "") -> str:
    code = str(code or "").strip().upper()
    type_label = str(type_label or "").strip()
    
    m = re.search(r"^[A-Z]{3}([PSA])\d+$", code)
    if m:
        letter = m.group(1)
        if letter == "P": return "Post Office"
        if letter == "S": return "Showroom"
        if letter == "A": return "Agent"

    if re.search(r"[A-Z]{3}P\d+", code): return "Post Office"
    if re.search(r"[A-Z]{3}S\d+", code): return "Showroom"
    if re.search(r"[A-Z]{3}A\d+", code): return "Agent"

    if "Warehouse" in type_label or "Hub" in type_label or "Operations" in type_label:
        return "Warehouse / Hub"

    if "Authorized" in type_label or "Agent" in type_label or "Dealer" in type_label:
        return "Agent"

    if "Post office" in type_label or "Post Office" in type_label:
        return "Post Office"

    return "Post Office"


def _post_office_export_row(item, fallback_branch=""):
    branch = item.get("branch") if isinstance(item.get("branch"), dict) else {}
    code = str(item.get("code") or item.get("Pickup Branch") or item.get("Post code") or "").strip().upper()
    branch_code = str(
        item.get("main_branch_code")
        or item.get("parentDepartmentCode")
        or item.get("Branch Code")
        or branch.get("code")
        or fallback_branch
        or ""
    ).strip().upper()

    if len(branch_code) == 3:
        branch_code = PROVINCE_TO_MAIN_BRANCH.get(branch_code, branch_code)

    commune_en = _strip_department_code(
        item.get("Commune EN") or item.get("Post office name") or item.get("enUsName") or item.get("name") or item.get("viVnName"),
        code,
    )
    commune_khmer = _strip_department_code(
        item.get("Commune Khmer") or item.get("kmKhmName") or item.get("enUsName") or item.get("name"),
        code,
    )
    branch_en = _strip_department_code(
        item.get("Branch EN") or branch.get("enUsName") or branch.get("name") or item.get("branch_name"),
        branch_code,
    )
    branch_khmer = _strip_department_code(
        item.get("Branch Khmer") or branch.get("kmKhmName") or branch.get("enUsName") or branch.get("name"),
        branch_code,
    )
    phone = _clean_export_phone(item.get("Phone Number") or item.get("Phone") or item.get("phone") or item.get("phone_detail")) or _get_fallback_location_phone(code, branch_code)
    
    from speed_report import MAIN_36_BRANCHES as _36
    main_b = PROVINCE_TO_MAIN_BRANCH.get(branch_code) or PROVINCE_TO_MAIN_BRANCH.get(code[:3]) or (code if code in _36 else "PNPP001")
    main_branch_phone = _clean_export_phone(item.get("Main Branch Phone") or item.get("main_branch_phone")) or _get_fallback_location_phone(main_b)

    category = _classify_facility(code, item.get("Category") or item.get("Post office level") or item.get("typeLabel") or item.get("type"))

    search_parts = [
        code,
        commune_en,
        commune_khmer,
        phone,
        branch_code,
        branch_en,
        branch_khmer,
        category,
    ]

    return {
        "Pickup Branch": code,
        "Post code": code,
        "Post office name": commune_en,
        "Commune EN": commune_en,
        "Commune Khmer": commune_khmer,
        "Phone": phone,
        "Phone Number": phone,
        "Main Branch Phone": main_branch_phone,
        "Branch Code": branch_code,
        "Branch EN": branch_en,
        "Branch Khmer": branch_khmer,
        "Category": category,
        "Post office level": category,
        "Type": str(item.get("Category") or item.get("typeLabel") or item.get("type") or "").strip(),
        "Status": str(item.get("Status") or item.get("statusLabel") or item.get("status") or "In effect").strip(),
        "Latitude": item.get("Latitude") or item.get("latitude"),
        "Longitude": item.get("Longitude") or item.get("longitude"),
        "Search Text": " | ".join(part for part in search_parts if part),
    }


def _safe_excel_label(value, fallback="Export", max_len=80):
    label = re.sub(r"[^A-Za-z0-9_-]+", "_", str(value or "").strip())
    label = re.sub(r"_+", "_", label).strip("_")
    return (label or fallback)[:max_len]


BRANCH_TO_PROVINCE_EN = {
    "PRE": "Prey Veng",
    "PNP": "Phnom Penh",
    "SVA": "Svay Rieng",
    "KAN": "Kandal",
    "KAM": "Kampot",
    "KOH": "Koh Kong",
    "SIH": "Preah Sihanouk",
    "SPE": "Kampong Speu",
    "TAK": "Takeo",
    "BAN": "Banteay Meanchey",
    "BAT": "Battambang",
    "CHH": "Kampong Chhnang",
    "PUR": "Pursat",
    "SIE": "Siem Reap",
    "PRH": "Preah Vihear",
    "ODD": "Oddar Meanchey",
    "THO": "Kampong Thom",
    "CHA": "Kampong Cham",
    "KRA": "Kratie",
    "TBK": "Tboung Khmum",
    "ROT": "Ratanak Kiri",
    "MON": "Mondul Kiri",
    "STU": "Stung Treng",
    "KEP": "Kep",
    "PAI": "Pailin"
}

PROVINCE_MAP_KH = {
    "PRE": "ព្រៃវែង",
    "PNP": "ភ្នំពេញ",
    "SVA": "ស្វាយរៀង",
    "KAN": "កណ្តាល",
    "KAM": "កំពត",
    "KOH": "កោះកុង",
    "SIH": "ព្រះសីហនុ",
    "SPE": "កំពង់ស្ពឺ",
    "TAK": "តាកែវ",
    "BAN": "បន្ទាយមានជ័យ",
    "BAT": "បាត់ដំបង",
    "CHH": "កំពង់ឆ្នាំង",
    "PUR": "ពោធិ៍សាត់",
    "SIE": "សៀមរាប",
    "PRH": "ព្រះវិហារ",
    "ODD": "ឧត្តរមានជ័យ",
    "THO": "កំពង់ធំ",
    "CHA": "កំពង់ចាម",
    "KRA": "ក្រចេះ",
    "TBK": "ត្បូងឃ្មុំ",
    "ROT": "រតនគិរី",
    "MON": "មណ្ឌលគិរី",
    "STU": "ស្ទឹងត្រែង",
    "KEP": "កែប",
    "PAI": "ប៉ៃលិន"
}

DISTRICT_FALLBACK_KH = {
    "PRE": "ព្រៃវែង",
    "PNP": "ចំការមន",
    "SVA": "ស្វាយរៀង",
    "KAN": "តាខ្មៅ",
    "KAM": "កំពត",
    "KOH": "ខេមរភូមិន្ទ",
    "SIH": "ព្រះសីហនុ",
    "SPE": "ច្បារមន",
    "TAK": "ដូនកែវ",
    "BAN": "សិរីសោភ័ណ",
    "BAT": "បាត់ដំបង",
    "CHH": "កំពង់ឆ្នាំង",
    "PUR": "ពោធិ៍សាត់",
    "SIE": "សៀមរាប",
    "PRH": "ព្រះវិហារ",
    "ODD": "សំរោង",
    "THO": "ស្ទឹងសែន",
    "CHA": "កំពង់ចាម",
    "KRA": "ក្រចេះ",
    "TBK": "សួង",
    "ROT": "បានលុង",
    "MON": "សែនមនោរម្យ",
    "STU": "ស្ទឹងត្រែង",
}

DISTRICT_FALLBACK_EN = {
    "PRE": "Prey Veng",
    "PNP": "Chamkar Mon",
    "SVA": "Svay Rieng",
    "KAN": "Ta Khmau",
    "KAM": "Kampot",
    "KOH": "Khemarak Phoumin",
    "SIH": "Preah Sihanouk",
    "SPE": "Chbar Mon",
    "TAK": "Doun Kaev",
    "BAN": "Serei Saophoan",
    "BAT": "Battambang",
    "CHH": "Kampong Chhnang",
    "PUR": "Pursat",
    "SIE": "Siem Reap",
    "PRH": "Preah Vihear",
    "ODD": "Samraong",
    "THO": "Steung Saen",
    "CHA": "Kampong Cham",
    "KRA": "Kratie",
    "TBK": "Suong",
    "ROT": "Banlung",
    "MON": "Senmonorom",
    "STU": "Stung Treng",
}

_gazetteer_data = None

def _get_gazetteer():
    global _gazetteer_data
    if _gazetteer_data is not None:
        return _gazetteer_data
    path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "cambodia_gazetteer.json")
    if os.path.exists(path):
        try:
            with open(path, "r", encoding="utf-8") as f:
                _gazetteer_data = json.load(f)
        except Exception:
            _gazetteer_data = []
    else:
        _gazetteer_data = []
    return _gazetteer_data

def _map_to_administrative_division(branch_code, commune_en_raw, commune_kh_raw):
    gazetteer = _get_gazetteer()
    prov_target = BRANCH_TO_PROVINCE_EN.get(branch_code, "")
    
    target_en = str(commune_en_raw).lower().replace(" ", "").replace("-", "")
    target_kh = str(commune_kh_raw).lower().replace(" ", "").replace("-", "")
    
    for item in gazetteer:
        if prov_target and item["prov_en"].lower().replace(" ", "") != prov_target.lower().replace(" ", ""):
            continue
        
        c_en = item["comm_en"].lower().replace(" ", "").replace("-", "")
        c_kh = item["comm_kh"].lower().replace(" ", "").replace("-", "")
        
        if c_en == target_en or c_kh == target_kh or c_en in target_en or target_en in c_en:
            return item["prov_kh"], item["dist_en"], item["dist_kh"], item["comm_kh"]
            
    prov_kh = PROVINCE_MAP_KH.get(branch_code, "បាត់ដំបង")
    dist_en = DISTRICT_FALLBACK_EN.get(branch_code, "Battambang")
    dist_kh = DISTRICT_FALLBACK_KH.get(branch_code, "បាត់ដំបង")
    comm_kh = commune_kh_raw or commune_en_raw
    return prov_kh, dist_en, dist_kh, comm_kh


def _write_post_office_export_excel(df, out_path, sheet_label, title):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    DARK_BLUE = "172033"
    WHITE = "FFFFFF"
    BORDER_CLR = "CCCCCC"
    
    thin_border = Border(
        left=Side(style="thin", color=BORDER_CLR),
        right=Side(style="thin", color=BORDER_CLR),
        top=Side(style="thin", color=BORDER_CLR),
        bottom=Side(style="thin", color=BORDER_CLR),
    )
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Stores"
    ws.views.sheetView[0].showGridLines = True
    
    data_headers = ["Province *", "District *", "District KH", "Delivery Store *", "Category *", "Phone Number", "Main Branch Phone", "Latitude", "Longitude"]
    
    for col_idx, col_name in enumerate(data_headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(name="Calibri", bold=True, color=WHITE, size=11)
        cell.fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws.row_dimensions[1].height = 28
    
    for idx, row in df.iterrows():
        branch_code = str(row.get("Branch Code", "")).strip().upper()
        commune_en = str(row.get("Commune EN", ""))
        commune_kh = str(row.get("Commune Khmer", ""))
        code = str(row.get("Pickup Branch", ""))
        category = str(row.get("Category") or _classify_facility(code, row.get("Type"))).strip()
        phone = _clean_export_phone(row.get("Phone Number") or row.get("Phone") or row.get("phone"))
        main_branch_phone = _clean_export_phone(row.get("Main Branch Phone") or row.get("main_branch_phone"))
        lat = row.get("Latitude")
        lon = row.get("Longitude")
        
        prov_kh, dist_en, dist_kh, comm_kh = _map_to_administrative_division(branch_code, commune_en, commune_kh)
        store_name = f"{code} - {commune_en}"
        
        row_idx = idx + 2
        
        ws.cell(row=row_idx, column=1, value=prov_kh).font = Font(name="Calibri", size=10)
        ws.cell(row=row_idx, column=2, value=dist_en).font = Font(name="Calibri", size=10)
        ws.cell(row=row_idx, column=3, value=dist_kh).font = Font(name="Calibri", size=10)
        ws.cell(row=row_idx, column=4, value=store_name).font = Font(name="Calibri", size=10)
        ws.cell(row=row_idx, column=5, value=category).font = Font(name="Calibri", size=10, bold=True)
        
        c_ph = ws.cell(row=row_idx, column=6, value=phone)
        c_ph.font = Font(name="Calibri", size=10)
        c_ph.number_format = "@"
        
        c_mb = ws.cell(row=row_idx, column=7, value=main_branch_phone)
        c_mb.font = Font(name="Calibri", size=10)
        c_mb.number_format = "@"
        
        ws.cell(row=row_idx, column=8, value=lat).font = Font(name="Calibri", size=10)
        ws.cell(row=row_idx, column=9, value=lon).font = Font(name="Calibri", size=10)
        
        for col_idx in range(1, 10):
            ws.cell(row=row_idx, column=col_idx).border = thin_border
            
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 45
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 20
    ws.column_dimensions["H"].width = 15
    ws.column_dimensions["I"].width = 15
    
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:I{len(df)+1}"
    
    wb.save(out_path)



async def send_pickup_branch_export(update, context, cfg, raw_args):
    first_arg = raw_args[0].lower() if raw_args else ""
    all_mode = first_arg in ("all", "pickup", "pickups", "search", "branches")
    branch_args = raw_args[1:] if all_mode else raw_args
    branch_codes = _parse_export_branches(branch_args, cfg)
    if not branch_codes:
        await private_or_current_reply(update, context, "No branch codes configured for export.")
        return

    label = "ALL_PICKUP" if all_mode and not branch_args else ",".join(branch_codes)
    description = "all pickup branches" if all_mode and not branch_args else ", ".join(branch_codes)
    msg = await send_requester_text(update, context, f"Fetching post offices for {description} in parallel...")

    try:
        import pandas as pd

        post_offices = []
        branch_errors = []
        
        # Concurrency limit of 4 to avoid hitting server limits/500 errors
        sem = asyncio.Semaphore(4)
        
        async def sem_download(code):
            async with sem:
                # Stagger the requests slightly to prevent a burst of 4 requests at the exact same millisecond
                await asyncio.sleep(0.3)
                return await asyncio.to_thread(
                    downloader.download_post_offices,
                    cfg["api"],
                    code,
                )
        
        # Launch parallel downloads with semaphore limit
        tasks = [sem_download(bc) for bc in branch_codes]
        results = await asyncio.gather(*tasks, return_exceptions=True)
        
        for branch_code, res in zip(branch_codes, results):
            if isinstance(res, Exception):
                branch_errors.append(f"{branch_code}: {res}")
                log.warning("Export failed for branch %s: %s", branch_code, res)
            else:
                for item in res:
                    if isinstance(item, dict):
                        item = dict(item)
                        item["_export_branch_query"] = branch_code
                    post_offices.append(item)

        if not post_offices:
            extra = "\n".join(branch_errors[:5])
            await edit_or_send_requester_text(
                msg,
                update,
                context,
                f"No post offices found for {description}." + (f"\n{extra}" if extra else "")
            )
            return

        # Fetch coordinates in parallel
        unique_codes = list(set(
            str(item.get("code", "")).strip().upper()
            for item in post_offices
            if isinstance(item, dict) and item.get("code")
        ))
        
        if unique_codes:
            await edit_or_send_requester_text(
                msg,
                update,
                context,
                f"Fetched {len(post_offices)} offices. Retrieving coordinates..."
            )
            detail_sem = asyncio.Semaphore(15)
            detail_tasks = [fetch_lat_long(code, cfg["api"]["bearer_token"], detail_sem) for code in unique_codes]
            detail_results = await asyncio.gather(*detail_tasks, return_exceptions=True)
            
            coords_map = {}
            for res in detail_results:
                if isinstance(res, tuple) and len(res) == 3:
                    code, lat, lon = res
                    coords_map[code] = (lat, lon)
            
            for item in post_offices:
                if isinstance(item, dict):
                    code = str(item.get("code", "")).strip().upper()
                    lat, lon = coords_map.get(code, (None, None))
                    item["latitude"] = lat
                    item["longitude"] = lon

        rows = [
            _post_office_export_row(item, item.get("_export_branch_query", ""))
            for item in post_offices
            if isinstance(item, dict)
        ]
        df = pd.DataFrame(rows)
        if "Pickup Branch" in df.columns:
            df = df[df["Pickup Branch"].astype(str).str.strip() != ""].copy()
            df = df.drop_duplicates(subset=["Pickup Branch"], keep="first")
        sort_cols = [c for c in ("Branch Code", "Pickup Branch") if c in df.columns]
        if sort_cols:
            df = df.sort_values(sort_cols).reset_index(drop=True)

        if df.empty:
            await edit_or_send_requester_text(
                msg,
                update,
                context,
                f"No valid post office rows found for {description}."
            )
            return

        if all_mode:
            df.to_csv(PICKUP_BRANCH_LOOKUP_PATH, index=False, encoding="utf-8-sig")

        await edit_or_send_requester_text(
            msg,
            update,
            context,
            f"Found {len(df)} pickup branches. Building modern Excel..."
        )

        tmpdir = tempfile.mkdtemp(prefix="export_po_")
        stamp = datetime.now().strftime("%d.%m_%HH%M")
        safe_label = _safe_excel_label(label)
        filename = f"PickupBranches_{safe_label}_{stamp}.xlsx"
        out_path = os.path.join(tmpdir, filename)
        title = f"Pickup Branch Search Export - {description} ({len(df)} offices)"

        _write_post_office_export_excel(df, out_path, label, title)

        with open(out_path, "rb") as f:
            await send_requester_document(update, context, f, filename)

        done_lines = [
            f"Exported {len(df)} pickup branches for {description}.",
            f"Time: {datetime.now().strftime('%d.%m.%Y %H:%M')}",
        ]
        if all_mode:
            done_lines.append(f"Updated local lookup: {os.path.basename(PICKUP_BRANCH_LOOKUP_PATH)}")
        if branch_errors:
            done_lines.append("Some branches failed: " + "; ".join(branch_errors[:3]))
            if len(branch_errors) > 3:
                done_lines.append(f"...and {len(branch_errors) - 3} more.")

        await edit_or_send_requester_text(msg, update, context, "\n".join(done_lines))

    except Exception as e:
        log.exception("Error in pickup branch export")
        await edit_or_send_requester_text(msg, update, context, f"Export failed: {e}")


@pm_required_handler
async def cmd_export(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/export <branch_code> — export post office list for a branch.
    Examples: /export KAM  |  /export PNP  |  /export SVA  |  /export all
    """
    await delete_group_command(update, context)
    cfg = load_config()

    raw_args = [a.strip() for a in (context.args or []) if a.strip()]
    if not raw_args:
        await private_or_current_reply(
            update,
            context,
            "Usage: /export <branch_code>\n"
            "Downloads the post office list for a branch.\n\n"
            "Examples:\n"
            "  /export KAM — all Kampot post offices\n"
            "  /export PNP — all Phnom Penh post offices\n"
            "  /export SVA — all Svay Rieng post offices\n"
            "  /export all — export all configured branches"
        )
        return

    first_arg = raw_args[0].lower()
    if first_arg in ("all", "pickup", "pickups", "search", "branches"):
        await send_pickup_branch_export(update, context, cfg, raw_args)
        return

    args = [a.strip().upper() for a in raw_args]
    branch_code = args[0]

    msg = await send_requester_text(
        update, context,
        f"📥 Fetching post offices for branch: {branch_code}..."
    )

    try:
        import pandas as pd

        post_offices = downloader.download_post_offices(cfg["api"], branch_code)

        if not post_offices:
            await edit_or_send_requester_text(
                msg, update, context,
                f"No post offices found for branch {branch_code}."
            )
            return

        # If the result is a list of dicts, convert to DataFrame
        if isinstance(post_offices, list) and isinstance(post_offices[0], dict):
            df = pd.json_normalize(post_offices, sep='_')
        else:
            df = pd.DataFrame(post_offices)

        # Select and rename useful columns to match management page
        col_map = {
            'code': 'Department Code',
            'name': 'Department Name',
            'parentDepartmentCode': 'Branch Code',
            'branch_name': 'Branch Name',
            'type': 'Type',
            'phone': 'Phone',
        }
        # Only keep columns that exist
        keep = [c for c in col_map if c in df.columns]
        df = df[keep].rename(columns=col_map)

        # Strip code prefix from Department Name (e.g. "KAMA001 - Chamnaom" → "Chamnaom")
        if 'Department Name' in df.columns:
            df['Department Name'] = df['Department Name'].str.replace(
                r'^[A-Z0-9]+ - ', '', regex=True)

        # Build "Branch" column as "KAM - Kampot"
        if 'Branch Code' in df.columns and 'Branch Name' in df.columns:
            df.insert(
                df.columns.get_loc('Branch Name') + 1,
                'Branch',
                df['Branch Code'] + ' - ' + df['Branch Name']
            )

        await edit_or_send_requester_text(
            msg, update, context,
            f"Found {len(df)} post offices. Building Excel..."
        )

        tmpdir = tempfile.mkdtemp(prefix="export_po_")
        stamp = datetime.now().strftime("%d.%m_%HH%M")
        filename = f"PostOffices_{branch_code}_{stamp}.xlsx"
        out_path = os.path.join(tmpdir, filename)

        # ── Styled Excel with Metfone Green theme ──
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        PRIMARY = "00A651"      # Metfone Green
        SECONDARY = "EAF7EF"    # Light Green
        WHITE = "FFFFFF"
        DARK_TEXT = "333333"
        BORDER_CLR = "B2D8B2"

        thin_border = Border(
            left=Side(style='thin', color=BORDER_CLR),
            right=Side(style='thin', color=BORDER_CLR),
            top=Side(style='thin', color=BORDER_CLR),
            bottom=Side(style='thin', color=BORDER_CLR),
        )
        header_font = Font(name='Calibri', bold=True, color=WHITE, size=11)
        header_fill = PatternFill(start_color=PRIMARY, end_color=PRIMARY, fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        alt_fill = PatternFill(start_color=SECONDARY, end_color=SECONDARY, fill_type='solid')
        data_font = Font(name='Calibri', color=DARK_TEXT, size=10)
        data_align = Alignment(vertical='center')
        title_font = Font(name='Calibri', bold=True, color=PRIMARY, size=14)

        wb = Workbook()
        ws = wb.active
        ws.title = branch_code

        # Title row
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
        title_cell = ws.cell(row=1, column=1,
                             value=f"📋 Post Office List — {branch_code} ({len(df)} offices)")
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[1].height = 30

        # Header row (row 2)
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=2, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
        ws.row_dimensions[2].height = 28

        # Data rows
        for row_idx, row_data in enumerate(df.itertuples(index=False), 3):
            is_alt = (row_idx % 2 == 1)
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = data_font
                cell.alignment = data_align
                cell.border = thin_border
                if is_alt:
                    cell.fill = alt_fill

        # Auto-fit column widths
        for col_idx, col_name in enumerate(df.columns, 1):
            max_len = len(str(col_name))
            for row_idx in range(3, len(df) + 3):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[ws.cell(row=2, column=col_idx).column_letter].width = min(max_len + 3, 45)

        # Freeze header
        ws.freeze_panes = 'A3'
        ws.auto_filter.ref = f"A2:{ws.cell(row=2, column=len(df.columns)).column_letter}2"

        wb.save(out_path)

        with open(out_path, "rb") as f:
            await send_requester_document(
                update, context, f, filename,
            )

        await edit_or_send_requester_text(
            msg, update, context,
            f"✅ Exported {len(df)} post offices for {branch_code} — "
            f"{datetime.now().strftime('%d.%m.%Y %H:%M')}"
        )

    except Exception as e:
        log.exception("Error in /export")
        await edit_or_send_requester_text(
            msg, update, context,
            f"❌ Export failed: {e}"
        )


@pm_required_handler
async def cmd_find(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/find <query> — search for orders by phone number or order ID."""
    await delete_group_command(update, context)
    args = [a.strip() for a in (context.args or []) if a.strip()]
    if not args:
        await private_or_current_reply(
            update,
            context,
            "Usage: `/find <phone_number>` or `/find <order_id>`\n"
            "Example: `/find 0712345678`",
        )
        return

    force_refresh = False
    if "force" in [a.lower() for a in args]:
        force_refresh = True
        args = [a for a in args if a.lower() != "force"]

    if not args:
        await private_or_current_reply(
            update,
            context,
            "Please specify a phone number or order ID to find.",
        )
        return

    query = args[0].lower()
    msg = await send_requester_text(
        update, context,
        f"🔍 Searching for '{query}' in recent orders (up to 7 days)..."
    )

    cfg = load_config()
    tmpdir = tempfile.mkdtemp(prefix="find_")
    stamp = datetime.now().strftime("%d.%m_%HH%M")
    src = os.path.join(tmpdir, f"export_{stamp}.xlsx")

    try:
        import downloader
        downloader.download_detail(cfg["api"], src, force_refresh=force_refresh)
        
        import pandas as pd
        xl = pd.ExcelFile(src)
        sheet = xl.sheet_names[0]
        if len(xl.sheet_names) > 1:
            for sname in xl.sheet_names:
                try:
                    s = xl.parse(sname, nrows=3)
                    if 'ORDER ID' in s.columns:
                        sheet = sname
                        break
                except Exception:
                    continue
        df = xl.parse(sheet)

        search_cols = ['ORDER ID', 'SENDER', 'RECEIVER', 'Cus name', 'Phone']
        match_idx = pd.Series(False, index=df.index)

        for col in search_cols:
            if col in df.columns:
                match_idx = match_idx | df[col].astype(str).str.lower().str.contains(query, na=False)

        results = df[match_idx]

        if results.empty:
            await edit_or_send_requester_text(
                msg, update, context,
                f"❌ No orders found matching '{query}'."
            )
            return

        out_lines = [f"✅ Found {len(results)} order(s) for '{query}':\n"]
        # Limit to first 20 to avoid blowing up Telegram message length too much
        limit = 20
        for _, row in results.head(limit).iterrows():
            order_id = row.get('ORDER ID', 'Unknown')
            status = row.get('CURRENT STATUS', 'Unknown')
            po = row.get('CURRENT POST OFFICE', 'Unknown')
            date = row.get('CREATED DATE', '')
            if pd.isna(date): date = ''
            sender = row.get('SENDER', '')
            if pd.isna(sender): sender = ''

            out_lines.append(f"📦 *Order ID:* `{order_id}`")
            out_lines.append(f"📍 *Status:* {status}")
            out_lines.append(f"🏢 *Post Office:* {po}")
            if date: out_lines.append(f"📅 *Created:* {date}")
            if sender: out_lines.append(f"👤 *Sender:* {sender}")
            out_lines.append("━━━━━━━━━━━━━━━━━━━━━━")

        if len(results) > limit:
            out_lines.append(f"... and {len(results) - limit} more. Showing first {limit}.")

        final_text = "\n".join(out_lines)
        # Final safety for telegram limits
        if len(final_text) > 4000:
            final_text = final_text[:4000] + "\n... (Message truncated)"

        await edit_or_send_requester_text(msg, update, context, final_text)

    except Exception as e:
        log.exception("Error in /find")
        await edit_or_send_requester_text(
            msg, update, context,
            f"❌ Search failed: {e}"
        )

@pm_required_handler
async def cmd_ask(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/ask <order_id> [kh] - find an order, identify the responsible post office to scan it, and describe the next step."""
    await delete_group_command(update, context)
    args = [a.strip() for a in (context.args or []) if a.strip()]
    if not args:
        await private_or_current_reply(
            update,
            context,
            "Usage: `/ask <order_id> [kh]`\n"
            "Example: `/ask 2900492262 kh`"
        )
        return

    use_khmer = False
    query = ""
    for arg in args:
        if arg.lower() in ("kh", "-kh"):
            use_khmer = True
        else:
            query = arg.lower()

    if not query:
        await private_or_current_reply(
            update,
            context,
            "Usage: `/ask <order_id> [kh]`\n"
            "Example: `/ask 2900492262 kh`"
        )
        return

    order_id = query
    loading_msg = f"🔍 កំពុងទាញយកព័ត៌មាន និងវិភាគការបញ្ជាទិញ '{query}'..." if use_khmer else f"🔍 Fetching details and analyzing order '{query}'..."
    msg = await send_requester_text(
        update, context,
        loading_msg
    )

    cfg = load_config()
    token = cfg["api"]["bearer_token"]

    try:
        import requests
        import pandas as pd
        
        headers = {
            "Authorization": f"Bearer {token}",
            "Accept": "application/json, text/plain, */*",
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        }
        
        # 1. Always attempt to query order search for rich details first
        order_data = None
        search_url = "https://gw-express.metfone.com.kh/tms-receiving/api/v1/orders/search"
        try:
            r_search = requests.get(search_url, params={"order_code": query}, headers=headers, timeout=15)
            if r_search.status_code == 200:
                order_data = r_search.json()
        except Exception as e:
            log.warning("Failed to query orders/search: %s", e)

        # 2. Query tracking trips API
        url = "https://gw-express.metfone.com.kh/tms-tracking/api/v1/order-tracking"
        params = {"order_id": query}
        
        trips = []
        is_synthetic = False
        
        try:
            r = requests.get(url, params=params, headers=headers, timeout=15)
            if r.status_code == 401:
                await edit_or_send_requester_text(
                    msg, update, context,
                    "❌ API Token expired. Please update config.json with a valid Bearer token."
                )
                return
            elif r.status_code == 200:
                data = r.json()
                trips = data.get("trackingTrips", [])
        except Exception as e:
            log.warning("Failed to query order-tracking: %s", e)

        # 3. Fallback to synthetic trip if tracking trips empty (e.g. no scans yet or branch restriction)
        if not trips:
            if order_data:
                status_val = str(order_data.get("order_status", ""))
                is_synthetic = True
                
                # Map status code to status name (-99 = Metfone undefined/unscanned)
                status_name_map = {
                    "-99": "REGISTERED",
                    "0": "REGISTERED",
                    "100": "REGISTERED",
                    "110": "CONFIRMED",
                    "120": "PICKING",
                    "200": "RECEIVED",
                    "201": "CANCELLED",
                    "210": "TRANSITING",
                    "302": "TRANSITING",
                    "306": "TRANSITING",
                    "310": "TRANSITING",
                    "309": "TRANSITING",
                    "311": "TRANSITING",
                    "500": "TRANSITING",
                    "300": "TRANSITING",
                    "400": "Assigned to deliver - Not completed",
                    "401": "DELIVERING",
                    "402": "DELIVERING",
                    "420": "PTC_STORE",
                    "430": "PTC_RE_DELIVERY",
                    "460": "PTC_REDELIVERING",
                    "472": "STORED",
                    "480": "CHANGE_ADDRESS",
                    "405": "DELIVERED",
                    "410": "DELIVERED_COMPLETED",
                    "512": "RETURN_EXCEPTION",
                    "520": "RETURNED",
                    "540": "RETURN_FAILED"
                }
                status_name_str = status_name_map.get(status_val, "CREATED")
                
                # Format time nicely from epoch millis
                created_date_epoch = order_data.get("order_created_date") or order_data.get("created_at") or 0
                created_time_str = ""
                if created_date_epoch:
                    try:
                        from datetime import timezone
                        dt_created = datetime.fromtimestamp(created_date_epoch / 1000.0, tz=timezone.utc)
                        dt_created_local = dt_created.astimezone(timezone(timedelta(hours=7)))
                        created_time_str = dt_created_local.strftime("%Y-%m-%dT%H:%M:%S+07:00")
                    except Exception:
                        pass
                
                synthetic_trip = {
                    "status": f"S{status_val}" if not status_val.startswith("S") else status_val,
                    "statusName": status_name_str,
                    "postcode": order_data.get("post_code", ""),
                    "deliveryPostcode": order_data.get("delivery_post_code", ""),
                    "updatedAt": created_time_str,
                    "updatedBy": {
                        "name": order_data.get("shipper", {}).get("name") or "System",
                        "phone": order_data.get("shipper", {}).get("phone") or ""
                    },
                    "desc": order_data.get("order_name") or "Order registered",
                    "shipperName": order_data.get("shipper", {}).get("name") or "Express",
                    "is_synthetic": True,
                    # Force milestone 1 to resolve from creation date for -99 / undefined orders
                    "_created_epoch": created_date_epoch,
                }
                trips = [synthetic_trip]
            else:
                # Neither tracking nor search found the order
                await edit_or_send_requester_text(
                    msg, update, context,
                    f"❌ មិនរកឃើញការបញ្ជាទិញដែលត្រូវនឹងលេខសម្គាល់ '{query}' ទេ។"
                    if use_khmer else
                    f"❌ No order found matching ID '{query}'."
                )
                return

        # Get latest event (first in list)
        latest = trips[0]
        status = latest.get("status", "").lstrip("S")
        status_name = latest.get("statusName", "")
        po = latest.get("postcode", "")
        delivery_po = latest.get("deliveryPostcode", "")

        # /ask flow buckets are intentionally fixed and separate from report generation.
        pickup_codes = {"110", "120", "200"}
        delivery_codes = {"401", "400", "430", "420", "402", "460"}
        pending_codes = {"210", "302", "306", "310", "309", "311", "500", "300"}
        done_codes = {"410", "201", "520"}

        def get_app_action(status_code):
            actions = cfg.get("ask_app_actions", {})
            if not isinstance(actions, dict):
                return {}
            default = actions.get("default", {})
            specific = actions.get(status_code, {})
            if not isinstance(default, dict):
                default = {}
            if not isinstance(specific, dict):
                specific = {}
            merged = {**default, **specific}
            return {
                key: str(value).strip()
                for key, value in merged.items()
                if value is not None and str(value).strip()
            }

        def get_status_flow(status_code):
            flows = cfg.get("ask_status_flow", {})
            if not isinstance(flows, dict):
                return {}
            default = flows.get("default", {})
            specific = flows.get(status_code, {})
            if not isinstance(default, dict):
                default = {}
            if not isinstance(specific, dict):
                specific = {}
            merged = {**default, **specific}
            next_statuses = merged.get("next_statuses", [])
            if not isinstance(next_statuses, list):
                next_statuses = [next_statuses] if next_statuses else []
            return {
                "meaning": str(merged.get("meaning", "")).strip(),
                "next_flow": str(merged.get("next_flow", "")).strip(),
                "next_statuses": [str(s).strip() for s in next_statuses if str(s).strip()],
            }

        app_action = get_app_action(status)
        status_flow = get_status_flow(status)
        responsible_handle = po.upper() if po else (delivery_po.upper() if delivery_po else "Unknown")

        # Load reference post office mapping
        ref_df = pd.DataFrame()
        if os.path.exists(REF_PATH):
            try:
                ref_df = pd.read_csv(REF_PATH, dtype=str)
                ref_df['current_post_office'] = ref_df['current_post_office'].astype(str).str.strip().str.upper()
                ref_df['post_office_handle'] = ref_df['post_office_handle'].astype(str).str.strip().str.upper()
            except Exception as e:
                log.warning("Could not load post office lookup ref: %s", e)

        def clean_phone_helper(ph):
            if not ph or str(ph).lower() == 'nan': return "Unknown Phone"
            ph_str = str(ph).strip()
            if ph_str.startswith("+855"):
                ph_str = "0" + ph_str[4:]
            if ph_str.startswith("+"):
                ph_str = "+" + "".join(c for c in ph_str[1:] if c.isdigit())
                if ph_str.startswith("+855"):
                    ph_str = "0" + ph_str[4:]
            if ph_str.isdigit() and not ph_str.startswith("0"):
                ph_str = "0" + ph_str
            return ph_str

        handle_name = "Unknown Name"
        handle_phone = "Unknown Phone"
        
        # 1. Lookup in reference CSV
        if not ref_df.empty and responsible_handle != "Unknown":
            match_ref = ref_df[ref_df['current_post_office'] == responsible_handle]
            if not match_ref.empty:
                name_val = match_ref.iloc[0].get('customer_name', '')
                phone_val = match_ref.iloc[0].get('phone', '')
                if name_val and str(name_val).lower() != 'nan':
                    handle_name = name_val
                if phone_val:
                    handle_phone = clean_phone_helper(phone_val)
            else:
                match_handle = ref_df[ref_df['post_office_handle'] == responsible_handle]
                if not match_handle.empty:
                    name_val = match_handle.iloc[0].get('customer_name', '')
                    phone_val = match_handle.iloc[0].get('phone', '')
                    if name_val and str(name_val).lower() != 'nan':
                        handle_name = name_val
                    if phone_val:
                        handle_phone = clean_phone_helper(phone_val)

        # 2. If still Unknown, search in tracking trips for postOffice info
        if handle_name == "Unknown Name" or handle_phone == "Unknown Phone":
            for trip in trips:
                po_info = trip.get("postOffice")
                if po_info and po_info.get("postcode", "").upper() == responsible_handle:
                    if handle_name == "Unknown Name" and po_info.get("name"):
                        handle_name = po_info.get("name")
                    if handle_phone == "Unknown Phone" and po_info.get("phone"):
                        handle_phone = clean_phone_helper(po_info.get("phone"))
                    break
                    
        # 3. If still Unknown, fall back to last scan operator's info
        if handle_name == "Unknown Name" or handle_phone == "Unknown Phone":
            upd = latest.get("updatedBy")
            if upd:
                if handle_name == "Unknown Name" and upd.get("name"):
                    handle_name = upd.get("name")
                if handle_phone == "Unknown Phone" and upd.get("phone"):
                    handle_phone = clean_phone_helper(upd.get("phone"))

        next_step = ""
        stage = "Unknown"
        # Undefined / unregistered (-99, 0, 100) are treated same as pickup stage
        registered_codes = ["-99", "0", "100"]
        if status in pickup_codes or status in registered_codes:
            stage = "📥 វគ្គទទួលអីវ៉ាន់ (Pickup Stage)" if use_khmer else "📥 Pickup Stage"
            stage = "📥 វគ្គទទួលអីវ៉ាន់ (Pickup Stage)" if use_khmer else "📥 Pickup Stage"
            next_step = (
                f"ការបញ្ជាទិញកំពុងរង់ចាំការទទួលអីវ៉ាន់។\n👉 **សកម្មភាពដែលត្រូវធ្វើ:** បួគុកដែលទទួលខុសត្រូវ (**{responsible_handle}**) ត្រូវតែស្កេន និងទទួលយកកញ្ចប់អីវ៉ាន់។"
                if use_khmer else
                f"The order is waiting to be picked up.\n👉 **Action needed:** The responsible post office (**{responsible_handle}**) must scan and pick up the parcel."
            )
        elif status == "400":
            stage = "🚚 វគ្គដឹកជញ្ជូន (Delivery Stage)" if use_khmer else "🚚 Delivery Stage"
            next_step = (
                f"ការបញ្ជាទិញត្រូវបានចាត់ឲ្យបុគ្គលិកដឹកជញ្ជូន ប៉ុន្តែមិនទាន់បញ្ចប់។\n👉 **សកម្មភាពដែលត្រូវធ្វើ:** បុគ្គលិកដឹកជញ្ជូនដែលបានចាត់តាំងត្រូវបើក Delivery tab/function ហើយស្កេន ឬធ្វើបច្ចុប្បន្នភាពការដឹកជញ្ជូនឲ្យបានបញ្ចប់។"
                if use_khmer else
                "The order has been assigned to a delivery staff but is not completed.\n👉 **Action needed:** The assigned delivery staff must open the Delivery tab/function, then scan or update the order after delivery is completed."
            )
        elif status in delivery_codes:
            stage = "🚚 វគ្គដឹកជញ្ជូន (Delivery Stage)" if use_khmer else "🚚 Delivery Stage"
            next_step = (
                f"ការបញ្ជាទិញកំពុងស្ថិតក្នុងការដឹកជញ្ជូន ឬនៅហាង។\n👉 **សកម្មភាពដែលត្រូវធ្វើ:** អ្នកទទួល/អតិថិជនត្រូវមកទទួលកញ្ចប់អីវ៉ាន់ និង **ស្កេនកូដ QR** លើកម្មវិធី mExpress App ដើម្បីបង់ថ្លៃដឹកជញ្ជូន/បញ្ចប់ការដឹកជញ្ជូន។"
                if use_khmer else
                f"The order is out for delivery or at the store.\n👉 **Action needed:** The receiver/customer needs to pick up the parcel and **scan the QR code** on the mExpress App to pay the fee/complete the delivery."
            )
        elif status in pending_codes:
            stage = "⏳ វគ្គបញ្ជូន/រង់ចាំ (Pending/Transit Stage)" if use_khmer else "⏳ Pending/Transit Stage"
            next_step = (
                f"ការបញ្ជាទិញកំពុងស្ថិតក្នុងការរង់ចាំ កំពុងបញ្ជូន ឬពន្យារពេល។\n👉 **សកម្មភាពដែលត្រូវធ្វើ:** បួគុកដែលទទួលខុសត្រូវ (**{responsible_handle}**) ត្រូវតែស្កេនកញ្ចប់អីវ៉ាន់ ដើម្បីធ្វើបច្ចុប្បន្នភាពស្ថានភាពបញ្ជូន ឬដោះស្រាយការផ្អាកដឹកជញ្ជូន។"
                if use_khmer else
                f"The order is currently pending, in transit, or delayed.\n👉 **Action needed:** The responsible post office (**{responsible_handle}**) must scan the package to update its transit status or resolve the delivery hold."
            )
        elif status == "410":
            stage = "✅ បានប្រគល់ជោគជ័យ (Successfully Delivered)" if use_khmer else "✅ Successfully Delivered"
            next_step = (
                "🎉 ការបញ្ជាទិញត្រូវបានប្រគល់ជូនដោយជោគជ័យ និងបានបញ្ចប់! មិនត្រូវការស្កេនបន្ថែមទៀតទេ។"
                if use_khmer else
                "🎉 The order has been successfully delivered and completed! No further scanning is needed."
            )
        elif status in {"201", "520"}:
            stage = "✅ បានបិទ/បញ្ចប់ (Closed/Done)" if use_khmer else "✅ Closed / Done"
            next_step = (
                "ការបញ្ជាទិញត្រូវបានបិទ/បញ្ចប់។ មិនត្រូវការស្កេនដឹកជញ្ជូនបន្ថែមទៀតទេ។"
                if use_khmer else
                "The order is closed/done. No further delivery scan is needed."
            )
        else:
            stage = f"ស្ថានភាព {status}" if use_khmer else f"Status {status}"
            next_step = (
                f"ការបញ្ជាទិញបច្ចុប្បន្នស្ថិតក្នុងស្ថានភាព: *{status_name}*។\n👉 **សកម្មភាពដែលត្រូវធ្វើ:** ពិនិត្យជាមួយបួគុកដែលទទួលខុសត្រូវ (**{responsible_handle}**) ដើម្បីបញ្ជាក់ថាតើត្រូវស្កេនដែរឬទេ។"
                if use_khmer else
                f"The order is currently in status: *{status_name}*.\n👉 **Action needed:** Check with the responsible post office (**{responsible_handle}**) to verify if scanning is required."
            )

        # Translations
        T = {
            "order_tracking": "ការតាមដានការបញ្ជាទិញ" if use_khmer else "Order tracking",
            "current_stage": "វគ្គបច្ចុប្បន្ន" if use_khmer else "Current Stage",
            "current_status": "ស្ថានភាពបច្ចុប្បន្ន" if use_khmer else "Current Status",
            "current_po": "បច្ចុប្បន្ននៅបួគុក" if use_khmer else "Current Post Office",
            "latest_tracking": "Tracking ចុងក្រោយ" if use_khmer else "Latest Tracking",
            "tracking_time": "ពេលវេលា" if use_khmer else "Time",
            "tracking_by": "អ្នកធ្វើ" if use_khmer else "By",
            "tracking_note": "កំណត់ចំណាំ" if use_khmer else "Note",
            "order_flow": "វឌ្ឍនភាពលំហូរការងារ" if use_khmer else "Order Flow Progress",
            "next_flow": "លំហូរបន្ទាប់" if use_khmer else "Next Flow",
            "meaning": "ន័យស្ថានភាព" if use_khmer else "Meaning",
            "expected_next": "ស្ថានភាពបន្ទាប់ដែលរំពឹង" if use_khmer else "Expected next status",
            "who_scan": "តើនរណាត្រូវស្កេន/ចាត់ចែងការងារនេះ?" if use_khmer else "Who needs to scan/handle this?",
            "resp_po": "បួគុកទទួលខុសត្រូវ" if use_khmer else "Responsible Post Office",
            "phone": "លេខទូរស័ព្ទ" if use_khmer else "Phone",
            "app_action": "ជំហានក្នុង App" if use_khmer else "App action",
            "actor": "អ្នកធ្វើ" if use_khmer else "Actor",
            "app": "App" if use_khmer else "App",
            "tab": "Tab" if use_khmer else "Tab",
            "function": "Function" if use_khmer else "Function",
            "receiver": "អ្នកទទួល" if use_khmer else "Receiver",
            "assigned_delivery_staff": "បុគ្គលិកដឹកជញ្ជូនដែលបានចាត់តាំង" if use_khmer else "Assigned delivery staff",
            "what_to_do": "ត្រូវធ្វើ" if use_khmer else "What to do",
            "next_step_title": "ជំហានបន្ទាប់ដើម្បីជោគជ័យ" if use_khmer else "Next Step to Success",
            "pending": "រង់ចាំ" if use_khmer else "Pending",
            "created": "បានបង្កើត" if use_khmer else "Created",
            "picked_up": "បានទទួល & កំពុងបញ្ជូន" if use_khmer else "Picked Up & In Transit",
            "rec_store": "ដល់បណ្តាញទទួល/ភ្នាក់ងារ" if use_khmer else "At Receiving Store/Agent",
            "orig_hub": "ដល់មជ្ឈមណ្ឌលដើមទី" if use_khmer else "At Origin Hub",
            "dest_store": "ដល់បណ្តាញចែកចាយ/ភ្នាក់ងារ" if use_khmer else "At Destination Store/Agent",
            "delivered": "បានប្រគល់ & ជោគជ័យ" if use_khmer else "Delivered & Successful",
        }

        # Helpers
        def format_time(ts_str):
            if not ts_str: return ""
            try:
                clean_ts = ts_str.split("+")[0]
                dt = datetime.strptime(clean_ts.split(".")[0], "%Y-%m-%dT%H:%M:%S")
                return dt.strftime("%d/%m/%Y %H:%M:%S")
            except Exception:
                return ts_str

        def format_epoch_ms(epoch_ms):
            if not epoch_ms: return ""
            try:
                from datetime import timezone
                dt = datetime.fromtimestamp(epoch_ms / 1000.0, tz=timezone.utc)
                dt_local = dt.astimezone(timezone(timedelta(hours=7)))
                return dt_local.strftime("%d/%m/%Y %H:%M:%S")
            except Exception:
                return ""

        def is_hub_postcode(pc):
            if not pc: return False
            pc = str(pc).upper()
            return "MEGA" in pc or "HUB" in pc or "DVC" in pc

        def get_trip_user(t):
            upd = t.get("updatedBy", {})
            name = upd.get("name") or ""
            ho = t.get("handoverInfo", {})
            staff_code = ""
            if ho:
                staff = ho.get("handoverStaff", {})
                staff_cre = ho.get("handoverStaffCreation", {})
                if staff.get("code"):
                    staff_code = staff.get("code")
                elif staff_cre.get("code"):
                    staff_code = staff_cre.get("code")
            if not name:
                name = t.get("shipperName") or ""
            if staff_code and name:
                return f"{staff_code} - {name}"
            return name

        def get_trip_phone(t):
            upd = t.get("updatedBy", {})
            phone = upd.get("phone") or upd.get("mobile") or upd.get("telephone") or ""
            return clean_phone_helper(phone) if phone else ""

        def get_trip_note(t):
            for key in ("desc", "description", "remark", "message", "statusDesc", "content", "reason", "note"):
                value = t.get(key)
                if value and str(value).strip() and str(value).strip() != status_name:
                    return str(value).strip()

            ignored = {
                str(t.get("status", "")).strip(),
                str(t.get("statusName", "")).strip(),
                str(t.get("postcode", "")).strip(),
                str(t.get("deliveryPostcode", "")).strip(),
                str(t.get("updatedAt", "")).strip(),
            }

            def iter_strings(value):
                if isinstance(value, str):
                    yield value
                elif isinstance(value, dict):
                    for child in value.values():
                        yield from iter_strings(child)
                elif isinstance(value, list):
                    for child in value:
                        yield from iter_strings(child)

            for text in iter_strings(t):
                clean = text.strip()
                if not clean or clean in ignored:
                    continue
                lowered = clean.lower()
                if any(word in lowered for word in ("assign", "deliver", "complete", "pickup", "receive", "transit")):
                    return clean
            return ""

        chrono_trips = trips[::-1]
        
        m1 = None  # Created
        m2 = None  # Picked up & In Transit
        m3 = None  # At Receiving Store/Agent
        m4 = None  # At Origin Hub
        m5 = None  # At Destination Store/Agent
        m6 = None  # Delivered & Successful
        
        # Milestone 1: Created (S110/S100 trip or creation date from order data)
        for t in chrono_trips:
            st = t.get("status", "").lstrip("S")
            if st in ("110", "100") or t.get("is_synthetic"):
                m1 = {
                    "updatedAt": format_epoch_ms(t.get("_created_epoch", 0)) or t.get("updatedAt", ""),
                    "is_fallback": True
                } if t.get("is_synthetic") else t
                break
        if not m1 and order_data:
            created_date_epoch = order_data.get("order_created_date") or order_data.get("created_at") or 0
            if created_date_epoch:
                m1 = {
                    "updatedAt": format_epoch_ms(created_date_epoch),
                    "is_fallback": True
                }
        if not m1 and chrono_trips:
            m1 = chrono_trips[0]

        # Milestone 2: Picked up & In Transit (status 200, 210, 120)
        for t in chrono_trips:
            st = t.get("status", "").lstrip("S")
            if st in ("200", "210", "120"):
                m2 = t
                break

        # Milestone 3: At Receiving Store/Agent (status 302, 310)
        for t in chrono_trips:
            st = t.get("status", "").lstrip("S")
            if st in ("302", "310"):
                m3 = t
                break

        # Milestone 4: At Origin Hub (status 306 at hub postcode)
        for t in chrono_trips:
            st = t.get("status", "").lstrip("S")
            pc = t.get("postcode", "")
            if st == "306" and is_hub_postcode(pc):
                m4 = t
                break

        # Milestone 5: At Destination Store/Agent (status 306 or 309 at non-hub postcode)
        dest_pc = order_data.get("delivery_post_code") if order_data else None
        if not dest_pc and chrono_trips:
            dest_pc = chrono_trips[0].get("deliveryPostcode")
            
        for t in chrono_trips:
            st = t.get("status", "").lstrip("S")
            pc = t.get("postcode", "")
            if st in ("306", "309", "311", "300", "302", "420", "430", "400"):
                if dest_pc and str(pc).upper() == str(dest_pc).upper():
                    if not m5:
                        m5 = t
                        break
                elif not is_hub_postcode(pc):
                    if m4:
                        t_time = t.get("updatedAt")
                        m4_time = m4.get("updatedAt")
                        if t_time and m4_time and t_time > m4_time:
                            if not m5:
                                m5 = t
                                break
                    else:
                        if not m5:
                            m5 = t
                            break

        # Milestone 6: Delivered & Successful
        for t in chrono_trips:
            st = t.get("status", "").lstrip("S")
            if st in done_codes:
                m6 = t
                break

        def format_milestone(m, label, num):
            if not m:
                return f"⏳ {num}. {label}: {T['pending']}"
            if m.get("is_fallback"):
                time_str = m.get("updatedAt")
                return f"✅ {num}. {label}: {time_str}"
            time_str = format_time(m.get("updatedAt"))
            po = m.get("postcode")
            user = get_trip_user(m)
            po_str = f" at {po}" if po else ""
            user_str = f" by {user}" if user else ""
            return f"✅ {num}. {label}: {time_str}{po_str}{user_str}"

        timeline = [
            format_milestone(m1, T["created"], 1),
            format_milestone(m2, T["picked_up"], 2),
            format_milestone(m3, T["rec_store"], 3),
            format_milestone(m4, T["orig_hub"], 4),
            format_milestone(m5, T["dest_store"], 5),
            format_milestone(m6, T["delivered"], 6),
        ]

        latest_time = format_time(latest.get("updatedAt", ""))
        latest_user = get_trip_user(latest)
        latest_phone = get_trip_phone(latest)
        latest_note = get_trip_note(latest)

        # Extract shipper, consignee and payment details if available
        shipper_info = ""
        consignee_info = ""
        payment_info = ""
        receiver_name = "Unknown Receiver"
        receiver_phone = "Unknown Phone"
        assigned_delivery_name = "Unknown delivery staff"
        assigned_delivery_phone = "Unknown Phone"

        def extract_assigned_delivery_staff(tracking_trips):
            pattern = re.compile(
                r"Assign\s+(.+?)\s*[-–—]\s*([+0-9][+0-9\s().-]*)\s+to\s+deliver",
                re.IGNORECASE,
            )

            def iter_strings(value):
                if isinstance(value, str):
                    yield value
                elif isinstance(value, dict):
                    for child in value.values():
                        yield from iter_strings(child)
                elif isinstance(value, list):
                    for child in value:
                        yield from iter_strings(child)

            for trip in tracking_trips:
                for text in iter_strings(trip):
                    match = pattern.search(text)
                    if match:
                        return match.group(1).strip(), clean_phone_helper(match.group(2))
            return assigned_delivery_name, assigned_delivery_phone

        if status == "400":
            assigned_delivery_name, assigned_delivery_phone = extract_assigned_delivery_staff(trips)

        if order_data:
            s_name = order_data.get("shipper", {}).get("name") or "N/A"
            s_phone = order_data.get("shipper", {}).get("phone") or "N/A"
            c_name = order_data.get("consignee", {}).get("name") or "N/A"
            c_phone = order_data.get("consignee", {}).get("phone") or "N/A"
            
            def clean_phone(ph):
                if not ph: return "N/A"
                ph_str = str(ph).strip()
                if ph_str.startswith("+855"):
                    ph_str = "0" + ph_str[4:]
                return ph_str
                
            s_phone = clean_phone(s_phone)
            c_phone = clean_phone(c_phone)
            if c_name and c_name != "N/A":
                receiver_name = c_name
            if c_phone and c_phone != "N/A":
                receiver_phone = c_phone
            
            serv = order_data.get("service_name") or order_data.get("service_code") or "N/A"
            w = order_data.get("calculated_weight") or 0.0
            cod_val = order_data.get("cod_money") or 0.0
            fee_val = order_data.get("total_fees") or 0.0
            payer_val = order_data.get("payer_type") or "N/A"
            
            # Format weight: if >= 1000g, format as kg, else g
            w_str = f"{w/1000.0:.2f} kg" if w >= 1000 else f"{w:.0f} g"
            
            if use_khmer:
                shipper_info = f"👤 **អ្នកផ្ញើ:** {s_name} ({s_phone})"
                consignee_info = f"👤 **អ្នកទទួល:** {c_name} ({c_phone})"
                payment_info = f"💰 **សេវាកម្ម:** {serv} | **ទម្ងន់:** {w_str}\n💵 **COD:** {cod_val} USD | **ថ្លៃសេវា:** {fee_val} USD ({payer_val})"
            else:
                shipper_info = f"👤 **Sender:** {s_name} ({s_phone})"
                consignee_info = f"👤 **Consignee:** {c_name} ({c_phone})"
                payment_info = f"💰 **Service:** {serv} | **Weight:** {w_str}\n💵 **COD:** {cod_val} USD | **Fee:** {fee_val} USD ({payer_val})"

        def md_escape(text):
            return str(text).replace("\\", "\\\\").replace("_", "\\_").replace("*", "\\*").replace("`", "\\`")

        qr_url = f"https://api.qrserver.com/v1/create-qr-code/?size=350x350&data={order_id}"
        out_lines = [
            f"📦 **{T['order_tracking']}: `{md_escape(order_id)}`**  ([📱 View QR to Scan]({qr_url}))",
        ]
        if shipper_info:
            out_lines.append(shipper_info)
        if consignee_info:
            out_lines.append(consignee_info)
        if payment_info:
            out_lines.append(payment_info)

        out_lines.extend([
            f"📊 **{T['current_stage']}:** {stage}",
            f"📍 **{T['current_status']}:** {md_escape(status_name)} ({md_escape(status)})",
            f"🏢 **{T['current_po']}:** `{md_escape(po)}`\n",
        ])

        out_lines.append(f"📈 **{T['order_flow']}:**\n" + "\n".join(timeline) + "\n")

        out_lines.append(f"🏢 **{T['resp_po']}:** `{md_escape(responsible_handle)}` ({md_escape(handle_name)}) - 📞 {md_escape(handle_phone)}")

        if app_action:
            actor = app_action.get("actor", "")
            actor_lower = actor.lower()
            if "delivery staff" in actor_lower or "courier" in actor_lower or "driver" in actor_lower:
                out_lines.append(f"🚚 **{T['assigned_delivery_staff']}:** {md_escape(assigned_delivery_name)} - 📞 {md_escape(assigned_delivery_phone)}")

        out_lines.append("")
        out_lines.append(f"📋 **{T['next_step_title']}:**\n{md_escape(next_step)}")

        await edit_or_send_requester_text(msg, update, context, "\n".join(out_lines), parse_mode="Markdown")

    except Exception as e:
        log.exception("Error in /ask")
        await edit_or_send_requester_text(
            msg, update, context,
            f"❌ បរាជ័យក្នុងការវិភាគការតាមដាន: {e}" if use_khmer else f"❌ Tracking analysis failed: {e}"
        )


@pm_required_handler
async def cmd_check(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/check <order_id1>, <order_id2>, ... — check scans and status details of multiple order IDs."""
    await delete_group_command(update, context)
    
    # Parse arguments
    raw_args = context.args or []
    # If the user pasted a single long string or multiple lines, let's extract all numbers
    full_text = " ".join(raw_args)
    
    # Check if the message is a reply to another message containing IDs
    if not full_text and update.message and update.message.reply_to_message:
        reply = update.message.reply_to_message
        full_text = reply.text or reply.caption or ""
        
    # Extract digit sequences of length >= 8
    order_ids = []
    for piece in re.split(r'[\s,;\n]+', full_text):
        piece = piece.strip()
        if piece.isdigit() and len(piece) >= 8:
            if piece not in order_ids:
                order_ids.append(piece)
                
    if not order_ids:
        await private_or_current_reply(
            update,
            context,
            "Usage: `/check <order_id1> <order_id2> ...`\n"
            "Example: `/check 3003568063 3003568385`"
        )
        return
        
    # Limit to 50 IDs to avoid rate limiting
    if len(order_ids) > 50:
        await private_or_current_reply(
            update,
            context,
            f"⚠️ Too many order IDs! Checked first 50 out of {len(order_ids)}."
        )
        order_ids = order_ids[:50]
        
    msg = await send_requester_text(
        update, context,
        f"🔍 Querying {len(order_ids)} orders from Metfone API in parallel..."
    )
    
    cfg = load_config()
    token = cfg["api"]["bearer_token"]
    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/json, text/plain, */*",
        "User-Agent": "Mozilla/5.0",
    }
    
    import requests
    # Query in parallel with concurrency semaphore
    sem = asyncio.Semaphore(5)
    
    async def fetch_order_info(oid):
        async with sem:
            await asyncio.sleep(0.1) # Stagger slightly
            
            # Fetch tracking
            url_track = "https://gw-express.metfone.com.kh/tms-tracking/api/v1/order-tracking"
            url_search = "https://gw-express.metfone.com.kh/tms-receiving/api/v1/orders/search"
            
            status_name = "Unknown"
            cod = 0.0
            fee = 0.0
            s402_scans = []
            s410_scans = []
            branch_code = ""
            commune_en = ""
            commune_kh = ""
            
            try:
                # 1. Search detail info
                r_search = await asyncio.to_thread(
                    requests.get, url_search, params={"order_code": oid}, headers=headers, timeout=15
                )
                if r_search.status_code == 200:
                    sdata = r_search.json()
                    cod = sdata.get("cod_money") or 0.0
                    fee = sdata.get("total_fees") or 0.0
                    branch_code = str(sdata.get("delivery_post_code") or sdata.get("post_code") or "").strip()
                    # Mapped commune
                    consignee_addr = sdata.get("consignee", {}).get("address", {})
                    for comp in consignee_addr.get("components", []):
                        if comp.get("type") in ("VILLAGE", "WARD"):
                            commune_en = comp.get("name")
                            commune_kh = comp.get("name")
                
                # 2. Tracking details
                r_track = await asyncio.to_thread(
                    requests.get, url_track, params={"order_id": oid}, headers=headers, timeout=15
                )
                if r_track.status_code == 200:
                    tdata = r_track.json()
                    trips = tdata.get("trackingTrips", [])
                    if trips:
                        latest = trips[0]
                        status_name = latest.get("statusName", "Unknown")
                        
                        # Find S402 and S410 details
                        # trips list is descending in time
                        for t in trips:
                            status = str(t.get("status", "")).lstrip("S")
                            user_name = t.get("updatedBy", {}).get("name") or t.get("shipperName") or ""
                            user_phone = t.get("updatedBy", {}).get("phone") or ""
                            user_info = f"{user_name} ({user_phone})" if user_phone else user_name
                            
                            ts = t.get("updatedAt", "")
                            formatted_ts = ts
                            if ts:
                                try:
                                    clean_ts = ts.split("+")[0]
                                    dt = datetime.strptime(clean_ts.split(".")[0], "%Y-%m-%dT%H:%M:%S")
                                    formatted_ts = dt.strftime("%d/%m/%Y %H:%M")
                                except Exception:
                                    pass
                            
                            if status == "402":
                                s402_scans.append(f"{formatted_ts} by {user_info}")
                            elif status == "410":
                                s410_scans.append(f"{formatted_ts} by {user_info}")
                                
            except Exception as e:
                status_name = f"Error: {e}"
                
            prov_kh = ""
            dist_kh = ""
            if branch_code:
                # Strip department code to get province prefix
                b_code = re.sub(r'^[A-Z]{3,5}P?\d*', '', branch_code).strip()
                if len(branch_code) >= 3:
                    prefix = branch_code[:3].upper()
                    prov_kh, _, dist_kh, _ = _map_to_administrative_division(prefix, commune_en, commune_kh)
                    
            # Join multiple scans if they exist (newest first)
            s402_str = "; ".join(s402_scans) if s402_scans else ""
            s410_str = "; ".join(s410_scans) if s410_scans else ""
            
            return {
                "Order ID": oid,
                "Status": status_name,
                "Province": prov_kh,
                "District": dist_kh,
                "COD": cod,
                "Fee": fee,
                "S402 Scans": s402_str,
                "S410 Scans": s410_str
            }

    tasks = [fetch_order_info(oid) for oid in order_ids]
    results = await asyncio.gather(*tasks)
    
    # Send text summary of the orders
    summary_lines = [f"📦 **Scan & Commission Summary ({len(order_ids)} orders):**\n"]
    for r in results:
        oid = r["Order ID"]
        status = r["Status"]
        s402 = r["S402 Scans"]
        s410 = r["S410 Scans"]
        
        status_emoji = "✅" if s410 else "🚚" if s402 else "⏳"
        line = f"{status_emoji} `{oid}`: {status}"
        if r["Province"] or r["District"]:
            line += f" ({r['Province']} - {r['District']})"
        if s402:
            # Split multiple scans for clean bullet points
            for scan in r["S402 Scans"].split("; "):
                line += f"\n   • 🚚 S402: {scan}"
        if s410:
            for scan in r["S410 Scans"].split("; "):
                line += f"\n   • ✅ S410: {scan}"
        summary_lines.append(line)
        summary_lines.append("")
        
    final_text = "\n".join(summary_lines)
    if len(final_text) > 4000:
        final_text = final_text[:4000] + "\n... (Summary truncated)"
        
    await edit_or_send_requester_text(msg, update, context, final_text)
    
    # Build Excel document for download
    try:
        import pandas as pd
        df_out = pd.DataFrame(results)
        
        tmpdir = tempfile.mkdtemp(prefix="check_")
        stamp = datetime.now().strftime("%d.%m_%HH%M")
        filename = f"CheckOrders_Details_{stamp}.xlsx"
        out_path = os.path.join(tmpdir, filename)
        
        # Write to Excel beautifully
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        DARK_BLUE = "172033"
        WHITE = "FFFFFF"
        BORDER_CLR = "CCCCCC"
        
        thin_border = Border(
            left=Side(style="thin", color=BORDER_CLR),
            right=Side(style="thin", color=BORDER_CLR),
            top=Side(style="thin", color=BORDER_CLR),
            bottom=Side(style="thin", color=BORDER_CLR),
        )
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Order Scans"
        ws.views.sheetView[0].showGridLines = True
        
        headers_excel = list(results[0].keys())
        for col_idx, col_name in enumerate(headers_excel, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = Font(name="Calibri", bold=True, color=WHITE, size=11)
            cell.fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        ws.row_dimensions[1].height = 28
        
        for row_idx, row_data in enumerate(results, 2):
            for col_idx, key in enumerate(headers_excel, 1):
                val = row_data[key]
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = Font(name="Calibri", size=10)
                cell.border = thin_border
                
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 50)
            
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers_excel))}{len(results)+1}"
        
        wb.save(out_path)
        
        # Send Document
        with open(out_path, "rb") as f:
            await send_requester_document(update, context, f, filename)
            
    except Exception as e:
        log.exception("Failed to build Excel for /check")
        await send_requester_text(update, context, f"❌ Failed to build Excel file: {e}")


@pm_required_handler
async def cmd_trace(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/trace <bill_id> — diagnose and track errors/status for a specific bill ID."""
    await delete_group_command(update, context)
    
    args = [a.strip() for a in (context.args or []) if a.strip()]
    bill_id = " ".join(args)
    
    if not bill_id and update.message and update.message.reply_to_message:
        reply = update.message.reply_to_message
        bill_id = (reply.text or reply.caption or "").strip()
        
    # Extract the first digit sequence of length >= 8 if it's a long text
    if bill_id:
        match = re.search(r'\d{8,}', bill_id)
        if match:
            bill_id = match.group(0)

    # Clean up non-digits just in case
    bill_id = "".join(c for c in bill_id if c.isdigit())
    
    if not bill_id:
        await private_or_current_reply(
            update,
            context,
            "Usage: `/trace <bill_id>`\n"
            "Example: `/trace 3003568063`"
        )
        return
        
    msg = await send_requester_text(
        update, context,
        f"🔍 Initiating trace for Bill ID `{bill_id}`..."
    )
    
    trace_results = []
    trace_results.append(f"🔍 **Trace Report for Bill ID:** `{bill_id}`\n")
    
    # ── 1. Check ignore / test lists ──
    trace_results.append("📋 **1. Settings & Config Check:**")
    is_ignored = False
    if os.path.exists("test_bills.txt"):
        try:
            with open("test_bills.txt", "r", encoding="utf-8") as f:
                for line in f:
                    if line.strip() == bill_id:
                        is_ignored = True
                        break
        except Exception as e:
            trace_results.append(f"   • ❌ Failed to read `test_bills.txt`: {e}")
            
    if is_ignored:
        trace_results.append(f"   • ⚠️ Listed in `test_bills.txt` (This bill is marked as TEST/IGNORED).")
    else:
        trace_results.append(f"   • ✅ Not listed in `test_bills.txt` (Active/not ignored).")
        
    # Check delayed_bills.json
    is_delayed = False
    if os.path.exists("delayed_bills.json"):
        try:
            with open("delayed_bills.json", "r", encoding="utf-8") as f:
                delayed = json.load(f)
                if bill_id in delayed:
                    is_delayed = True
                    trace_results.append(f"   • ⚠️ Listed in `delayed_bills.json` (Delayed until: `{delayed[bill_id]}`).")
        except Exception as e:
            trace_results.append(f"   • ❌ Failed to read `delayed_bills.json`: {e}")
            
    if not is_delayed:
        trace_results.append("   • ✅ Not listed in `delayed_bills.json` (No delay filter applied).")
    
    trace_results.append("")
    
    # ── 2. Search local Excel cache ──
    trace_results.append("📂 **2. Local Excel Cache Search:**")
    found_in_cache = False
    cache_file = os.path.join(HERE, "cache", "latest_detail.xlsx")
    if os.path.exists(cache_file):
        try:
            import pandas as pd
            xl = pd.ExcelFile(cache_file)
            for sheet in xl.sheet_names:
                df = xl.parse(sheet)
                order_col = None
                for col in df.columns:
                    if "order" in str(col).lower() or "id" in str(col).lower():
                        order_col = col
                        break
                if not order_col and len(df.columns) > 3:
                    for col in df.columns:
                        if df[col].astype(str).str.contains(bill_id, na=False).any():
                            order_col = col
                            break
                if order_col is not None:
                    matches = df[df[order_col].astype(str).str.strip() == bill_id]
                    if not matches.empty:
                        found_in_cache = True
                        row = matches.iloc[0]
                        trace_results.append(f"   • ✅ Found in cached Excel (`{sheet}` sheet):")
                        status = row.get("CURRENT STATUS") or row.get("Current Status") or row.get("Trạng thái hiện tại") or "N/A"
                        po = row.get("CURRENT POST OFFICE") or row.get("Current Post Office") or row.get("Bưu cục hiện tại") or "N/A"
                        sender = row.get("SENDER") or row.get("Sender") or "N/A"
                        receiver = row.get("RECEIVER") or row.get("Receiver") or "N/A"
                        trace_results.append(f"     - Status: `{status}`")
                        trace_results.append(f"     - Post Office: `{po}`")
                        trace_results.append(f"     - Sender: `{sender}` | Receiver: `{receiver}`")
                        break
            if not found_in_cache:
                trace_results.append("   • ℹ️ Bill ID not found in the latest cached Excel data.")
        except Exception as e:
            trace_results.append(f"   • ❌ Error reading cached Excel: {e}")
    else:
        trace_results.append("   • ℹ️ No cached Excel data found.")
        
    trace_results.append("")
    
    # ── 3. Live API Diagnostics ──
    trace_results.append("⚡ **3. Live TMS API Diagnostics:**")
    cfg = load_config()
    token = cfg.get("api", {}).get("bearer_token")
    if not token:
        trace_results.append("   • ❌ API Token is missing in config.json")
    else:
        import requests
        headers = {
            "Authorization": f"Bearer {token}",
            "Accept": "application/json, text/plain, */*",
            "User-Agent": "Mozilla/5.0",
        }
        
        # Live Search call
        url_search = "https://gw-express.metfone.com.kh/tms-receiving/api/v1/orders/search"
        trace_results.append("   *Querying Search API...*")
        try:
            r_search = await asyncio.to_thread(
                requests.get, url_search, params={"order_code": bill_id}, headers=headers, timeout=15
            )
            trace_results.append(f"   • Status: HTTP `{r_search.status_code}`")
            if r_search.status_code == 200:
                sdata = r_search.json()
                cod = sdata.get("cod_money") or 0.0
                fee = sdata.get("total_fees") or 0.0
                bp = sdata.get("delivery_post_code") or sdata.get("post_code") or "N/A"
                trace_results.append(f"     - COD: `{cod}` | Fee: `{fee}` | Post Code: `{bp}`")
            else:
                trace_results.append(f"     - Error Response: `{r_search.text[:250]}`")
        except Exception as e:
            trace_results.append(f"     - Search API Connection Failed: `{e}`")
            
        # Live Tracking call
        url_track = "https://gw-express.metfone.com.kh/tms-tracking/api/v1/order-tracking"
        trace_results.append("   *Querying Tracking API...*")
        try:
            r_track = await asyncio.to_thread(
                requests.get, url_track, params={"order_id": bill_id}, headers=headers, timeout=15
            )
            trace_results.append(f"   • Status: HTTP `{r_track.status_code}`")
            if r_track.status_code == 200:
                tdata = r_track.json()
                trips = tdata.get("trackingTrips", [])
                if trips:
                    latest = trips[0]
                    status_name = latest.get("statusName", "Unknown")
                    trace_results.append(f"     - Current Status Name: `{status_name}`")
                    trace_results.append(f"     - Recent Trip Scans (max 3):")
                    for idx, t in enumerate(trips[:3]):
                        status = t.get("status", "")
                        desc = t.get("description", "")
                        user = t.get("updatedBy", {}).get("name") or t.get("shipperName") or "System"
                        ts = t.get("updatedAt") or ""
                        trace_results.append(f"       {idx+1}. [{status}] {desc} (by {user} at {ts})")
                else:
                    trace_results.append("     - No tracking trips found.")
            else:
                trace_results.append(f"     - Error Response: `{r_track.text[:250]}`")
        except Exception as e:
            trace_results.append(f"     - Tracking API Connection Failed: `{e}`")
            
    trace_results.append("")
    
    # ── 4. Log Scan ──
    trace_results.append("📝 **4. Execution Logs (`bot.log`):**")
    log_file = os.path.join(HERE, "bot.log")
    if os.path.exists(log_file):
        try:
            matching_lines = []
            with open(log_file, "r", encoding="utf-8", errors="ignore") as lf:
                for line in lf:
                    if bill_id in line:
                        matching_lines.append(line.strip())
            if matching_lines:
                trace_results.append(f"   • Found {len(matching_lines)} matching log entry/entries:")
                for ml in matching_lines[-15:]:
                    trace_results.append(f"     - `{ml[:150]}`")
            else:
                trace_results.append("   • No logs found matching this Bill ID.")
        except Exception as e:
            trace_results.append(f"   • ❌ Failed to read log file: {e}")
    else:
        trace_results.append("   • ℹ/ No log file `bot.log` exists yet.")
        
    final_text = "\n".join(trace_results)
    if len(final_text) > 4000:
        final_text = final_text[:4000] + "\n... (Trace truncated due to size)"
        
    await edit_or_send_requester_text(msg, update, context, final_text)


@user_guard
async def cmd_qr(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/qr <order_id> - generate a QR code for the order ID to scan on the phone screen."""
    await delete_group_command(update, context)
    args = [a.strip() for a in (context.args or []) if a.strip()]
    if not args:
        await private_or_current_reply(
            update,
            context,
            "Usage: `/qr <order_id>`\n"
            "Example: `/qr 3003516618`"
        )
        return

    order_id = args[0]
    chat_id = requester_chat_id(update)
    if not chat_id:
        return

    qr_url = f"https://api.qrserver.com/v1/create-qr-code/?size=350x350&data={order_id}"
    caption = (
        f"📦 **QR Code for Order:** `{order_id}`\n"
        f"👉 *Scan this directly off your screen using your mExpress App.*"
    )

    try:
        await safe_api_call(
            context.bot.send_photo,
            chat_id=chat_id,
            photo=qr_url,
            caption=caption,
            parse_mode="Markdown"
        )
    except Exception as e:
        log.warning("Could not send QR photo to private chat %s: %s", chat_id, e)
        if is_group_chat(update) and update.effective_chat:
            try:
                await safe_api_call(
                    context.bot.send_photo,
                    chat_id=update.effective_chat.id,
                    photo=qr_url,
                    caption=caption,
                    parse_mode="Markdown",
                )
            except Exception as e2:
                log.warning("Fallback send QR to group failed: %s", e2)
        else:
            await private_or_current_reply(
                update,
                context,
                f"❌ Failed to send QR code image: {e}\n"
                f"🔗 Click here to view QR code: {qr_url}"
            )


# ── Push handler ───────────────────────────────────────────────────────────────

@pm_required_handler
async def run_push(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
    force_test: bool = False,
):
    await delete_group_command(update, context)

    cfg = load_config()
    allowed = cfg["telegram"].get("allowed_chat_ids") or []
    chat_id = update.effective_chat.id
    if allowed and chat_id not in allowed:
        await private_or_current_reply(update, context, "This chat is not allowed to use the bot.")
        return

    paused = is_paused(cfg)
    test_mode = force_test or paused
    if force_test:
        await private_or_current_reply(
            update,
            context,
            "🧪 TEST MODE.\n"
            "Reports will only be sent to you. Nothing will be forwarded to groups."
        )
    elif paused:
        await private_or_current_reply(
            update,
            context,
            "⏸ Bot is paused — running in TEST mode.\n"
            "Reports will only be sent to you (no group forwarding)."
        )

    msg = await send_requester_text(update, context, "Downloading data from OPS...")
    if is_group_chat(update) and msg is None:
        log.warning(
            "Stopping push because private requester messages are unavailable."
        )
        return

    tmpdir = tempfile.mkdtemp(prefix="push_")
    track_report_dir(tmpdir)
    stamp  = datetime.now().strftime("%d.%m_%HH%M")
    src    = os.path.join(tmpdir, f"export_{stamp}.xlsx")

    try:
        # ── Determine zone/mode before download so we fetch only needed branches ──
        raw_args = [arg.strip() for arg in getattr(context, "args", []) if arg.strip()]

        # ── Extract inline remark: `push remark: some text` ──────────────────
        # Supports: `push remark: text`, `push zone remark: text`, etc.
        inline_remark = None
        clean_args = []
        remark_started = False
        remark_parts = []
        for tok in raw_args:
            if remark_started:
                remark_parts.append(tok)
            elif tok.lower().startswith("remark:"):
                # e.g. "remark:Please" or "remark: Please" (next tokens)
                remark_started = True
                suffix = tok[len("remark:"):].strip()
                if suffix:
                    remark_parts.append(suffix)
            elif tok.lower() == "remark":
                # next token should be ":" or text
                remark_started = True
            else:
                clean_args.append(tok)
        if remark_parts:
            inline_remark = " ".join(remark_parts).strip()
        raw_args = clean_args  # remove remark tokens before zone parsing

        # Check for force refresh in arguments
        force_refresh = False
        if "force" in [a.lower() for a in raw_args]:
            force_refresh = True
            raw_args = [a for a in raw_args if a.lower() != "force"]

        zone_key_arg = raw_args[0].lower() if raw_args else None
        total_zones = cfg.get("total_zones", {})
        zone_branches_map = cfg.get("zone_branches", {})

        zone_override_branch = None
        zone_mode = None
        target_handles = []

        if zone_key_arg == "all":
            # "push all" = ALL 23 groups + 5 zones → fetch ALL branches
            all_branches = []
            for zb in zone_branches_map.values():
                for b in zb.split(","):
                    if b.strip() and b.strip() not in all_branches:
                        all_branches.append(b.strip())
            # Also include default branch_code branches
            for b in cfg["api"].get("branch_code", "").split(","):
                if b.strip() and b.strip() not in all_branches:
                    all_branches.append(b.strip())
            zone_override_branch = ",".join(all_branches)
            zone_mode = "ALL"

        elif zone_key_arg == "zone":
            # "push zone" = 5 zone groups only → fetch ALL zone branches
            all_branches = []
            for zb in zone_branches_map.values():
                for b in zb.split(","):
                    if b.strip() and b.strip() not in all_branches:
                        all_branches.append(b.strip())
            zone_override_branch = ",".join(all_branches)
            zone_mode = "ZONE"

        elif zone_key_arg and zone_key_arg in total_zones:
            # "push zone5" = specific zone only
            if zone_key_arg in zone_branches_map:
                zone_override_branch = zone_branches_map[zone_key_arg]
            target_handles = [h.upper() for h in total_zones[zone_key_arg]]
            zone_mode = zone_key_arg.upper()

        else:
            # plain "push" = 23 groups → uses default branch_code from config
            # exclude any leftover remark tokens (already stripped above)
            target_handles = [arg.upper() for arg in raw_args if arg]

        downloader.download_detail(cfg["api"], src, branch_code=zone_override_branch, force_refresh=force_refresh)
        msg = await edit_or_send_requester_text(
            msg, update, context, "Download done. Generating reports..."
        )

        if not os.path.exists(REF_PATH):
            await edit_or_send_requester_text(
                msg, update, context, f"Error: Missing reference file {REF_PATH}"
            )
            return

        if force_test and not target_handles and not zone_mode and is_group_chat(update):
            group_handles = get_forward_mapping(cfg).get(str(update.effective_chat.id), [])
            group_handles = [h.upper() for h in group_handles if h and h != "*"]
            if group_handles:
                target_handles = group_handles

        mode = get_mode(cfg)
        result = generate_report.generate_reports_from_data(
            src, REF_PATH, tmpdir, return_metadata=True, mode=mode, target_handles=target_handles
        )
        update_webapp_cache(result)
        save_highlight_history(result)

        # ── Apply handle filters ──────────────────────────────────────────
        if target_handles:
            result["handle_results"] = [
                hr for hr in result["handle_results"]
                if hr["handle"] in target_handles
            ]

        # Recalculate overall counts after filtering
        if target_handles or zone_mode:
            overall = {"Pickup": 0, "Delivery": 0, "Pending": 0}
            for hr in result["handle_results"]:
                for k in overall:
                    overall[k] += hr["handle_counts"].get(k, 0)
            result["overall_counts"] = overall
            grand_total = sum(overall.values())
            label = f"{zone_mode} " if zone_mode else ""
            result["summary_caption"] = "\n".join([
                f"📋 {label}Daily Report  {datetime.now().strftime('%d/%m/%Y %H:%M')}",
                f"Pickup: {overall['Pickup']}  |  Delivery: {overall['Delivery']}  |  Pending: {overall['Pending']}",
                f"Grand Total: {grand_total}",
            ])

        if inline_remark:
            for hr in result["handle_results"]:
                hr["remark"] += f"  |  Remark: {inline_remark}"
            if "summary_caption" in result:
                result["summary_caption"] += f"\n📝 Remark: {inline_remark}"

        if not result["handle_results"]:
            await edit_or_send_requester_text(
                msg, update, context, "No data found.",
            )
            return

        msg = await edit_or_send_requester_text(msg, update, context, "Sending images...")

        # ── Pre-calculate covered handles for hiding ─────────────────────────
        forward_groups = get_all_forward_groups(cfg)
        forward_mapping = get_forward_mapping(cfg)
        covered_handles = set()
        actual_forward_groups = []

        for group_id_str in forward_groups:
            allowed = forward_mapping.get(str(group_id_str), ["*"])
            wants_all = "*" in allowed
            will_receive = False
            for hr in result["handle_results"]:
                if wants_all or hr["handle"] in allowed:
                    will_receive = True
                    break
            if will_receive:
                actual_forward_groups.append(group_id_str)

        forward_groups = actual_forward_groups
        
        if not test_mode and forward_groups:
            for group_id_str in forward_groups:
                allowed = forward_mapping.get(str(group_id_str), ["*"])
                if "*" in allowed:
                    covered_handles.add("*")
                else:
                    covered_handles.update(allowed)
            # Also cover all handles mapped in any zone so they aren't sent to the requester
            for zone_key, zone_hlist in total_zones.items():
                covered_handles.update([h.upper() for h in zone_hlist if h])

        # ── Send to the requester ─────────────────────────────────────────────
        for hr in result["handle_results"]:
            handle_str = hr["handle"]
            
            # Hide from requester in production mode (only show in test mode)
            if not test_mode:
                continue
                
            for hf in hr["handle_files"]:
                try:
                    img_buf = excel_to_image.excel_to_image(hf["path"])
                    img_buf.name = f"{hr['handle']}.png"
                    await send_requester_photo(update, context, img_buf)
                except Exception as e:
                    log.warning(f"Image render failed {hr['handle']}: {e}")
            await send_requester_text(update, context, hr["remark"])

        with open(result["final_xlsx"], "rb") as f:
            await send_requester_document(
                update,
                context,
                f,
                os.path.basename(result["final_xlsx"]),
            )
        await send_requester_text(update, context, result["summary_caption"])

        # ── Forward to groups only outside test mode ──────────────────────────
        if test_mode:
            if force_test:
                test_text = (
                    f"🧪 TEST MODE — not forwarded to groups.\n"
                    f"Data: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
                )
            else:
                test_text = (
                    f"⏸ TEST MODE — not forwarded to groups.\n"
                    f"Use /resume to enable group forwarding.\n"
                    f"Data: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
                )
            await edit_or_send_requester_text(msg, update, context, test_text)
            return

        # ── Zone group forwarding (push zone / push all) ────────────────────
        zone_fwd_map = cfg.get("zone_forward_mapping", {})
        send_to_regular = zone_mode is None or zone_mode == "ALL"
        send_to_zones = zone_mode in ("ZONE", "ALL") or (
            zone_mode and zone_mode.startswith("ZONE") and zone_mode != "ALL"
        )

        # Forward to regular groups (push / push all)
        if send_to_regular and forward_groups:
            context.user_data["pending_forward"] = {
                "result": result,
                "forward_groups": forward_groups,
                "forward_mapping": forward_mapping,
                "created_at": datetime.now().isoformat(timespec="seconds"),
            }
            await edit_or_send_requester_text(
                msg, update, context,
                f"Forwarding to {len(forward_groups)} group(s)...",
            )
            payload = context.user_data.pop("pending_forward", None)
            sent_groups = await forward_result_to_groups(context, payload)
            await send_requester_text(
                update, context,
                f"✅ Forwarded to {sent_groups} group(s).",
            )

        # Forward to zone groups (push zone / push all / push zone5)
        # Each zone gets a summary image + total Excel (same as /total)
        if send_to_zones and zone_fwd_map:
            import pandas as pd
            zone_sent = 0
            for group_id_str, zone_key in zone_fwd_map.items():
                # For "push zone5", only forward to zone5 group
                if zone_mode and zone_mode not in ("ZONE", "ALL"):
                    if zone_key != zone_mode.lower():
                        continue

                zone_handles = [h.upper() for h in total_zones.get(zone_key, [])]
                if not zone_handles:
                    continue

                # Filter handle_results for this zone
                zone_results = [
                    hr for hr in result["handle_results"]
                    if hr["handle"] in zone_handles
                ]
                if not zone_results:
                    continue

                try:
                    group_id = int(group_id_str)
                except ValueError:
                    group_id = group_id_str

                # Calculate zone totals
                zone_overall = {"Pickup": 0, "Delivery": 0, "Pending": 0}
                for hr in zone_results:
                    for k in zone_overall:
                        zone_overall[k] += hr["handle_counts"].get(k, 0)
                zone_grand = sum(zone_overall.values())

                zone_label = zone_key.upper()
                zone_caption = "\n".join([
                    f"📋 {zone_label} Report  {datetime.now().strftime('%d/%m/%Y %H:%M')}",
                    f"Pickup: {zone_overall['Pickup']}  |  Delivery: {zone_overall['Delivery']}  |  Pending: {zone_overall['Pending']}",
                    f"Grand Total: {zone_grand}",
                ])
                if inline_remark:
                    zone_caption += f"\n📝 Remark: {inline_remark}"

                # Build zone-filtered result for total Excel
                zone_result = {**result, "handle_results": zone_results, "overall_counts": zone_overall}
                # Filter type_data DataFrames
                zone_result["type_data"] = {}
                for rn in ["Pickup", "Delivery", "Pending"]:
                    df = result.get("type_data", {}).get(rn)
                    if df is not None and not df.empty:
                        filter_col = "POST OFFICE HANDLE"
                        if filter_col in df.columns:
                            zone_result["type_data"][rn] = df[df[filter_col].isin(zone_handles)].copy()
                        else:
                            zone_result["type_data"][rn] = df.copy()
                    else:
                        zone_result["type_data"][rn] = pd.DataFrame()

                try:
                    # ── Build day_date_counts and urgent_counts for zone image ──
                    zone_day_date_counts = {}
                    zone_urgent_counts   = {}
                    today_date = datetime.now().date()

                    for rn in ["Pickup", "Delivery", "Pending"]:
                        df_z = zone_result["type_data"].get(rn)
                        if df_z is None or df_z.empty:
                            continue
                        date_col_z = result.get("date_col") or (
                            "CREATED DATE" if "CREATED DATE" in df_z.columns else
                            "CURRENT TIME"  if "CURRENT TIME"  in df_z.columns else None
                        )
                        if date_col_z and date_col_z in df_z.columns:
                            parsed_z = pd.to_datetime(df_z[date_col_z], dayfirst=True,
                                                      format="mixed", errors="coerce")
                            df_z = df_z.copy()
                            df_z["_zdate"] = parsed_z.dt.date

                        handle_col = "POST OFFICE HANDLE"
                        if handle_col not in df_z.columns:
                            continue

                        for _, row_z in df_z.iterrows():
                            h = str(row_z.get(handle_col, "")).strip().upper()
                            if not h:
                                continue
                            # date counts
                            d_val = row_z.get("_zdate") if "_zdate" in df_z.columns else None
                            if d_val and not pd.isna(d_val):
                                zone_day_date_counts.setdefault(h, {})
                                zone_day_date_counts[h][d_val] = zone_day_date_counts[h].get(d_val, 0) + 1
                            # urgent = overdue (created > 1 day ago)
                            created_d = None
                            if "CREATED DATE" in df_z.columns:
                                cd = pd.to_datetime(row_z.get("CREATED DATE"), dayfirst=True,
                                                    format="mixed", errors="coerce")
                                if not pd.isna(cd):
                                    created_d = cd.date()
                            if created_d and (today_date - created_d).days > 1:
                                zone_urgent_counts[h] = zone_urgent_counts.get(h, 0) + 1

                    # 1. Summary image
                    img_buf = generate_summary.build_summary_image(
                        zone_results,
                        zone_overall,
                        zone_label=zone_label,
                        day_date_counts=zone_day_date_counts if zone_day_date_counts else None,
                        urgent_counts=zone_urgent_counts if zone_urgent_counts else None,
                    )
                    img_buf.name = f"{zone_key}_summary.png"
                    await safe_api_call(context.bot.send_photo, chat_id=group_id, photo=img_buf)
                    await asyncio.sleep(0.5)

                    # 2. Total Excel
                    zone_xlsx = os.path.join(tmpdir, f"Total_{zone_label}_{stamp}.xlsx")
                    generate_summary.build_total_excel(zone_result, zone_xlsx)
                    with open(zone_xlsx, "rb") as f:
                        await safe_api_call(
                            context.bot.send_document,
                            chat_id=group_id,
                            document=f,
                            filename=os.path.basename(zone_xlsx),
                            caption=zone_caption,
                        )
                    zone_sent += 1
                    await asyncio.sleep(1.0)
                except Exception as e:
                    log.error(f"Zone forward to {group_id} ({zone_key}): {e}")

            if zone_sent:
                await send_requester_text(
                    update, context,
                    f"✅ Forwarded to {zone_sent} zone group(s).",
                )

        if not send_to_regular and not send_to_zones:
            # Specific zone push (push zone5) — already handled above via send_to_zones
            pass

        await edit_or_send_requester_text(
            msg, update, context,
            f"Done. {datetime.now().strftime('%d.%m.%Y %H:%M')}",
        )

        # (Remark prompt removed as requested)

    except Exception as e:
        log.exception("Error in run_push")
        # 1. Try to send the detailed error privately to the user who requested it.
        user_id = requester_chat_id(update)
        if user_id:
            try:
                await safe_api_call(
                    context.bot.send_message,
                    chat_id=user_id,
                    text=f"❌ Error in run_push: {e}"
                )
            except Exception as pm_err:
                log.warning(f"Failed to send private error message to user {user_id}: {pm_err}")
                
        # 2. Clean up or update the status message in the group so there is no lingering status or error message there.
        if msg and is_group_chat(update):
            try:
                await context.bot.delete_message(
                    chat_id=update.effective_chat.id,
                    message_id=msg.message_id
                )
            except Exception as del_err:
                log.warning(f"Failed to delete status message in group: {del_err}")
                # Fallback: Edit it to a simple status so it doesn't look stuck, without leaking the detailed error
                try:
                    await safe_api_call(
                        msg.edit_text,
                        "❌ Push failed."
                    )
                except Exception:
                    pass


@user_guard
async def cmd_test(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Run a private test push without forwarding anything to groups."""
    await run_push(update, context, force_test=True)


@user_guard
async def cmd_add(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Add test bill(s) directly to the ignore list and the configured Excel/CSV file."""
    await delete_group_command(update, context)
    if not context.args:
        # Check if they used it as `/testbill add <id>` fallback
        if update.message and update.message.text and update.message.text.startswith("/testbill"):
            args = context.args[1:] if len(context.args) > 1 else []
            if not args:
                await private_or_current_reply(update, context, "Usage: `/add <bill_id>`", parse_mode="Markdown")
                return
        else:
            await private_or_current_reply(update, context, "Usage: `/add <bill_id1> <bill_id2> ...`", parse_mode="Markdown")
            return
    else:
        if context.args[0].lower() == "add":
            args = context.args[1:]
        else:
            args = context.args

    txt_path = "test_bills.txt"
    current_ids = set()
    if os.path.exists(txt_path):
        with open(txt_path, "r", encoding="utf-8") as f:
            for line in f:
                val = line.strip()
                if val:
                    current_ids.add(val)

    added = []
    already_exist = []
    for bill_id in args:
        bill_id = bill_id.strip()
        if bill_id:
            if bill_id in current_ids:
                already_exist.append(bill_id)
            else:
                current_ids.add(bill_id)
                added.append(bill_id)
    
    if added:
        # 1. Update text file
        with open(txt_path, "w", encoding="utf-8") as f:
            for bid in sorted(current_ids):
                f.write(bid + "\n")
        
        # 2. Update configured Excel/CSV file
        cfg = load_config()
        tr_cfg = cfg.get("test_receipts", {})
        if tr_cfg.get("enabled"):
            path = tr_cfg.get("path")
            if path:
                try:
                    import pandas as pd
                    df_test = None
                    order_col = "ORDER ID"
                    
                    if os.path.exists(path):
                        if path.lower().endswith((".xlsx", ".xls")):
                            df_test = pd.read_excel(path, dtype=str)
                        else:
                            df_test = pd.read_csv(path, dtype=str, keep_default_na=False)
                        
                        df_test = df_test.fillna("")
                        if not df_test.empty:
                            found_col = next(
                                (
                                    c for c in df_test.columns
                                    if "order" in str(c).lower()
                                    or "phi" in str(c).lower()
                                    or "shipment" in str(c).lower()
                                ),
                                df_test.columns[0]
                            )
                            if found_col:
                                order_col = found_col
                    
                    if df_test is None or df_test.empty:
                        df_test = pd.DataFrame(columns=[order_col])
                    
                    # Read current items in the file to avoid duplicates
                    file_ids = set(df_test[order_col].astype(str).str.strip().tolist())
                    new_rows = []
                    for bid in added:
                        if bid not in file_ids:
                            new_rows.append({order_col: bid})
                        else:
                            if bid not in already_exist:
                                already_exist.append(bid)
                    
                    if new_rows:
                        df_new = pd.DataFrame(new_rows)
                        df_test = pd.concat([df_test, df_new], ignore_index=True)
                        
                        if path.lower().endswith((".xlsx", ".xls")):
                            df_test.to_excel(path, index=False)
                        else:
                            df_test.to_csv(path, index=False, encoding="utf-8-sig")
                except Exception as e:
                    log.error(f"Could not update test file at {path}: {e}")

        cfg = load_config()
        all_ids = generate_report.load_test_order_ids(cfg)
        msg_text = f"✅ Added test bills: {', '.join(added)}\n"
        if already_exist:
            msg_text += f"⚠️ Already in the list: {', '.join(already_exist)}\n"
        msg_text += f"Total ignored test bills: {len(all_ids)} (synchronised with your test file)"
        await private_or_current_reply(update, context, msg_text)
    else:
        if already_exist:
            await private_or_current_reply(
                update, 
                context, 
                f"⚠️ All specified bills ({', '.join(already_exist)}) are **already** in your ignore list."
            )
        else:
            await private_or_current_reply(update, context, "No bills were specified.")


@user_guard
async def cmd_remove(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Remove test bill(s) from the ignore list and the configured Excel/CSV file."""
    await delete_group_command(update, context)
    
    if context.args and context.args[0].lower() in ("remove", "delete", "rm", "del"):
        args = context.args[1:]
    else:
        args = context.args

    if not args:
        await private_or_current_reply(update, context, "Usage: `/remove <bill_id>`", parse_mode="Markdown")
        return

    txt_path = "test_bills.txt"
    current_ids = set()
    if os.path.exists(txt_path):
        with open(txt_path, "r", encoding="utf-8") as f:
            for line in f:
                val = line.strip()
                if val:
                    current_ids.add(val)

    removed = []
    for bill_id in args:
        bill_id = bill_id.strip()
        if bill_id in current_ids:
            current_ids.remove(bill_id)
        # We track all arguments as potentially removed from the excel/csv file
        removed.append(bill_id)
            
    # 1. Update text file
    with open(txt_path, "w", encoding="utf-8") as f:
        for bid in sorted(current_ids):
            f.write(bid + "\n")
            
    # 2. Update Excel/CSV file
    cfg = load_config()
    tr_cfg = cfg.get("test_receipts", {})
    file_removed_count = 0
    if tr_cfg.get("enabled"):
        path = tr_cfg.get("path")
        if path and os.path.exists(path):
            try:
                import pandas as pd
                if path.lower().endswith((".xlsx", ".xls")):
                    df_test = pd.read_excel(path, dtype=str)
                else:
                    df_test = pd.read_csv(path, dtype=str, keep_default_na=False)
                
                df_test = df_test.fillna("")
                if not df_test.empty:
                    order_col = next(
                        (
                            c for c in df_test.columns
                            if "order" in str(c).lower()
                            or "phi" in str(c).lower()
                            or "shipment" in str(c).lower()
                        ),
                        df_test.columns[0]
                    )
                    if order_col:
                        initial_len = len(df_test)
                        # Remove the matched IDs
                        df_test = df_test[~df_test[order_col].astype(str).str.strip().isin(removed)].copy()
                        file_removed_count = initial_len - len(df_test)
                        
                        if file_removed_count > 0:
                            if path.lower().endswith((".xlsx", ".xls")):
                                df_test.to_excel(path, index=False)
                            else:
                                df_test.to_csv(path, index=False, encoding="utf-8-sig")
            except Exception as e:
                log.error(f"Could not update test file at {path}: {e}")

    cfg = load_config()
    all_ids = generate_report.load_test_order_ids(cfg)
    await private_or_current_reply(
        update, 
        context, 
        f"❌ Removed test bills from ignore list: {', '.join(removed)}\n"
        f"Total ignored test bills remaining: {len(all_ids)} (removed {file_removed_count} directly from your test file)"
    )


@user_guard
async def cmd_list(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """List all test bills currently ignored by exporting them to an Excel file."""
    await delete_group_command(update, context)
    cfg = load_config()
    current_ids = generate_report.load_test_order_ids(cfg)
                    
    if not current_ids:
        await private_or_current_reply(update, context, "No test bills in the ignore list.")
        return

    import pandas as pd
    import io
    
    # Create Excel in memory
    df = pd.DataFrame(sorted(current_ids), columns=["Ignored Bill ID"])
    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Test Bills")
    bio.seek(0)
    
    msg = await send_requester_text(
        update, 
        context, 
        f"Generating Excel list of ignored test bills ({len(current_ids)})..."
    )
    
    try:
        await send_requester_document(
            update,
            context,
            bio,
            filename=f"ignored_test_bills_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        await edit_or_send_requester_text(
            msg, 
            update, 
            context, 
            f"✅ Exported {len(current_ids)} ignored test bills to Excel."
        )
    except Exception as e:
        log.error(f"Error sending test bills Excel: {e}")
        await edit_or_send_requester_text(
            msg, 
            update, 
            context, 
            f"Error generating Excel: {e}"
        )


@user_guard
async def cmd_delay(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Delay a bill (ignore it) until a certain date or for a number of days.
    Usage: /delay <bill_id> <date_or_days>
    Examples: /delay 2900492262 25.06.2026
              /delay 2900492262 3
    """
    await delete_group_command(update, context)
    if not context.args or len(context.args) < 2:
        await private_or_current_reply(
            update, 
            context, 
            "Usage:\n`/delay <bill_id> <date_or_days>`\n\n"
            "Examples:\n"
            "`/delay 2900492262 25.06.2026` (ignore until June 25)\n"
            "`/delay 2900492262 3` (ignore for 3 days)",
            parse_mode="Markdown"
        )
        return

    bill_id = context.args[0].strip()
    val = context.args[1].strip()
    delay_path = "delayed_bills.json"

    # Try parsing value
    exp_date_str = None
    
    # 1. Parse as number of days
    try:
        days = int(val)
        exp_date_str = (datetime.now() + timedelta(days=days)).date().isoformat()
    except (ValueError, OverflowError):
        pass

    # 2. Parse as DD.MM.YYYY or DD/MM/YYYY or YYYY-MM-DD
    if not exp_date_str:
        for fmt in ("%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d"):
            try:
                dt = datetime.strptime(val, fmt)
                exp_date_str = dt.date().isoformat()
                break
            except ValueError:
                pass

    if not exp_date_str:
        await private_or_current_reply(
            update,
            context,
            "❌ Invalid date or days format.\n"
            "Please use a number of days (e.g., `3`) or a date format like `DD.MM.YYYY` (e.g., `25.06.2026`)."
        )
        return

    # Load current delayed bills
    delayed = {}
    if os.path.exists(delay_path):
        try:
            with open(delay_path, "r", encoding="utf-8") as f:
                delayed = json.load(f)
        except Exception:
            pass

    delayed[bill_id] = exp_date_str

    with open(delay_path, "w", encoding="utf-8") as f:
        json.dump(delayed, f, ensure_ascii=False, indent=2)

    # Format expiration date nicely for user
    exp_date_nice = datetime.strptime(exp_date_str, "%Y-%m-%d").strftime("%d.%m.%Y")
    await private_or_current_reply(
        update,
        context,
        f"✅ Bill `{bill_id}` is now delayed (ignored) until **{exp_date_nice}**.",
        parse_mode="Markdown"
    )


@user_guard
async def cmd_undelay(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Remove the delay on a bill so it stops being ignored.
    Usage: /undelay <bill_id>
    """
    await delete_group_command(update, context)
    if not context.args:
        await private_or_current_reply(update, context, "Usage: `/undelay <bill_id>`", parse_mode="Markdown")
        return

    bill_id = context.args[0].strip()
    delay_path = "delayed_bills.json"

    if not os.path.exists(delay_path):
        await private_or_current_reply(update, context, "No delayed bills exist.")
        return

    try:
        with open(delay_path, "r", encoding="utf-8") as f:
            delayed = json.load(f)
    except Exception:
        delayed = {}

    if bill_id in delayed:
        del delayed[bill_id]
        with open(delay_path, "w", encoding="utf-8") as f:
            json.dump(delayed, f, ensure_ascii=False, indent=2)
        await private_or_current_reply(update, context, f"✅ Removed delay on bill `{bill_id}`. It will now be processed in reports.", parse_mode="Markdown")
    else:
        await private_or_current_reply(update, context, f"Bill `{bill_id}` was not found in the delayed bills list.", parse_mode="Markdown")


@pm_required_handler
async def cmd_delaylist(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """List all currently delayed bills and their expiration dates."""
    await delete_group_command(update, context)
    delay_path = "delayed_bills.json"

    if not os.path.exists(delay_path):
        await private_or_current_reply(update, context, "No delayed bills in the list.")
        return

    try:
        with open(delay_path, "r", encoding="utf-8") as f:
            delayed = json.load(f)
    except Exception:
        delayed = {}

    if not delayed:
        await private_or_current_reply(update, context, "No delayed bills in the list.")
        return

    lines = []
    today = datetime.now().date()
    
    # Sort by expiration date
    sorted_delayed = sorted(delayed.items(), key=lambda x: x[1])
    
    for bill_id, exp_date_str in sorted_delayed:
        try:
            exp_date = datetime.strptime(exp_date_str, "%Y-%m-%d").date()
            days_left = (exp_date - today).days
            exp_date_nice = exp_date.strftime("%d.%m.%Y")
            lines.append(f"• `{bill_id}`: Resumes on {exp_date_nice} ({days_left} days left)")
        except Exception:
            lines.append(f"• `{bill_id}`: exp {exp_date_str}")

    text = f"📋 **Currently Delayed Bills ({len(delayed)}):**\n" + "\n".join(lines)
    await private_or_current_reply(update, context, text, parse_mode="Markdown")


@user_guard
async def cmd_app(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/app — Open the interactive Mini App explorer."""
    await delete_group_command(update, context)
    cfg = load_config()
    webapp_url = cfg["telegram"].get("webapp_url", "http://localhost:8080")
    
    from telegram import InlineKeyboardMarkup, InlineKeyboardButton, WebAppInfo
    keyboard = [
        [InlineKeyboardButton("📱 Open Data Explorer", web_app=WebAppInfo(url=webapp_url))]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await private_or_current_reply(
        update,
        context,
        "🔍 **Interactive Data Explorer**\n\n"
        "Click the button below to search, filter, and view delivery details directly inside Telegram.",
        reply_markup=reply_markup,
        parse_mode="Markdown"
    )


@pm_required_handler
async def cmd_report(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/report — generate Excel with all active bills (excludes done statuses)."""
    await delete_group_command(update, context)
    cfg = load_config()
    msg = await send_requester_text(update, context, "Building active bills report...")
    tmpdir = tempfile.mkdtemp(prefix="report_")
    track_report_dir(tmpdir)
    stamp = datetime.now().strftime("%d.%m_%HH%M")
    src = os.path.join(tmpdir, f"export_{stamp}.xlsx")
    try:
        downloader.download_detail(cfg["api"], src)
        out_xlsx = os.path.join(tmpdir, f"Active_Bills_{stamp}.xlsx")
        count = report_cmd.build_report_excel(src, out_xlsx, cfg)
        caption = f"Active Bills Report {datetime.now().strftime('%d/%m/%Y %H:%M')}\nTotal: {count} bills"
        with open(out_xlsx, "rb") as f:
            await send_requester_document(update, context, f, os.path.basename(out_xlsx), caption=caption)
        await edit_or_send_requester_text(msg, update, context, f"Done. {count} active bills exported.")
    except Exception as e:
        log.exception("Error in /report")
        await edit_or_send_requester_text(msg, update, context, f"Error: {e}")


def build_master_daily_report_excel(template_path, raw_excel_path, output_path, target_date, cutoff_time):
    import win32com.client
    import os
    import pandas as pd
    import datetime
    import time
    import shutil
    
    # 1. Load raw data using pandas with dynamic header detection
    xl = pd.ExcelFile(raw_excel_path)
    sheet_name = xl.sheet_names[0]
    df_preview = xl.parse(sheet_name, header=None, nrows=10)
    header_row = 4
    for r_idx, row in df_preview.iterrows():
        if any('ORDER_NUMBER' in str(x).upper() or 'ORDER ID' in str(x).upper() for x in row):
            header_row = r_idx
            break
    df = xl.parse(sheet_name, skiprows=header_row)
    
    # Column mapping & normalizations
    col_map = {}
    for c in df.columns:
        cu = str(c).strip().upper()
        if 'ORDER_NUMBER' in cu or 'ORDER ID' in cu:
            col_map['ORDER_NUMBER'] = c
        elif cu == 'CUSTOMER' or 'SENDER' in cu:
            col_map['CUSTOMER'] = c
        elif 'ORIGIN_BRANCH' in cu or 'RECEIVE BRANCH' in cu:
            col_map['ORIGIN_BRANCH'] = c
        elif 'ORIGIN_POST' in cu or 'RECEIVE POST' in cu:
            col_map['ORIGIN_POST'] = c
        elif 'DESTINATION_BRANCH' in cu or 'DELIVERY BRANCH' in cu:
            col_map['DESTINATION_BRANCH'] = c
        elif 'DESTINATION_POST' in cu or 'DELIVERY POST' in cu:
            col_map['DESTINATION_POST'] = c
        elif 'CREATED_BY' in cu or 'ACTION USER' in cu:
            col_map['CREATED_BY'] = c
        elif 'CREATED_AT' in cu or 'CREATED DATE' in cu:
            col_map['CREATED_AT'] = c
        elif 'PAYMENT_TERM' in cu or 'PAYMENT METHOD' in cu:
            col_map['PAYMENT_TERM'] = c
        elif 'BASE_FEE' in cu or 'BASE FEE' in cu:
            col_map['BASE_FEE'] = c
        elif 'VAS_FEE' in cu or 'VAS FEE' in cu:
            col_map['VAS_FEE'] = c
        elif 'DISCOUNT' in cu:
            col_map['DISCOUNT'] = c
        elif 'TOTAL_AMOUNT' in cu or 'TOTAL FEE' in cu:
            col_map['TOTAL_AMOUNT'] = c
        elif 'COD' in cu:
            col_map['COD'] = c
        elif 'ACTUAL_WEIGHT' in cu or 'WEIGHT' in cu:
            col_map['ACTUAL_WEIGHT'] = c
        elif 'SIZE' in cu:
            col_map['SIZE'] = c
        elif cu == 'BASE_SERVICE' or cu == 'SERVICE':
            col_map['BASE_SERVICE'] = c
        elif 'VAS_SERVICE' in cu:
            col_map['VAS_SERVICE'] = c
        elif 'VAS_INFO' in cu:
            col_map['VAS_INFO'] = c
        elif 'CARGO_TYPE' in cu or 'GOOD TYPE' in cu:
            col_map['CARGO_TYPE'] = c
        elif 'FROM_SOURCE' in cu:
            col_map['FROM_SOURCE'] = c

    # Vectorized lookups
    service_map = {"CLT": "Tiêu chuẩn", "CCN": "Nhanh", "CNT": "Tiêu chuẩn"}
    source_map = {
        "APP - OMS": "App KH",
        "WEB - OMS": "Web KH",
        "WEB - TMS": "Web nhân viên",
        "APP - TMS": "App nhân viên"
    }
    zone_map = {
        "PNP": "ZONE 1", "KAN": "ZONE 1", "PRE": "ZONE 1", "SVA": "ZONE 1",
        "TAK": "ZONE 2", "SPE": "ZONE 2", "KAM": "ZONE 2", "KEP": "ZONE 2", "KOH": "ZONE 2", "SIH": "ZONE 2",
        "BAN": "ZONE 3", "BAT": "ZONE 3", "PUR": "ZONE 3", "CHH": "ZONE 3",
        "SIE": "ZONE 4", "PRH": "ZONE 4", "THO": "ZONE 4", "ODD": "ZONE 4",
        "CHA": "ZONE 5", "TBK": "ZONE 5", "KRA": "ZONE 5", "MON": "ZONE 5", "ROT": "ZONE 5", "STU": "ZONE 5"
    }
    branch_map = {
        "KEP": "TAK",
        "PAI": "BAT",
        "TBK": "CHA"
    }
    internal_phones = {"884589745", "315555236", "886663766", "716560202", "977898616", "716718617", "716787878"}

    created_col = col_map.get('CREATED_AT', 'CREATED_AT')
    if created_col in df.columns:
        df['parsed_dt'] = pd.to_datetime(df[created_col], dayfirst=True, format='mixed', errors='coerce')
    else:
        df['parsed_dt'] = pd.NaT

    df = df.sort_values(by='parsed_dt', ascending=True, na_position='last').reset_index(drop=True)
    
    df['month'] = df['parsed_dt'].dt.month.fillna(0).astype(int)
    df['day'] = df['parsed_dt'].dt.day.fillna(0).astype(int)
    df['hour'] = df['parsed_dt'].dt.hour.fillna(0).astype(int)
    df['minute'] = df['parsed_dt'].dt.minute.fillna(0).astype(int)

    cust_col = col_map.get('CUSTOMER', 'CUSTOMER')
    def extract_phone(cust):
        s = str(cust).strip()
        if " - " in s:
            p = s.split(" - ")[0].strip()
            digits = "".join(ch for ch in p if ch.isdigit())
            return digits
        return ""

    if cust_col in df.columns:
        df['phone'] = df[cust_col].apply(extract_phone)
        df['cum_count'] = df.groupby(cust_col).cumcount() + 1
    else:
        df['phone'] = ""
        df['cum_count'] = 1

    # Prepare computed rows in memory
    rows_data = []
    for idx, row in df.iterrows():
        r = idx + 2
        m = int(row['month'])
        d = int(row['day'])
        h = int(row['hour'])
        mi = int(row['minute'])
        
        date_formula = f"=DATE(2026,A{r},B{r})"
        time_formula = f"=TIME(C{r},D{r},1)"
        
        base_svc = str(row.get(col_map.get('BASE_SERVICE', 'BASE_SERVICE'), '')).strip()
        svc_text = service_map.get(base_svc, "Tiêu chuẩn")
        
        src_val = str(row.get(col_map.get('FROM_SOURCE', 'FROM_SOURCE'), '')).strip()
        src_text = source_map.get(src_val, "Không xác định")
        
        try:
            cod_num = float(row.get(col_map.get('COD', 'COD (USD)'), 0) or 0)
        except Exception:
            cod_num = 0.0
        cod_text = "COD" if cod_num > 0 else "Non-COD"
        
        phone_str = row['phone']
        cust_type = "Nội bộ" if phone_str in internal_phones else "KH ngoài"
        
        try:
            tot_fee = float(row.get(col_map.get('TOTAL_AMOUNT', 'TOTAL_AMOUNT (USD) (4) = (1) + (2) - (3)'), 0) or 0)
        except Exception:
            tot_fee = 0.0
            
        branch_orig = str(row.get(col_map.get('ORIGIN_BRANCH', 'ORIGIN_BRANCH'), '')).strip().upper()
        zone_text = zone_map.get(branch_orig, "")
        branch_text = branch_map.get(branch_orig, branch_orig)
        inter_prov = "Nội Tỉnh" if base_svc == "CNT" else "Liên Tỉnh"
        
        try:
            phone_val = int(phone_str) if phone_str else ""
        except Exception:
            phone_val = phone_str
            
        cum_cnt = int(row['cum_count'])
        
        po_orig = str(row.get(col_map.get('ORIGIN_POST', 'ORIGIN_POST'), '')).strip().upper()
        channel_code = po_orig[3] if len(po_orig) > 3 else "P"
        
        try:
            weight_g = float(row.get(col_map.get('ACTUAL_WEIGHT', 'ACTUAL_WEIGHT (G)'), 0) or 0)
            weight_kg = weight_g / 1000.0
        except Exception:
            weight_kg = 0.0
            weight_g = 0.0
            
        oid_val = str(row.get(col_map.get('ORDER_NUMBER', 'ORDER_NUMBER'), '')).strip()
        if oid_val.endswith('.0'):
            oid_val = oid_val[:-2]
            
        row_cells = [
            m, d, h, mi, date_formula, time_formula,
            svc_text, src_text, cod_text, cust_type, tot_fee,
            zone_text, branch_text, inter_prov, phone_val, cum_cnt,
            channel_code, weight_kg,
            row.get('No', idx+1),
            oid_val,
            str(row.get(col_map.get('CUSTOMER', 'CUSTOMER'), '')),
            branch_orig,
            po_orig,
            str(row.get(col_map.get('DESTINATION_BRANCH', 'DESTINATION_BRANCH'), '')).strip().upper(),
            str(row.get(col_map.get('DESTINATION_POST', 'DESTINATION_POST'), '')).strip().upper(),
            str(row.get(col_map.get('CREATED_BY', 'CREATED_BY'), '')),
            str(row.get(col_map.get('CREATED_AT', 'CREATED_AT'), '')),
            str(row.get(col_map.get('PAYMENT_TERM', 'PAYMENT_TERM'), '')),
            row.get(col_map.get('BASE_FEE', 'BASE_FEE (USD) (1)'), 0),
            row.get(col_map.get('VAS_FEE', 'VAS_FEE (USD) (2)'), 0),
            row.get(col_map.get('DISCOUNT', 'DISCOUNT (USD) (3)'), 0),
            tot_fee,
            cod_num,
            weight_g,
            str(row.get(col_map.get('SIZE', 'SIZE (L*W*H) (CM)'), '')),
            base_svc,
            str(row.get(col_map.get('VAS_SERVICE', 'VAS_SERVICE'), '')),
            str(row.get(col_map.get('VAS_INFO', 'VAS_INFO'), '')),
            str(row.get(col_map.get('CARGO_TYPE', 'CARGO_TYPE'), '')),
            src_val
        ]
        rows_data.append(row_cells)

    # Copy template to output path
    shutil.copy2(template_path, output_path)
    
    import pythoncom
    pythoncom.CoInitialize()
    
    excel = None
    wb = None
    try:
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
        except Exception:
            excel = win32com.client.Dispatch("Excel.Application")
            
        try:
            excel.Visible = False
        except Exception:
            pass
        try:
            excel.DisplayAlerts = False
            excel.ScreenUpdating = False
        except Exception:
            pass
            
        abs_output_path = os.path.abspath(output_path)
        wb = excel.Workbooks.Open(abs_output_path)
        
        # 1. Update Date and Cutoff in Zone_Report (C1, C2)
        try:
            ws_zone = wb.Worksheets("Zone_Report")
            ws_zone.Cells(1, 3).Value = target_date.strftime("%Y-%m-%d")
            ws_zone.Cells(2, 3).Value = cutoff_time.strftime("%H:%M:%S")
        except Exception:
            pass
            
        # Update Date and Cutoff in Province_Report (B1, B2)
        for name in ["Province_Report", "Province_Report (2)"]:
            try:
                ws = wb.Worksheets(name)
                ws.Cells(1, 2).Value = target_date.strftime("%Y-%m-%d")
                ws.Cells(2, 2).Value = cutoff_time.strftime("%H:%M:%S")
            except Exception:
                pass
                
        # 2. Write to Data Revenue sheet
        try:
            ws_rev = wb.Worksheets("Data Revenue")
            
            if rows_data:
                # Write in single block assignment
                rng_write = ws_rev.Range(ws_rev.Cells(2, 1), ws_rev.Cells(len(rows_data) + 1, 40))
                rng_write.Value = tuple(tuple(r) for r in rows_data)
                
                # Clear leftover old rows if previous data was longer
                old_last = ws_rev.Cells(ws_rev.Rows.Count, "AA").End(-4162).Row
                if old_last > len(rows_data) + 1:
                    ws_rev.Range(ws_rev.Cells(len(rows_data) + 2, 1), ws_rev.Cells(old_last, 40)).Value = None
        except Exception as e:
            import logging
        # Ensure Zone_Report customer analysis rows (U175:Y{last_r}) have live formulas
        try:
            ws_z = wb.Worksheets("Zone_Report")
            last_cust_r = ws_z.Cells(ws_z.Rows.Count, "R").End(-4162).Row # xlUp
            if last_cust_r >= 175:
                ws_z.Range(f"U175:Y{last_cust_r}").Formula = ws_z.Range("U175:Y175").Formula
        except Exception:
            pass

        # Recalculate
        excel.CalculateFull()
        
        # Extract exact values directly from calculated Province_Report
        metrics = None
        try:
            ws_p = wb.Worksheets("Province_Report")
            metrics = {
                "total": int(ws_p.Range("AG6").Value or 0),
                "diff_n1": int(ws_p.Range("AI6").Value or 0),
                "sp": int(ws_p.Range("AL6").Value or 0),
                "agent": int(ws_p.Range("AP6").Value or 0),
                "showroom": int(ws_p.Range("AT6").Value or 0),
                "zero_sp_pnp": [],
                "zero_sp_prov": [],
                "zero_branches": []
            }
            for r in range(8, 45):
                po_code = str(ws_p.Range(f"B{r}").Value or '').strip().upper()
                if not po_code or len(po_code) <= 3:
                    continue
                po_orders = int(ws_p.Range(f"K{r}").Value or 0)
                if po_orders == 0:
                    if po_code.startswith("PNP"):
                        metrics["zero_sp_pnp"].append(po_code)
                    else:
                        metrics["zero_sp_prov"].append(po_code)
                        
            for r in range(7, 29):
                b_name = str(ws_p.Range(f"AC{r}").Value or '').strip()
                if not b_name:
                    continue
                b_orders = int(ws_p.Range(f"AG{r}").Value or 0)
                if b_orders == 0:
                    metrics["zero_branches"].append(b_name)
        except Exception:
            import logging
            logging.exception("Error extracting metrics from Province_Report")
            
        wb.Save()
        return metrics
        
    finally:
        if wb:
            try:
                wb.Close(SaveChanges=True)
            except Exception:
                pass
        if excel:
            try:
                excel.Quit()
            except Exception:
                pass
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass


@pm_required_handler
async def cmd_daily_report(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/dailyreport [date] — generate text daily report (volume, comparison, zero-order offices/branches), render screenshot image and send populated Master Daily Excel."""
    await delete_group_command(update, context)
    cfg = load_config()
    
    # Parse optional date argument
    args = [a.strip() for a in (context.args or []) if a.strip()]
    target_date = None
    today = datetime.now().date()
    
    if args:
        date_str = args[0]
        # Support DD/MM, DD/MM/YYYY, YYYY-MM-DD
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d/%m"):
            try:
                dt = datetime.strptime(date_str, fmt)
                if "%d/%m" in fmt:
                    dt = dt.replace(year=today.year)
                target_date = dt.date()
                break
            except ValueError:
                continue
        if target_date is None:
            await private_or_current_reply(
                update, context,
                f"❌ Invalid date format: '{date_str}'.\n"
                f"Please use DD/MM (e.g. 10/08), DD/MM/YYYY (e.g. 10/08/2026), or YYYY-MM-DD."
            )
            return
    else:
        target_date = today

    date_display = target_date.strftime("%d/%m")
    msg = await send_requester_text(update, context, f"Generating daily report for {date_display}...")
    
    tmpdir = tempfile.mkdtemp(prefix="daily_report_")
    track_report_dir(tmpdir)
    src = os.path.join(tmpdir, f"export_pickup_revenue_{target_date.strftime('%Y%m%d')}.xlsx")
    
    last_week_date = target_date - timedelta(days=7)
    
    try:
        # Date range: from Day 1 of the month to target_date
        # If target_date is within first 7 days, we still count from Day 1 (or last week if earlier)
        month_start_date = target_date.replace(day=1)
        earliest_date = min(month_start_date, last_week_date)
        from_date_str = earliest_date.strftime("%Y%m%d")
        to_date_str = target_date.strftime("%Y%m%d")
        
        # Download pickup revenue detail from API
        downloader.download_pickup_revenue(
            cfg["api"], src,
            from_date=from_date_str,
            to_date=to_date_str
        )
            
        # Parse Excel data
        import pandas as pd
        xl = pd.ExcelFile(src)
        sheet_name = xl.sheet_names[0]
        df_preview = xl.parse(sheet_name, header=None, nrows=10)
        header_row = 4
        for r_idx, row in df_preview.iterrows():
            if any('ORDER_NUMBER' in str(x).upper() or 'ORDER ID' in str(x).upper() for x in row):
                header_row = r_idx
                break
        df = xl.parse(sheet_name, skiprows=header_row)
        
        # Determine created date col
        created_col = None
        for c in df.columns:
            cu = str(c).strip().upper()
            if 'CREATED_AT' in cu or 'CREATED DATE' in cu:
                created_col = c
                break
        if not created_col:
            raise ValueError("Excel file is missing CREATED_AT/CREATED DATE column.")
            
        # Parse datetime
        df['parsed_datetime'] = pd.to_datetime(df[created_col], dayfirst=True, format='mixed', errors='coerce')
        df['parsed_date'] = df['parsed_datetime'].dt.date
        
        # Exclude test orders
        test_order_ids = set()
        from generate_report import load_test_order_ids
        try:
            test_order_ids = load_test_order_ids(cfg)
        except Exception:
            pass
            
        test_keywords = cfg.get("pivot", {}).get("test_keywords", ["test"])
        
        oid_col = None
        cust_col = None
        po_col = None
        branch_col = None
        for c in df.columns:
            cu = str(c).strip().upper()
            if 'ORDER_NUMBER' in cu or 'ORDER ID' in cu:
                oid_col = c
            elif cu == 'CUSTOMER' or 'SENDER' in cu:
                cust_col = c
            elif 'ORIGIN_POST' in cu or 'RECEIVE POST' in cu:
                po_col = c
            elif 'ORIGIN_BRANCH' in cu or 'RECEIVE BRANCH' in cu:
                branch_col = c

        # Comprehensive test exclusion matching generate_report
        test_col = next((c for c in df.columns if str(c).strip().upper() in ['TEST', 'IS_TEST', 'IS TEST', 'TEST ORDER']), None)
        if test_col:
            df = df[df[test_col].isna() | df[test_col].astype(str).str.strip().isin(['', 'nan', 'NaN', '#N/A'])].copy()
            
        test_check_cols = [c for c in df.columns if any(k in str(c).lower() for k in ('name', 'note', 'customer', 'sender', 'receiver', 'remark', 'address'))]
        
        def is_test_row(row):
            if oid_col:
                oid = str(row.get(oid_col, '')).strip()
                if oid.endswith('.0'):
                    oid = oid[:-2]
                if oid in test_order_ids:
                    return True
            for col in test_check_cols:
                val = str(row.get(col, '')).lower()
                if any(kw in val for kw in test_keywords):
                    return True
            return False
            
        df['is_test_order'] = df.apply(is_test_row, axis=1)
        df_clean = df[~df['is_test_order']].copy()
        
        # Sort chronologically and compute cumulative order count per customer in month
        df_clean = df_clean.sort_values(by='parsed_datetime', ascending=True, na_position='last').reset_index(drop=True)
        if cust_col and cust_col in df_clean.columns:
            df_clean['cum_count'] = df_clean.groupby(cust_col).cumcount() + 1
        else:
            df_clean['cum_count'] = 1
        
        # Determine cutoff time
        if target_date == today:
            cutoff_time = datetime.now().time()
        else:
            cutoff_time = datetime.max.time()
            
        # Filter target date orders
        df_target_all = df_clean[df_clean['parsed_date'] == target_date]
        df_target = df_target_all[df_target_all['parsed_datetime'].dt.time <= cutoff_time].copy()
        total_target = len(df_target)
        
        # Filter same day yesterday orders (cùng kỳ ngày)
        yesterday_date = target_date - timedelta(days=1)
        df_yesterday_all = df_clean[df_clean['parsed_date'] == yesterday_date]
        df_yesterday = df_yesterday_all[df_yesterday_all['parsed_datetime'].dt.time <= cutoff_time]
        total_yesterday = len(df_yesterday)
        
        diff = total_target - total_yesterday
        if diff < 0:
            diff_text = f"giảm {-diff} đơn"
        elif diff > 0:
            diff_text = f"tăng {diff} đơn"
        else:
            diff_text = "bằng 0 đơn"
            
        # Channels breakdown
        def get_channel_type(po):
            po_str = str(po).strip().upper()
            if len(po_str) > 3:
                c = po_str[3]
                if c == 'P':
                    return 'Service Point'
                elif c == 'A':
                    return 'Agent'
                elif c == 'S':
                    return 'Showroom'
            return 'Service Point'
            
        df_target = df_target.copy()
        target_po_col = po_col or 'ORIGIN_POST'
        df_target['channel'] = df_target[target_po_col].apply(get_channel_type) if target_po_col in df_target.columns else 'Service Point'
        channel_counts = df_target['channel'].value_counts()
        
        count_sp = channel_counts.get('Service Point', 0)
        count_agent = channel_counts.get('Agent', 0)
        count_showroom = channel_counts.get('Showroom', 0)
        
        # Zero order post offices (Service Points only)
        total_zones = cfg.get("total_zones", {})
        monitored_sp_pos = []
        for zone, pos in total_zones.items():
            for po in pos:
                po_upper = po.strip().upper()
                if len(po_upper) > 3 and po_upper[3] == 'P':
                    monitored_sp_pos.append(po_upper)
        monitored_sp_pos = sorted(list(set(monitored_sp_pos)))
        
        active_pos_today = set()
        if target_po_col in df_target.columns:
            active_pos_today = set(df_target[target_po_col].dropna().astype(str).str.strip().str.upper().unique())
        zero_sp_pos = [po for po in monitored_sp_pos if po not in active_pos_today]
        
        # Group into Phnom Penh and Provinces
        zero_pnp = [po for po in zero_sp_pos if po.startswith("PNP")]
        zero_province = [po for po in zero_sp_pos if not po.startswith("PNP")]
        
        # Zero order branches
        BRANCH_NAMES = {
            "PNP": "Phnom Penh",
            "KAN": "Kandal",
            "PRE": "Prey Veng",
            "SVA": "Svay Rieng",
            "KAM": "Kampot",
            "KOH": "Koh Kong",
            "SIH": "Sihanoukville",
            "SPE": "Kampong Speu",
            "TAK": "Takeo",
            "BAN": "Banteay Meanchey",
            "BAT": "Battambang",
            "CHH": "Kampong Chhnang",
            "PUR": "Pursat",
            "PRH": "Preah Vihear",
            "SIE": "Siem Reap",
            "THO": "Kampong Thom",
            "ODD": "Otdar Meanchey",
            "CHA": "Kampong Cham",
            "KRA": "Kratie",
            "MON": "Mondul Kiri",
            "ROT": "Ratanak Kiri",
            "STU": "Stung Treng"
        }
        
        branch_map_ref = {"KEP": "TAK", "PAI": "BAT", "TBK": "CHA"}
        def get_branch_code(po):
            po_str = str(po).strip().upper()
            prefix = po_str[:3]
            return branch_map_ref.get(prefix, prefix)
            
        active_branches = set()
        for po in active_pos_today:
            b_code = get_branch_code(po)
            if b_code in BRANCH_NAMES:
                active_branches.add(b_code)
                
        zero_branches = [b for b in BRANCH_NAMES.keys() if b not in active_branches]
        zero_branch_names = [BRANCH_NAMES[b] for b in zero_branches]
        
        # Zero New Customer Inday calculation from real data
        df_target_new_cust = df_target[df_target.get('cum_count', 0) == 1] if 'cum_count' in df_target.columns else pd.DataFrame()
        active_new_cust_pos = set()
        if not df_target_new_cust.empty and target_po_col in df_target_new_cust.columns:
            active_new_cust_pos = set(df_target_new_cust[target_po_col].dropna().astype(str).str.strip().str.upper().unique())
            
        zero_new_cust_by_zone = {}
        for z_idx in range(1, 6):
            z_key = f"zone{z_idx}"
            pos = total_zones.get(z_key, [])
            zero_in_z = [po.strip().upper() for po in pos if po.strip().upper() not in active_new_cust_pos]
            zero_new_cust_by_zone[z_key] = zero_in_z

        def format_zone_zero_new_cust(zero_pos_by_zone):
            lines = []
            for z_idx in range(1, 6):
                z_key = f"zone{z_idx}"
                pos = zero_pos_by_zone.get(z_key, [])
                if not pos:
                    lines.append(f"*Zone {z_idx}: Không có")
                    continue
                    
                pnp_nums = []
                other_pos = []
                for po in sorted(pos):
                    po_u = po.strip().upper()
                    if po_u.startswith("PNPP"):
                        num_part = po_u[4:]
                        pnp_nums.append(num_part)
                    else:
                        other_pos.append(po_u)
                        
                parts = []
                if pnp_nums:
                    parts.append(f"PNPP( {', '.join(pnp_nums)} )")
                if other_pos:
                    parts.extend(other_pos)
                    
                lines.append(f"*Zone {z_idx}: {', '.join(parts)}")
            return "\n".join(lines)

        no_new_cust_text = format_zone_zero_new_cust(zero_new_cust_by_zone)

        # Formatting report output
        time_str = datetime.now().strftime("%H:%M") if target_date == today else "23:59"
        date_formatted = target_date.strftime("%d/%m")
        
        report_text = (
            f"📦 BÁO CÁO SẢN LƯỢNG {date_formatted}-{time_str}\n\n"
            f"Báo cáo PTGĐ Anh @Trungnh2 và các anh GĐCN, GĐV @everyone PKD kính gửi kết quả sán lượng đến hiện tại:\n\n"
            f"Tổng sản lượng: {total_target} đơn,  {diff_text} so với cùng kỳ ngày  .\n\n"
            f"📌 Xét theo kênh:\n\n"
            f"Bưu cục : {count_sp} đơn\n"
            f"Đại lý : {count_agent} đơn \n"
            f"Showroom: {count_showroom} đơn \n\n"
            f"📌{len(zero_sp_pos)}  bưu cục chưa phát sinh đơn:\n"
            f"Khu vực Phnom Penh ({len(zero_pnp)} bưu cục): {', '.join(zero_pnp) if zero_pnp else 'Không có'}\n\n"
            f"Khu vực Tỉnh ({len(zero_province)} bưu cục): {', '.join(zero_province) if zero_province else 'Không có'}\n\n"
            f"📌{len(zero_branches)} Chi nhánh chưa phát sinh đơn:\n"
        )
        if zero_branches:
            report_text += '\n'.join([f"+{name} " for name in sorted(zero_branch_names)]) + "\n\n"
        else:
            report_text += "Không có\n\n"
            
        report_text += (
            f"📌 Service Point no result New customer Inday:\n"
            f"{no_new_cust_text}\n\n"
            f"Trân trọng."
        )
        
        # Generate populated excel report
        import io
        template_dir = os.path.dirname(os.path.abspath(__file__))
        template_files = [f for f in os.listdir(template_dir) if f.startswith("0.Master Daily Report") and f.endswith(".xlsx")]
        
        template_path = None
        if template_files:
            template_files.sort(reverse=True)
            template_path = os.path.join(template_dir, template_files[0])
            template_name = template_files[0]
        else:
            template_name = "0.Master Daily Report - new - Aug.xlsx"
            template_path = os.path.join(template_dir, template_name)
            
        if os.path.exists(template_path):
            output_xlsx_name = f"Master_Daily_Report_{date_formatted.replace('/', '_')}_{time_str.replace(':', 'H')}.xlsx"
            output_xlsx_path = os.path.join(tmpdir, output_xlsx_name)
            
            try:
                msg = await edit_or_send_requester_text(msg, update, context, report_text + f"\n\nGenerating master Excel report using {template_name}...")
                metrics = await asyncio.to_thread(build_master_daily_report_excel, template_path, src, output_xlsx_path, target_date, cutoff_time)
                
                if metrics:
                    t_val = metrics.get("total", total_target)
                    diff_v = metrics.get("diff_n1", diff)
                    if diff_v < 0:
                        diff_t = f"giảm {-diff_v} đơn"
                    elif diff_v > 0:
                        diff_t = f"tăng {diff_v} đơn"
                    else:
                        diff_t = "bằng 0 đơn"
                    
                    sp_v = metrics.get("sp", count_sp)
                    ag_v = metrics.get("agent", count_agent)
                    sh_v = metrics.get("showroom", count_showroom)
                    
                    z_pnp = metrics.get("zero_sp_pnp", zero_pnp)
                    z_prov = metrics.get("zero_sp_prov", zero_province)
                    z_br = metrics.get("zero_branches", zero_branch_names)
                    
                    report_text = (
                        f"📦 BÁO CÁO SẢN LƯỢNG {date_formatted}-{time_str}\n\n"
                        f"Báo cáo PTGĐ Anh @Trungnh2 và các anh GĐCN, GĐV @everyone PKD kính gửi kết quả sán lượng đến hiện tại:\n\n"
                        f"Tổng sản lượng: {t_val} đơn,  {diff_t} so với cùng kỳ ngày  .\n\n"
                        f"📌 Xét theo kênh:\n\n"
                        f"Bưu cục : {sp_v} đơn\n"
                        f"Đại lý : {ag_v} đơn \n"
                        f"Showroom: {sh_v} đơn \n\n"
                        f"📌{len(z_pnp) + len(z_prov)}  bưu cục chưa phát sinh đơn:\n"
                        f"Khu vực Phnom Penh ({len(z_pnp)} bưu cục): {', '.join(z_pnp) if z_pnp else 'Không có'}\n\n"
                        f"Khu vực Tỉnh ({len(z_prov)} bưu cục): {', '.join(z_prov) if z_prov else 'Không có'}\n\n"
                        f"📌{len(z_br)} Chi nhánh chưa phát sinh đơn:\n"
                    )
                    if z_br:
                        report_text += '\n'.join([f"+{name} " for name in sorted(z_br)]) + "\n\n"
                    else:
                        report_text += "Không có\n\n"
                        
                    report_text += (
                        f"📌 Service Point no result New customer Inday:\n"
                        f"{no_new_cust_text}\n\n"
                        f"Trân trọng."
                    )
                
                msg = await edit_or_send_requester_text(msg, update, context, report_text + "\n\nRendering report images...")
                
                # Render the reports
                from excel_to_image import render_excel_reports
                reports_map = await asyncio.to_thread(render_excel_reports, output_xlsx_path, target_date, tmpdir)
                
                # Send the generated reports in a swipeable Telegram album
                from telegram import InputMediaPhoto
                media_list = []
                
                report_order = [
                    "zone_summary", "customer_report", "day_report", "month_report",
                    "sp_order_express_all", "sp_customer_development",
                    "sp_zone_1", "sp_zone_2", "sp_zone_3", "sp_zone_4", "sp_zone_5"
                ]
                captions = {
                    "zone_summary": f"📊 Báo cáo kết quả SXKD ({date_formatted})",
                    "customer_report": f"👥 Báo cáo khách hàng mới ({date_formatted})",
                    "day_report": f"📅 Branch Day Report ({date_formatted})",
                    "month_report": f"📅 Branch Month Report ({date_formatted})",
                    "sp_order_express_all": f"📦 [SERVICE POINT] Report of Order Express ({date_formatted})",
                    "sp_customer_development": f"👥 [SERVICE POINT] Customer Development ({date_formatted})",
                    "sp_zone_1": f"📍 Service Point Zone 1 ({date_formatted})",
                    "sp_zone_2": f"📍 Service Point Zone 2 ({date_formatted})",
                    "sp_zone_3": f"📍 Service Point Zone 3 ({date_formatted})",
                    "sp_zone_4": f"📍 Service Point Zone 4 ({date_formatted})",
                    "sp_zone_5": f"📍 Service Point Zone 5 ({date_formatted})"
                }
                
                for rep_name in report_order:
                    if rep_name in reports_map and os.path.exists(reports_map[rep_name]):
                        with open(reports_map[rep_name], "rb") as f:
                            photo_data = f.read()
                            media_list.append(InputMediaPhoto(io.BytesIO(photo_data), caption=captions.get(rep_name, "")))
                            
                if media_list:
                    # Telegram supports up to 10 photos per media group
                    for chunk_idx in range(0, len(media_list), 10):
                        chunk = media_list[chunk_idx:chunk_idx+10]
                        await send_requester_media_group(update, context, chunk)
                    
                # Send the Excel file
                with open(output_xlsx_path, "rb") as f:
                    await send_requester_document(
                        update, context, f,
                        output_xlsx_name,
                        caption=f"📊 Master Daily Excel {date_formatted} {time_str}"
                    )
                await edit_or_send_requester_text(msg, update, context, report_text)
            except Exception as exc:
                log.exception("Error generating populated master Excel")
                await edit_or_send_requester_text(msg, update, context, report_text + f"\n\n⚠️ Error generating Excel/Image: {exc}")
        else:
            log.warning(f"Template not found at {template_path}")
            await edit_or_send_requester_text(msg, update, context, report_text + "\n\n⚠️ Note: Master Daily Report template was not found, so no Excel file was attached.")
        
    except Exception as e:
        log.exception("Error in /dailyreport")
        await edit_or_send_requester_text(msg, update, context, f"Error: {e}")


async def on_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not update.message:
        return
    cfg = load_config()
    if not is_user_allowed(update, cfg):
        log.info("Ignoring message from unauthorized user %s", update.effective_user and update.effective_user.id)
        return
    keyword = cfg["telegram"].get("keyword", "push").lower().strip()
    text    = (update.message.text or "").strip()
    parts   = text.split()
    lower_parts = [p.lower() for p in parts]

    pending = context.user_data.get("pending_forward")
    answer = lower_parts[0].strip(".,!?") if lower_parts else ""
    yes_answers = {"yes", "y", "ok", "send", "confirm", "forward"}
    no_answers = {"no", "n", "cancel", "stop"}

    if pending and answer in yes_answers | no_answers:
        await delete_group_command(update, context)
        if answer in no_answers:
            context.user_data.pop("pending_forward", None)
            await private_or_current_reply(
                update,
                context,
                "Cancelled. Report was not forwarded to any group.",
            )
            return

        payload = context.user_data.pop("pending_forward", None)
        if not payload:
            await private_or_current_reply(
                update,
                context,
                "No pending report to forward.",
            )
            return

        await private_or_current_reply(
            update,
            context,
            f"Confirmed. Forwarding to {len(payload['forward_groups'])} group(s)...",
        )
        sent_groups = await forward_result_to_groups(context, payload)
        await private_or_current_reply(
            update,
            context,
            f"Done. Forwarded to {sent_groups} group(s). "
            f"{datetime.now().strftime('%d.%m.%Y %H:%M')}",
        )
        return
    
    # (Pending remark handling removed)

    if keyword in lower_parts:
        idx = lower_parts.index(keyword)
        # Anything typed after "push" becomes the args (e.g., "push PNPP014")
        context.args = parts[idx + 1:]
        await run_push(update, context)


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    # Start the WebApp HTTP Server in a background daemon thread
    server_thread = threading.Thread(target=start_webapp_server, daemon=True)
    server_thread.start()

    cfg   = load_config()
    tg    = cfg["telegram"]
    token = tg["bot_token"]
    if "DIEN_" in token:
        raise SystemExit("Set bot_token in config.json first.")

    builder = (
        Application.builder()
        .token(token)
        .connect_timeout(30.0)
        .read_timeout(30.0)
        .write_timeout(30.0)
        .pool_timeout(30.0)
        .get_updates_connect_timeout(30.0)
        .get_updates_read_timeout(30.0)
    )

    proxy_url = tg.get("proxy_url")
    if proxy_url:
        builder = builder.proxy(proxy_url).get_updates_proxy(proxy_url)
        log.info("Using proxy: %s", proxy_url)

    app = builder.build()
    app.add_handler(CommandHandler("app",          cmd_app))
    app.add_handler(CommandHandler("push",         run_push))
    app.add_handler(CommandHandler("total",        cmd_total))
    app.add_handler(CommandHandler("penalty",      cmd_penalty))
    app.add_handler(CommandHandler("speed",        cmd_speed))
    app.add_handler(CommandHandler("vs",           cmd_vs))
    app.add_handler(CommandHandler("vs2",          cmd_vs2))
    app.add_handler(CommandHandler("help",         cmd_help))
    app.add_handler(CommandHandler("pause",        cmd_pause))
    app.add_handler(CommandHandler("stop",         cmd_pause))
    app.add_handler(CommandHandler("resume",       cmd_resume))
    app.add_handler(CommandHandler("start",        cmd_resume))
    app.add_handler(CommandHandler("status",       cmd_status))
    app.add_handler(CommandHandler("statues",      cmd_statues))
    app.add_handler(CommandHandler("statuses",     cmd_statues))
    app.add_handler(CommandHandler("mode",         cmd_mode))
    app.add_handler(CommandHandler("register",     cmd_register))
    app.add_handler(CommandHandler("unregister",   cmd_unregister))
    app.add_handler(CommandHandler("groups",       cmd_groups))
    app.add_handler(CommandHandler("export",       cmd_export))
    app.add_handler(CommandHandler("find",         cmd_find))
    app.add_handler(CommandHandler("ask",          cmd_ask))
    app.add_handler(CommandHandler("check",        cmd_check))
    app.add_handler(CommandHandler("qr",           cmd_qr))
    app.add_handler(CommandHandler("trace",        cmd_trace))
    app.add_handler(CommandHandler("add",          cmd_add))
    app.add_handler(CommandHandler("remove",       cmd_remove))
    app.add_handler(CommandHandler("del",          cmd_remove))
    app.add_handler(CommandHandler("list",         cmd_list))
    app.add_handler(CommandHandler("delay",        cmd_delay))
    app.add_handler(CommandHandler("undelay",      cmd_undelay))
    app.add_handler(CommandHandler("delaylist",    cmd_delaylist))
    app.add_handler(CommandHandler("clean",        cmd_clean))
    app.add_handler(CommandHandler("report",       cmd_report))
    app.add_handler(CommandHandler("deletereport", cmd_delete_report))
    app.add_handler(CommandHandler("delreport",    cmd_delete_report))
    app.add_handler(CommandHandler("dailyreport",  cmd_daily_report))
    app.add_handler(CommandHandler("test",         cmd_test))
    app.add_handler(CommandHandler("testmode",     cmd_test_mode))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, on_text))

    log.info("Bot running. Commands: push, /total, /vs, /vs2, /export, /find, /ask, /check, /trace, /statues, /help, /pause, /resume, /status, /mode, /register, /groups, /add, /remove, /list, /delay, /undelay, /delaylist, /clean, /qr, /deletereport, /dailyreport")
    try:
        app.run_polling(allowed_updates=Update.ALL_TYPES)
    except Exception as e:
        if "Conflict" in str(e):
            log.error("Another bot instance is already running. Stop it first, then restart.")
        else:
            raise


if __name__ == "__main__":
    main()
