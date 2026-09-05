import os, sys, json
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

DONE_BILLS_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "done_bills.json")

def load_done_bills() -> dict:
    if os.path.exists(DONE_BILLS_PATH):
        try:
            with open(DONE_BILLS_PATH, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}
    return {}

def save_done_bills(data: dict) -> None:
    try:
        with open(DONE_BILLS_PATH, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
    except Exception as e:
        print(f"Error saving done_bills: {e}")

def add_done_bill(order_id: str, user_name: str = "", remark: str = "") -> bool:
    oid = str(order_id).strip()
    if not oid:
        return False
    data = load_done_bills()
    data[oid] = {
        "order_id": oid,
        "marked_by": user_name or "Unknown",
        "marked_at": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        "remark": remark or ""
    }
    save_done_bills(data)
    return True

def remove_done_bill(order_id: str) -> bool:
    oid = str(order_id).strip()
    data = load_done_bills()
    if oid in data:
        del data[oid]
        save_done_bills(data)
        return True
    return False

def get_done_bill_ids() -> set:
    data = load_done_bills()
    return set(str(k).strip() for k in data.keys())

STATUS_MAPPING_EN_KH = {
    '100': '100 - Create Order (បង្កើតការបញ្ជាទិញ)',
    '110': '110 - Assigned Pickup (ចាត់អ្នកទៅទទួល)',
    '200': '200 - Picking Up (កំពុងទៅទទួល)',
    '201': '201 - Pickup Failed (ទទួលមិនបានសម្រេច)',
    '210': '210 - Received Order (ស្កេនទទួលបញ្ញើ)',
    '302': '302 - Complete Sack (បញ្ចប់ការវេចខ្ចប់បាវ)',
    '306': '306 - Accept Handover (ទទួលការប្រគល់)',
    '309': '309 - Receive Bill (ស្កេនទទួលកញ្ចប់)',
    '310': '310 - Transiting (កំពុងឆ្លងកាត់)',
    '311': '311 - Manual Receiving (ទទួលដោយផ្ទាល់)',
    '400': '400 - Ready for Delivery (ត្រៀមដឹកជញ្ជូន)',
    '401': '401 - Out for Delivery Today (ដឹកជូនថ្ងៃនេះ)',
    '402': '402 - Assignment Confirmation (ចាត់អ្នកដឹក)',
    '410': '410 - Delivered Successfully (ដឹកជូនជោគជ័យ)',
    '420': '420 - Customer Postponed (ណាត់ដឹកឡើងវិញ)',
    '430': '430 - Out of Delivery Area (ក្រៅតំបន់សេវាដឹក)',
    '460': '460 - Delivery Failed (ដឹកមិនបានសម្រេច)',
    '470': '470 - Damaged / Lost (កញ្ចប់ខូចខាត/បាត់បង់)',
    '471': '471 - Customs Inspection (ត្រួតពិនិត្យគយ)',
    '472': '472 - Customer Refused (ភ្ញៀវបដិសេធទទួល)',
    '480': '480 - Wrong Address / Forwarding (ប្តូរអាសយដ្ឋានដឹក)',
    '500': '500 - Return Confirmation (បញ្ជាក់ការត្រឡប់)',
    '510': '510 - Returning to Sender (កំពុងត្រឡប់ទៅអ្នកផ្ញើ)',
    '511': '511 - Return in Transit (កំពុងបញ្ជូនត្រឡប់)',
    '512': '512 - Return at Post Office (ទំនិញត្រឡប់ដល់សាខា)',
    '520': '520 - Returned Successfully (ត្រឡប់ជោគជ័យ)',
    '540': '540 - Return Cancelled (បោះបង់ការត្រឡប់)',
}

def get_status_en_kh(sc: str, raw_status: str = '') -> str:
    sc_clean = str(sc).strip()
    if sc_clean in STATUS_MAPPING_EN_KH:
        return STATUS_MAPPING_EN_KH[sc_clean]
    if raw_status and str(raw_status).strip() and str(raw_status).strip() != 'nan':
        return f"{sc_clean} - {raw_status}"
    return sc_clean

def build_delayed_ge3_report(src_xlsx: str, out_xlsx: str, min_days: int = 3) -> tuple[int, int, int, float, float]:
    """
    Builds structured day-by-day NOT_ASSIGN_AND_DELIVERY_GE_3DAYS.xlsx workbook:
    - ALL OVERDUE (>= 3 DAYS)
    - TODAY (3 DAYS) -> Turned 3 days today
    - YESTERDAY (4 DAYS) -> Turned 3 days yesterday
    - OLDER (>= 5 DAYS) -> Backlog >= 5 days
    - TOMORROW RISK (2 DAYS) -> Warning list turning 3 days tomorrow
    - NOT ASSIGN (>= 3 DAYS)
    - DELIVERY (>= 3 DAYS)
    """
    df = pd.read_excel(src_xlsx)
    df.columns = [str(c).strip().upper() for c in df.columns]

    col_order = next((c for c in df.columns if 'ORDER ID' in c or 'ORDER_NUMBER' in c), 'ORDER ID')
    col_sender = next((c for c in df.columns if 'SENDER' in c or 'CUSTOMER' in c), 'SENDER')
    col_receiver = next((c for c in df.columns if 'RECEIVER' in c), 'RECEIVER')
    col_status = next((c for c in df.columns if 'CURRENT STATUS' in c or 'STATUS' in c or 'STATUES' in c), 'CURRENT STATUS')
    col_current_time = next((c for c in df.columns if 'CURRENT TIME' in c or 'ACTION TIME' in c or 'ACTION_TIME' in c), 'CURRENT TIME')
    col_created_date = next((c for c in df.columns if 'CREATED DATE' in c or 'CREATED_AT' in c or 'CREATED' in c), 'CREATED DATE')
    col_current_po = next((c for c in df.columns if 'CURRENT POST' in c or 'ACTION POST' in c), 'CURRENT POST OFFICE')
    col_dest_po = next((c for c in df.columns if 'DELIVERY POST' in c or 'DESTINATION_POST' in c), 'DELIVERY POST OFFICE')
    col_weight = next((c for c in df.columns if 'WEIGHT' in c), 'WEIGHT (G)')
    col_cod = next((c for c in df.columns if 'COD' in c), 'COD (USD)')
    col_action_user = next((c for c in df.columns if 'ACTION USER' in c or 'CREATED_BY' in c), 'ACTION USER')

    df['sc'] = df[col_status].astype(str).str.extract(r'^(\d{3})')[0]

    # Target statuses: Not Assign (306, 402) & Delivery (400, 401, 420, 430, 470, 471, 472, 480)
    target_statuses = ['306', '402', '400', '401', '420', '430', '470', '471', '472', '480']
    df_filtered = df[df['sc'].isin(target_statuses)].copy()

    # ── EXCLUDE DONE BILLS & LIVE REAL-TIME DELIVERED ────────────────────────
    done_ids = get_done_bill_ids()
    if done_ids and col_order in df_filtered.columns:
        df_filtered = df_filtered[~df_filtered[col_order].astype(str).str.strip().isin(done_ids)].copy()

    # Real-time tracking check against live transactional DB
    try:
        from downloader import batch_verify_delivered
        cfg_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config.json")
        if os.path.exists(cfg_path):
            with open(cfg_path, "r", encoding="utf-8") as f_cfg:
                cfg_obj = json.load(f_cfg)
            cand_ids = [str(x).strip() for x in df_filtered[col_order].dropna().unique() if str(x).strip() and str(x).strip() != 'nan']
            if cand_ids:
                confirmed_delivered = batch_verify_delivered(cfg_obj.get("api", {}), cand_ids, max_workers=40, timeout=10)
                if confirmed_delivered:
                    df_filtered = df_filtered[~df_filtered[col_order].astype(str).str.strip().isin(confirmed_delivered)].copy()
    except Exception as e_live:
        print(f"[DELAYED LIVE CHECK] Warning: {e_live}")

    # ── GUARANTEE 100% UNIQUE ORDER IDs (NO DUPLICATES) ──────────────────────
    if col_order in df_filtered.columns:
        df_filtered = df_filtered.drop_duplicates(subset=[col_order], keep='first').copy()

    # Date calculation (days pending since last action)
    df_filtered['dt_last_action'] = pd.to_datetime(df_filtered[col_current_time], dayfirst=True, format='mixed', errors='coerce')
    df_filtered['dt_created'] = pd.to_datetime(df_filtered[col_created_date], dayfirst=True, format='mixed', errors='coerce')
    df_filtered['dt_last_action'] = df_filtered['dt_last_action'].fillna(df_filtered['dt_created'])

    now = datetime.now()
    df_filtered['day_pending'] = (now - df_filtered['dt_last_action']).dt.total_seconds() / (24 * 3600)
    df_filtered['day_pending_int'] = df_filtered['day_pending'].fillna(0).astype(int)

    def classify_table(sc):
        if sc in ('306', '402'):
            return 'ASSIGN DELIVER'
        return 'DELIVERY'

    def classify_aging_timeline(dp):
        if dp == 2:
            return 'Tomorrow (Turn 3 Days)'
        elif dp == 3:
            return 'Today (Turned 3 Days)'
        elif dp == 4:
            return 'Yesterday (4 Days Old)'
        elif 5 <= dp <= 7:
            return '5 - 7 Days Old'
        else:
            return '>= 8 Days (Critical)'

    df_filtered['TABLE'] = df_filtered['sc'].apply(classify_table)
    df_filtered['STATUS_EN_KH'] = df_filtered.apply(lambda r: get_status_en_kh(r['sc'], r[col_status]), axis=1)
    df_filtered['AGING_TIMELINE'] = df_filtered['day_pending_int'].apply(classify_aging_timeline)

    # Subsets by day
    df_ge = df_filtered[df_filtered['day_pending_int'] >= min_days].sort_values(by='day_pending_int', ascending=False).copy()
    df_today = df_filtered[df_filtered['day_pending_int'] == 3].sort_values(by='day_pending_int', ascending=False).copy()
    df_ytd = df_filtered[df_filtered['day_pending_int'] == 4].sort_values(by='day_pending_int', ascending=False).copy()
    df_older = df_filtered[df_filtered['day_pending_int'] >= 5].sort_values(by='day_pending_int', ascending=False).copy()
    df_tmr = df_filtered[df_filtered['day_pending_int'] == 2].sort_values(by='day_pending_int', ascending=False).copy()

    df_not_assign = df_ge[df_ge['TABLE'] == 'ASSIGN DELIVER'].copy()
    df_delivery = df_ge[df_ge['TABLE'] == 'DELIVERY'].copy()

    # Build Excel Workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )
    hdr_font = Font(name='Segoe UI', size=10, bold=True, color='FFFFFF')
    hdr_fill = PatternFill('solid', fgColor='0F172A')
    zebra_fill = PatternFill('solid', fgColor='F8FAFC')
    white_fill = PatternFill('solid', fgColor='FFFFFF')
    urg_fill = PatternFill('solid', fgColor='FEE2E2')
    urg_font = Font(name='Segoe UI', size=9, bold=True, color='991B1B')
    data_font = Font(name='Segoe UI', size=9)

    def add_sheet(sheet_title, data_df, banner_text, banner_bg='1E293B'):
        ws = wb.create_sheet(title=sheet_title)
        ws.views.sheetView[0].showGridLines = True
        
        export_df = pd.DataFrame({
            'NO': range(1, len(data_df) + 1),
            'BILL (ORDER ID)': data_df[col_order],
            'SENDER': data_df[col_sender].fillna(''),
            'RECEIVER': data_df[col_receiver].fillna(''),
            'LAST PERSON ACTION': data_df[col_action_user].fillna('N/A'),
            'DAY PENDING': data_df['day_pending_int'],
            'TIMELINE': data_df['AGING_TIMELINE'],
            'TABLE': data_df['TABLE'],
            'CURRENT STATUS (EN & KH)': data_df['STATUS_EN_KH'],
            'LAST ACTION TIME': data_df[col_current_time].astype(str),
            'CURRENT POST OFFICE': data_df[col_current_po].fillna(''),
            'DELIVERY POST OFFICE': data_df[col_dest_po].fillna(''),
            'WEIGHT (KG)': pd.to_numeric(data_df[col_weight], errors='coerce').fillna(0).div(1000.0).round(2),
            'COD (USD)': pd.to_numeric(data_df[col_cod], errors='coerce').fillna(0).round(2)
        })
        
        tot_weight = float(export_df['WEIGHT (KG)'].sum()) if len(export_df) else 0.0
        tot_cod = float(export_df['COD (USD)'].sum()) if len(export_df) else 0.0
        
        ws.merge_cells('A1:N1')
        ws['A1'] = banner_text
        ws['A1'].font = Font(name='Segoe UI', size=12, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill('solid', fgColor=banner_bg)
        ws['A1'].alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws.row_dimensions[1].height = 30
        
        ws.merge_cells('A2:N2')
        ws['A2'] = f"Total Orders: {len(data_df):,} bills | Total Weight: {tot_weight:,.2f} kg | Total COD: ${tot_cod:,.2f}"
        ws['A2'].font = Font(name='Segoe UI', size=10, bold=True, color='1E293B')
        ws['A2'].fill = PatternFill('solid', fgColor='F1F5F9')
        ws['A2'].alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws.row_dimensions[2].height = 22
        
        headers = list(export_df.columns)
        ws.row_dimensions[3].height = 24
        for c_idx, h in enumerate(headers, 1):
            cell = ws.cell(3, c_idx, h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            
        for r_idx, row in enumerate(export_df.itertuples(index=False), 4):
            ws.row_dimensions[r_idx].height = 19
            is_zebra = (r_idx % 2 == 0)
            bg = zebra_fill if is_zebra else white_fill
            
            for c_idx, val in enumerate(row, 1):
                cell = ws.cell(r_idx, c_idx, val)
                cell.font = data_font
                cell.fill = bg
                cell.border = thin_border
                
                h_name = headers[c_idx - 1]
                if h_name == 'NO':
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif h_name in ('BILL (ORDER ID)', 'TIMELINE', 'TABLE', 'CURRENT POST OFFICE', 'DELIVERY POST OFFICE', 'LAST ACTION TIME'):
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif h_name == 'DAY PENDING':
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.font = urg_font
                    cell.fill = urg_fill
                elif h_name in ('WEIGHT (KG)', 'COD (USD)'):
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                    cell.number_format = '#,##0.00'
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')

        ws.column_dimensions['A'].width = 8   # NO
        ws.column_dimensions['B'].width = 16  # BILL ID
        ws.column_dimensions['C'].width = 28  # SENDER
        ws.column_dimensions['D'].width = 28  # RECEIVER
        ws.column_dimensions['E'].width = 30  # LAST PERSON ACTION
        ws.column_dimensions['F'].width = 14  # DAY PENDING
        ws.column_dimensions['G'].width = 24  # TIMELINE
        ws.column_dimensions['H'].width = 16  # TABLE
        ws.column_dimensions['I'].width = 44  # CURRENT STATUS (EN & KH)
        ws.column_dimensions['J'].width = 20  # LAST ACTION TIME
        ws.column_dimensions['K'].width = 20  # CURRENT PO
        ws.column_dimensions['L'].width = 20  # DELIVERY PO
        ws.column_dimensions['M'].width = 14  # WEIGHT (KG)
        ws.column_dimensions['N'].width = 14  # COD (USD)

    tot_weight_kg = pd.to_numeric(df_ge[col_weight], errors='coerce').fillna(0).sum() / 1000.0
    tot_cod_usd = pd.to_numeric(df_ge[col_cod], errors='coerce').fillna(0).sum()

    # Sheet 1: ALL OVERDUE (>= 3 DAYS)
    add_sheet(f'ALL GE 3DAYS ({len(df_ge)})', df_ge, f"ALL OVERDUE ORDERS (>= 3 DAYS / 72+ HOURS) — {now.strftime('%d/%m/%Y %H:%M')}", banner_bg='0F172A')

    # Sheet 2: TODAY (TURNED 3 DAYS TODAY)
    add_sheet(f'TODAY 3DAYS ({len(df_today)})', df_today, f"TURNED 3 DAYS TODAY (FRESH OVERDUE) — {now.strftime('%d/%m/%Y %H:%M')}", banner_bg='D97706')

    # Sheet 3: YESTERDAY (TURNED 3 DAYS YESTERDAY / DAY 4)
    add_sheet(f'YTD 4DAYS ({len(df_ytd)})', df_ytd, f"TURNED 3 DAYS YESTERDAY (4 DAYS PENDING) — {now.strftime('%d/%m/%Y %H:%M')}", banner_bg='EA580C')

    # Sheet 4: OLDER (>= 5 DAYS)
    add_sheet(f'OLDER >=5DAYS ({len(df_older)})', df_older, f"CRITICAL BACKLOG (>= 5 DAYS OVERDUE) — {now.strftime('%d/%m/%Y %H:%M')}", banner_bg='991B1B')

    # Sheet 5: TOMORROW WARNING (2 DAYS PENDING)
    add_sheet(f'TMR RISK 2DAYS ({len(df_tmr)})', df_tmr, f"HIGH-RISK WARNING (TURNING 3 DAYS TOMORROW) — {now.strftime('%d/%m/%Y %H:%M')}", banner_bg='2563EB')

    # Sheet 6: ASSIGN DELIVER (>= 3 DAYS)
    add_sheet(f'ASSIGN DELIVER ({len(df_not_assign)})', df_not_assign, f"TABLE 1: ASSIGN DELIVER ORDERS (>= 3 DAYS) — {now.strftime('%d/%m/%Y %H:%M')}", banner_bg='C2410C')

    # Sheet 7: DELIVERY (>= 3 DAYS)
    add_sheet(f'DELIVERY ({len(df_delivery)})', df_delivery, f"TABLE 2: DELIVERY PENDING ORDERS (>= 3 DAYS) — {now.strftime('%d/%m/%Y %H:%M')}", banner_bg='1D4ED8')

    wb.save(out_xlsx)
    return len(df_ge), len(df_not_assign), len(df_delivery), float(tot_weight_kg), float(tot_cod_usd)
